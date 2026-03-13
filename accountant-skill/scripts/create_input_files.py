"""
Create input files from reference General Ledger and General Journal.
- Cash Receipts Journal: Where Cash at Bank (10100) is DEBITED
- Cash Payments Journal: Where Cash at Bank (10100) is CREDITED
- General Journal: Adjustments from reference General Journal
"""
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# Styling
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
NORMAL_FONT = Font(size=11, name='Arial')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def read_general_ledger(filepath):
    """Read reference General Ledger and extract transactions."""
    print(f"Reading: {filepath}")
    wb = load_workbook(filepath)
    ws = wb['Gen_Ledger']

    transactions = []
    for row in range(5, ws.max_row + 1):
        date = ws.cell(row=row, column=1).value
        account_code = ws.cell(row=row, column=7).value
        account_name = ws.cell(row=row, column=8).value
        description = ws.cell(row=row, column=9).value
        debit = ws.cell(row=row, column=10).value
        credit = ws.cell(row=row, column=11).value

        if date and account_code:
            try:
                debit_val = float(debit) if debit and not str(debit).startswith('=') else 0
            except:
                debit_val = 0
            try:
                credit_val = float(credit) if credit and not str(credit).startswith('=') else 0
            except:
                credit_val = 0

            transactions.append({
                'date': date,
                'account_code': int(account_code) if account_code else None,
                'account_name': account_name,
                'description': description,
                'debit': debit_val,
                'credit': credit_val
            })

    print(f"  Found {len(transactions)} transactions")
    return transactions

def read_general_journal(filepath):
    """Read reference General Journal (adjustments)."""
    print(f"Reading: {filepath}")
    wb = load_workbook(filepath)
    ws = wb['Gen_Journal']

    adjustments = []
    for row in range(8, ws.max_row + 1):  # Data starts at row 8
        date = ws.cell(row=row, column=1).value
        jv_id = ws.cell(row=row, column=2).value
        account_code = ws.cell(row=row, column=4).value
        account_name = ws.cell(row=row, column=5).value
        description = ws.cell(row=row, column=6).value
        debit = ws.cell(row=row, column=7).value
        credit = ws.cell(row=row, column=8).value

        if date and account_code:
            try:
                debit_val = float(debit) if debit and not str(debit).startswith('=') else None
            except:
                debit_val = None
            try:
                credit_val = float(credit) if credit and not str(credit).startswith('=') else None
            except:
                credit_val = None

            # Skip if both are None (empty row)
            if debit_val is None and credit_val is None:
                continue

            adjustments.append({
                'date': date,
                'jv_id': jv_id,
                'account_code': int(account_code) if account_code else None,
                'account_name': account_name,
                'description': description,
                'debit': debit_val if debit_val else 0,
                'credit': credit_val if credit_val else 0
            })

    print(f"  Found {len(adjustments)} adjustment entries")
    return adjustments

def create_cash_journals(transactions, output_dir):
    """Create Cash Receipts and Cash Payments journals from GL transactions.

    The reference GL has transactions in sequential order where:
    - Most cash entries (10100) have a single counterparty as immediate neighbor
    - Some cash entries have multiple counterparties (grouped between cash entries)
    """

    cash_receipts = []
    cash_payments = []

    # Find all cash entry indices (both debit and credit) for boundary calculation
    all_cash_indices = [i for i, t in enumerate(transactions) if t['account_code'] == 10100]
    # Find cash debit indices for iterating receipts
    cash_debit_indices = [i for i, t in enumerate(transactions) if t['account_code'] == 10100 and t['debit'] > 0]
    # Find cash credit indices for iterating payments
    cash_credit_indices = [i for i, t in enumerate(transactions) if t['account_code'] == 10100 and t['credit'] > 0]

    # Process Cash Receipts
    for idx, i in enumerate(cash_debit_indices):
        t = transactions[i]
        cash_amount = t['debit']

        # First try: immediate next entry
        if i + 1 < len(transactions):
            next_t = transactions[i + 1]
            if next_t['credit'] == cash_amount and next_t['account_code'] != 10100:
                cash_receipts.append({
                    'date': t['date'],
                    'description': t['description'],
                    'amount': cash_amount,
                    'contra_account': next_t['account_code'],
                    'contra_name': next_t['account_name']
                })
                continue

        # Second try: immediate previous entry
        if i - 1 >= 0:
            prev_t = transactions[i - 1]
            if prev_t['credit'] == cash_amount and prev_t['account_code'] != 10100:
                cash_receipts.append({
                    'date': t['date'],
                    'description': t['description'],
                    'amount': cash_amount,
                    'contra_account': prev_t['account_code'],
                    'contra_name': prev_t['account_name']
                })
                continue

        # Third try: multiple counterparties (look both before and after)
        # Use ALL cash entries as boundaries (not just debits)
        # Find the position of this entry in all_cash_indices
        pos_in_all = all_cash_indices.index(i)
        prev_cash_idx = all_cash_indices[pos_in_all - 1] if pos_in_all > 0 else 0
        next_cash_idx = all_cash_indices[pos_in_all + 1] if pos_in_all + 1 < len(all_cash_indices) else len(transactions)

        # Look AFTER this cash entry
        counterparties_after = []
        for j in range(i + 1, next_cash_idx):
            if transactions[j]['credit'] > 0:
                counterparties_after.append(transactions[j])

        total_after = sum(cp['credit'] for cp in counterparties_after)
        if abs(total_after - cash_amount) < 1:
            for cp in counterparties_after:
                cash_receipts.append({
                    'date': t['date'],
                    'description': t['description'],
                    'amount': cp['credit'],
                    'contra_account': cp['account_code'],
                    'contra_name': cp['account_name']
                })
            continue

        # Look BEFORE this cash entry
        counterparties_before = []
        for j in range(prev_cash_idx + 1, i):
            if transactions[j]['credit'] > 0:
                counterparties_before.append(transactions[j])

        total_before = sum(cp['credit'] for cp in counterparties_before)
        if abs(total_before - cash_amount) < 1:
            for cp in counterparties_before:
                cash_receipts.append({
                    'date': t['date'],
                    'description': t['description'],
                    'amount': cp['credit'],
                    'contra_account': cp['account_code'],
                    'contra_name': cp['account_name']
                })
            continue

        # Fourth try: description-based matching for monthly summaries
        # Some entries have descriptions like "Sale revenue from 140 ml for August"
        # and counterparties with descriptions like "Sale revenue from 140 ml"
        if t['description']:
            desc_lower = t['description'].lower()
            # Check for monthly summary patterns
            if 'sale revenue' in desc_lower or 'sales revenue' in desc_lower:
                # Extract product type (140 ml, 175 ml, etc.)
                product_match = None
                for product in ['140 ml', '140ml', '175 ml', '175ml', '500 ml', '500ml', '1l', '1 l']:
                    if product in desc_lower:
                        product_match = product.replace(' ', '')  # Normalize
                        break

                if product_match:
                    # Find all revenue entries (40000) with matching product in description
                    matching_credits = []
                    for j in range(prev_cash_idx + 1, i):
                        if transactions[j]['credit'] > 0 and transactions[j]['account_code'] == 40000:
                            cp_desc = transactions[j]['description'] or ''
                            cp_desc_lower = cp_desc.lower().replace(' ', '')
                            if product_match in cp_desc_lower:
                                matching_credits.append(transactions[j])

                    total_matching = sum(cp['credit'] for cp in matching_credits)
                    if abs(total_matching - cash_amount) < 1 and matching_credits:
                        for cp in matching_credits:
                            cash_receipts.append({
                                'date': t['date'],
                                'description': t['description'],
                                'amount': cp['credit'],
                                'contra_account': cp['account_code'],
                                'contra_name': cp['account_name']
                            })
                        continue

    # Process Cash Payments
    for idx, i in enumerate(cash_credit_indices):
        t = transactions[i]
        cash_amount = t['credit']

        # First try: immediate previous entry
        if i - 1 >= 0:
            prev_t = transactions[i - 1]
            if prev_t['debit'] == cash_amount and prev_t['account_code'] != 10100:
                cash_payments.append({
                    'date': t['date'],
                    'description': t['description'],
                    'amount': cash_amount,
                    'contra_account': prev_t['account_code'],
                    'contra_name': prev_t['account_name']
                })
                continue

        # Second try: immediate next entry
        if i + 1 < len(transactions):
            next_t = transactions[i + 1]
            if next_t['debit'] == cash_amount and next_t['account_code'] != 10100:
                cash_payments.append({
                    'date': t['date'],
                    'description': t['description'],
                    'amount': cash_amount,
                    'contra_account': next_t['account_code'],
                    'contra_name': next_t['account_name']
                })
                continue

        # Third try: multiple counterparties (look both before and after)
        # Use ALL cash entries as boundaries
        pos_in_all = all_cash_indices.index(i)
        prev_cash_idx = all_cash_indices[pos_in_all - 1] if pos_in_all > 0 else 0
        next_cash_idx = all_cash_indices[pos_in_all + 1] if pos_in_all + 1 < len(all_cash_indices) else len(transactions)

        # Look BEFORE this cash entry
        counterparties_before = []
        for j in range(prev_cash_idx + 1, i):
            if transactions[j]['debit'] > 0:
                counterparties_before.append(transactions[j])

        total_before = sum(cp['debit'] for cp in counterparties_before)
        if abs(total_before - cash_amount) < 1:
            for cp in counterparties_before:
                cash_payments.append({
                    'date': t['date'],
                    'description': t['description'],
                    'amount': cp['debit'],
                    'contra_account': cp['account_code'],
                    'contra_name': cp['account_name']
                })
            continue

        # Look AFTER this cash entry
        counterparties_after = []
        for j in range(i + 1, next_cash_idx):
            if transactions[j]['debit'] > 0:
                counterparties_after.append(transactions[j])

        total_after = sum(cp['debit'] for cp in counterparties_after)
        if abs(total_after - cash_amount) < 1:
            for cp in counterparties_after:
                cash_payments.append({
                    'date': t['date'],
                    'description': t['description'],
                    'amount': cp['debit'],
                    'contra_account': cp['account_code'],
                    'contra_name': cp['account_name']
                })

    # Sort by date
    cash_receipts.sort(key=lambda x: x['date'])
    cash_payments.sort(key=lambda x: x['date'])

    print(f"\nCash Receipts: {len(cash_receipts)}")
    print(f"Cash Payments: {len(cash_payments)}")

    # Create Cash Receipts Journal
    create_cash_receipts_journal(cash_receipts, output_dir / 'cash_receipts_journal.xlsx')

    # Create Cash Payments Journal
    create_cash_payments_journal(cash_payments, output_dir / 'cash_payments_journal.xlsx')

    return cash_receipts, cash_payments

def create_cash_receipts_journal(receipts, filepath):
    """Create Cash Receipts Journal Excel file."""
    print(f"\nCreating: {filepath}")

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Headers
    headers = ['Date', 'Receipt No', 'Received From', 'Description', 'Amount', 'Bank Account', 'Debit Account', 'Credit Account']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Data
    for idx, r in enumerate(receipts):
        row = idx + 2
        date_str = r['date'].strftime('%Y-%m-%d') if hasattr(r['date'], 'strftime') else str(r['date'])
        month = r['date'].strftime('%m') if hasattr(r['date'], 'strftime') else '01'
        receipt_no = f"CR-{month}{idx+1:03d}"

        ws.cell(row=row, column=1, value=date_str).border = THIN_BORDER
        ws.cell(row=row, column=2, value=receipt_no).border = THIN_BORDER
        ws.cell(row=row, column=3, value=r['contra_name'] if r['contra_name'] else '').border = THIN_BORDER
        ws.cell(row=row, column=4, value=r['description'] if r['description'] else '').border = THIN_BORDER
        cell = ws.cell(row=row, column=5, value=r['amount'])
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'
        ws.cell(row=row, column=6, value='Main').border = THIN_BORDER
        ws.cell(row=row, column=7, value=10100).border = THIN_BORDER  # Cash at Bank (Debit)
        ws.cell(row=row, column=8, value=r['contra_account']).border = THIN_BORDER

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    wb.save(filepath)
    print(f"  Saved {len(receipts)} cash receipts")

def create_cash_payments_journal(payments, filepath):
    """Create Cash Payments Journal Excel file."""
    print(f"\nCreating: {filepath}")

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Headers
    headers = ['Date', 'Payment No', 'Paid To', 'Description', 'Amount', 'Bank Account', 'Debit Account', 'Credit Account']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Data
    for idx, p in enumerate(payments):
        row = idx + 2
        date_str = p['date'].strftime('%Y-%m-%d') if hasattr(p['date'], 'strftime') else str(p['date'])
        month = p['date'].strftime('%m') if hasattr(p['date'], 'strftime') else '01'
        payment_no = f"CP-{month}{idx+1:03d}"

        ws.cell(row=row, column=1, value=date_str).border = THIN_BORDER
        ws.cell(row=row, column=2, value=payment_no).border = THIN_BORDER
        ws.cell(row=row, column=3, value=p['contra_name'] if p['contra_name'] else '').border = THIN_BORDER
        ws.cell(row=row, column=4, value=p['description'] if p['description'] else '').border = THIN_BORDER
        cell = ws.cell(row=row, column=5, value=p['amount'])
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'
        ws.cell(row=row, column=6, value='Main').border = THIN_BORDER
        ws.cell(row=row, column=7, value=p['contra_account']).border = THIN_BORDER
        ws.cell(row=row, column=8, value=10100).border = THIN_BORDER  # Cash at Bank (Credit)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    wb.save(filepath)
    print(f"  Saved {len(payments)} cash payments")

def create_general_journal(adjustments, filepath):
    """Create General Journal from adjustments."""
    print(f"\nCreating: {filepath}")

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Headers
    headers = ['Date', 'JV No', 'Description', 'Debit Account', 'Credit Account', 'Debit Amount', 'Credit Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Group adjustments by JV ID to create proper entries
    jv_groups = defaultdict(list)
    for a in adjustments:
        jv_groups[a['jv_id']].append(a)

    # Data
    row = 2
    for jv_id in sorted(jv_groups.keys(), key=lambda x: str(x)):
        group = jv_groups[jv_id]
        for a in group:
            date_str = a['date'].strftime('%Y-%m-%d') if hasattr(a['date'], 'strftime') else str(a['date'])

            ws.cell(row=row, column=1, value=date_str).border = THIN_BORDER
            ws.cell(row=row, column=2, value=a['jv_id']).border = THIN_BORDER
            ws.cell(row=row, column=3, value=a['description'] if a['description'] else '').border = THIN_BORDER
            ws.cell(row=row, column=4, value=a['account_code']).border = THIN_BORDER
            ws.cell(row=row, column=5, value='').border = THIN_BORDER

            if a['debit'] > 0:
                cell = ws.cell(row=row, column=6, value=a['debit'])
                cell.border = THIN_BORDER
                cell.number_format = '#,##0'
                ws.cell(row=row, column=7, value='').border = THIN_BORDER
            else:
                ws.cell(row=row, column=6, value='').border = THIN_BORDER
                cell = ws.cell(row=row, column=7, value=a['credit'])
                cell.border = THIN_BORDER
                cell.number_format = '#,##0'

            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    wb.save(filepath)
    print(f"  Saved {len(adjustments)} adjustment entries")

def main():
    # Paths
    ref_dir = Path('Exisitng Accounting Workflow _ reference files')
    gl_path = ref_dir / 'Ledger Accounts' / 'General_Ledger_edited.xlsx'
    gj_path = ref_dir / 'Books of Prime Entry' / 'General_Journal.xlsx'
    output_dir = Path('data/input/journals')

    # Read reference files
    transactions = read_general_ledger(gl_path)
    adjustments = read_general_journal(gj_path)

    # Create cash journals from GL transactions
    create_cash_journals(transactions, output_dir)

    # Create general journal from adjustments
    create_general_journal(adjustments, output_dir / 'general_journal.xlsx')

    print("\n=== SUMMARY ===")
    print(f"Input files created in: {output_dir}")
    print("  - cash_receipts_journal.xlsx")
    print("  - cash_payments_journal.xlsx")
    print("  - general_journal.xlsx")

if __name__ == '__main__':
    main()