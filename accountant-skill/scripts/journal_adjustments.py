"""
Module 4: Journal Adjustments
Shwe Mandalay Cafe / K&K Finance Team

Generates period-end adjusting entries and produces a formal adjusting
entries workbook before the trial balance is prepared.

Auto-generated entries:
  1. Depreciation      — calculated from fixed_assets_ledger.xlsx (Straight-Line)
  2. Bank Recon        — imported from the Adjusting Entries sheet of Module 3 output

Informational reference sheets (already posted via journals, shown for completeness):
  3. Accruals          — reads GL for movements in accrued expense accounts
  4. Prepayments       — reads GL for movements in prepaid asset accounts

Usage:
    python journal_adjustments.py <data_dir> <period_start> <period_end> <output_file>

    python journal_adjustments.py \\
        data/Jan2026 \\
        2026-01-01 \\
        2026-01-31 \\
        data/Jan2026/adjusting_entries_Jan2026.xlsx

Output sheets:
    Dashboard             — summary by type, totals, double-entry validation
    Depreciation Schedule — per-asset SL calculation and grouped journal entries
    Bank Recon Entries    — entries imported from Module 3 output
    Accruals              — already-posted accruals from journals (reference)
    Prepayments           — already-posted prepaid recognitions (reference)
    All Entries           — master journal: all new ADJ- entries
    Account Impact        — pre/post balances for every affected account
    Exceptions            — (only if errors found)
"""

import sys
import os
from pathlib import Path
from datetime import date

import pandas as pd
import numpy as np

sys.path.insert(0, str(Path(__file__).parent))
from utils.excel_reader import read_xlsx, read_all_sheets, filter_by_period
from utils.excel_writer import (
    create_workbook, add_sheet, write_title, write_header_row,
    write_data_row, write_section_header, write_total_row,
    write_validation_result, auto_fit_columns, freeze_panes,
    save_workbook, NORMAL_FONT, TOTAL_FONT, NEGATIVE_FONT,
    THIN_BORDER, PASS_FILL, FAIL_FILL, WARNING_FILL,
    NUMBER_FORMAT_NEG, DATE_FORMAT, HEADER_FILL, HEADER_FONT
)
from utils.coa_mapper import COAMapper
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

DEPR_EXPENSE_CODE = '5300'
DEPR_EXPENSE_NAME = 'Depreciation Expense'

# Maps asset account code → (accum depr code, accum depr name)
ACCUM_DEPR_MAP = {
    1610: ('1611', 'Accum. Depr. — Buildings'),
    1620: ('1621', 'Accum. Depr. — Plant & Machinery'),
    1630: ('1631', 'Accum. Depr. — Furniture & Fixtures'),
    1640: ('1641', 'Accum. Depr. — Vehicles'),
    1650: ('1651', 'Accum. Depr. — Office Equipment'),
    1700: ('1701', 'Accum. Depr. — Intangibles'),
}

# Prepaid and accrual account ranges (for informational sheets)
PREPAID_ACCOUNT_RANGE = (1300, 1499)
ACCRUAL_ACCOUNT_CODES = {2020: 'Accrued Expenses', 2030: 'Accrued Wages'}


# ─────────────────────────────────────────────────────────────────────────────
# 1. DEPRECIATION
# ─────────────────────────────────────────────────────────────────────────────

def compute_depreciation(data_dir):
    """
    Read fixed_assets_ledger.xlsx and compute monthly straight-line depreciation.

    Returns:
        asset_rows: list of dicts — per-asset depreciation detail
        journal_entries: list of dicts — grouped by accum depr account
        error: str or None
    """
    path = Path(data_dir) / 'fixed_assets_ledger.xlsx'
    result = read_xlsx(
        path,
        required_columns=['Asset ID', 'Description', 'Cost'],
        optional_columns=['Account Code', 'Category', 'Date Acquired',
                          'Useful Life (Years)', 'Salvage Value',
                          'Depreciation Method', 'Accumulated Depreciation',
                          'Net Book Value', 'Status']
    )
    if result['error']:
        return None, None, result['error']

    df = result['data']
    asset_rows = []
    # Accumulate depreciation by (asset_account_code, accum_depr_code)
    grouped = {}   # key: (asset_code_int, accum_code_str) → total_monthly_depr

    for _, row in df.iterrows():
        status = str(row.get('Status', 'Active')).strip().lower()
        if status not in ('active', '', 'nan'):
            continue  # skip disposed/inactive assets

        asset_id = row.get('Asset ID', '')
        description = row.get('Description', '')
        method = str(row.get('Depreciation Method', 'Straight-Line')).strip()
        category = str(row.get('Category', '')).strip()

        try:
            cost = float(row.get('Cost', 0) or 0)
            salvage = float(row.get('Salvage Value', 0) or 0)
            useful_life = float(row.get('Useful Life (Years)', 0) or 0)
            account_code = int(float(row.get('Account Code', 0) or 0))
            accum_depr_existing = float(row.get('Accumulated Depreciation', 0) or 0)
        except (ValueError, TypeError):
            continue

        if useful_life <= 0 or cost <= salvage:
            continue

        # Only support Straight-Line for now
        if 'straight' in method.lower() or method == 'SL':
            annual_depr = (cost - salvage) / useful_life
        else:
            # Reducing balance: rate = 1 - (salvage/cost)^(1/life)
            if cost > 0 and salvage > 0:
                rate = 1 - (salvage / cost) ** (1 / useful_life)
            else:
                rate = 1 / useful_life
            nbv = cost - accum_depr_existing
            annual_depr = nbv * rate
            method = 'Reducing Balance'

        monthly_depr = round(annual_depr / 12, 2)

        # Look up accum depr account
        accum_code, accum_name = ACCUM_DEPR_MAP.get(
            account_code, ('1699', 'Accum. Depr. — Other')
        )

        asset_rows.append({
            'asset_id': asset_id,
            'description': description,
            'category': category,
            'account_code': account_code,
            'method': method,
            'cost': cost,
            'salvage': salvage,
            'useful_life': useful_life,
            'annual_depr': round(annual_depr, 2),
            'monthly_depr': monthly_depr,
            'accum_code': accum_code,
            'accum_name': accum_name,
        })

        key = (account_code, accum_code, accum_name)
        grouped[key] = grouped.get(key, 0) + monthly_depr

    if not asset_rows:
        return [], [], None

    # Build one journal entry per asset category (grouped by accum depr account)
    journal_entries = []
    for (asset_code, accum_code, accum_name), total_depr in sorted(grouped.items()):
        # Find category name
        cat_assets = [a for a in asset_rows if a['account_code'] == asset_code]
        category_label = cat_assets[0]['category'] if cat_assets else str(asset_code)
        journal_entries.append({
            'type': 'Depreciation',
            'category': f'Depreciation — {category_label}',
            'dr_code': DEPR_EXPENSE_CODE,
            'dr_name': DEPR_EXPENSE_NAME,
            'cr_code': accum_code,
            'cr_name': accum_name,
            'amount': round(total_depr, 2),
            'supporting': f'Monthly SL depreciation — {category_label}',
        })

    return asset_rows, journal_entries, None


# ─────────────────────────────────────────────────────────────────────────────
# 2. BANK RECONCILIATION ENTRIES
# ─────────────────────────────────────────────────────────────────────────────

def load_bank_recon_entries(data_dir, period_end):
    """
    Read the 'Adjusting Entries' sheet from bank_reconciliation_*.xlsx.
    Falls back to searching for any bank_reconciliation*.xlsx in data_dir.

    Returns: list of entry dicts, error string or None
    """
    data_dir = Path(data_dir)

    # Find the bank reconciliation file
    candidates = list(data_dir.glob('bank_reconciliation*.xlsx'))
    if not candidates:
        return [], f"No bank_reconciliation*.xlsx found in {data_dir} — bank recon entries skipped."

    recon_file = candidates[0]

    try:
        # The sheet has a title block above the actual headers — find the header row first
        df_raw = pd.read_excel(recon_file, sheet_name='Adjusting Entries', header=None)
        header_row_idx = None
        for i, row_vals in df_raw.iterrows():
            if 'Date' in row_vals.values:
                header_row_idx = i
                break
        if header_row_idx is None:
            return [], f"Could not find header row in 'Adjusting Entries' sheet of {recon_file.name}"
        df = pd.read_excel(recon_file, sheet_name='Adjusting Entries', header=header_row_idx)
    except Exception as e:
        return [], f"Could not read 'Adjusting Entries' sheet from {recon_file.name}: {e}"

    df.columns = [str(c).strip() for c in df.columns]

    # Required columns
    needed = ['Date', 'Reference', 'Description', 'Dr Code', 'Dr Amount', 'Cr Code', 'Cr Amount']
    for col in needed:
        if col not in df.columns:
            return [], f"'Adjusting Entries' sheet in {recon_file.name} missing column '{col}'. Found: {list(df.columns)}"

    def _norm_code(val):
        """Convert float-string codes like '1020.0' to clean '1020'."""
        try:
            return str(int(float(str(val).strip())))
        except (ValueError, TypeError):
            return str(val).strip()

    entries = []
    for _, row in df.iterrows():
        try:
            dr_amt = float(row['Dr Amount'])
            cr_amt = float(row['Cr Amount'])
        except (ValueError, TypeError):
            continue   # skip total/validation rows

        if pd.isna(dr_amt) or pd.isna(cr_amt) or dr_amt <= 0:
            continue

        entries.append({
            'type': 'Bank Recon',
            'category': str(row.get('Category', 'Bank Reconciliation Adjustment')).strip(),
            'dr_code': _norm_code(row['Dr Code']),
            'dr_name': str(row.get('Dr Account', '')).strip(),
            'cr_code': _norm_code(row['Cr Code']),
            'cr_name': str(row.get('Cr Account', '')).strip(),
            'amount': round(dr_amt, 2),
            'supporting': f"Bank recon adj — {row.get('Description', '')}",
        })

    return entries, None


# ─────────────────────────────────────────────────────────────────────────────
# 3. GL-BASED INFORMATIONAL SECTIONS (accruals + prepayments already posted)
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_cols(df):
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def _find_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def load_gl_reference(data_dir, period_start, period_end, coa=None):
    """
    Read general_ledger.xlsx and return period transactions for:
      - Accrual accounts (2020, 2030)
      - Prepaid accounts (1300–1499)
    Used for informational/reference sheets only.

    Returns:
        accrual_rows: list of dicts
        prepaid_rows: list of dicts
        gl_balances:  dict[account_code → closing_balance]
    """
    path = Path(data_dir) / 'general_ledger.xlsx'
    if not path.exists():
        return [], [], {}

    result = read_all_sheets(path)
    if result['error'] or result['data'] is None:
        return [], [], {}

    sheets = result['data']
    if len(sheets) != 1:
        return [], [], {}

    df = _normalize_cols(list(sheets.values())[0])
    code_col  = _find_col(df, ['account code', 'code', 'acct code'])
    date_col  = _find_col(df, ['date', 'trans date'])
    debit_col = _find_col(df, ['debit', 'dr'])
    cr_col    = _find_col(df, ['credit', 'cr'])
    bal_col   = _find_col(df, ['balance', 'bal'])
    desc_col  = _find_col(df, ['description', 'narration', 'details'])
    ref_col   = _find_col(df, ['reference', 'ref'])

    if code_col is None or date_col is None:
        return [], [], {}

    df[code_col] = pd.to_numeric(df[code_col], errors='coerce')
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

    accrual_rows = []
    prepaid_rows = []
    gl_balances = {}

    for code in df[code_col].dropna().unique():
        code_int = int(code)
        acct_df = df[df[code_col] == code_int].copy()

        # Get normal balance from COA (needed to compute closing correctly)
        normal_bal = 'debit'
        if coa:
            info = coa.get_account(code_int)
            if info:
                normal_bal = info['normal_balance'].lower()

        # Opening balance: last Balance value before period_start
        pre_rows = acct_df[acct_df[date_col] < pd.Timestamp(period_start)]
        opening = 0.0
        if not pre_rows.empty and bal_col:
            last_val = pd.to_numeric(pre_rows.iloc[-1].get(bal_col), errors='coerce')
            if not pd.isna(last_val):
                opening = float(last_val)

        # Period rows
        period_df = acct_df[
            acct_df[date_col].notna() &
            (acct_df[date_col] >= pd.Timestamp(period_start)) &
            (acct_df[date_col] <= pd.Timestamp(period_end))
        ]

        # Compute closing balance: try Balance column last row, else compute from movements
        closing = None
        if bal_col and not period_df.empty:
            last_val = pd.to_numeric(period_df.iloc[-1].get(bal_col), errors='coerce')
            if not pd.isna(last_val):
                closing = float(last_val)

        if closing is None:
            period_dr = pd.to_numeric(period_df[debit_col], errors='coerce').fillna(0).sum() \
                if debit_col else 0.0
            period_cr = pd.to_numeric(period_df[cr_col], errors='coerce').fillna(0).sum() \
                if cr_col else 0.0
            if normal_bal == 'debit':
                closing = opening + period_dr - period_cr
            else:
                closing = opening - period_dr + period_cr

        if closing is not None:
            gl_balances[code_int] = round(closing, 2)

        for _, row in period_df.iterrows():
            debit  = float(pd.to_numeric(row.get(debit_col, 0), errors='coerce') or 0) \
                if debit_col else 0.0
            credit = float(pd.to_numeric(row.get(cr_col, 0), errors='coerce') or 0) \
                if cr_col else 0.0
            if debit == 0 and credit == 0:
                continue
            entry = {
                'account_code': code_int,
                'date': row[date_col],
                'reference': str(row.get(ref_col, '') or '') if ref_col else '',
                'description': str(row.get(desc_col, '') or '') if desc_col else '',
                'debit': debit,
                'credit': credit,
            }
            if code_int in ACCRUAL_ACCOUNT_CODES:
                entry['account_name'] = ACCRUAL_ACCOUNT_CODES[code_int]
                accrual_rows.append(entry)
            if PREPAID_ACCOUNT_RANGE[0] <= code_int <= PREPAID_ACCOUNT_RANGE[1]:
                entry['account_name'] = f'Account {code_int}'
                prepaid_rows.append(entry)

    return accrual_rows, prepaid_rows, gl_balances


# ─────────────────────────────────────────────────────────────────────────────
# 4. ASSEMBLE ALL ENTRIES + ACCOUNT IMPACT
# ─────────────────────────────────────────────────────────────────────────────

def assign_entry_numbers(all_entries, period_end):
    """Assign sequential ADJ- reference numbers and set date to period_end."""
    for i, entry in enumerate(all_entries, start=1):
        entry['ref'] = f'ADJ-{i:03d}'
        entry['date'] = period_end
    return all_entries


def compute_account_impact(all_entries, gl_balances, coa):
    """
    For each account affected by the new adjusting entries, compute:
      - pre-adjustment GL balance
      - total adjustment (net)
      - post-adjustment balance
    """
    # Aggregate adjustments by account code
    adj_by_account = {}   # code_str → net_adjustment (Dr positive, Cr negative from account perspective)

    for entry in all_entries:
        dr_code = entry['dr_code']
        cr_code = entry['cr_code']
        amt = entry['amount']

        # Debit side: increases debit-normal accounts, decreases credit-normal
        adj_by_account[dr_code] = adj_by_account.get(dr_code, 0) + amt

        # Credit side: decreases debit-normal accounts, increases credit-normal
        adj_by_account[cr_code] = adj_by_account.get(cr_code, 0) - amt

    impacts = []
    for code_str, net_adj in adj_by_account.items():
        try:
            code_int = int(float(code_str))
        except (ValueError, TypeError):
            code_int = 0

        info = coa.get_account(code_int) if code_int else None
        acct_name = info['name'] if info else code_str
        acct_type = info['type'] if info else 'Unknown'
        normal_bal = info['normal_balance'] if info else 'debit'

        pre_balance = gl_balances.get(code_int, 0.0)

        # net_adj is positive for debits, negative for credits
        # For a debit-normal account: Dr increases balance
        # For a credit-normal account: Dr decreases balance (so net_adj sign is reversed)
        if normal_bal.lower() == 'debit':
            post_balance = pre_balance + net_adj
            adj_display = net_adj
        else:
            post_balance = pre_balance - net_adj
            adj_display = -net_adj  # show as positive for credit adjustments

        impacts.append({
            'code': code_str,
            'name': acct_name,
            'type': acct_type,
            'normal_balance': normal_bal,
            'pre_balance': round(pre_balance, 2),
            'adjustment': round(adj_display, 2),
            'post_balance': round(post_balance, 2),
        })

    impacts.sort(key=lambda x: x['code'])
    return impacts


# ─────────────────────────────────────────────────────────────────────────────
# 5. EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_date(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ''
    try:
        return pd.Timestamp(val).strftime('%Y-%m-%d')
    except Exception:
        return str(val)


def _n(val):
    """Return numeric or None."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    return float(val)


def write_dashboard(wb, all_entries, depr_entries, bank_entries,
                    period_start, period_end, exceptions):
    ws = add_sheet(wb, 'Dashboard', tab_color='00B050')
    period_str = f"{period_start}  to  {period_end}"
    row = write_title(ws, 'Adjusting Entries — Dashboard',
                      'Shwe Mandalay Cafe', period_str)

    # Summary by type
    row = write_section_header(ws, 'ENTRIES BY TYPE', row, col_span=4)
    row = write_header_row(ws, ['Type', 'Entries', 'Total Dr', 'Total Cr'], row)

    type_groups = {}
    for e in all_entries:
        t = e['type']
        type_groups.setdefault(t, []).append(e)

    grand_dr = grand_cr = 0.0
    for t, entries in type_groups.items():
        total = sum(e['amount'] for e in entries)
        row = write_data_row(ws, [t, len(entries), _n(total), _n(total)], row,
                             number_cols=[3, 4])
        grand_dr += total
        grand_cr += total

    row = write_total_row(ws, 'GRAND TOTAL',
                          [len(all_entries), _n(grand_dr), _n(grand_cr)],
                          row, double_line=True)
    row += 1

    # Validation
    row = write_section_header(ws, 'VALIDATION', row, col_span=4)
    balanced = round(grand_dr - grand_cr, 2) == 0
    row = write_header_row(ws, ['Check', 'Value', 'Result'], row)
    ws.cell(row=row, column=1, value='Total Debits = Total Credits').font = NORMAL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2,
            value=f"{grand_dr:,.2f} = {grand_cr:,.2f}").font = NORMAL_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    write_validation_result(ws, row, 3, balanced)
    ws.cell(row=row, column=3).border = THIN_BORDER
    row += 2

    # Entry list
    row = write_section_header(ws, 'ALL ENTRIES', row, col_span=4)
    row = write_header_row(ws, ['Reference', 'Type', 'Description', 'Amount'], row)
    for e in all_entries:
        row = write_data_row(ws, [e['ref'], e['type'],
                                   e.get('supporting', e['category']),
                                   _n(e['amount'])], row, number_cols=[4])
    row += 1

    # Exceptions
    if exceptions:
        row = write_section_header(ws, 'WARNINGS', row, col_span=4)
        for exc in exceptions:
            row = write_data_row(ws, ['Warning', exc], row)

    auto_fit_columns(ws)
    ws.column_dimensions['C'].width = 45
    freeze_panes(ws, row=2, col=1)


def write_depreciation_schedule(wb, asset_rows, depr_entries, period_end):
    ws = add_sheet(wb, 'Depreciation Schedule', tab_color='4472C4')
    row = write_title(ws, 'Depreciation Schedule',
                      'Monthly Straight-Line Depreciation', f'Period ending {period_end}')

    # Per-asset detail
    row = write_section_header(ws, 'PER-ASSET DETAIL', row, col_span=8)
    headers = ['Asset ID', 'Description', 'Category', 'Method',
               'Cost', 'Salvage Value', 'Useful Life (yrs)',
               'Annual Depr.', 'Monthly Depr.', 'Depr. Account', 'Accum. Depr. Account']
    row = write_header_row(ws, headers, row)

    total_monthly = 0.0
    for a in asset_rows:
        row = write_data_row(ws, [
            a['asset_id'], a['description'], a['category'], a['method'],
            _n(a['cost']), _n(a['salvage']), _n(a['useful_life']),
            _n(a['annual_depr']), _n(a['monthly_depr']),
            DEPR_EXPENSE_CODE, a['accum_code'],
        ], row, number_cols=[5, 6, 7, 8, 9])
        total_monthly += a['monthly_depr']

    row = write_total_row(ws, 'Total Monthly Depreciation',
                          [None, None, None, None, None, None,
                           None, _n(total_monthly)],
                          row, double_line=True)
    row += 2

    # Journal entries grouped by category
    row = write_section_header(ws, 'JOURNAL ENTRIES (grouped by asset category)', row, col_span=8)
    headers2 = ['Reference', 'Date', 'Description',
                'Debit Account', 'Dr Code', 'Debit Amount',
                'Credit Account', 'Cr Code', 'Credit Amount']
    row = write_header_row(ws, headers2, row)

    for e in depr_entries:
        row = write_data_row(ws, [
            e['ref'], _fmt_date(e['date']), e['category'],
            e['dr_name'], e['dr_code'], _n(e['amount']),
            e['cr_name'], e['cr_code'], _n(e['amount']),
        ], row, number_cols=[6, 9])

    total_depr = sum(e['amount'] for e in depr_entries)
    row = write_total_row(ws, 'Total Depreciation',
                          [None, None, None, None, _n(total_depr), None, None, _n(total_depr)],
                          row, double_line=True)

    auto_fit_columns(ws)
    ws.column_dimensions['B'].width = 32
    freeze_panes(ws)


def write_bank_recon_entries_sheet(wb, bank_entries, period_end):
    ws = add_sheet(wb, 'Bank Recon Entries', tab_color='4472C4')
    row = write_title(ws, 'Bank Reconciliation — Adjusting Entries',
                      'Entries identified in Module 3 bank reconciliation',
                      f'Period ending {period_end}')

    if not bank_entries:
        ws.cell(row=row, column=1,
                value='No bank reconciliation adjusting entries found.').font = NORMAL_FONT
        return

    headers = ['Reference', 'Date', 'Category', 'Description',
               'Debit Account', 'Dr Code', 'Debit Amount',
               'Credit Account', 'Cr Code', 'Credit Amount']
    row = write_header_row(ws, headers, row)

    total = 0.0
    for e in bank_entries:
        row = write_data_row(ws, [
            e['ref'], _fmt_date(e['date']), e['category'],
            e.get('supporting', ''),
            e['dr_name'], e['dr_code'], _n(e['amount']),
            e['cr_name'], e['cr_code'], _n(e['amount']),
        ], row, number_cols=[7, 10])
        total += e['amount']

    row = write_total_row(ws, 'Total',
                          [None, None, None, None, None, _n(total), None, None, _n(total)],
                          row, double_line=True)

    auto_fit_columns(ws)
    ws.column_dimensions['D'].width = 40
    freeze_panes(ws)


def write_accruals_sheet(wb, accrual_rows, period_start, period_end):
    ws = add_sheet(wb, 'Accruals', tab_color='4472C4')
    row = write_title(ws, 'Accruals — Reference',
                      'Accrued expenses already posted in journals this period',
                      f'{period_start}  to  {period_end}')

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    note = ws.cell(row=row, column=1,
                   value='These entries were already recorded via the Payroll Journal and General Journal. '
                         'No further action required — shown here for completeness.')
    note.font = Font(italic=True, size=10, name='Arial', color='595959')
    row += 2

    if not accrual_rows:
        ws.cell(row=row, column=1,
                value='No accrual account movements found in the General Ledger for this period.'
                ).font = NORMAL_FONT
        return

    headers = ['Date', 'Reference', 'Account Code', 'Account Name',
               'Description', 'Debit', 'Credit']
    row = write_header_row(ws, headers, row)

    total_dr = total_cr = 0.0
    for r in sorted(accrual_rows, key=lambda x: x['date']):
        row = write_data_row(ws, [
            _fmt_date(r['date']), r['reference'], r['account_code'],
            r['account_name'], r['description'],
            _n(r['debit']), _n(r['credit']),
        ], row, number_cols=[6, 7])
        total_dr += r['debit']
        total_cr += r['credit']

    row = write_total_row(ws, 'Total', [None, None, None, None,
                                         _n(total_dr), _n(total_cr)],
                          row, double_line=True)
    auto_fit_columns(ws)
    ws.column_dimensions['E'].width = 40
    freeze_panes(ws)


def write_prepayments_sheet(wb, prepaid_rows, period_start, period_end):
    ws = add_sheet(wb, 'Prepayments', tab_color='4472C4')
    row = write_title(ws, 'Prepayments — Reference',
                      'Prepaid expense recognitions already posted in journals this period',
                      f'{period_start}  to  {period_end}')

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    note = ws.cell(row=row, column=1,
                   value='These entries were already recorded via the General Journal. '
                         'No further action required — shown here for completeness.')
    note.font = Font(italic=True, size=10, name='Arial', color='595959')
    row += 2

    if not prepaid_rows:
        ws.cell(row=row, column=1,
                value='No prepaid account movements found in the General Ledger for this period.'
                ).font = NORMAL_FONT
        return

    headers = ['Date', 'Reference', 'Account Code', 'Account Name',
               'Description', 'Debit', 'Credit']
    row = write_header_row(ws, headers, row)

    total_dr = total_cr = 0.0
    for r in sorted(prepaid_rows, key=lambda x: x['date']):
        row = write_data_row(ws, [
            _fmt_date(r['date']), r['reference'], r['account_code'],
            r['account_name'], r['description'],
            _n(r['debit']), _n(r['credit']),
        ], row, number_cols=[6, 7])
        total_dr += r['debit']
        total_cr += r['credit']

    row = write_total_row(ws, 'Total', [None, None, None, None,
                                         _n(total_dr), _n(total_cr)],
                          row, double_line=True)
    auto_fit_columns(ws)
    ws.column_dimensions['E'].width = 40
    freeze_panes(ws)


def write_all_entries(wb, all_entries, period_end):
    ws = add_sheet(wb, 'All Entries', tab_color='4472C4')
    row = write_title(ws, 'All Adjusting Entries — Master Journal',
                      'Period-end adjusting entries generated by Module 4',
                      f'Period ending {period_end}')

    headers = ['Entry No.', 'Date', 'Type', 'Category / Description',
               'Debit Account', 'Dr Code', 'Debit Amount',
               'Credit Account', 'Cr Code', 'Credit Amount',
               'Supporting Reference']
    row = write_header_row(ws, headers, row)

    total_dr = total_cr = 0.0
    for e in all_entries:
        row = write_data_row(ws, [
            e['ref'], _fmt_date(e['date']), e['type'],
            e.get('category', ''),
            e['dr_name'], e['dr_code'], _n(e['amount']),
            e['cr_name'], e['cr_code'], _n(e['amount']),
            e.get('supporting', ''),
        ], row, number_cols=[7, 10])
        total_dr += e['amount']
        total_cr += e['amount']

    row = write_total_row(ws, 'TOTALS',
                          [None, None, None, None, None,
                           _n(total_dr), None, None, _n(total_cr), None],
                          row, double_line=True)
    row += 1

    # Double-entry validation
    balanced = round(total_dr - total_cr, 2) == 0
    ws.cell(row=row, column=1, value='Double-Entry Check').font = NORMAL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=6,
            value=f"Dr {total_dr:,.2f} = Cr {total_cr:,.2f}").font = NORMAL_FONT
    ws.cell(row=row, column=6).border = THIN_BORDER
    write_validation_result(ws, row, 7, balanced)
    ws.cell(row=row, column=7).border = THIN_BORDER

    auto_fit_columns(ws)
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['K'].width = 40
    freeze_panes(ws)


def write_account_impact(wb, impacts, period_end):
    ws = add_sheet(wb, 'Account Impact', tab_color='70AD47')
    row = write_title(ws, 'Account Impact Analysis',
                      'Pre- and post-adjustment balances for all affected accounts',
                      f'Period ending {period_end}')

    headers = ['Account Code', 'Account Name', 'Type', 'Normal Balance',
               'Pre-Adj Balance', 'Adjustment', 'Post-Adj Balance', 'Change %']
    row = write_header_row(ws, headers, row)

    for imp in impacts:
        pre = imp['pre_balance']
        adj = imp['adjustment']
        post = imp['post_balance']
        # Change %
        if pre != 0:
            change_pct = adj / abs(pre)
        else:
            change_pct = None

        row = write_data_row(ws, [
            imp['code'], imp['name'], imp['type'], imp['normal_balance'],
            _n(pre), _n(adj), _n(post),
            change_pct,
        ], row, number_cols=[5, 6, 7])

        # Format change % cell
        pct_cell = ws.cell(row=row - 1, column=8)
        if change_pct is not None:
            pct_cell.value = change_pct
            pct_cell.number_format = '0.0%'
            pct_cell.alignment = Alignment(horizontal='right')
            pct_cell.font = NORMAL_FONT
        else:
            pct_cell.value = 'N/A'
            pct_cell.alignment = Alignment(horizontal='center')
        pct_cell.border = THIN_BORDER

    auto_fit_columns(ws)
    ws.column_dimensions['B'].width = 35
    freeze_panes(ws)


def write_exceptions_sheet(wb, exceptions):
    ws = add_sheet(wb, 'Exceptions', tab_color='FF0000')
    row = write_title(ws, 'Exceptions & Warnings')
    row = write_header_row(ws, ['#', 'Exception / Warning'], row)
    for i, exc in enumerate(exceptions, 1):
        row = write_data_row(ws, [i, exc], row)
    auto_fit_columns(ws)
    freeze_panes(ws)


# ─────────────────────────────────────────────────────────────────────────────
# 6. MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)

    data_dir   = sys.argv[1]
    period_start = sys.argv[2]
    period_end   = sys.argv[3]
    output_file  = sys.argv[4]

    print(f"\n{'='*60}")
    print(f"  MODULE 4 -- JOURNAL ADJUSTMENTS")
    print(f"  Period : {period_start}  to  {period_end}")
    print(f"  Data   : {data_dir}")
    print(f"  Output : {output_file}")
    print(f"{'='*60}\n")

    coa_path = Path(data_dir) / 'chart_of_accounts.xlsx'
    coa = COAMapper(str(coa_path)) if coa_path.exists() else COAMapper()
    exceptions = []

    # ── 1. Depreciation ──────────────────────────────────────────────────────
    print("Computing depreciation...")
    asset_rows, depr_entries, err = compute_depreciation(data_dir)
    if err:
        print(f"  ERROR: {err}")
        exceptions.append(err)
        asset_rows, depr_entries = [], []
    else:
        total_depr = sum(e['amount'] for e in depr_entries)
        print(f"  Assets processed   : {len(asset_rows)}")
        print(f"  Depreciation entries: {len(depr_entries)}")
        print(f"  Total monthly depr : {total_depr:,.2f}")
        for e in depr_entries:
            print(f"    Dr {e['dr_code']} / Cr {e['cr_code']}  {e['amount']:>10,.2f}"
                  f"  ({e['category']})")

    # ── 2. Bank recon entries ────────────────────────────────────────────────
    print("\nLoading bank reconciliation entries...")
    bank_entries, warn = load_bank_recon_entries(data_dir, period_end)
    if warn:
        print(f"  WARNING: {warn}")
        exceptions.append(warn)
    else:
        total_bank = sum(e['amount'] for e in bank_entries)
        print(f"  Bank recon entries  : {len(bank_entries)}")
        print(f"  Total              : {total_bank:,.2f}")
        for e in bank_entries:
            print(f"    Dr {e['dr_code']} / Cr {e['cr_code']}  {e['amount']:>10,.2f}"
                  f"  ({e['category']})")

    # ── 3. GL reference data ─────────────────────────────────────────────────
    print("\nLoading GL reference data...")
    accrual_rows, prepaid_rows, gl_balances = load_gl_reference(
        data_dir, period_start, period_end, coa)
    print(f"  Accrual movements  : {len(accrual_rows)}")
    print(f"  Prepaid movements  : {len(prepaid_rows)}")
    print(f"  GL accounts loaded : {len(gl_balances)}")

    # ── 4. Combine + number entries ──────────────────────────────────────────
    all_entries = depr_entries + bank_entries
    all_entries = assign_entry_numbers(all_entries, period_end)

    total_dr = total_cr = sum(e['amount'] for e in all_entries)
    balanced = True   # by construction (each entry has amount for both Dr and Cr)

    print(f"\n{'-'*50}")
    print(f"  Total new entries  : {len(all_entries)}")
    print(f"  Grand total debits : {total_dr:,.2f}")
    print(f"  Grand total credits: {total_cr:,.2f}")
    print(f"  Balanced           : {balanced}")
    print(f"{'-'*50}")

    for e in all_entries:
        print(f"  {e['ref']}  {e['type']:<12}  "
              f"Dr {e['dr_code']}  Cr {e['cr_code']}  {e['amount']:>10,.2f}")

    # ── 5. Account impact ────────────────────────────────────────────────────
    impacts = compute_account_impact(all_entries, gl_balances, coa)

    print(f"\nAccount impact ({len(impacts)} accounts affected):")
    for imp in impacts:
        print(f"  {imp['code']}  {imp['name']:<30}  "
              f"Pre: {imp['pre_balance']:>10,.2f}  "
              f"Adj: {imp['adjustment']:>+10,.2f}  "
              f"Post: {imp['post_balance']:>10,.2f}")

    # ── 6. Write Excel ───────────────────────────────────────────────────────
    print(f"\nWriting output to: {output_file}")
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    wb = create_workbook()
    write_dashboard(wb, all_entries, depr_entries, bank_entries,
                    period_start, period_end, exceptions)
    write_depreciation_schedule(wb, asset_rows, depr_entries, period_end)
    write_bank_recon_entries_sheet(wb, bank_entries, period_end)
    write_accruals_sheet(wb, accrual_rows, period_start, period_end)
    write_prepayments_sheet(wb, prepaid_rows, period_start, period_end)
    write_all_entries(wb, all_entries, period_end)
    write_account_impact(wb, impacts, period_end)
    if exceptions:
        write_exceptions_sheet(wb, exceptions)

    save_workbook(wb, output_file)

    print(f"\n{'='*60}")
    print(f"  OUTPUT: {output_file}")
    print(f"  Sheets: Dashboard | Depreciation Schedule | Bank Recon Entries |")
    print(f"          Accruals | Prepayments | All Entries | Account Impact"
          + (" | Exceptions" if exceptions else ""))
    print(f"  New adjusting entries : {len(all_entries)}")
    print(f"  Grand total (Dr = Cr) : {total_dr:,.2f}")
    print(f"  Double-entry balanced : {'YES' if balanced else 'NO'}")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
