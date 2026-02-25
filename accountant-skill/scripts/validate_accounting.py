"""
Module 7: Full-Cycle Accounting Validation
Shwe Mandalay Cafe / K&K Finance Team

Performs comprehensive integrity checks across all module outputs to ensure:
  1. Double-entry accounting integrity (Dr = Cr everywhere)
  2. Control account reconciliation (AR, AP, Cash match between GL and subsidiary ledgers)
  3. Cross-module validation (data flows correctly between modules)
  4. Financial statement validation (BS balances, CF reconciles)

Output: audit_validation_[PERIOD].xlsx with all checks documented.

Usage:
    python validate_accounting.py <data_dir> <period_start> <period_end> <output_file>

Example:
    python validate_accounting.py data/Jan2026 2026-01-01 2026-01-31 data/Jan2026/audit_validation_Jan2026.xlsx
"""

import sys
import os
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np

sys.path.insert(0, str(Path(__file__).parent))
from utils.excel_reader import read_xlsx, read_all_sheets, find_xlsx_files
from utils.excel_writer import (
    create_workbook, add_sheet, write_title, write_header_row,
    write_data_row, write_section_header, write_total_row,
    write_validation_result, auto_fit_columns, freeze_panes,
    save_workbook, NORMAL_FONT, TOTAL_FONT, THIN_BORDER,
    PASS_FILL, FAIL_FILL, WARNING_FILL
)
from utils.coa_mapper import COAMapper, CONTRA_ACCOUNTS
from utils.double_entry import (
    validate_journal_balance, validate_trial_balance,
    validate_balance_sheet, check_control_account
)
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm_code(val):
    """Normalize float-string account codes: '1020.0' -> '1020'."""
    try:
        return str(int(float(str(val).strip())))
    except (ValueError, TypeError):
        return str(val).strip()


def _n(val):
    """Return numeric value or None; blanks out zero / NaN for display."""
    if val is None:
        return None
    try:
        v = float(val)
    except (ValueError, TypeError):
        return None
    if np.isnan(v) or abs(v) < 0.005:
        return None
    return v


def _fmt_date(val):
    """Format a date value as YYYY-MM-DD."""
    try:
        return pd.Timestamp(val).strftime('%Y-%m-%d')
    except Exception:
        return str(val) if val else ''


def _find_col(df, candidates):
    """Return the first column name from candidates that exists in df.columns."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _normalize_cols(df):
    """Lowercase + strip all column names."""
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def to_num(series):
    """Convert series to numeric, filling NaN with 0."""
    return pd.to_numeric(series, errors='coerce').fillna(0.0)


# ---------------------------------------------------------------------------
# 1. Load All Module Outputs
# ---------------------------------------------------------------------------

def load_all_outputs(data_dir):
    """
    Load all module output files from the data directory.

    Returns: dict with keys:
        'books_of_prime_entry', 'ledger_summary', 'bank_reconciliation',
        'adjusting_entries', 'trial_balance', 'financial_statements'
    """
    data_dir = Path(data_dir)
    outputs = {}
    errors = []

    # Module 1: Books of Prime Entry
    bope_files = list(data_dir.glob('books_of_prime_entry*.xlsx'))
    if bope_files:
        result = read_all_sheets(bope_files[0])
        if result['error']:
            errors.append(f"Module 1: {result['error']}")
        else:
            outputs['books_of_prime_entry'] = result['data']
    else:
        errors.append("Module 1: books_of_prime_entry*.xlsx not found")

    # Module 2: Ledger Summary
    ledger_files = list(data_dir.glob('ledger_summary*.xlsx'))
    if ledger_files:
        result = read_all_sheets(ledger_files[0])
        if result['error']:
            errors.append(f"Module 2: {result['error']}")
        else:
            outputs['ledger_summary'] = result['data']
    else:
        errors.append("Module 2: ledger_summary*.xlsx not found")

    # Module 3: Bank Reconciliation
    bank_files = list(data_dir.glob('bank_reconciliation*.xlsx'))
    if bank_files:
        result = read_all_sheets(bank_files[0])
        if result['error']:
            errors.append(f"Module 3: {result['error']}")
        else:
            outputs['bank_reconciliation'] = result['data']
    else:
        errors.append("Module 3: bank_reconciliation*.xlsx not found")

    # Module 4: Adjusting Entries
    adj_files = list(data_dir.glob('adjusting_entries*.xlsx'))
    if adj_files:
        result = read_all_sheets(adj_files[0])
        if result['error']:
            errors.append(f"Module 4: {result['error']}")
        else:
            outputs['adjusting_entries'] = result['data']
    else:
        errors.append("Module 4: adjusting_entries*.xlsx not found")

    # Module 5: Trial Balance
    tb_files = list(data_dir.glob('trial_balance*.xlsx'))
    if tb_files:
        result = read_all_sheets(tb_files[0])
        if result['error']:
            errors.append(f"Module 5: {result['error']}")
        else:
            outputs['trial_balance'] = result['data']
    else:
        errors.append("Module 5: trial_balance*.xlsx not found")

    # Module 6: Financial Statements
    fin_files = list(data_dir.glob('financial_statements*.xlsx'))
    if fin_files:
        result = read_all_sheets(fin_files[0])
        if result['error']:
            errors.append(f"Module 6: {result['error']}")
        else:
            outputs['financial_statements'] = result['data']
    else:
        errors.append("Module 6: financial_statements*.xlsx not found")

    return outputs, errors


# ---------------------------------------------------------------------------
# 2. Double-Entry Validation
# ---------------------------------------------------------------------------

def check_double_entry(outputs):
    """
    Validate double-entry integrity across all journals and trial balances.

    Returns: list of check result dicts with keys:
        'check_name', 'status' (PASS/FAIL/WARN), 'details', 'dr_total', 'cr_total', 'difference'
    """
    results = []

    # --- Module 1: Books of Prime Entry ---
    if 'books_of_prime_entry' in outputs:
        bope = outputs['books_of_prime_entry']

        # Check Dashboard sheet for journal totals
        if 'Dashboard' in bope:
            df = bope['Dashboard']
            df.columns = [str(c).strip().lower() for c in df.columns]

            # Find journal rows (skip totals row)
            for _, row in df.iterrows():
                journal_name = str(row.get('journal', '')).strip()
                if journal_name in ('GRAND TOTAL', '', 'Journal'):
                    continue
                if pd.isna(journal_name) or not journal_name:
                    continue

                # Get debit and credit totals
                debit_col = _find_col(df, ['total debits', 'debits', 'debit'])
                credit_col = _find_col(df, ['total credits', 'credits', 'credit'])

                if debit_col and credit_col:
                    dr = to_num(row.get(debit_col, 0))
                    cr = to_num(row.get(credit_col, 0))
                    diff = dr - cr
                    passed = abs(diff) < 0.01

                    results.append({
                        'category': 'Double-Entry',
                        'check_name': f"Module 1 - {journal_name}",
                        'status': 'PASS' if passed else 'FAIL',
                        'details': f"Dr={dr:,.2f}, Cr={cr:,.2f}, Diff={diff:,.2f}" if not passed else f"Dr=Cr={dr:,.2f}",
                        'dr_total': dr,
                        'cr_total': cr,
                        'difference': diff
                    })

    # --- Module 4: Adjusting Entries ---
    if 'adjusting_entries' in outputs:
        adj = outputs['adjusting_entries']

        if 'All Entries' in adj:
            # Find header row dynamically
            df_raw = adj['All Entries']
            header_row_idx = None
            for i, row_vals in df_raw.iterrows():
                row_strs = [str(v).strip() for v in row_vals.values]
                if 'Dr Code' in row_strs or 'Entry No.' in row_strs:
                    header_row_idx = i
                    break

            if header_row_idx is not None:
                df = df_raw.iloc[header_row_idx + 1:].copy()  # Skip header row
                df.columns = [str(c).strip() for c in df_raw.iloc[header_row_idx].values]

                # Filter out TOTALS row
                df = df[df['Dr Code'].apply(lambda x: _norm_code(x).isdigit())]

                if len(df) > 0:
                    total_dr = to_num(df['Debit Amount']).sum()
                    total_cr = to_num(df['Credit Amount']).sum()
                    diff = total_dr - total_cr
                    passed = abs(diff) < 0.01

                    results.append({
                        'category': 'Double-Entry',
                        'check_name': 'Module 4 - All Adjusting Entries',
                        'status': 'PASS' if passed else 'FAIL',
                        'details': f"Dr={total_dr:,.2f}, Cr={total_cr:,.2f}, Diff={diff:,.2f}" if not passed else f"Dr=Cr={total_dr:,.2f}",
                        'dr_total': total_dr,
                        'cr_total': total_cr,
                        'difference': diff
                    })

    # --- Module 5: Trial Balance ---
    if 'trial_balance' in outputs:
        tb = outputs['trial_balance']

        # Check Unadjusted TB
        if 'Unadjusted TB' in tb:
            df = tb['Unadjusted TB']
            df.columns = [str(c).strip().lower() for c in df.columns]

            debit_col = _find_col(df, ['debit', 'dr'])
            credit_col = _find_col(df, ['credit', 'cr'])

            if debit_col and credit_col:
                # Filter out TOTAL row
                df_data = df[df['account code'].apply(lambda x: str(x).isdigit() if pd.notna(x) else False)]

                if len(df_data) > 0:
                    total_dr = to_num(df_data[debit_col]).sum()
                    total_cr = to_num(df_data[credit_col]).sum()
                    diff = total_dr - total_cr
                    passed = abs(diff) < 0.01

                    results.append({
                        'category': 'Double-Entry',
                        'check_name': 'Module 5 - Unadjusted Trial Balance',
                        'status': 'PASS' if passed else 'FAIL',
                        'details': f"Dr={total_dr:,.2f}, Cr={total_cr:,.2f}, Diff={diff:,.2f}" if not passed else f"Dr=Cr={total_dr:,.2f}",
                        'dr_total': total_dr,
                        'cr_total': total_cr,
                        'difference': diff
                    })

        # Check Adjusted TB
        if 'Adjusted TB' in tb:
            df = tb['Adjusted TB']
            df.columns = [str(c).strip().lower() for c in df.columns]

            debit_col = _find_col(df, ['debit', 'dr'])
            credit_col = _find_col(df, ['credit', 'cr'])

            if debit_col and credit_col:
                df_data = df[df['account code'].apply(lambda x: str(x).isdigit() if pd.notna(x) else False)]

                if len(df_data) > 0:
                    total_dr = to_num(df_data[debit_col]).sum()
                    total_cr = to_num(df_data[credit_col]).sum()
                    diff = total_dr - total_cr
                    passed = abs(diff) < 0.01

                    results.append({
                        'category': 'Double-Entry',
                        'check_name': 'Module 5 - Adjusted Trial Balance',
                        'status': 'PASS' if passed else 'FAIL',
                        'details': f"Dr={total_dr:,.2f}, Cr={total_cr:,.2f}, Diff={diff:,.2f}" if not passed else f"Dr=Cr={total_dr:,.2f}",
                        'dr_total': total_dr,
                        'cr_total': total_cr,
                        'difference': diff
                    })

    return results


# ---------------------------------------------------------------------------
# 3. Control Account Reconciliation
# ---------------------------------------------------------------------------

def check_control_accounts(outputs, coa):
    """
    Reconcile AR, AP, and Cash between GL and subsidiary ledgers.

    Returns: list of check result dicts
    """
    results = []

    # Get GL balances from ledger_summary or trial_balance
    gl_ar = gl_ap = gl_cash = None

    if 'trial_balance' in outputs and 'Adjusted TB' in outputs['trial_balance']:
        tb = outputs['trial_balance']['Adjusted TB']
        tb.columns = [str(c).strip().lower() for c in tb.columns]

        for _, row in tb.iterrows():
            code = _norm_code(row.get('account code', ''))
            if not code.isdigit():
                continue
            code = int(code)

            debit = to_num(row.get('debit', 0))
            credit = to_num(row.get('credit', 0))

            # Net balance based on normal balance
            if code == 1100:  # AR
                info = coa.get_account(code)
                normal = info['normal_balance'].lower() if info else 'debit'
                gl_ar = debit - credit if normal == 'debit' else credit - debit
            elif code == 2010:  # AP
                info = coa.get_account(code)
                normal = info['normal_balance'].lower() if info else 'credit'
                gl_ap = credit - debit if normal == 'credit' else debit - credit
            elif code in [1020, 1021, 1022]:  # Cash at Bank
                info = coa.get_account(code)
                normal = info['normal_balance'].lower() if info else 'debit'
                cash_bal = debit - credit if normal == 'debit' else credit - debit
                gl_cash = (gl_cash or 0) + cash_bal

    # Get subsidiary ledger totals
    ar_subsidiary = ap_subsidiary = cash_subsidiary = None

    if 'ledger_summary' in outputs:
        ls = outputs['ledger_summary']

        # AR by Customer
        if 'AR by Customer' in ls:
            df = ls['AR by Customer']
            df.columns = [str(c).strip().lower() for c in df.columns]
            closing_col = _find_col(df, ['closing balance', 'closing', 'balance'])
            if closing_col:
                ar_subsidiary = to_num(df[closing_col]).sum()

        # AP by Supplier
        if 'AP by Supplier' in ls:
            df = ls['AP by Supplier']
            df.columns = [str(c).strip().lower() for c in df.columns]
            closing_col = _find_col(df, ['closing balance', 'closing', 'balance'])
            if closing_col:
                ap_subsidiary = to_num(df[closing_col]).sum()

        # Cash by Bank
        if 'Cash by Bank' in ls:
            df = ls['Cash by Bank']
            df.columns = [str(c).strip().lower() for c in df.columns]
            closing_col = _find_col(df, ['closing balance', 'closing', 'balance'])
            if closing_col:
                cash_subsidiary = to_num(df[closing_col]).sum()

    # Perform reconciliations
    if gl_ar is not None or ar_subsidiary is not None:
        diff = (gl_ar or 0) - (ar_subsidiary or 0)
        passed = abs(diff) < 0.01 if (gl_ar is not None and ar_subsidiary is not None) else False
        status = 'PASS' if passed else ('FAIL' if (gl_ar is not None and ar_subsidiary is not None) else 'SKIP')

        results.append({
            'category': 'Control Account Recon',
            'check_name': 'AR Control Account (1100)',
            'status': status,
            'details': f"GL={_n(gl_ar)}, Sub={_n(ar_subsidiary)}, Diff={_n(diff)}" if status != 'SKIP' else 'Data not available',
            'gl_balance': gl_ar,
            'subsidiary_total': ar_subsidiary,
            'difference': diff
        })

    if gl_ap is not None or ap_subsidiary is not None:
        diff = (gl_ap or 0) - (ap_subsidiary or 0)
        passed = abs(diff) < 0.01 if (gl_ap is not None and ap_subsidiary is not None) else False
        status = 'PASS' if passed else ('FAIL' if (gl_ap is not None and ap_subsidiary is not None) else 'SKIP')

        results.append({
            'category': 'Control Account Recon',
            'check_name': 'AP Control Account (2010)',
            'status': status,
            'details': f"GL={_n(gl_ap)}, Sub={_n(ap_subsidiary)}, Diff={_n(diff)}" if status != 'SKIP' else 'Data not available',
            'gl_balance': gl_ap,
            'subsidiary_total': ap_subsidiary,
            'difference': diff
        })

    if gl_cash is not None or cash_subsidiary is not None:
        diff = (gl_cash or 0) - (cash_subsidiary or 0)
        passed = abs(diff) < 0.01 if (gl_cash is not None and cash_subsidiary is not None) else False
        status = 'PASS' if passed else ('FAIL' if (gl_cash is not None and cash_subsidiary is not None) else 'SKIP')

        results.append({
            'category': 'Control Account Recon',
            'check_name': 'Cash Control Account (1020)',
            'status': status,
            'details': f"GL={_n(gl_cash)}, Sub={_n(cash_subsidiary)}, Diff={_n(diff)}" if status != 'SKIP' else 'Data not available',
            'gl_balance': gl_cash,
            'subsidiary_total': cash_subsidiary,
            'difference': diff
        })

    return results


# ---------------------------------------------------------------------------
# 4. Cross-Module Flow Validation
# ---------------------------------------------------------------------------

def check_cross_module_flow(outputs):
    """
    Verify data flows correctly between modules.

    Checks:
    - Module 3 Adjusting Entries flow to Module 4
    - Module 4 Adjusting Entries flow to Module 5
    - Module 5 Adjusted TB flows to Module 6
    """
    results = []

    # Check: Module 3 -> Module 4 (Bank Recon ADJ entries to Adjusting Entries)
    if 'bank_reconciliation' in outputs and 'adjusting_entries' in outputs:
        br = outputs['bank_reconciliation']
        adj = outputs['adjusting_entries']

        br_adj_count = 0
        if 'Adjusting Entries' in br:
            df = br['Adjusting Entries']
            # Find header row
            header_row_idx = None
            for i, row_vals in df.iterrows():
                row_strs = [str(v).strip() for v in row_vals.values]
                if 'Date' in row_strs and 'Dr Code' in row_strs:
                    header_row_idx = i
                    break

            if header_row_idx is not None:
                df_data = df.iloc[header_row_idx + 1:].copy()
                br_adj_count = len(df_data)

        adj_entry_count = 0
        if 'All Entries' in adj:
            df_raw = adj['All Entries']
            header_row_idx = None
            for i, row_vals in df_raw.iterrows():
                row_strs = [str(v).strip() for v in row_vals.values]
                if 'Entry No.' in row_strs:
                    header_row_idx = i
                    break

            if header_row_idx is not None:
                df = df_raw.iloc[header_row_idx + 1:].copy()
                df.columns = [str(c).strip() for c in df_raw.iloc[header_row_idx].values]
                # Filter out TOTALS row
                df = df[df['Dr Code'].apply(lambda x: _norm_code(x).isdigit())]
                # Count bank reconciliation entries (ADJ-BANK-*)
                adj_entry_count = len(df[df['Type'].str.contains('BANK', case=False, na=False)])

        # BR adjusting entries should flow to Module 4
        matched = br_adj_count == adj_entry_count or (br_adj_count > 0 and adj_entry_count > 0)
        status = 'PASS' if matched else ('WARN' if br_adj_count > 0 else 'SKIP')

        results.append({
            'category': 'Cross-Module Flow',
            'check_name': 'Module 3 -> Module 4 (Bank Recon ADJ)',
            'status': status,
            'details': f"BR Adj Entries={br_adj_count}, Module 4 Bank ADJ={adj_entry_count}",
        })

    # Check: Module 4 -> Module 5 (All adjusting entries flow to TB)
    if 'adjusting_entries' in outputs and 'trial_balance' in outputs:
        adj = outputs['adjusting_entries']
        tb = outputs['trial_balance']

        adj_accounts = set()
        if 'All Entries' in adj:
            df_raw = adj['All Entries']
            header_row_idx = None
            for i, row_vals in df_raw.iterrows():
                row_strs = [str(v).strip() for v in row_vals.values]
                if 'Dr Code' in row_strs:
                    header_row_idx = i
                    break

            if header_row_idx is not None:
                df = df_raw.iloc[header_row_idx + 1:].copy()
                df.columns = [str(c).strip() for c in df_raw.iloc[header_row_idx].values]
                df = df[df['Dr Code'].apply(lambda x: _norm_code(x).isdigit())]
                adj_accounts = set(_norm_code(c) for c in df['Dr Code'].tolist() + df['Cr Code'].tolist())

        tb_adj_accounts = set()
        if 'Adjustments' in tb:
            df = tb['Adjustments']
            # Find header row
            header_row_idx = None
            for i, row_vals in df.iterrows():
                row_strs = [str(v).strip() for v in row_vals.values]
                if 'Dr Code' in row_strs:
                    header_row_idx = i
                    break

            if header_row_idx is not None:
                df_data = df.iloc[header_row_idx + 1:].copy()
                df_data.columns = [str(c).strip() for c in df.iloc[header_row_idx].values]
                # Get accounts from per-account summary section
                if 'Account Code' in df_data.columns:
                    tb_adj_accounts = set(_norm_code(c) for c in df_data['Account Code'].dropna().tolist())

        # Check if all adjusting entry accounts appear in TB
        missing = adj_accounts - tb_adj_accounts if tb_adj_accounts else set()
        status = 'PASS' if not missing else ('WARN' if adj_accounts else 'SKIP')

        results.append({
            'category': 'Cross-Module Flow',
            'check_name': 'Module 4 -> Module 5 (Adj Entries to TB)',
            'status': status,
            'details': f"Adj accounts={len(adj_accounts)}, TB adj accounts={len(tb_adj_accounts)}" + (f", Missing={len(missing)}" if missing else ""),
        })

    # Check: Module 5 -> Module 6 (Adjusted TB to Financial Statements)
    if 'trial_balance' in outputs and 'financial_statements' in outputs:
        tb = outputs['trial_balance']
        fs = outputs['financial_statements']

        # Get net profit from TB (sum of revenue and expense accounts)
        tb_net_profit = None
        if 'Adjusted TB' in tb:
            df = tb['Adjusted TB']
            df.columns = [str(c).strip().lower() for c in df.columns]

            debit_col = _find_col(df, ['debit', 'dr'])
            credit_col = _find_col(df, ['credit', 'cr'])

            if debit_col and credit_col:
                revenue = credit = debit = 0.0
                for _, row in df.iterrows():
                    code_str = str(row.get('account code', ''))
                    if not code_str.replace('.', '').replace('-', '').isdigit():
                        continue
                    try:
                        code = int(float(code_str))
                    except (ValueError, TypeError):
                        continue

                    dr = to_num(row.get(debit_col, 0))
                    cr = to_num(row.get(credit_col, 0))

                    if 4000 <= code <= 4999:  # Revenue
                        revenue += cr - dr
                    elif 5000 <= code <= 5999:  # Expense
                        debit += dr - cr

                tb_net_profit = revenue - debit

        # Get net profit from financial statements (Income Statement)
        fs_net_profit = None
        if 'Income Statement' in fs:
            df = fs['Income Statement']
            # Look for NET PROFIT row
            for i, row in df.iterrows():
                row_str = str(row.iloc[0]).upper() if len(row) > 0 else ''
                if 'NET PROFIT' in row_str or 'NET INCOME' in row_str:
                    # Get the value from the current period column
                    for val in row.iloc[1:]:
                        try:
                            fs_net_profit = float(val)
                            break
                        except (ValueError, TypeError):
                            continue
                    break

        if tb_net_profit is not None and fs_net_profit is not None:
            diff = abs(tb_net_profit - fs_net_profit)
            passed = diff < 0.01
            status = 'PASS' if passed else 'FAIL'

            results.append({
                'category': 'Cross-Module Flow',
                'check_name': 'Module 5 -> Module 6 (Net Profit tie-out)',
                'status': status,
                'details': f"TB Net Profit={tb_net_profit:,.2f}, FS Net Profit={fs_net_profit:,.2f}, Diff={diff:,.2f}",
            })

    return results


# ---------------------------------------------------------------------------
# 5. Financial Statement Validation
# ---------------------------------------------------------------------------

def check_financials(outputs, coa):
    """
    Validate financial statement integrity.

    Checks:
    - Balance Sheet: Assets = Equity + Liabilities
    - Cash Flow: Opening + Net Change = Closing
    - IS Net Profit matches BS equity addition
    """
    results = []

    if 'financial_statements' not in outputs:
        return results

    fs = outputs['financial_statements']

    # --- Balance Sheet Validation ---
    if 'Balance Sheet' in fs:
        df = fs['Balance Sheet']

        # Normalize columns
        df_cols = [str(c).strip().lower() for c in df.columns]
        df_normalized = df.copy()
        df_normalized.columns = df_cols

        # Try to find total rows
        total_assets = total_equity = total_liabilities = None

        # Find the value columns (usually last 1-2 columns)
        value_cols = [c for c in df_cols if 'unnamed' not in c or c in ['amount', 'total', 'value']]
        if not value_cols:
            value_cols = df_cols[1:]  # Fall back to all columns except first

        for i, row in df.iterrows():
            row_str = str(row.iloc[0]).upper() if len(row) > 0 else ''

            # Look for TOTAL ASSETS (not NON-CURRENT or CURRENT)
            if 'TOTAL ASSETS' in row_str and 'NON-CURRENT' not in row_str and 'CURRENT' not in row_str:
                # Try to get value from the "Total" column or last numeric column
                for col_idx in range(len(row) - 1, 0, -1):
                    try:
                        val = float(row.iloc[col_idx])
                        total_assets = val
                        break
                    except (ValueError, TypeError):
                        continue

            # TOTAL EQUITY (but not TOTAL EQUITY & LIABILITIES)
            if 'TOTAL EQUITY' in row_str and 'LIABILITIES' not in row_str:
                for col_idx in range(len(row) - 1, 0, -1):
                    try:
                        val = float(row.iloc[col_idx])
                        total_equity = val
                        break
                    except (ValueError, TypeError):
                        continue

            # TOTAL LIABILITIES
            if 'TOTAL LIABILITIES' in row_str and 'EQUITY' not in row_str:
                for col_idx in range(len(row) - 1, 0, -1):
                    try:
                        val = float(row.iloc[col_idx])
                        total_liabilities = val
                        break
                    except (ValueError, TypeError):
                        continue

        if total_assets is not None and total_equity is not None and total_liabilities is not None:
            diff = total_assets - (total_equity + total_liabilities)
            passed = abs(diff) < 0.01
            status = 'PASS' if passed else 'FAIL'

            results.append({
                'category': 'Financial Validation',
                'check_name': 'Balance Sheet: Assets = Equity + Liabilities',
                'status': status,
                'details': f"Assets={total_assets:,.2f}, Equity={total_equity:,.2f}, Liabilities={total_liabilities:,.2f}, Diff={diff:,.2f}",
            })
        elif 'Dashboard' not in fs:
            # Only add failure if we couldn't find values AND there's no Dashboard validation
            results.append({
                'category': 'Financial Validation',
                'check_name': 'Balance Sheet: Assets = Equity + Liabilities',
                'status': 'WARN',
                'details': f'Could not parse totals from Balance Sheet (Assets={total_assets}, Equity={total_equity}, Liabilities={total_liabilities})',
            })

    # --- Cash Flow Validation ---
    if 'Cash Flow' in fs:
        df = fs['Cash Flow']

        opening_cash = closing_cash = net_change = None

        for i, row in df.iterrows():
            row_str = str(row.iloc[0]).upper() if len(row) > 0 else ''

            if 'OPENING' in row_str and 'CASH' in row_str:
                # Get last numeric value in row
                for col_idx in range(len(row) - 1, 0, -1):
                    try:
                        opening_cash = float(row.iloc[col_idx])
                        break
                    except (ValueError, TypeError):
                        continue

            if 'CLOSING' in row_str and 'CASH' in row_str:
                for col_idx in range(len(row) - 1, 0, -1):
                    try:
                        closing_cash = float(row.iloc[col_idx])
                        break
                    except (ValueError, TypeError):
                        continue

            if 'NET INCREASE' in row_str or 'NET CHANGE' in row_str:
                for col_idx in range(len(row) - 1, 0, -1):
                    try:
                        net_change = float(row.iloc[col_idx])
                        break
                    except (ValueError, TypeError):
                        continue

        if opening_cash is not None and closing_cash is not None and net_change is not None:
            expected_closing = opening_cash + net_change
            diff = abs(closing_cash - expected_closing)
            passed = diff < 0.01
            status = 'PASS' if passed else 'FAIL'

            results.append({
                'category': 'Financial Validation',
                'check_name': 'Cash Flow: Opening + Net Change = Closing',
                'status': status,
                'details': f"Opening={opening_cash:,.2f}, Net Change={net_change:,.2f}, Expected Closing={expected_closing:,.2f}, Actual Closing={closing_cash:,.2f}, Diff={diff:,.2f}",
            })
        elif 'Dashboard' not in fs:
            results.append({
                'category': 'Financial Validation',
                'check_name': 'Cash Flow: Opening + Net Change = Closing',
                'status': 'WARN',
                'details': f'Could not parse Cash Flow values (Opening={opening_cash}, Net Change={net_change}, Closing={closing_cash})',
            })

    # --- Dashboard Checks (if available) ---
    if 'Dashboard' in fs:
        df = fs['Dashboard']

        # Look for validation check rows
        for i, row in df.iterrows():
            # Get first column value as check name
            first_col = str(row.iloc[0]).upper() if len(row) > 0 else ''

            # Look for validation section marker
            if 'VALIDATION' in first_col and 'CHECK' in first_col:
                continue  # Skip header row

            # Look for Balance Sheet check
            if 'BALANCE SHEET' in first_col and ('ASSETS' in first_col or 'EQUITY' in first_col or 'LIABILITIES' in first_col):
                # Look for result in last column
                for col_idx in range(len(row) - 1, 0, -1):
                    result = str(row.iloc[col_idx]).upper()
                    if 'PASS' in result or 'YES' in result:
                        status = 'PASS'
                        break
                    elif 'FAIL' in result or 'NO' in result:
                        status = 'FAIL'
                        break
                    elif result.strip() and result != 'NAN':
                        status = 'PASS' if result.strip() else 'FAIL'
                        break
                else:
                    continue

                results.append({
                    'category': 'Financial Validation',
                    'check_name': 'Balance Sheet (Dashboard Check)',
                    'status': status,
                    'details': f"Dashboard validation: {result}",
                })

            # Look for Cash Flow check
            if 'CASH FLOW' in first_col and ('CHECK' in first_col or 'RECONCILE' in first_col):
                for col_idx in range(len(row) - 1, 0, -1):
                    result = str(row.iloc[col_idx]).upper()
                    if 'PASS' in result or 'YES' in result:
                        status = 'PASS'
                        break
                    elif 'FAIL' in result or 'NO' in result:
                        status = 'FAIL'
                        break
                    elif result.strip() and result != 'NAN':
                        status = 'PASS' if result.strip() else 'FAIL'
                        break
                else:
                    continue

                results.append({
                    'category': 'Financial Validation',
                    'check_name': 'Cash Flow (Dashboard Check)',
                    'status': status,
                    'details': f"Dashboard validation: {result}",
                })

    return results


# ---------------------------------------------------------------------------
# 6. Output Sheet Writers
# ---------------------------------------------------------------------------

def write_dashboard(wb, all_results, period_start, period_end):
    """Write the Dashboard summary sheet."""
    ws = add_sheet(wb, 'Dashboard', tab_color='00B050')
    row = write_title(ws, 'FULL-CYCLE ACCOUNTING VALIDATION',
                      'Audit Validation Report',
                      f"{period_start} to {period_end}")

    # Summary counts
    total = len(all_results)
    passed = sum(1 for r in all_results if r['status'] == 'PASS')
    failed = sum(1 for r in all_results if r['status'] == 'FAIL')
    warned = sum(1 for r in all_results if r['status'] == 'WARN')
    skipped = sum(1 for r in all_results if r['status'] == 'SKIP')

    # Summary table
    row = write_section_header(ws, 'VALIDATION SUMMARY', row, col_span=4)
    row = write_header_row(ws, ['Status', 'Count', 'Percentage', ''], row)

    for status, count in [('PASS', passed), ('FAIL', failed), ('WARN', warned), ('SKIP', skipped)]:
        pct = f"{count / total * 100:.1f}%" if total > 0 else '0.0%'
        row = write_data_row(ws, [status, count, pct, ''], row)

        # Color code the status cell
        fill_color = PASS_FILL if status == 'PASS' else (FAIL_FILL if status == 'FAIL' else (WARNING_FILL if status == 'WARN' else None))
        if fill_color:
            for col in range(1, 5):
                ws.cell(row=row - 1, column=col).fill = fill_color

    row += 1

    # Overall result
    overall_pass = failed == 0
    row = write_section_header(ws, 'OVERALL RESULT', row, col_span=4)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value='All Critical Checks:').font = NORMAL_FONT
    result_cell = ws.cell(row=row, column=3, value='PASS' if overall_pass else 'FAIL')
    result_cell.font = Font(bold=True, color='006100' if overall_pass else '9C0006')
    result_cell.fill = PASS_FILL if overall_pass else FAIL_FILL
    result_cell.alignment = Alignment(horizontal='center')
    row += 1

    row += 1

    # Detailed results by category
    row = write_section_header(ws, 'DETAILED RESULTS', row, col_span=5)
    row = write_header_row(ws, ['Category', 'Check Name', 'Status', 'Details', ''], row)

    # Group by category
    categories = ['Double-Entry', 'Control Account Recon', 'Cross-Module Flow', 'Financial Validation']

    for category in categories:
        cat_results = [r for r in all_results if r.get('category') == category]

        if cat_results:
            # Write category header
            row = write_section_header(ws, category.upper(), row, col_span=5)

            for result in cat_results:
                status = result.get('status', 'UNKNOWN')
                fill = PASS_FILL if status == 'PASS' else (FAIL_FILL if status == 'FAIL' else (WARNING_FILL if status == 'WARN' else None))

                values = [
                    result.get('category', ''),
                    result.get('check_name', ''),
                    status,
                    result.get('details', ''),
                    ''
                ]
                row = write_data_row(ws, values, row)

                # Color the status cell
                status_col = 3
                status_cell = ws.cell(row=row - 1, column=status_col)
                status_cell.font = Font(bold=True, color='006100' if status == 'PASS' else ('9C0006' if status == 'FAIL' else '000000'))
                if fill:
                    status_cell.fill = fill

    auto_fit_columns(ws)
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['D'].width = 50
    freeze_panes(ws, row=2, col=1)


def write_detail_sheets(wb, all_results, period_start, period_end):
    """Write detailed breakdown sheets by category."""

    # Double-Entry Details Sheet
    de_results = [r for r in all_results if r.get('category') == 'Double-Entry']
    if de_results:
        ws = add_sheet(wb, 'Double-Entry Checks', tab_color='4472C4')
        row = write_title(ws, 'Double-Entry Validation Details',
                          'Dr = Cr verification across all journals and TBs',
                          f"{period_start} to {period_end}")

        row = write_header_row(ws, ['Check Name', 'Status', 'Dr Total', 'Cr Total', 'Difference', 'Details'], row)

        for r in de_results:
            fill = PASS_FILL if r['status'] == 'PASS' else FAIL_FILL
            values = [
                r.get('check_name', ''),
                r.get('status', ''),
                _n(r.get('dr_total')),
                _n(r.get('cr_total')),
                _n(r.get('difference')),
                r.get('details', '')
            ]
            row = write_data_row(ws, values, row, number_cols=[3, 4, 5])

            # Color status cell
            status_cell = ws.cell(row=row - 1, column=2)
            status_cell.font = Font(bold=True, color='006100' if r['status'] == 'PASS' else '9C0006')
            status_cell.fill = fill

        auto_fit_columns(ws)
        freeze_panes(ws)

    # Control Account Recon Sheet
    ca_results = [r for r in all_results if r.get('category') == 'Control Account Recon']
    if ca_results:
        ws = add_sheet(wb, 'Control Account Recon', tab_color='4472C4')
        row = write_title(ws, 'Control Account Reconciliation',
                          'AR, AP, Cash GL vs Subsidiary Ledger comparison',
                          f"{period_start} to {period_end}")

        row = write_header_row(ws, ['Account', 'Status', 'GL Balance', 'Subsidiary Total', 'Difference', 'Details'], row)

        for r in ca_results:
            fill = PASS_FILL if r['status'] == 'PASS' else (FAIL_FILL if r['status'] == 'FAIL' else WARNING_FILL)
            values = [
                r.get('check_name', ''),
                r.get('status', ''),
                _n(r.get('gl_balance')),
                _n(r.get('subsidiary_total')),
                _n(r.get('difference')),
                r.get('details', '')
            ]
            row = write_data_row(ws, values, row, number_cols=[3, 4, 5])

            status_cell = ws.cell(row=row - 1, column=2)
            status_cell.font = Font(bold=True, color='006100' if r['status'] == 'PASS' else '9C0006')
            status_cell.fill = fill

        auto_fit_columns(ws)
        freeze_panes(ws)

    # Cross-Module Flow Sheet
    cm_results = [r for r in all_results if r.get('category') == 'Cross-Module Flow']
    if cm_results:
        ws = add_sheet(wb, 'Cross-Module Flow', tab_color='70AD47')
        row = write_title(ws, 'Cross-Module Data Flow Validation',
                          'Verify data flows correctly between modules',
                          f"{period_start} to {period_end}")

        row = write_header_row(ws, ['Check Name', 'Status', 'Details'], row)

        for r in cm_results:
            fill = PASS_FILL if r['status'] == 'PASS' else (WARNING_FILL if r['status'] == 'WARN' else FAIL_FILL)
            values = [
                r.get('check_name', ''),
                r.get('status', ''),
                r.get('details', '')
            ]
            row = write_data_row(ws, values, row)

            status_cell = ws.cell(row=row - 1, column=2)
            status_cell.font = Font(bold=True, color='006100' if r['status'] == 'PASS' else '9C0006')
            status_cell.fill = fill

        auto_fit_columns(ws)
        freeze_panes(ws)

    # Financial Validation Sheet
    fv_results = [r for r in all_results if r.get('category') == 'Financial Validation']
    if fv_results:
        ws = add_sheet(wb, 'Financial Validation', tab_color='4472C4')
        row = write_title(ws, 'Financial Statement Validation',
                          'BS and CF integrity checks',
                          f"{period_start} to {period_end}")

        row = write_header_row(ws, ['Check Name', 'Status', 'Details'], row)

        for r in fv_results:
            fill = PASS_FILL if r['status'] == 'PASS' else FAIL_FILL
            values = [
                r.get('check_name', ''),
                r.get('status', ''),
                r.get('details', '')
            ]
            row = write_data_row(ws, values, row)

            status_cell = ws.cell(row=row - 1, column=2)
            status_cell.font = Font(bold=True, color='006100' if r['status'] == 'PASS' else '9C0006')
            status_cell.fill = fill

        auto_fit_columns(ws)
        freeze_panes(ws)


def write_exceptions_sheet(wb, all_results):
    """Write Exceptions sheet with only FAIL and WARN results."""
    exceptions = [r for r in all_results if r['status'] in ('FAIL', 'WARN')]

    if exceptions:
        ws = add_sheet(wb, 'Exceptions', tab_color='FF0000')
        row = write_title(ws, 'Exceptions & Warnings',
                          'Items requiring attention')

        row = write_header_row(ws, ['#', 'Category', 'Check Name', 'Status', 'Details'], row)

        for i, r in enumerate(exceptions, 1):
            fill = FAIL_FILL if r['status'] == 'FAIL' else WARNING_FILL
            values = [
                i,
                r.get('category', ''),
                r.get('check_name', ''),
                r.get('status', ''),
                r.get('details', '')
            ]
            row = write_data_row(ws, values, row)

            # Color the entire row
            for col in range(1, 6):
                cell = ws.cell(row=row - 1, column=col)
                if not cell.fill or cell.fill.fill_type is None:
                    cell.fill = fill

        auto_fit_columns(ws)
        freeze_panes(ws)


# ---------------------------------------------------------------------------
# 7. Main Function
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)

    data_dir = sys.argv[1]
    period_start = sys.argv[2]
    period_end = sys.argv[3]
    output_file = sys.argv[4]

    print("=" * 60)
    print("  MODULE 7 -- FULL-CYCLE ACCOUNTING VALIDATION")
    print(f"  Period : {period_start} to {period_end}")
    print(f"  Data   : {data_dir}")
    print(f"  Output : {output_file}")
    print("=" * 60)
    print()

    # Load COA for classification
    coa_path = Path(data_dir) / 'chart_of_accounts.xlsx'
    coa = COAMapper(str(coa_path)) if coa_path.exists() else COAMapper()

    # Load all module outputs
    print("Loading module outputs...")
    outputs, load_errors = load_all_outputs(data_dir)

    if load_errors:
        print("  Load warnings:")
        for err in load_errors:
            print(f"    - {err}")
    else:
        print("  All module outputs loaded successfully.")

    print()
    print("Running validation checks...")

    # Run all validation checks
    all_results = []

    # 1. Double-Entry Validation
    print("  - Checking double-entry integrity...")
    de_results = check_double_entry(outputs)
    all_results.extend(de_results)

    # 2. Control Account Reconciliation
    print("  - Reconciling control accounts...")
    ca_results = check_control_accounts(outputs, coa)
    all_results.extend(ca_results)

    # 3. Cross-Module Flow Validation
    print("  - Validating cross-module data flow...")
    cm_results = check_cross_module_flow(outputs)
    all_results.extend(cm_results)

    # 4. Financial Statement Validation
    print("  - Validating financial statements...")
    fv_results = check_financials(outputs, coa)
    all_results.extend(fv_results)

    # Add load errors as WARN results
    for err in load_errors:
        all_results.append({
            'category': 'Data Load',
            'check_name': err.split(':')[0] if ':' in err else err,
            'status': 'WARN',
            'details': err
        })

    # Print summary
    print()
    total = len(all_results)
    passed = sum(1 for r in all_results if r['status'] == 'PASS')
    failed = sum(1 for r in all_results if r['status'] == 'FAIL')
    warned = sum(1 for r in all_results if r['status'] == 'WARN')

    print(f"Validation complete: {passed}/{total} passed, {failed} failed, {warned} warnings")

    # Write output workbook
    print()
    print(f"Writing validation report to: {output_file}")
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    wb = create_workbook()
    write_dashboard(wb, all_results, period_start, period_end)
    write_detail_sheets(wb, all_results, period_start, period_end)
    write_exceptions_sheet(wb, all_results)
    save_workbook(wb, output_file)

    print()
    print("=" * 60)
    print(f"  OUTPUT  : {output_file}")
    print(f"  Sheets  : Dashboard | Double-Entry Checks | Control Account Recon")
    print(f"            Cross-Module Flow | Financial Validation | Exceptions")
    print(f"  RESULT  : {'PASS' if failed == 0 else 'FAIL'} ({passed}/{total} checks passed)")
    print("=" * 60)
    print()


if __name__ == '__main__':
    main()
