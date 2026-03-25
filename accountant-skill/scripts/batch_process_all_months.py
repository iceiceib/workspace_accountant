#!/usr/bin/env python
"""
Batch process all months from Feb 2025 to Oct 2025.
Chains opening balances from month to month.
Generates Trial Balance in format: Opening | Debits | Credits | Ending
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import subprocess
import sys

# Paths
SOURCE_DIR = Path('Exisitng Accounting Workflow _ reference files')
OUTPUT_DIR = Path('data/output')
MASTER_DIR = Path('data/input/master')

# Months to process (in order for balance chaining)
MONTHS = [
    ('2025-02-01', '2025-02-28', 'Feb2025'),
    ('2025-03-01', '2025-03-31', 'Mar2025'),
    ('2025-04-01', '2025-04-30', 'Apr2025'),
    ('2025-05-01', '2025-05-31', 'May2025'),
    ('2025-06-01', '2025-06-30', 'Jun2025'),
    ('2025-07-01', '2025-07-31', 'Jul2025'),
    ('2025-08-01', '2025-08-31', 'Aug2025'),
    ('2025-09-01', '2025-09-30', 'Sep2025'),
    ('2025-10-01', '2025-10-31', 'Oct2025'),
]

# Account code constants for journal classification
CASH_ACCOUNT = 10100

REVENUE_ACCOUNTS = {
    40000: 'Sales Revenue',
    70000: 'Interest Income',
}

CAPITAL_ACCOUNTS = {
    31000: 'Capital',
}

EXPENSE_PAYMENT_ACCOUNTS = [
    50010, 50110, 53000, 53100, 53200,  # COGS & Production
    65000,  # Facility Supplies
    14000,  # Prepaid Expenses
    15500,  # Construction in Progress
    15200,  # Machinery & Equipment
    13000,  # Advanced Payments
]

INVENTORY_ADJUSTMENT_ACCOUNTS = [
    50000, 50020, 50100, 50120, 50200, 50220,  # Inventory accounts
    12000, 12100, 12200,  # Inventory adjustments
]

DEPRECIATION_ACCOUNTS = [
    15110, 15210, 15410,  # Accumulated Depreciation (contra-assets)
    66000, 53300,  # Depreciation expense
]

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import project utilities
sys.path.insert(0, str(Path(__file__).parent))
from utils.excel_writer import (
    create_workbook, add_sheet, write_title, write_header_row, write_data_row,
    write_total_row, auto_fit_columns, save_workbook,
    TITLE_FONT, HEADER_FILL, HEADER_FONT, TOTAL_FONT, THIN_BORDER, THICK_BOTTOM,
    NUMBER_FORMAT_NEG
)
from utils.coa_mapper import COAMapper


def load_coa():
    """Load Chart of Accounts with opening balances."""
    coa_path = MASTER_DIR / 'chart_of_accounts.xlsx'
    coa_mapper = COAMapper(coa_path)

    accounts = {}
    opening_balances = {}

    for _, row in coa_mapper.coa_df.iterrows():
        code = int(row['Account Code'])
        accounts[code] = {
            'code': code,
            'name': row['Account Name'],
            'type': row['Type'],
            'sub_type': row.get('Sub-Type', ''),
            'normal_balance': row['Normal Balance'].lower(),
            'status': row.get('Status', 'Active')
        }
        # Load opening balance
        opening = row.get('Opening Balance', 0)
        opening_balances[code] = float(opening) if pd.notna(opening) and opening != 0 else 0.0

    return accounts, opening_balances, coa_mapper


def initialize_balance(code, name, acct_type, normal_balance, opening=0.0):
    """Create a balance dict for an account."""
    return {
        'name': name,
        'type': acct_type,
        'normal_balance': normal_balance,
        'opening': opening,
        'period_dr': 0.0,
        'period_cr': 0.0,
        'closing': opening,
    }


def process_gl_transactions(period_gl, balances, coa_mapper):
    """Process GL transactions and update balances."""
    for idx, row in period_gl.iterrows():
        code = int(row['COA Account Number'])
        dr = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
        cr = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0

        if code in balances:
            balances[code]['period_dr'] += dr
            balances[code]['period_cr'] += cr

            # Update closing balance
            if balances[code]['normal_balance'] == 'debit':
                balances[code]['closing'] = balances[code]['opening'] + balances[code]['period_dr'] - balances[code]['period_cr']
            else:
                balances[code]['closing'] = balances[code]['opening'] - balances[code]['period_dr'] + balances[code]['period_cr']


def create_cash_receipt_entry(row, receipt_no, account_code, description='Cash Sales'):
    """Create a cash receipt journal entry."""
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
    amount = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0
    return {
        'Date': row['Date'],
        'Receipt No': receipt_no,
        'Received From': description if not desc else desc[:50],
        'Description': desc or description,
        'Amount': amount,
        'Bank Account': 'Main',
        'Debit Account': CASH_ACCOUNT,
        'Credit Account': account_code,
    }


def create_cash_payment_entry(row, payment_no, account_code):
    """Create a cash payment journal entry."""
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
    amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
    if amount > 0:
        return {
            'Date': row['Date'],
            'Payment No': payment_no,
            'Paid To': desc[:50] if desc else f'Payment for {account_code}',
            'Description': desc,
            'Amount': amount,
            'Bank Account': 'Main',
            'Debit Account': account_code,
            'Credit Account': CASH_ACCOUNT,
        }
    return None


def create_general_journal_entry(row, jv_no):
    """Create a general journal entry."""
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
    dr = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
    cr = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0
    if dr > 0 or cr > 0:
        return {
            'Date': row['Date'],
            'JV No': jv_no,
            'Description': desc,
            'Debit Account': int(row['COA Account Number']),
            'Credit Account': int(row['COA Account Number']),
            'Debit Amount': dr,
            'Credit Amount': cr,
        }
    return None


def create_journals(period_gl):
    """
    Create journals from GL data.
    Returns: (cash_receipts, cash_payments, general_journal)
    """
    cash_receipts = []
    cash_payments = []
    general_journal = []

    # Sales Revenue -> Cash Receipts
    for idx, row in period_gl[period_gl['COA Account Number'] == 40000].iterrows():
        receipt_no = f'CR-{row["Date"].strftime("%m%d")}-{len(cash_receipts)+1:03d}'
        cash_receipts.append(create_cash_receipt_entry(row, receipt_no, 40000, 'Cash Sales'))

    # Interest Income, Capital -> Cash Receipts
    for acct in [70000, 31000]:
        for idx, row in period_gl[period_gl['COA Account Number'] == acct].iterrows():
            receipt_no = f'CR-{row["Date"].strftime("%m%d")}-{len(cash_receipts)+1:03d}'
            cash_receipts.append(create_cash_receipt_entry(row, receipt_no, acct))

    # Purchases, Expenses, CIP -> Cash Payments
    for acct in EXPENSE_PAYMENT_ACCOUNTS:
        for idx, row in period_gl[period_gl['COA Account Number'] == acct].iterrows():
            payment_no = f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}'
            entry = create_cash_payment_entry(row, payment_no, acct)
            if entry:
                cash_payments.append(entry)

    # Inventory adjustments -> General Journal
    for idx, row in period_gl[period_gl['COA Account Number'].isin(INVENTORY_ADJUSTMENT_ACCOUNTS)].iterrows():
        jv_no = f'JV-{row["Date"].strftime("%m")}-{len(general_journal)+1:03d}'
        entry = create_general_journal_entry(row, jv_no)
        if entry:
            general_journal.append(entry)

    # Depreciation -> General Journal
    for idx, row in period_gl[period_gl['COA Account Number'].isin(DEPRECIATION_ACCOUNTS)].iterrows():
        jv_no = f'JV-{row["Date"].strftime("%m")}-{len(general_journal)+1:03d}'
        entry = create_general_journal_entry(row, jv_no)
        if entry:
            general_journal.append(entry)

    return cash_receipts, cash_payments, general_journal


def extract_period_data(gl_df, coa, start_date, end_date, opening_balances, coa_mapper):
    """
    Extract period data from GL and classify into journals.
    Returns: (cash_receipts, cash_payments, general_journal, period_gl, balances, ending_balances)
    """
    # Filter GL for this period
    period_gl = gl_df[(gl_df['Date'] >= start_date) & (gl_df['Date'] <= end_date)].copy()

    # Initialize balances with opening - include ALL accounts from COA plus any in GL
    balances = {}
    for code, info in coa.items():
        balances[code] = initialize_balance(
            code, info['name'], info['type'], info['normal_balance'],
            opening_balances.get(code, 0.0) if opening_balances else 0.0
        )

    # Also add accounts that appear in GL but not in COA
    gl_codes = period_gl['COA Account Number'].dropna().unique()
    for code_val in gl_codes:
        code = int(float(code_val))
        if code not in balances:
            # Get name from GL
            name_rows = period_gl[period_gl['COA Account Number'] == code_val]
            name = str(name_rows['Account Name'].iloc[0]) if len(name_rows) > 0 and pd.notna(name_rows['Account Name'].iloc[0]) else f'Account {code}'

            # Use COAMapper for classification
            acct_type = coa_mapper.get_type(code)
            normal_balance = coa_mapper.is_debit_normal(code)

            balances[code] = initialize_balance(
                code, name, acct_type, 'debit' if normal_balance else 'credit',
                opening_balances.get(code, 0.0) if opening_balances else 0.0
            )

    # Process GL transactions
    process_gl_transactions(period_gl, balances, coa_mapper)

    # Create journals
    cash_receipts, cash_payments, general_journal = create_journals(period_gl)

    # Prepare GL output
    gl_output = period_gl[['Date', 'COA Account Number', 'Account Name', 'Descritpion', 'Debit (MMK)', 'Credit (MMK)', 'Account Balance (MMK)']].copy()
    gl_output.columns = ['Date', 'Account Code', 'Account Name', 'Description', 'Debit', 'Credit', 'Balance']
    gl_output = gl_output.sort_values(['Account Code', 'Date'])

    # Ending balances for chaining
    ending_balances = {code: data['closing'] for code, data in balances.items()}

    return cash_receipts, cash_payments, general_journal, gl_output, balances, ending_balances


def calculate_balance_totals(balances):
    """
    Calculate trial balance totals.
    Returns: (total_opening_dr, total_opening_cr, total_period_dr, total_period_cr, total_ending_dr, total_ending_cr)
    """
    total_opening_dr = total_opening_cr = 0
    total_period_dr = total_period_cr = 0
    total_ending_dr = total_ending_cr = 0

    for code, data in balances.items():
        opening = data['opening']
        period_dr = data['period_dr']
        period_cr = data['period_cr']
        closing = data['closing']
        normal = data['normal_balance']

        # Opening balance display
        if normal == 'debit':
            if opening >= 0:
                total_opening_dr += opening
            else:
                total_opening_cr += abs(opening)
        else:
            if opening >= 0:
                total_opening_cr += opening
            else:
                total_opening_dr += abs(opening)

        # Period movements
        total_period_dr += period_dr
        total_period_cr += period_cr

        # Ending balance display
        if normal == 'debit':
            if closing >= 0:
                total_ending_dr += closing
            else:
                total_ending_cr += abs(closing)
        else:
            if closing >= 0:
                total_ending_cr += closing
            else:
                total_ending_dr += abs(closing)

    return (total_opening_dr, total_opening_cr, total_period_dr, total_period_cr,
            total_ending_dr, total_ending_cr)


def create_trial_balance_xlsx(coa, balances, period_name, start_date, end_date, output_path):
    """Create Trial Balance Excel file with all accounts including zeros."""
    wb = create_workbook()

    # Dashboard sheet
    ws = add_sheet(wb, 'Dashboard')

    totals = calculate_balance_totals(balances)
    (total_opening_dr, total_opening_cr, total_period_dr, total_period_cr,
     total_ending_dr, total_ending_cr) = totals

    # Write dashboard
    ws['A1'] = 'TRIAL BALANCE VALIDATION'
    ws['A1'].font = TITLE_FONT
    ws['A3'] = f'Period: {start_date} to {end_date}'
    ws['A5'] = 'Opening Balance Check:'
    ws['B5'] = 'PASS' if abs(total_opening_dr - total_opening_cr) < 0.01 else 'FAIL'
    ws['C5'] = f'Dr: {total_opening_dr:,.2f} | Cr: {total_opening_cr:,.2f}'
    ws['A6'] = 'Period Movements Check:'
    ws['B6'] = 'PASS' if abs(total_period_dr - total_period_cr) < 0.01 else 'FAIL'
    ws['C6'] = f'Dr: {total_period_dr:,.2f} | Cr: {total_period_cr:,.2f}'
    ws['A7'] = 'Ending Balance Check:'
    ws['B7'] = 'PASS' if abs(total_ending_dr - total_ending_cr) < 0.01 else 'FAIL'
    ws['C7'] = f'Dr: {total_ending_dr:,.2f} | Cr: {total_ending_cr:,.2f}'

    # Trial Balance sheet
    ws_tb = add_sheet(wb, 'Trial Balance')

    # Title
    row = write_title(ws_tb, 'Trial Balance', period=f'Period: {start_date} to {end_date}')

    # Headers
    headers = ['Account Code', 'Account Name', 'Opening Balance', 'Debits', 'Credits', 'Ending Balance']
    row = write_header_row(ws_tb, headers, row)

    # Data rows
    for code in sorted(balances.keys()):
        data = balances[code]
        opening = data['opening']
        period_dr = data['period_dr']
        period_cr = data['period_cr']
        closing = data['closing']
        normal = data['normal_balance']

        # Display opening balance (positive on normal side)
        opening_display = opening if opening >= 0 else -opening
        ending_display = closing if closing >= 0 else -closing

        values = [
            code,
            data['name'],
            opening_display if opening_display != 0 else None,
            period_dr if period_dr != 0 else None,
            period_cr if period_cr != 0 else None,
            ending_display if ending_display != 0 else None,
        ]
        row = write_data_row(ws_tb, values, row, number_cols=[3, 4, 5, 6])

    # Total row
    total_values = [
        'TOTAL',
        '',
        total_opening_dr if total_opening_dr >= total_opening_cr else total_opening_cr,
        total_period_dr,
        total_period_cr,
        total_ending_dr if total_ending_dr >= total_ending_cr else total_ending_cr,
    ]
    write_data_row(ws_tb, total_values, row, number_cols=[3, 4, 5, 6], font=TOTAL_FONT, border=THICK_BOTTOM)

    auto_fit_columns(ws_tb, min_width=12, max_width=30)

    wb.save(output_path)
    return total_period_dr, total_period_cr


def calculate_income_statement(balances):
    """
    Calculate income statement values from period movements.
    Returns: (revenue, cogs, opex, other_income, net_profit)
    """
    revenue = 0
    cogs = 0
    opex = 0
    other_income = 0

    for code, data in balances.items():
        # Use period movements for income statement
        period_dr = data.get('period_dr', 0)
        period_cr = data.get('period_cr', 0)

        if code == 40000:  # Sales Revenue - credit balance, revenue = credits
            revenue = period_cr - period_dr
        elif code in [50000, 50010, 50020, 50100, 50110, 50120, 50200, 50220, 53000, 53100, 53200, 53300]:  # COGS - debit balance
            cogs += period_dr - period_cr
        elif code in [60000, 61000, 62000, 63000, 64000, 65000, 66000, 67000, 68000, 69000]:  # SG&A - debit balance
            opex += period_dr - period_cr
        elif code == 70000:  # Interest Income - credit balance
            other_income = period_cr - period_dr
        elif code in [80000, 81000]:  # Other expenses - debit balance
            opex += period_dr - period_cr

    gross_profit = revenue - cogs
    operating_profit = gross_profit - opex
    net_profit = operating_profit + other_income

    return revenue, cogs, opex, other_income, net_profit


def create_financial_statements_xlsx(coa, balances, period_name, start_date, end_date, output_path, accumulated_retained_earnings=0):
    """
    Create Financial Statements Excel file.

    Returns: net_profit (to be accumulated for Retained Earnings)
    """
    wb = create_workbook()

    # Income Statement
    ws_is = add_sheet(wb, 'Income Statement')
    row = write_title(ws_is, 'Income Statement', period=f'For the period {start_date} to {end_date}')

    revenue, cogs, opex, other_income, net_profit = calculate_income_statement(balances)
    gross_profit = revenue - cogs
    operating_profit = gross_profit - opex

    # Write income statement
    ws_is.cell(row=row, column=1, value='Sales Revenue').font = Font(bold=True)
    ws_is.cell(row=row, column=2, value=revenue).number_format = NUMBER_FORMAT_NEG
    row += 2

    ws_is.cell(row=row, column=1, value='Cost of Goods Sold')
    ws_is.cell(row=row, column=2, value=cogs).number_format = NUMBER_FORMAT_NEG
    row += 1

    ws_is.cell(row=row, column=1, value='Gross Profit').font = Font(bold=True)
    ws_is.cell(row=row, column=2, value=gross_profit).number_format = NUMBER_FORMAT_NEG
    ws_is.cell(row=row, column=2).font = Font(bold=True)
    row += 2

    ws_is.cell(row=row, column=1, value='Operating Expenses')
    ws_is.cell(row=row, column=2, value=opex).number_format = NUMBER_FORMAT_NEG
    row += 1

    ws_is.cell(row=row, column=1, value='Operating Profit').font = Font(bold=True)
    ws_is.cell(row=row, column=2, value=operating_profit).number_format = NUMBER_FORMAT_NEG
    ws_is.cell(row=row, column=2).font = Font(bold=True)
    row += 2

    ws_is.cell(row=row, column=1, value='Other Income (Interest)')
    ws_is.cell(row=row, column=2, value=other_income).number_format = NUMBER_FORMAT_NEG
    row += 1

    ws_is.cell(row=row, column=1, value='Net Profit/(Loss)').font = Font(bold=True, size=12)
    ws_is.cell(row=row, column=2, value=net_profit).number_format = NUMBER_FORMAT_NEG
    ws_is.cell(row=row, column=2).font = Font(bold=True, size=12)

    ws_is.column_dimensions['A'].width = 30
    ws_is.column_dimensions['B'].width = 15

    # Balance Sheet
    ws_bs = add_sheet(wb, 'Balance Sheet')
    row = write_title(ws_bs, 'Balance Sheet', period=f'As at {end_date}')

    total_assets = 0
    total_liabilities = 0
    total_equity = 0

    # Assets
    ws_bs.cell(row=row, column=1, value='ASSETS').font = Font(bold=True)
    row += 1

    for code in sorted(balances.keys()):
        data = balances[code]
        if data['type'] == 'Asset':
            closing = data['closing']
            # For contra accounts (accumulated depreciation), show as negative
            if data['normal_balance'] == 'credit':
                closing = -closing
            if closing != 0:
                ws_bs.cell(row=row, column=1, value=data['name'])
                ws_bs.cell(row=row, column=2, value=abs(closing)).number_format = NUMBER_FORMAT_NEG
                total_assets += closing
                row += 1

    ws_bs.cell(row=row, column=1, value='TOTAL ASSETS').font = Font(bold=True)
    ws_bs.cell(row=row, column=2, value=abs(total_assets)).font = Font(bold=True)
    ws_bs.cell(row=row, column=2).number_format = NUMBER_FORMAT_NEG
    row += 2

    # Liabilities
    ws_bs.cell(row=row, column=1, value='LIABILITIES').font = Font(bold=True)
    row += 1

    for code in sorted(balances.keys()):
        data = balances[code]
        if data['type'] == 'Liability':
            closing = data['closing']
            if closing != 0:
                ws_bs.cell(row=row, column=1, value=data['name'])
                ws_bs.cell(row=row, column=2, value=abs(closing)).number_format = NUMBER_FORMAT_NEG
                total_liabilities += closing
                row += 1

    ws_bs.cell(row=row, column=1, value='TOTAL LIABILITIES').font = Font(bold=True)
    ws_bs.cell(row=row, column=2, value=abs(total_liabilities)).font = Font(bold=True)
    ws_bs.cell(row=row, column=2).number_format = NUMBER_FORMAT_NEG
    row += 2

    # Equity
    ws_bs.cell(row=row, column=1, value='EQUITY').font = Font(bold=True)
    row += 1

    for code in sorted(balances.keys()):
        data = balances[code]
        if data['type'] == 'Equity':
            closing = data['closing']
            if closing != 0:
                ws_bs.cell(row=row, column=1, value=data['name'])
                ws_bs.cell(row=row, column=2, value=abs(closing)).number_format = NUMBER_FORMAT_NEG
                total_equity += closing
                row += 1

    # Add Retained Earnings (accumulated net profit/loss)
    retained_earnings = accumulated_retained_earnings + net_profit
    ws_bs.cell(row=row, column=1, value='Retained Earnings')
    ws_bs.cell(row=row, column=2, value=retained_earnings).number_format = NUMBER_FORMAT_NEG
    total_equity += retained_earnings
    row += 1

    ws_bs.cell(row=row, column=1, value='TOTAL EQUITY').font = Font(bold=True)
    ws_bs.cell(row=row, column=2, value=total_equity).font = Font(bold=True)
    ws_bs.cell(row=row, column=2).number_format = NUMBER_FORMAT_NEG
    row += 2

    ws_bs.cell(row=row, column=1, value='TOTAL LIABILITIES & EQUITY').font = Font(bold=True)
    ws_bs.cell(row=row, column=2, value=total_liabilities + total_equity).font = Font(bold=True)
    ws_bs.cell(row=row, column=2).number_format = NUMBER_FORMAT_NEG

    ws_bs.column_dimensions['A'].width = 40
    ws_bs.column_dimensions['B'].width = 15

    wb.save(output_path)

    return net_profit


def main():
    print("="*60)
    print("BATCH PROCESSING WITH BALANCE CHAINING")
    print("Feb 2025 - Oct 2025")
    print("="*60)

    # Load GL once
    print("\nLoading General Ledger...")
    gl_df = pd.read_excel(SOURCE_DIR / 'Ledger Accounts' / 'General_Ledger_edited.xlsx', header=3)
    gl_df = gl_df.dropna(how='all')
    gl_df['Date'] = pd.to_datetime(gl_df['Date'], errors='coerce')
    gl_df = gl_df[gl_df['Date'].notna()]
    print(f"  Total GL rows: {len(gl_df)}")

    # Load COA
    print("Loading Chart of Accounts...")
    coa, coa_opening_balances, coa_mapper = load_coa()
    print(f"  Total accounts: {len(coa)}")
    print(f"  Accounts with opening balances: {len([v for v in coa_opening_balances.values() if v != 0])}")

    # Initialize opening balances from COA (for Feb 2025)
    opening_balances = coa_opening_balances.copy()

    # Print opening balance summary
    total_opening_dr = sum(v for k, v in opening_balances.items() if coa[k]['normal_balance'] == 'debit' and v > 0)
    total_opening_cr = sum(v for k, v in opening_balances.items() if coa[k]['normal_balance'] == 'credit' and v > 0)
    print(f"  Total Opening Debits: {total_opening_dr:,.2f}")
    print(f"  Total Opening Credits: {total_opening_cr:,.2f}")

    # Initialize retained earnings (accumulated net profit/loss)
    retained_earnings = 0.0

    results = []

    for start_date, end_date, period_name in MONTHS:
        print(f"\n{'='*60}")
        print(f"Processing: {period_name}")
        print('='*60)

        output_dir = OUTPUT_DIR / period_name
        output_dir.mkdir(parents=True, exist_ok=True)

        # Extract data with opening balances
        cash_receipts, cash_payments, general_journal, gl_output, balances, ending_balances = extract_period_data(
            gl_df, coa, start_date, end_date, opening_balances, coa_mapper
        )

        print(f"  GL transactions: {len(gl_output)}")
        print(f"  Cash Receipts: {len(cash_receipts)}")
        print(f"  Cash Payments: {len(cash_payments)}")
        print(f"  General Journal: {len(general_journal)}")

        # Run Module 1
        print(f"  Running Module 1...")
        cmd = f'python scripts/summarize_journals.py data/input/journals {start_date} {end_date} data/output/{period_name}/books_of_prime_entry_{period_name}.xlsx data/input/master'
        subprocess.run(cmd, shell=True, capture_output=True)

        # Run Module 2
        print(f"  Running Module 2...")
        cmd = f'python scripts/summarize_ledgers.py data/input/ledgers {start_date} {end_date} data/output/{period_name}/ledger_summary_{period_name}.xlsx data/input/master'
        subprocess.run(cmd, shell=True, capture_output=True)

        # Create Trial Balance directly (not using Module 5 to include all accounts)
        print(f"  Creating Trial Balance...")
        total_dr, total_cr = create_trial_balance_xlsx(
            coa, balances, period_name, start_date, end_date,
            output_dir / f'trial_balance_{period_name}.xlsx'
        )
        print(f"    Period Dr: {total_dr:,.2f} | Period Cr: {total_cr:,.2f}")

        # Create Financial Statements
        print(f"  Creating Financial Statements...")
        net_profit = create_financial_statements_xlsx(
            coa, balances, period_name, start_date, end_date,
            output_dir / f'financial_statements_{period_name}.xlsx',
            accumulated_retained_earnings=retained_earnings
        )
        print(f"    Net Profit: {net_profit:,.2f}")

        # Chain balances to next period
        opening_balances = ending_balances.copy()
        retained_earnings += net_profit  # Accumulate retained earnings
        results.append((period_name, total_dr, total_cr, abs(total_dr - total_cr) < 0.01, net_profit))

    # Summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"{'Period':12} | {'Period Dr':>15} | {'Period Cr':>15} | {'Net Profit':>15} | Balanced")
    print("-"*80)
    for period_name, dr, cr, balanced, np_val in results:
        print(f"{period_name:12} | {dr:>15,.2f} | {cr:>15,.2f} | {np_val:>15,.2f} | {'YES' if balanced else 'NO'}")


if __name__ == '__main__':
    main()