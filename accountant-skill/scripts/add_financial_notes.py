"""
Add Financial Notes Sheet to Existing Financial Statements
Reads existing Income Statement and Balance Sheet data and creates a Financial Notes sheet.

Usage:
    python scripts/add_financial_notes.py <financial_statements_file>

Example:
    python scripts/add_financial_notes.py data/output/Feb2025/financial_statements_Feb2025.xlsx
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Styling constants
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
TITLE_FONT = Font(bold=True, size=14, name='Arial')
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
SECTION_FONT = Font(bold=True, size=11, name='Arial', color='1F4E79')
NORMAL_FONT = Font(size=11, name='Arial')
TOTAL_FONT = Font(bold=True, size=11, name='Arial')
NEGATIVE_FONT = Font(size=11, name='Arial', color='FF0000')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
BOTTOM_BORDER = Border(bottom=Side(style='medium'))
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))

NUMBER_FORMAT_NEG = '#,##0;(#,##0);"-"'


def _n(val):
    """Return float or None; treat zero/NaN as None for display."""
    if val is None:
        return None
    try:
        v = float(val)
    except (ValueError, TypeError):
        return None
    if abs(v) < 0.005:
        return None
    return v


# Map account names to codes and classifications
ACCOUNT_NAME_MAP = {
    # Current Assets
    'cash in hand': {'code': '10000', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Cash & Equivalents'},
    'cash at bank': {'code': '10100', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Cash & Equivalents'},
    'accounts receivable': {'code': '11000', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Accounts Receivable'},
    'inventory - raw material': {'code': '12000', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Inventory'},
    'inventory - packaging': {'code': '12100', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Inventory'},
    'inventory - finished good': {'code': '12200', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Inventory'},
    'inventory - finished goods': {'code': '12200', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Inventory'},
    'inventory adjustments': {'code': '12300', 'type': 'Asset', 'sub_type': 'Current Asset (Contra)', 'group': 'Inventory'},
    'work-in-progress': {'code': '12400', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Inventory'},
    'advanced payments': {'code': '13000', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Prepayments'},
    'deferred preliminary expenses': {'code': '14000', 'type': 'Asset', 'sub_type': 'Current Asset', 'group': 'Prepayments'},
    # Non-Current Assets
    'land': {'code': '15000', 'type': 'Asset', 'sub_type': 'Non-Current Asset', 'group': 'Land'},
    'buildings & structures': {'code': '15100', 'type': 'Asset', 'sub_type': 'Non-Current Asset', 'group': 'Buildings & Structures'},
    'accumulated depreciation - buildings': {'code': '15110', 'type': 'Asset', 'sub_type': 'Non-Current Asset (Contra)', 'group': 'Accumulated Depreciation'},
    'machinery & equipment': {'code': '15200', 'type': 'Asset', 'sub_type': 'Non-Current Asset', 'group': 'Machinery & Equipment'},
    'accumulated depreciation - machinery': {'code': '15210', 'type': 'Asset', 'sub_type': 'Non-Current Asset (Contra)', 'group': 'Accumulated Depreciation'},
    'office & facility equipment': {'code': '15300', 'type': 'Asset', 'sub_type': 'Non-Current Asset', 'group': 'Office & Facility Equipment'},
    'accumulated depreciation - office': {'code': '15310', 'type': 'Asset', 'sub_type': 'Non-Current Asset (Contra)', 'group': 'Accumulated Depreciation'},
    'electrical & utility systems': {'code': '15400', 'type': 'Asset', 'sub_type': 'Non-Current Asset', 'group': 'Electrical & Utility Systems'},
    'accumulated depreciation - electrical': {'code': '15410', 'type': 'Asset', 'sub_type': 'Non-Current Asset (Contra)', 'group': 'Accumulated Depreciation'},
    'construction in progress': {'code': '15500', 'type': 'Asset', 'sub_type': 'Non-Current Asset', 'group': 'Construction in Progress'},
    'motor vehicles': {'code': '15600', 'type': 'Asset', 'sub_type': 'Non-Current Asset', 'group': 'Motor Vehicles'},
    'accumulated depreciation - vehicles': {'code': '15510', 'type': 'Asset', 'sub_type': 'Non-Current Asset (Contra)', 'group': 'Accumulated Depreciation'},
    # Liabilities
    'accounts payable': {'code': '20000', 'type': 'Liability', 'sub_type': 'Current Liability', 'group': 'Accounts Payable'},
    'short-term loans': {'code': '21000', 'type': 'Liability', 'sub_type': 'Current Liability', 'group': 'Short-term Loans'},
    'utility bills': {'code': '22000', 'type': 'Liability', 'sub_type': 'Current Liability', 'group': 'Accrued Expenses'},
    'wages payable': {'code': '22200', 'type': 'Liability', 'sub_type': 'Current Liability', 'group': 'Accrued Expenses'},
    'bank loan': {'code': '25000', 'type': 'Liability', 'sub_type': 'Non-Current Liability', 'group': 'Long-term Loans'},
    # Equity
    'paid-up capital': {'code': '31000', 'type': 'Equity', 'sub_type': 'Equity', 'group': 'Paid-up Capital'},
    'retained earnings': {'code': '32000', 'type': 'Equity', 'sub_type': 'Equity', 'group': 'Retained Earnings'},
    # Revenue
    'sales revenue': {'code': '40000', 'type': 'Revenue', 'sub_type': 'Operating Revenue', 'group': 'Sales Revenue'},
    # COGS
    'cost of goods sold': {'code': '50000', 'type': 'Expense', 'sub_type': 'COGS', 'group': 'Cost of Goods Sold'},
    # Operating Expenses
    'marketing & advertising': {'code': '60000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Marketing & Advertising'},
    'office salaries': {'code': '61000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Office Salaries'},
    'meal allowance': {'code': '62000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Employee Benefits'},
    'utilities': {'code': '63000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Utilities'},
    'transportation & distribution': {'code': '64000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Transportation'},
    'factory buildings & office supplies': {'code': '65000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Facility & Office Supplies'},
    'depreciation expenses - sg&a': {'code': '66000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Depreciation'},
    'inventory write-off': {'code': '67000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Inventory Write-offs'},
    'other expenses': {'code': '68000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Other Expenses'},
    'key management compensation': {'code': '69000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Management Compensation'},
    'operating expenses': {'code': '60000', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Operating Expenses'},
    # Other Income
    'interest income': {'code': '70000', 'type': 'Revenue', 'sub_type': 'Other Income', 'group': 'Interest Income'},
    'other income': {'code': '70000', 'type': 'Revenue', 'sub_type': 'Other Income', 'group': 'Other Income'},
}


def get_account_info(name):
    """Get account info by name from the mapping."""
    name_lower = name.strip().lower()
    # Direct match
    if name_lower in ACCOUNT_NAME_MAP:
        return ACCOUNT_NAME_MAP[name_lower]
    # Partial match
    for key, info in ACCOUNT_NAME_MAP.items():
        if key in name_lower or name_lower in key:
            return info
    # Default classification based on keywords
    if 'accumulated depreciation' in name_lower:
        return {'code': '15XXX', 'type': 'Asset', 'sub_type': 'Non-Current Asset (Contra)', 'group': 'Accumulated Depreciation'}
    if 'depreciation' in name_lower:
        return {'code': '66XXX', 'type': 'Expense', 'sub_type': 'Operating Expense', 'group': 'Depreciation'}
    return None


def read_simple_format_sheet(ws):
    """
    Read accounts from a simple 2-column format (Name, Amount).
    Returns list of dicts: {code, name, amount, type, sub_type, group}
    """
    accounts = []
    current_section = None

    for row in range(1, ws.max_row + 1):
        name_val = ws.cell(row=row, column=1).value
        amount_val = ws.cell(row=row, column=2).value

        if not name_val:
            continue

        name_str = str(name_val).strip()
        name_lower = name_str.lower()

        # Skip headers and totals
        if name_lower in ['income statement', 'balance sheet', '']:
            continue
        if 'for the period' in name_lower or 'as at' in name_lower:
            continue
        if name_lower.startswith('total ') and not name_lower == 'total':
            continue

        # Check if it's a section header (no amount)
        if amount_val is None or amount_val == '':
            if any(kw in name_lower for kw in ['assets', 'liabilities', 'equity', 'revenue', 'cogs', 'expense']):
                current_section = name_str
            continue

        # Try to get amount
        try:
            amount = float(amount_val)
        except (ValueError, TypeError):
            continue

        # Get account info
        info = get_account_info(name_str)
        if info:
            accounts.append({
                'code': info['code'],
                'name': name_str,
                'amount': amount,
                'type': info['type'],
                'sub_type': info['sub_type'],
                'group': info['group']
            })
        else:
            # Unknown account - try to classify by keywords
            if any(kw in name_lower for kw in ['cash', 'bank']):
                type_, sub_type, group = 'Asset', 'Current Asset', 'Cash & Equivalents'
            elif 'inventory' in name_lower:
                type_, sub_type, group = 'Asset', 'Current Asset', 'Inventory'
            elif 'receivable' in name_lower:
                type_, sub_type, group = 'Asset', 'Current Asset', 'Accounts Receivable'
            elif 'payable' in name_lower:
                type_, sub_type, group = 'Liability', 'Current Liability', 'Accounts Payable'
            elif 'capital' in name_lower:
                type_, sub_type, group = 'Equity', 'Equity', 'Paid-up Capital'
            elif 'retained' in name_lower:
                type_, sub_type, group = 'Equity', 'Equity', 'Retained Earnings'
            elif 'revenue' in name_lower or 'sales' in name_lower:
                type_, sub_type, group = 'Revenue', 'Operating Revenue', 'Sales Revenue'
            elif 'profit' in name_lower:
                type_, sub_type, group = 'Equity', 'Equity', 'Retained Earnings'
            else:
                type_, sub_type, group = 'Other', 'Other', 'Other'

            accounts.append({
                'code': 'XXXXX',
                'name': name_str,
                'amount': amount,
                'type': type_,
                'sub_type': sub_type,
                'group': group
            })

    return accounts


def write_financial_notes_sheet(wb, accounts, period_str):
    """Create the Financial Notes sheet from collected accounts."""
    # Remove existing Financial Notes sheet if present
    if 'Financial Notes' in wb.sheetnames:
        del wb['Financial Notes']

    ws = wb.create_sheet('Financial Notes')

    # Title
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value='Financial Notes').font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value='Detailed Breakdown of Financial Statement Accounts').font = Font(bold=True, size=12, name='Arial')
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=period_str).font = Font(italic=True, size=11, name='Arial')
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2

    # Organize accounts by type and sub_type
    structure = {}
    for acct in accounts:
        type_ = acct['type']
        sub_type = acct['sub_type']
        group = acct['group']

        if type_ not in structure:
            structure[type_] = {}
        if sub_type not in structure[type_]:
            structure[type_][sub_type] = {}
        if group not in structure[type_][sub_type]:
            structure[type_][sub_type][group] = []

        structure[type_][sub_type][group].append(acct)

    # Write sections
    type_order = ['Asset', 'Liability', 'Equity', 'Revenue', 'Expense']
    section_num = 0

    for type_ in type_order:
        if type_ not in structure:
            continue

        section_num += 1

        # Section header
        if type_ == 'Asset':
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value='BALANCE SHEET NOTES').font = SECTION_FONT
            ws.cell(row=row, column=1).fill = SECTION_FILL
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        if type_ == 'Revenue':
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value='INCOME STATEMENT NOTES').font = SECTION_FONT
            ws.cell(row=row, column=1).fill = SECTION_FILL
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        # Type header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=f'{section_num}. {type_.upper()}').font = Font(bold=True, size=11, name='Arial')
        row += 1

        type_total = 0.0

        for sub_type in sorted(structure[type_].keys()):
            groups = structure[type_][sub_type]
            sub_type_total = 0.0

            # Sub-type header
            ws.cell(row=row, column=2, value=sub_type).font = Font(bold=True, italic=True, size=11, name='Arial')
            row += 1

            for group_name in sorted(groups.keys()):
                group_accounts = groups[group_name]
                group_accounts.sort(key=lambda x: x['code'])

                for acct in group_accounts:
                    ws.cell(row=row, column=1, value=acct['code']).font = NORMAL_FONT
                    ws.cell(row=row, column=1).border = THIN_BORDER
                    ws.cell(row=row, column=2, value=f'  {acct["name"]}').font = NORMAL_FONT
                    ws.cell(row=row, column=2).border = THIN_BORDER

                    amount = acct['amount']
                    cell = ws.cell(row=row, column=3, value=_n(abs(amount)))
                    cell.font = NORMAL_FONT
                    cell.border = THIN_BORDER
                    cell.number_format = NUMBER_FORMAT_NEG
                    cell.alignment = Alignment(horizontal='right')

                    sub_type_total += abs(amount)
                    row += 1

            # Sub-type total
            if len(groups) > 0:
                ws.cell(row=row, column=2, value=f'Total {sub_type}').font = TOTAL_FONT
                ws.cell(row=row, column=2).border = BOTTOM_BORDER
                total_cell = ws.cell(row=row, column=3, value=_n(sub_type_total))
                total_cell.font = TOTAL_FONT
                total_cell.border = BOTTOM_BORDER
                total_cell.number_format = NUMBER_FORMAT_NEG
                total_cell.alignment = Alignment(horizontal='right')
                row += 1

            type_total += sub_type_total

        # Type total
        ws.cell(row=row, column=2, value=f'TOTAL {type_.upper()}').font = TOTAL_FONT
        ws.cell(row=row, column=2).border = DOUBLE_BOTTOM
        total_cell = ws.cell(row=row, column=3, value=_n(type_total))
        total_cell.font = TOTAL_FONT
        total_cell.border = DOUBLE_BOTTOM
        total_cell.number_format = NUMBER_FORMAT_NEG
        total_cell.alignment = Alignment(horizontal='right')
        row += 2

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # Freeze panes
    ws.freeze_panes = 'B5'

    # Set tab color
    ws.sheet_properties.tabColor = '70AD47'


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    filepath = Path(sys.argv[1])
    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    print(f"Processing: {filepath.name}")

    # Load workbook
    wb = load_workbook(filepath)

    # Read accounts from Income Statement and Balance Sheet
    accounts = []

    if 'Income Statement' in wb.sheetnames:
        ws_is = wb['Income Statement']
        is_accounts = read_simple_format_sheet(ws_is)
        accounts.extend(is_accounts)
        print(f"  Income Statement: {len(is_accounts)} accounts")

    if 'Balance Sheet' in wb.sheetnames:
        ws_bs = wb['Balance Sheet']
        bs_accounts = read_simple_format_sheet(ws_bs)
        accounts.extend(bs_accounts)
        print(f"  Balance Sheet: {len(bs_accounts)} accounts")

    if not accounts:
        print("ERROR: No account data found in Income Statement or Balance Sheet")
        sys.exit(1)

    print(f"  Total accounts: {len(accounts)}")

    # Create Financial Notes sheet
    period_str = filepath.stem.replace('financial_statements_', '')
    write_financial_notes_sheet(wb, accounts, period_str)

    # Save workbook
    wb.save(filepath)
    print(f"  Saved with Financial Notes sheet")


if __name__ == '__main__':
    main()