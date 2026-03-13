"""
Add Financial Notes Sheet to Financial Statements (Reference Format)
Creates Financial Notes sheet matching the reference file structure.

Usage:
    python scripts/add_financial_notes_v2.py <financial_statements_file>

Example:
    python scripts/add_financial_notes_v2.py data/output/Feb2025/financial_statements_Feb2025.xlsx
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styling constants
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
TITLE_FONT = Font(bold=True, size=14, name='Arial')
SECTION_FONT = Font(bold=True, size=11, name='Arial')
NORMAL_FONT = Font(size=11, name='Arial')
TOTAL_FONT = Font(bold=True, size=11, name='Arial')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
BOTTOM_BORDER = Border(bottom=Side(style='medium'))
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))

NUMBER_FORMAT = '#,##0;(#,##0);"-"'

# Column indices
COL_NOTE_NUM = 2      # Column B
COL_ITEM_NAME = 3     # Column C
COL_AMOUNT = 4        # Column D


def read_account_data(ws):
    """
    Read accounts from Income Statement or Balance Sheet.
    Returns dict: {account_name: amount}
    """
    accounts = {}
    for row in range(1, ws.max_row + 1):
        name_val = ws.cell(row=row, column=1).value
        amount_val = ws.cell(row=row, column=2).value

        if not name_val:
            continue

        name_str = str(name_val).strip()

        # Skip headers and totals
        if name_str.lower() in ['income statement', 'balance sheet', '']:
            continue
        if 'for the period' in name_str.lower() or 'as at' in name_str.lower():
            continue
        if name_str.lower().startswith('total ') or name_str.lower() == 'total':
            continue
        if name_str.lower().startswith('gross ') or name_str.lower().startswith('operating profit'):
            continue
        if name_str.lower().startswith('net '):
            continue
        if name_str.upper() in ['ASSETS', 'LIABILITIES', 'EQUITY']:
            continue

        # Try to get amount
        try:
            amount = float(amount_val) if amount_val is not None else 0
        except (ValueError, TypeError):
            continue

        accounts[name_str] = amount

    return accounts


def get_amount(accounts_dict, *names):
    """Get account amount by searching for matching names."""
    for name in names:
        # Try exact match
        if name in accounts_dict:
            return accounts_dict[name]
        # Try case-insensitive match
        for acc_name, amt in accounts_dict.items():
            if acc_name.lower().strip() == name.lower().strip():
                return amt
    return 0


def write_financial_notes(wb, is_accounts, bs_accounts, period_str):
    """
    Create Financial Notes sheet with reference format structure.
    """
    # Remove existing Financial Notes sheet if present
    if 'Financial Notes' in wb.sheetnames:
        del wb['Financial Notes']

    ws = wb.create_sheet('Financial Notes')
    row = 1

    # Merge all accounts for easier lookup
    all_accounts = {**is_accounts, **bs_accounts}

    # ========== NOTE 1: REVENUE ==========
    row += 1
    ws.cell(row=row, column=COL_NOTE_NUM, value='1').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Revenue').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    sales_revenue = get_amount(all_accounts, 'Sales Revenue')
    ws.cell(row=row, column=COL_ITEM_NAME, value='Sales Revenue').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=sales_revenue).number_format = NUMBER_FORMAT
    sales_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'=SUM({get_column_letter(COL_AMOUNT)}{sales_row})').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 2: COGS ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='2').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='COGS').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    cogs = get_amount(all_accounts, 'Cost of Goods Sold')
    ws.cell(row=row, column=COL_ITEM_NAME, value='Cost of Goods Sold').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=cogs).number_format = NUMBER_FORMAT
    cogs_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'=SUM({get_column_letter(COL_AMOUNT)}{cogs_row})').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 3: SG&A ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='3').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='SG&A').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    sga_items = [
        ('Marketing & Advertising Expense', 'Marketing & Advertising'),
        ('Office Salaries', 'Office Salaries'),
        ('Meal Allowance Expense', 'Meal Allowance'),
        ('Utilities (Electricity & Water)', 'Utilities'),
        ('Transportation & Distribution Expense', 'Transportation & Distribution'),
        ('Factory Buildings & Office Supplies', 'Factory Buildings & Office Supplies'),
        ('Depreciation Expenses - SG&A', 'Depreciation Expenses - SG&A'),
        ('Inventory Write-off', 'Inventory Write-off'),
        ('Other Expenses', 'Other Expenses'),
        ('Key Management Compensation', 'Key Management Compensation'),
    ]

    sga_start_row = row
    sga_has_items = False

    for display_name, search_name in sga_items:
        amt = get_amount(all_accounts, display_name, search_name)
        if amt != 0:  # Only show if non-zero
            ws.cell(row=row, column=COL_ITEM_NAME, value=display_name).font = NORMAL_FONT
            ws.cell(row=row, column=COL_AMOUNT, value=amt).number_format = NUMBER_FORMAT
            row += 1
            sga_has_items = True

    # If no detailed items found, show single "Operating Expenses" line
    if not sga_has_items:
        op_exp = get_amount(all_accounts, 'Operating Expenses')
        ws.cell(row=row, column=COL_ITEM_NAME, value='Operating Expenses').font = NORMAL_FONT
        ws.cell(row=row, column=COL_AMOUNT, value=op_exp).number_format = NUMBER_FORMAT
        row += 1

    ws.cell(row=row, column=COL_AMOUNT, value=f'=SUM({get_column_letter(COL_AMOUNT)}{sga_start_row}:{get_column_letter(COL_AMOUNT)}{row-1})').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 4: DEPRECIATION & AMORTIZATION ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='4').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Depreciation & Amortization').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    dep_cogs = get_amount(all_accounts, 'Depreciation Expenses - COGS')
    dep_sga = get_amount(all_accounts, 'Depreciation Expenses - SG&A')

    ws.cell(row=row, column=COL_ITEM_NAME, value='Depreciation Expenses - COGS').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=dep_cogs).number_format = NUMBER_FORMAT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value='Depreciation Expenses - SG&A').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=dep_sga).number_format = NUMBER_FORMAT
    dep_end_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'=SUM({get_column_letter(COL_AMOUNT)}{dep_end_row-1}:{get_column_letter(COL_AMOUNT)}{dep_end_row})').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 5: OTHER INCOME ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='5').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Other Income').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    interest_income = get_amount(all_accounts, 'Interest Income', 'Other Income (Interest)', 'Other Income')
    ws.cell(row=row, column=COL_ITEM_NAME, value='Interest Income').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=interest_income).number_format = NUMBER_FORMAT
    interest_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'=SUM({get_column_letter(COL_AMOUNT)}{interest_row})').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 11: CASH ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='11').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Cash').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    cash_in_hand = get_amount(all_accounts, 'Cash in hand', 'Cash in Hand')
    cash_at_bank = get_amount(all_accounts, 'Cash at Bank', 'Cash at bank')

    ws.cell(row=row, column=COL_ITEM_NAME, value='Cash in hand').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=cash_in_hand).number_format = NUMBER_FORMAT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value='Cash at Bank').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=cash_at_bank).number_format = NUMBER_FORMAT
    cash_end_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'=SUM({get_column_letter(COL_AMOUNT)}{cash_end_row-1}:{get_column_letter(COL_AMOUNT)}{cash_end_row})').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 12: ACCOUNTS RECEIVABLE ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='12').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Accounts Receivable').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    ar = get_amount(all_accounts, 'Accounts Receivable', 'Accounts receivable')
    ws.cell(row=row, column=COL_ITEM_NAME, value='Accounts Receivable').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=ar).number_format = NUMBER_FORMAT
    ar_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'={get_column_letter(COL_AMOUNT)}{ar_row}').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 13: INVENTORY ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='13').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Inventory').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    inv_items = [
        ('Inventory - Raw Material', 'Inventory - Raw Material'),
        ('Inventory - Packaging', 'Inventory - Packaging'),
        ('Inventory - Work-in-progress', 'Work-in-progress', 'Inventory - Work-in-progress'),
        ('Inventory - Finished Goods', 'Inventory - Finished Goods', 'Inventory - Finished Good'),
        ('Inventory Adjustments', 'Inventory Adjustments'),
    ]

    inv_start_row = row
    for display_name, *search_names in inv_items:
        amt = get_amount(all_accounts, *search_names)
        if amt != 0:  # Only show if non-zero
            ws.cell(row=row, column=COL_ITEM_NAME, value=display_name).font = NORMAL_FONT
            ws.cell(row=row, column=COL_AMOUNT, value=amt).number_format = NUMBER_FORMAT
            row += 1

    if row > inv_start_row:
        ws.cell(row=row, column=COL_AMOUNT, value=f'=SUM({get_column_letter(COL_AMOUNT)}{inv_start_row}:{get_column_letter(COL_AMOUNT)}{row-1})').font = TOTAL_FONT
        ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 14: ADVANCE PAYMENTS ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='14').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Advance Payments').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    adv_payments = get_amount(all_accounts, 'Advanced Payments', 'Advance Payments')
    ws.cell(row=row, column=COL_ITEM_NAME, value='Advanced Payments').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=adv_payments).number_format = NUMBER_FORMAT
    adv_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'={get_column_letter(COL_AMOUNT)}{adv_row}').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 15: DEFERRED PRELIMINARY EXPENSES ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='15').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Deferred Preliminary Expenses').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    deferred = get_amount(all_accounts, 'Deferred Preliminary Expenses')
    ws.cell(row=row, column=COL_ITEM_NAME, value='Deferred Preliminary Expenses').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=deferred).number_format = NUMBER_FORMAT
    deferred_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'={get_column_letter(COL_AMOUNT)}{deferred_row}').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 16: PROPERTY, PLANT AND EQUIPMENT ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='16').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Property, Plant and Equipment').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    ppe_start_row = row

    # Assets (positive)
    ppe_assets = [
        ('Land', 'Land'),
        ('Buildings & Structures', 'Buildings & Structures'),
        ('Machinery & Equipment', 'Machinery & Equipment'),
        ('Office & Facility Equipment', 'Office & Facility Equipment'),
        ('Electrical & Utility Systems', 'Electrical & Utility Systems'),
        ('Construction in Progress', 'Construction in Progress'),
        ('Motor Vehicles', 'Motor Vehicles'),
    ]

    for display_name, search_name in ppe_assets:
        amt = get_amount(all_accounts, search_name, display_name)
        if amt != 0:  # Only show if non-zero
            ws.cell(row=row, column=COL_ITEM_NAME, value=display_name).font = NORMAL_FONT
            ws.cell(row=row, column=COL_AMOUNT, value=amt).number_format = NUMBER_FORMAT
            row += 1

    # Accumulated Depreciation (negative values)
    acc_dep_items = [
        ('Accumulated Depreciation - Buildings & Structures', 'Accumulated Depreciation - Buildings'),
        ('Accumulated Depreciation - Machinery & Equipment', 'Accumulated Depreciation - Machinery'),
        ('Accumulated Depreciation - Office & Facility Equipment', 'Accumulated Depreciation - Office'),
        ('Accumulated Depreciation - Electrical & Utility Systems', 'Accumulated Depreciation - Electrical'),
        ('Accumulated Depreciation - Motor Vehicles', 'Accumulated Depreciation - Vehicles'),
    ]

    for display_name, search_name in acc_dep_items:
        amt = get_amount(all_accounts, search_name, display_name)
        if amt != 0:  # Only show if non-zero
            ws.cell(row=row, column=COL_ITEM_NAME, value=display_name).font = NORMAL_FONT
            # Display accumulated depreciation as negative
            ws.cell(row=row, column=COL_AMOUNT, value=-amt if amt > 0 else amt).number_format = NUMBER_FORMAT
            row += 1

    if row > ppe_start_row:
        ws.cell(row=row, column=COL_AMOUNT, value=f'=SUM({get_column_letter(COL_AMOUNT)}{ppe_start_row}:{get_column_letter(COL_AMOUNT)}{row-1})').font = TOTAL_FONT
        ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 17: PAID-UP CAPITAL ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='17').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Paid-up Capital').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    capital = get_amount(all_accounts, 'Paid-up Capital', 'Paid up Capital')
    ws.cell(row=row, column=COL_ITEM_NAME, value='Paid-up Capital').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=capital).number_format = NUMBER_FORMAT
    capital_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'={get_column_letter(COL_AMOUNT)}{capital_row}').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 18: RETAINED EARNINGS ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='18').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Retained Earnings').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    retained = get_amount(all_accounts, 'Retained Earnings')
    ws.cell(row=row, column=COL_ITEM_NAME, value='Retained Earnings').font = NORMAL_FONT
    ws.cell(row=row, column=COL_AMOUNT, value=retained).number_format = NUMBER_FORMAT
    retained_row = row
    row += 1
    ws.cell(row=row, column=COL_AMOUNT, value=f'={get_column_letter(COL_AMOUNT)}{retained_row}').font = TOTAL_FONT
    ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # ========== NOTE 19: LIABILITIES ==========
    ws.cell(row=row, column=COL_NOTE_NUM, value='19').font = SECTION_FONT
    ws.cell(row=row, column=COL_ITEM_NAME, value='Liabilities').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=COL_ITEM_NAME, value=period_str).font = NORMAL_FONT
    row += 1

    liab_items = [
        ('Accounts Payable', 'Accounts Payable'),
        ('Short-term Loans', 'Short-term Loans'),
        ('Utility Bills', 'Utility Bills'),
        ('Wages Payable', 'Wages Payable'),
        ('Bank Loan', 'Bank Loan'),
    ]

    liab_start_row = row
    for display_name, search_name in liab_items:
        amt = get_amount(all_accounts, search_name, display_name)
        ws.cell(row=row, column=COL_ITEM_NAME, value=display_name).font = NORMAL_FONT
        ws.cell(row=row, column=COL_AMOUNT, value=amt).number_format = NUMBER_FORMAT
        row += 1

    if row > liab_start_row:
        ws.cell(row=row, column=COL_AMOUNT, value=f'=SUM({get_column_letter(COL_AMOUNT)}{liab_start_row}:{get_column_letter(COL_AMOUNT)}{row-1})').font = TOTAL_FONT
        ws.cell(row=row, column=COL_AMOUNT).number_format = NUMBER_FORMAT
    row += 2

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 18

    # Tab color
    ws.sheet_properties.tabColor = '70AD47'


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    filepath = Path(sys.argv[1])
    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    print(f"Processing: {filepath.name}")

    # Load workbook
    wb = load_workbook(filepath)

    # Read accounts from Income Statement and Balance Sheet
    is_accounts = {}
    bs_accounts = {}

    if 'Income Statement' in wb.sheetnames:
        ws_is = wb['Income Statement']
        is_accounts = read_account_data(ws_is)
        print(f"  Income Statement: {len(is_accounts)} accounts")

    if 'Balance Sheet' in wb.sheetnames:
        ws_bs = wb['Balance Sheet']
        bs_accounts = read_account_data(ws_bs)
        print(f"  Balance Sheet: {len(bs_accounts)} accounts")

    if not is_accounts and not bs_accounts:
        print("ERROR: No account data found")
        sys.exit(1)

    # Create Financial Notes sheet
    period_str = filepath.stem.replace('financial_statements_', '')
    write_financial_notes(wb, is_accounts, bs_accounts, period_str)

    # Save workbook
    wb.save(filepath)
    print(f"  Saved with Financial Notes sheet")


if __name__ == '__main__':
    main()