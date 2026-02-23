"""
Module 1: Summarize Books of Prime Entry
Reads all 6 journals for a period and produces a consolidated summary.

Usage:
    python summarize_journals.py <input_dir> <period_start> <period_end> <output_file> [coa_file] [pcc_file]

Example:
    python summarize_journals.py data/Jan2026 2026-01-01 2026-01-31 data/Jan2026/books_of_prime_entry_Jan2026.xlsx data/Jan2026/chart_of_accounts.xlsx data/Jan2026/profit_cost_centers.xlsx
"""
import sys
import os
import pandas as pd
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))
from utils.excel_reader import read_xlsx, filter_by_period
from utils.excel_writer import (create_workbook, add_sheet, write_title, write_header_row,
                                 write_data_row, write_section_header, write_total_row,
                                 auto_fit_columns, freeze_panes, save_workbook, TOTAL_FONT,
                                 PASS_FILL, FAIL_FILL, write_validation_result)
from utils.double_entry import validate_journal_balance
from utils.coa_mapper import COAMapper
from utils.pc_cc_mapper import PCCCMapper


PC_CC_COLS = ['Profit Center', 'Cost Center']

JOURNAL_CONFIGS = {
    'Sales Journal': {
        'filename_patterns': ['sales_journal', 'sales journal'],
        'required': ['Date', 'Debit Account', 'Credit Account'],
        'optional': ['Invoice No', 'Customer', 'Description', 'Amount', 'Debit Amount', 'Credit Amount'] + PC_CC_COLS,
        'reference_col': 'Invoice No',
    },
    'Purchases Journal': {
        'filename_patterns': ['purchases_journal', 'purchases journal', 'purchase_journal'],
        'required': ['Date', 'Debit Account', 'Credit Account'],
        'optional': ['Reference', 'Supplier', 'Description', 'Amount', 'Debit Amount', 'Credit Amount'] + PC_CC_COLS,
        'reference_col': 'Reference',
    },
    'Cash Receipts Journal': {
        'filename_patterns': ['cash_receipts_journal', 'cash receipts journal', 'cash_receipts'],
        'required': ['Date', 'Debit Account', 'Credit Account'],
        'optional': ['Receipt No', 'Received From', 'Description', 'Amount', 'Debit Amount', 'Credit Amount', 'Bank Account'] + PC_CC_COLS,
        'reference_col': 'Receipt No',
    },
    'Cash Payments Journal': {
        'filename_patterns': ['cash_payments_journal', 'cash payments journal', 'cash_payments'],
        'required': ['Date', 'Debit Account', 'Credit Account'],
        'optional': ['Payment No', 'Paid To', 'Description', 'Amount', 'Debit Amount', 'Credit Amount', 'Bank Account'] + PC_CC_COLS,
        'reference_col': 'Payment No',
    },
    'Payroll Journal': {
        'filename_patterns': ['payroll_journal', 'payroll journal', 'payroll'],
        'required': ['Date', 'Debit Account', 'Credit Account'],
        'optional': ['Employee / Department', 'Description', 'Debit Amount', 'Credit Amount', 'Amount'] + PC_CC_COLS,
        'reference_col': 'Employee / Department',
    },
    'General Journal': {
        'filename_patterns': ['general_journal', 'general journal'],
        'required': ['Date', 'Debit Account', 'Credit Account'],
        'optional': ['JV No', 'Description', 'Debit Amount', 'Credit Amount', 'Amount'] + PC_CC_COLS,
        'reference_col': 'JV No',
    },
}


def find_journal_file(input_dir, patterns):
    """Find a journal file in the input directory."""
    input_dir = Path(input_dir)
    for f in input_dir.glob('*.xlsx'):
        fname = f.stem.lower().replace('-', '_').replace(' ', '_')
        for pattern in patterns:
            if pattern.replace(' ', '_') in fname:
                return f
    return None


def get_amounts(df):
    """Extract debit and credit amounts from a journal DataFrame."""
    if 'Debit Amount' in df.columns and 'Credit Amount' in df.columns:
        df['_debit'] = pd.to_numeric(df['Debit Amount'], errors='coerce').fillna(0)
        df['_credit'] = pd.to_numeric(df['Credit Amount'], errors='coerce').fillna(0)
    elif 'Amount' in df.columns:
        df['_debit'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)
        df['_credit'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)
    else:
        df['_debit'] = 0
        df['_credit'] = 0
    return df


def summarize_single_journal(df, journal_name):
    """Summarize a single journal's transactions by account."""
    df = get_amounts(df)
    
    # Summarize by debit account
    debit_summary = df.groupby('Debit Account').agg(
        debit_total=('_debit', 'sum'),
        debit_count=('_debit', 'count')
    ).reset_index()
    debit_summary.rename(columns={'Debit Account': 'Account Code'}, inplace=True)
    
    # Summarize by credit account
    credit_summary = df.groupby('Credit Account').agg(
        credit_total=('_credit', 'sum'),
        credit_count=('_credit', 'count')
    ).reset_index()
    credit_summary.rename(columns={'Credit Account': 'Account Code'}, inplace=True)
    
    # Merge
    summary = pd.merge(debit_summary, credit_summary, on='Account Code', how='outer').fillna(0)
    summary['Account Code'] = summary['Account Code'].astype(int)
    summary = summary.sort_values('Account Code')
    
    return {
        'summary': summary,
        'total_debits': df['_debit'].sum(),
        'total_credits': df['_credit'].sum(),
        'transaction_count': len(df),
        'journal_name': journal_name,
    }


def write_pc_summary_sheet(wb, pc_summary, cc_summary, period_start, period_end, pcc=None):
    """Write the Profit Center P&L summary sheet."""
    from openpyxl.styles import PatternFill, Font
    ws = add_sheet(wb, 'PC Summary', tab_color='70AD47')
    row = write_title(ws, 'Profit Center Summary — P&L View', period=f"{period_start} to {period_end}")

    profit_centers = pcc.profit_centers if pcc else {}
    pc_codes = list(profit_centers.keys())
    pc_names = [profit_centers[p] for p in pc_codes]

    # Header: PC names as columns
    row = write_header_row(ws, ['Category'] + pc_names + ['TOTAL'], row)

    def write_pl_row(ws, label, values, row, is_total=False, indent=False):
        prefix = '  ' if indent else ''
        ws.cell(row=row, column=1, value=prefix + label)
        total = sum(v for v in values if isinstance(v, (int, float)))
        all_vals = values + [total]
        for col_i, val in enumerate(all_vals, 2):
            cell = ws.cell(row=row, column=col_i, value=val)
            cell.number_format = '#,##0;(#,##0);"-"'
            from openpyxl.styles import Alignment
            cell.alignment = Alignment(horizontal='right')
            if is_total:
                cell.font = TOTAL_FONT
            if isinstance(val, (int, float)) and val < 0:
                from openpyxl.styles import Font as F
                cell.font = F(bold=is_total, color='FF0000', size=11, name='Arial')
        if is_total:
            ws.cell(row=row, column=1).font = TOTAL_FONT
        return row + 1

    # Revenue
    row = write_section_header(ws, 'REVENUE', row, col_span=len(pc_codes) + 2)
    rev_vals = [pc_summary[p]['revenue'] for p in pc_codes]
    row = write_pl_row(ws, 'Sales Revenue', rev_vals, row, indent=True)
    row = write_pl_row(ws, 'TOTAL REVENUE', rev_vals, row, is_total=True)

    row += 1
    # COGS
    row = write_section_header(ws, 'COST OF GOODS SOLD', row, col_span=len(pc_codes) + 2)
    cogs_vals = [-pc_summary[p]['cogs'] for p in pc_codes]
    row = write_pl_row(ws, 'Cost of Goods Sold', cogs_vals, row, indent=True)
    row = write_pl_row(ws, 'TOTAL COGS', cogs_vals, row, is_total=True)

    row += 1
    # Gross profit
    gp_vals = [rev_vals[i] + cogs_vals[i] for i in range(len(pc_codes))]
    row = write_pl_row(ws, 'GROSS PROFIT', gp_vals, row, is_total=True)
    gp_total = sum(gp_vals)
    rev_total = sum(rev_vals)
    gp_pct = f"{gp_total/rev_total*100:.1f}%" if rev_total else '-'
    ws.cell(row=row - 1, column=len(pc_codes) + 3, value=f'GM%: {gp_pct}')

    row += 1
    # Operating expenses
    row = write_section_header(ws, 'OPERATING EXPENSES', row, col_span=len(pc_codes) + 2)
    opex_vals = [-pc_summary[p]['opex'] for p in pc_codes]
    row = write_pl_row(ws, 'Operating Expenses', opex_vals, row, indent=True)
    row = write_pl_row(ws, 'TOTAL OPEX', opex_vals, row, is_total=True)

    row += 1
    # Operating profit
    op_vals = [gp_vals[i] + opex_vals[i] for i in range(len(pc_codes))]
    row = write_pl_row(ws, 'OPERATING PROFIT', op_vals, row, is_total=True)

    row += 1
    # Non-operating
    row = write_section_header(ws, 'NON-OPERATING EXPENSES', row, col_span=len(pc_codes) + 2)
    nonop_vals = [-pc_summary[p]['nonop'] for p in pc_codes]
    row = write_pl_row(ws, 'Non-Operating Expenses', nonop_vals, row, indent=True)

    row += 1
    # Net
    net_vals = [op_vals[i] + nonop_vals[i] for i in range(len(pc_codes))]
    row = write_pl_row(ws, 'NET PROFIT / (LOSS)', net_vals, row, is_total=True)

    auto_fit_columns(ws)
    freeze_panes(ws)


def write_cc_summary_sheet(wb, cc_summary, period_start, period_end, pcc=None):
    """Write the Cost Center breakdown sheet."""
    ws = add_sheet(wb, 'CC Summary', tab_color='70AD47')
    row = write_title(ws, 'Cost Center Summary', period=f"{period_start} to {period_end}")
    row = write_header_row(ws, ['CC Code', 'Cost Center Name', 'Total Costs (Dr)', 'Total Credits', 'Net Cost'], row)

    cost_centers = pcc.cost_centers if pcc else {}
    grand_dr = grand_cr = 0.0
    for cc_code in sorted(cc_summary.keys()):
        info = cost_centers.get(cc_code, {'name': cc_code})
        dr = cc_summary[cc_code]['debits']
        cr = cc_summary[cc_code]['credits']
        row = write_data_row(ws, [cc_code, info['name'], dr, cr, dr - cr], row)
        grand_dr += dr
        grand_cr += cr

    row += 1
    write_total_row(ws, 'TOTAL', [None, grand_dr, grand_cr, grand_dr - grand_cr], row, double_line=True)
    auto_fit_columns(ws)
    freeze_panes(ws)


def main(input_dir, period_start, period_end, output_file, coa_file=None, pcc_file=None):
    input_dir = Path(input_dir)
    wb = create_workbook()
    coa = COAMapper(coa_file) if coa_file else COAMapper()
    pcc = PCCCMapper(pcc_file) if pcc_file else PCCCMapper()

    all_summaries = []
    journal_results = {}
    journal_dfs = {}        # raw filtered DataFrames — used for PC/CC analysis
    exceptions = []

    # ── Process each journal ─────────────────────────────────────────────────
    for journal_name, config in JOURNAL_CONFIGS.items():
        filepath = find_journal_file(input_dir, config['filename_patterns'])
        if filepath is None:
            exceptions.append({'Journal': journal_name, 'Issue': 'File not found', 'Details': f"No file matching {config['filename_patterns']}"})
            continue

        result = read_xlsx(filepath, required_columns=config['required'], optional_columns=config['optional'], date_columns=['Date'])
        if result['error']:
            exceptions.append({'Journal': journal_name, 'Issue': 'Read error', 'Details': result['error']})
            continue

        df = result['data']
        df = filter_by_period(df, 'Date', period_start, period_end)

        if len(df) == 0:
            exceptions.append({'Journal': journal_name, 'Issue': 'No data in period', 'Details': f"No transactions between {period_start} and {period_end}"})
            continue

        # Validate double-entry balance
        if 'Debit Amount' in df.columns and 'Credit Amount' in df.columns:
            balance_check = validate_journal_balance(df, 'Debit Amount', 'Credit Amount', group_col=config.get('reference_col'))
        else:
            balance_check = validate_journal_balance(df)

        if not balance_check['balanced']:
            for ub in balance_check['unbalanced_entries']:
                exceptions.append({'Journal': journal_name, 'Issue': 'Unbalanced entry', 'Details': str(ub)})

        # Validate PC/CC codes
        pcc_issues = pcc.validate_journal_rows(df, journal_name)
        for issue in pcc_issues:
            exceptions.append({'Journal': issue['journal'], 'Issue': f"Row {issue['row']}", 'Details': issue['issue']})

        summary = summarize_single_journal(df, journal_name)
        journal_results[journal_name] = summary
        journal_dfs[journal_name] = get_amounts(df.copy())
        all_summaries.append(summary)

        # Write journal detail sheet
        short_name = journal_name.replace(' Journal', '')[:20]
        ws = add_sheet(wb, short_name, tab_color='4472C4')
        row = write_title(ws, journal_name, period=f"{period_start} to {period_end}")
        row = write_header_row(ws, ['Account Code', 'Debit Total', 'Debit Count', 'Credit Total', 'Credit Count'], row)

        for _, r in summary['summary'].iterrows():
            row = write_data_row(ws, [
                int(r['Account Code']), r['debit_total'], int(r['debit_count']),
                r['credit_total'], int(r['credit_count'])
            ], row)

        row += 1
        row = write_total_row(ws, 'TOTALS', [summary['total_debits'], summary['transaction_count'],
                                               summary['total_credits'], summary['transaction_count']], row, double_line=True)

        balance_ok = abs(summary['total_debits'] - summary['total_credits']) < 0.01
        row += 1
        ws.cell(row=row, column=1, value='Balance Check:')
        write_validation_result(ws, row, 2, balance_ok)

        auto_fit_columns(ws)
        freeze_panes(ws)

    # ── PC/CC Summary ────────────────────────────────────────────────────────
    pc_summary, cc_summary = pcc.build_pc_summary(journal_dfs)
    has_pcc_data = any(
        pc_summary[p]['revenue'] + pc_summary[p]['cogs'] + pc_summary[p]['opex'] > 0
        for p in pc_summary
    )

    # ── Dashboard sheet ──────────────────────────────────────────────────────
    ws_dash = add_sheet(wb, 'Dashboard', tab_color='00B050')
    wb.move_sheet('Dashboard', offset=-(len(wb.sheetnames) - 1))
    row = write_title(ws_dash, 'SHWE MANDALAY CAFE', 'Books of Prime Entry Summary', f"{period_start} to {period_end}")

    row = write_header_row(ws_dash, ['Journal', 'Transactions', 'Total Debits', 'Total Credits', 'Balanced'], row)
    grand_debits = 0
    grand_credits = 0
    grand_count = 0

    for jname, jresult in journal_results.items():
        balanced = abs(jresult['total_debits'] - jresult['total_credits']) < 0.01
        row = write_data_row(ws_dash, [
            jname, jresult['transaction_count'], jresult['total_debits'], jresult['total_credits'],
            'YES' if balanced else 'NO'
        ], row)
        grand_debits += jresult['total_debits']
        grand_credits += jresult['total_credits']
        grand_count += jresult['transaction_count']

    row += 1
    row = write_total_row(ws_dash, 'GRAND TOTAL', [grand_count, grand_debits, grand_credits,
                                                     'YES' if abs(grand_debits - grand_credits) < 0.01 else 'NO'], row, double_line=True)

    # Consolidated account summary
    row += 2
    row = write_section_header(ws_dash, 'CONSOLIDATED ACCOUNT SUMMARY', row, col_span=5)
    row = write_header_row(ws_dash, ['Account Code', 'Account Name', 'Total Debits', 'Total Credits', 'Net'], row)

    consolidated = {}
    for jresult in journal_results.values():
        for _, r in jresult['summary'].iterrows():
            code = int(r['Account Code'])
            if code not in consolidated:
                consolidated[code] = {'debit': 0, 'credit': 0}
            consolidated[code]['debit'] += r['debit_total']
            consolidated[code]['credit'] += r['credit_total']

    for code in sorted(consolidated.keys()):
        acct = coa.get_account(code)
        name = acct['name'] if acct else f'Account {code}'
        dr = consolidated[code]['debit']
        cr = consolidated[code]['credit']
        row = write_data_row(ws_dash, [code, name, dr, cr, dr - cr], row)

    auto_fit_columns(ws_dash)
    freeze_panes(ws_dash)

    # ── PC/CC sheets (only if PC/CC data exists) ─────────────────────────────
    if has_pcc_data:
        write_pc_summary_sheet(wb, pc_summary, cc_summary, period_start, period_end, pcc)
        write_cc_summary_sheet(wb, cc_summary, period_start, period_end, pcc)

    # ── Exceptions sheet ─────────────────────────────────────────────────────
    if exceptions:
        ws_exc = add_sheet(wb, 'Exceptions', tab_color='FF0000')
        row = write_title(ws_exc, 'Exceptions & Flags')
        row = write_header_row(ws_exc, ['Journal', 'Issue', 'Details'], row)
        for exc in exceptions:
            row = write_data_row(ws_exc, [exc['Journal'], exc['Issue'], exc['Details']], row)
        auto_fit_columns(ws_exc)

    save_workbook(wb, output_file)
    print(f"Summary saved to: {output_file}")
    print(f"Journals processed: {len(journal_results)}/{len(JOURNAL_CONFIGS)}")
    print(f"Grand Total — Debits: {grand_debits:,.0f} | Credits: {grand_credits:,.0f} | Balanced: {abs(grand_debits - grand_credits) < 0.01}")
    if has_pcc_data:
        print("PC/CC Summary sheets written.")
    if exceptions:
        print(f"Exceptions: {len(exceptions)}")


if __name__ == '__main__':
    if len(sys.argv) < 5:
        print("Usage: python summarize_journals.py <input_dir> <period_start> <period_end> <output_file> [coa_file] [pcc_file]")
        sys.exit(1)

    coa = sys.argv[5] if len(sys.argv) > 5 else None
    pcc = sys.argv[6] if len(sys.argv) > 6 else None
    main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], coa, pcc)
