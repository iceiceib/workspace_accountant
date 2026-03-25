"""
Creates sample ledger .xlsx files for testing Module 2.
Consistent with the journal data from create_test_data.py (Jan 2026).

IMPORTANT: Uses 5-digit account codes matching chart_of_accounts.xlsx:
  - 10100 = Cash at Bank
  - 11000 = Accounts Receivable
  - 12000 = Inventory - Raw Material
  - 15100 = Buildings & Structures
  - 20000 = Accounts Payable
  - 31000 = Paid-up Capital
  - 40000 = Sales Revenue
  - etc.

Run:
    python create_ledger_test_data.py <output_dir>
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Account code mapping: 5-digit codes matching COA
CODE = {
    # Assets - Current
    'CASH_IN_HAND': 10000,
    'CASH_AT_BANK': 10100,
    'AR': 11000,
    'INV_RM': 12000,
    'INV_PKG': 12100,
    'INV_FG': 12200,
    'INV_ADJ': 12300,
    'WIP': 12400,
    'ADVANCED_PAYMENTS': 13000,
    # Assets - Non-Current
    'LAND': 15000,
    'BUILDINGS': 15100,
    'ACCUM_DEPR_BUILDINGS': 15110,
    'MACHINERY': 15200,
    'ACCUM_DEPR_MACHINERY': 15210,
    'OFFICE_EQUIP': 15300,
    'ACCUM_DEPR_OFFICE': 15310,
    'ELECTRICAL': 15400,
    'ACCUM_DEPR_ELECTRICAL': 15410,
    'MOTOR_VEHICLES': 15600,
    'ACCUM_DEPR_VEHICLES': 15510,
    # Liabilities
    'AP': 20000,
    'SHORT_TERM_LOANS': 21000,
    'UTILITY_BILLS': 22000,
    'WAGES_PAYABLE': 22200,
    'BANK_LOAN': 25000,
    # Equity
    'PAID_UP_CAPITAL': 31000,
    'RETAINED_EARNINGS': 32000,
    # Revenue
    'SALES_REVENUE': 40000,
    # COGS / Expenses
    'OPENING_INV_RM': 50000,
    'PURCHASES_RM': 50010,
    'CLOSING_INV_RM': 50020,
    'OPENING_INV_PKG': 50100,
    'PURCHASES_PKG': 50110,
    'CLOSING_INV_PKG': 50120,
    'OPENING_INV_FG': 50200,
    'CLOSING_INV_FG': 50220,
    'DIRECT_MATERIALS': 50320,
    'DIRECT_LABOR_WAGES': 53000,
    'MACHINE_MAINTENANCE': 53100,
    'PRODUCTION_UTILITIES': 53200,
    'DEPR_COGS': 53300,
    'MARKETING': 60000,
    'OFFICE_SALARIES': 61000,
    'MEAL_ALLOWANCE': 62000,
    'UTILITIES_SGA': 63000,
    'TRANSPORTATION': 64000,
    'FACTORY_SUPPLIES': 65000,
    'DEPR_SGA': 66000,
    'INVENTORY_WRITEOFF': 67000,
    'OTHER_EXPENSES': 68000,
    # Other Income
    'INTEREST_INCOME': 70000,
}

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))


def style_sheet(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')
        c.border = THIN
        ws.column_dimensions[get_column_letter(i)].width = max(len(str(h)) + 4, 14)


def add_rows(ws, rows, start=2):
    for r, row in enumerate(rows, start):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')


# ─────────────────────────────────────────────────────────────────────────────
# 1. GENERAL LEDGER
# ─────────────────────────────────────────────────────────────────────────────
def create_general_ledger(out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "General Ledger"
    headers = ['Account Code', 'Account Name', 'Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    style_sheet(ws, headers)

    # Opening balances as of 2025-12-31 (one row per account, date before period)
    # Format: (code, name, date, ref, desc, debit, credit, balance)
    opening_rows = [
        # Assets — Debit normal
        (10100, 'Cash at Bank — Main', '2025-12-31', 'OB', 'Opening Balance', None, None, 1500000),
        (10000, 'Petty Cash',           '2025-12-31', 'OB', 'Opening Balance', None, None, 50000),
        (11000, 'Accounts Receivable',  '2025-12-31', 'OB', 'Opening Balance', None, None, 800000),
        (12300, 'Allowance for Doubtful Debts', '2025-12-31', 'OB', 'Opening Balance', None, None, 30000),
        (12000, 'Inventory — Raw Materials',    '2025-12-31', 'OB', 'Opening Balance', None, None, 600000),
        (13000, 'Prepaid Insurance',    '2025-12-31', 'OB', 'Opening Balance', None, None, 120000),
        # Non-current assets
        (15100, 'Buildings',            '2025-12-31', 'OB', 'Opening Balance', None, None, 5000000),
        (15110, 'Accum Depr — Buildings','2025-12-31','OB', 'Opening Balance', None, None, 500000),
        (15200, 'Plant & Machinery',    '2025-12-31', 'OB', 'Opening Balance', None, None, 3000000),
        (15210, 'Accum Depr — P&M',    '2025-12-31', 'OB', 'Opening Balance', None, None, 450000),
        (15300, 'Furniture & Fixtures', '2025-12-31', 'OB', 'Opening Balance', None, None, 800000),
        (15310, 'Accum Depr — F&F',    '2025-12-31', 'OB', 'Opening Balance', None, None, 160000),
        (15400, 'Office Equipment',     '2025-12-31', 'OB', 'Opening Balance', None, None, 500000),
        (15410, 'Accum Depr — Equipment','2025-12-31','OB', 'Opening Balance', None, None, 100000),
        # Liabilities — Credit normal (balance = credit balance)
        (20000, 'Accounts Payable',     '2025-12-31', 'OB', 'Opening Balance', None, None, 250000),
        (22200, 'Accrued Wages',        '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (21000, 'Short-term Loans',     '2025-12-31', 'OB', 'Opening Balance', None, None, 500000),
        (25000, 'Long-term Loans',      '2025-12-31', 'OB', 'Opening Balance', None, None, 2000000),
        # Equity
        (31000, "Owner's Capital",      '2025-12-31', 'OB', 'Opening Balance', None, None, 7180000),
        # Note: Owner's Drawings uses same account 31000 as contra-equity (handled in period transactions)
        (32000, 'Retained Earnings',    '2025-12-31', 'OB', 'Opening Balance', None, None, 1200000),
        # Revenue & Expense — zero opening (income statement resets)
        (40000, 'Sales Revenue',        '2025-12-31','OB', 'Opening Balance', None, None, 0),
        (50320, 'Raw Materials Used',   '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (50110, 'Packaging Costs',      '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (53000, 'Salaries & Wages',     '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (62000, 'Employee Benefits',    '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (68000, 'Rent Expense',         '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (63000, 'Utilities Expense',    '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (68000, 'Telephone & Internet', '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (53100, 'Repairs & Maintenance','2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (65000, 'Cleaning & Sanitation','2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (64000, 'Transportation',       '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (60000, 'Marketing & Advertising','2025-12-31','OB','Opening Balance', None, None, 0),
        (68000, 'Insurance Expense',    '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (67000, 'Bad Debt Expense',     '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
        (68000, 'Bank Charges & Fees',  '2025-12-31', 'OB', 'Opening Balance', None, None, 0),
    ]

    # Jan 2026 period transactions posted from journals
    period_rows = [
        # ── 1020 Cash at Bank ──────────────────────────────────────────────
        (10100,'Cash at Bank','2026-01-02','CRJ-001','Cash sales - Shop 1',          380000, None, None),
        (10100,'Cash at Bank','2026-01-03','CRJ-002','Cash sales - Shop 2',          290000, None, None),
        (10100,'Cash at Bank','2026-01-05','CRJ-003','AR receipt - Hla Min',         250000, None, None),
        (10100,'Cash at Bank','2026-01-05','CPJ-001','Payment - Golden Harvest',      None, 320000, None),
        (10100,'Cash at Bank','2026-01-06','CPJ-002','Rent - Shop 1',                None, 450000, None),
        (10100,'Cash at Bank','2026-01-07','CRJ-004','Cash sales - Shop 3',          420000, None, None),
        (10100,'Cash at Bank','2026-01-07','CPJ-003','Rent - Shop 2',                None, 380000, None),
        (10100,'Cash at Bank','2026-01-08','CPJ-004','Rent - Shop 3',                None, 350000, None),
        (10100,'Cash at Bank','2026-01-10','CRJ-005','AR receipt - Kyaw Zin',        180000, None, None),
        (10100,'Cash at Bank','2026-01-10','CPJ-005','Electricity bill',             None, 125000, None),
        (10100,'Cash at Bank','2026-01-10','CPJ-006','Payment - Fresh Market',       None, 185000, None),
        (10100,'Cash at Bank','2026-01-12','CRJ-006','Cash sales - all shops',       550000, None, None),
        (10100,'Cash at Bank','2026-01-12','CPJ-007','Internet & telephone',         None,  55000, None),
        (10100,'Cash at Bank','2026-01-14','CRJ-007','AR receipt - Win Htut',        450000, None, None),
        (10100,'Cash at Bank','2026-01-15','CPJ-008','Payment - Pack & Go',          None,  95000, None),
        (10100,'Cash at Bank','2026-01-15','CPJ-009','Transport costs',              None,  85000, None),
        (10100,'Cash at Bank','2026-01-17','CRJ-008','Cash sales - Shop 1',          315000, None, None),
        (10100,'Cash at Bank','2026-01-18','CPJ-010','Petty cash top-up',            None, 100000, None),
        (10100,'Cash at Bank','2026-01-20','CRJ-009','AR receipt - Thida Aye',       320000, None, None),
        (10100,'Cash at Bank','2026-01-20','CPJ-011','Payment - Supply Hub',         None, 198000, None),
        (10100,'Cash at Bank','2026-01-22','CRJ-010','Cash sales - Shop 2',          410000, None, None),
        (10100,'Cash at Bank','2026-01-22','CPJ-012','Insurance premium',            None,  65000, None),
        (10100,'Cash at Bank','2026-01-25','CRJ-011','AR receipt - Zaw Lin',         750000, None, None),
        (10100,'Cash at Bank','2026-01-25','CPJ-013','Equipment repair',             None,  95000, None),
        (10100,'Cash at Bank','2026-01-28','CRJ-012','Cash sales - all shops',       625000, None, None),
        (10100,'Cash at Bank','2026-01-28','CPJ-014','Marketing agency',             None, 120000, None),
        (10100,'Cash at Bank','2026-01-30','CRJ-013','Capital introduced',           500000, None, None),
        (10100,'Cash at Bank','2026-01-30','CPJ-015','Owner drawings',               None, 200000, None),
        (10100,'Cash at Bank','2026-01-31','PAY-001','Salaries - kitchen staff',     None, 750000, None),
        (10100,'Cash at Bank','2026-01-31','PAY-002','Salaries - service staff',     None, 960000, None),
        (10100,'Cash at Bank','2026-01-31','PAY-003','Salaries - management',        None, 600000, None),
        (10100,'Cash at Bank','2026-01-31','PAY-004','Salaries - delivery staff',    None, 360000, None),
        (10100,'Cash at Bank','2026-01-31','GJ-004', 'Bank service charges',         None,   8500, None),

        # ── 1030 Petty Cash ────────────────────────────────────────────────
        (10000,'Petty Cash','2026-01-18','CPJ-010','Petty cash top-up',              100000, None, None),

        # ── 1100 Accounts Receivable ───────────────────────────────────────
        (11000,'AR','2026-01-02','SJ-001','Credit sale - Hla Min',       250000, None, None),
        (11000,'AR','2026-01-03','SJ-002','Credit sale - Kyaw Zin',      180000, None, None),
        (11000,'AR','2026-01-05','SJ-003','Credit sale - Win Htut',      450000, None, None),
        (11000,'AR','2026-01-07','SJ-004','Credit sale - Thida Aye',     320000, None, None),
        (11000,'AR','2026-01-10','SJ-005','Credit sale - Mg Mg',         195000, None, None),
        (11000,'AR','2026-01-12','SJ-006','Credit sale - Su Su',         600000, None, None),
        (11000,'AR','2026-01-15','SJ-007','Credit sale - Aye Aye',       280000, None, None),
        (11000,'AR','2026-01-17','SJ-008','Credit sale - Zaw Lin',       750000, None, None),
        (11000,'AR','2026-01-20','SJ-009','Credit sale - Nay Chi',       215000, None, None),
        (11000,'AR','2026-01-22','SJ-010','Credit sale - Tun Tun',       340000, None, None),
        (11000,'AR','2026-01-25','SJ-011','Credit sale - Khin Khin',     290000, None, None),
        (11000,'AR','2026-01-28','SJ-012','Credit sale - Aung Aung',     480000, None, None),
        (11000,'AR','2026-01-30','SJ-013','Credit sale - Ye Naung',      175000, None, None),
        (11000,'AR','2026-01-05','CRJ-003','Receipt - Hla Min',          None, 250000, None),
        (11000,'AR','2026-01-10','CRJ-005','Receipt - Kyaw Zin',         None, 180000, None),
        (11000,'AR','2026-01-14','CRJ-007','Receipt - Win Htut',         None, 450000, None),
        (11000,'AR','2026-01-20','CRJ-009','Receipt - Thida Aye',        None, 320000, None),
        (11000,'AR','2026-01-25','CRJ-011','Receipt - Zaw Lin',          None, 750000, None),
        (11000,'AR','2026-01-31','GJ-003', 'Write-off bad debt',         None,  15000, None),

        # ── 1110 Allowance for Doubtful Debts ─────────────────────────────
        (12300,'Allowance for DD','2026-01-31','GJ-003','Write-off bad debt', 15000, None, None),

        # ── 1200 Inventory — Raw Materials ────────────────────────────────
        (12000,'Inventory','2026-01-31','GJ-001','Transfer to COGS', None, 450000, None),

        # ── 1320 Prepaid Insurance ─────────────────────────────────────────
        (13000,'Prepaid Insurance','2026-01-31','GJ-002','Insurance recognized - Jan', None, 20000, None),

        # ── 2010 Accounts Payable ──────────────────────────────────────────
        (20000,'AP','2026-01-03','PJ-001','Purchase - Golden Harvest rice/flour',  None, 320000, None),
        (20000,'AP','2026-01-05','PJ-002','Purchase - Fresh Market veg',           None, 185000, None),
        (20000,'AP','2026-01-08','PJ-003','Purchase - Pack & Go packaging',        None,  95000, None),
        (20000,'AP','2026-01-10','PJ-004','Purchase - Golden Harvest oil',         None, 145000, None),
        (20000,'AP','2026-01-13','PJ-005','Purchase - Fresh Market meat',          None, 265000, None),
        (20000,'AP','2026-01-15','PJ-006','Purchase - Pack & Go containers',       None,  72000, None),
        (20000,'AP','2026-01-18','PJ-007','Purchase - Supply Hub coffee',          None, 198000, None),
        (20000,'AP','2026-01-20','PJ-008','Purchase - Fresh Market produce',       None, 223000, None),
        (20000,'AP','2026-01-23','PJ-009','Purchase - Golden Harvest flour',       None, 310000, None),
        (20000,'AP','2026-01-27','PJ-010','Purchase - Pack & Go napkins',          None,  58000, None),
        (20000,'AP','2026-01-29','PJ-011','Purchase - Supply Hub cleaning',        None,  45000, None),
        (20000,'AP','2026-01-05','CPJ-001','Payment - Golden Harvest',             320000, None, None),
        (20000,'AP','2026-01-10','CPJ-006','Payment - Fresh Market',               185000, None, None),
        (20000,'AP','2026-01-15','CPJ-008','Payment - Pack & Go',                   95000, None, None),
        (20000,'AP','2026-01-20','CPJ-011','Payment - Supply Hub',                 198000, None, None),

        # ── 2030 Accrued Wages ────────────────────────────────────────────
        (22200,'Accrued Wages','2026-01-31','PAY-005','Overtime accrual - kitchen', None, 85000, None),
        (22200,'Accrued Wages','2026-01-31','PAY-006','Employee benefits accrual',  None, 120000, None),

        # ── 3010 Owner's Capital ───────────────────────────────────────────
        (31000,"Owner's Capital",'2026-01-30','CRJ-013','Capital introduced', None, 500000, None),

        # ── 3020 Owner's Drawings ──────────────────────────────────────────
        (31000,"Owner's Drawings",'2026-01-30','CPJ-015','Owner drawings Jan', 200000, None, None),

        # ── Revenue accounts ───────────────────────────────────────────────
        (40000,'Sales Rev - Shop 1','2026-01-02','CRJ-001','Cash sales',           None, 380000, None),
        (40000,'Sales Rev - Shop 1','2026-01-02','SJ-001', 'Credit sale',          None, 250000, None),
        (40000,'Sales Rev - Shop 1','2026-01-03','SJ-002', 'Credit sale',          None, 180000, None),
        (40000,'Sales Rev - Shop 1','2026-01-10','SJ-005', 'Credit sale',          None, 195000, None),
        (40000,'Sales Rev - Shop 1','2026-01-17','CRJ-008','Cash sales',           None, 315000, None),
        (40000,'Sales Rev - Shop 1','2026-01-20','SJ-009', 'Credit sale',          None, 215000, None),
        (40000,'Sales Rev - Shop 1','2026-01-30','SJ-013', 'Credit sale',          None, 175000, None),

        (40000,'Sales Rev - Shop 2','2026-01-03','CRJ-002','Cash sales',           None, 290000, None),
        (40000,'Sales Rev - Shop 2','2026-01-05','SJ-003', 'Credit sale',          None, 450000, None),
        (40000,'Sales Rev - Shop 2','2026-01-12','SJ-006', 'Credit sale',          None, 600000, None),
        (40000,'Sales Rev - Shop 2','2026-01-22','CRJ-010','Cash sales',           None, 410000, None),
        (40000,'Sales Rev - Shop 2','2026-01-22','SJ-010', 'Credit sale',          None, 340000, None),

        (40000,'Sales Rev - Shop 3','2026-01-07','CRJ-004','Cash sales',           None, 420000, None),
        (40000,'Sales Rev - Shop 3','2026-01-07','SJ-004', 'Credit sale',          None, 320000, None),
        (40000,'Sales Rev - Shop 3','2026-01-15','SJ-007', 'Credit sale',          None, 280000, None),
        (40000,'Sales Rev - Shop 3','2026-01-25','SJ-011', 'Credit sale',          None, 290000, None),

        (40000,'Sales Rev - General','2026-01-12','CRJ-006','Cash sales',          None, 550000, None),
        (40000,'Sales Rev - General','2026-01-17','SJ-008', 'Credit sale',         None, 750000, None),
        (40000,'Sales Rev - General','2026-01-28','CRJ-012','Cash sales',          None, 625000, None),
        (40000,'Sales Rev - General','2026-01-28','SJ-012', 'Credit sale',         None, 480000, None),

        # ── COGS & Expenses ────────────────────────────────────────────────
        (50320,'Raw Materials','2026-01-03','PJ-001','Rice/flour purchase',  320000, None, None),
        (50320,'Raw Materials','2026-01-05','PJ-002','Vegetables',           185000, None, None),
        (50320,'Raw Materials','2026-01-10','PJ-004','Oil/seasoning',        145000, None, None),
        (50320,'Raw Materials','2026-01-13','PJ-005','Meat/protein',         265000, None, None),
        (50320,'Raw Materials','2026-01-18','PJ-007','Coffee/tea',           198000, None, None),
        (50320,'Raw Materials','2026-01-20','PJ-008','Produce/dairy',        223000, None, None),
        (50320,'Raw Materials','2026-01-23','PJ-009','Flour/grains',         310000, None, None),
        (50320,'Raw Materials','2026-01-31','GJ-001','Inventory to COGS',    450000, None, None),

        (50110,'Packaging','2026-01-08','PJ-003','Take-away containers',      95000, None, None),
        (50110,'Packaging','2026-01-15','PJ-006','Containers',                72000, None, None),
        (50110,'Packaging','2026-01-27','PJ-010','Napkins/disposables',       58000, None, None),

        (53000,'Salaries','2026-01-31','PAY-001','Kitchen staff salaries',   750000, None, None),
        (53000,'Salaries','2026-01-31','PAY-002','Service staff salaries',   960000, None, None),
        (53000,'Salaries','2026-01-31','PAY-003','Management salaries',      600000, None, None),
        (53000,'Salaries','2026-01-31','PAY-004','Delivery salaries',        360000, None, None),
        (53000,'Salaries','2026-01-31','PAY-005','Kitchen overtime accrual',  85000, None, None),

        (62000,'Employee Benefits','2026-01-31','PAY-006','Jan benefits accrual', 120000, None, None),

        (68000,'Rent','2026-01-06','CPJ-002','Shop 1 rent',  450000, None, None),
        (68000,'Rent','2026-01-07','CPJ-003','Shop 2 rent',  380000, None, None),
        (68000,'Rent','2026-01-08','CPJ-004','Shop 3 rent',  350000, None, None),

        (63000,'Utilities','2026-01-10','CPJ-005','Electricity Jan', 125000, None, None),

        (68000,'Tel & Internet','2026-01-12','CPJ-007','Internet/tel Jan', 55000, None, None),

        (53100,'Repairs','2026-01-25','CPJ-013','Kitchen equipment repair', 95000, None, None),

        (65000,'Cleaning','2026-01-29','PJ-011','Cleaning supplies',  45000, None, None),

        (64000,'Transport','2026-01-15','CPJ-009','Delivery costs Jan', 85000, None, None),

        (60000,'Marketing','2026-01-28','CPJ-014','Social media Jan', 120000, None, None),

        (68000,'Insurance','2026-01-22','CPJ-012','Insurance premium',  65000, None, None),
        (68000,'Insurance','2026-01-31','GJ-002', 'Prepaid recognized', 20000, None, None),

        (68000,'Bank Charges','2026-01-31','GJ-004','Bank service charges', 8500, None, None),
    ]

    all_rows = opening_rows + period_rows
    add_rows(ws, all_rows)
    path = out / 'general_ledger.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. ACCOUNTS RECEIVABLE LEDGER
# ─────────────────────────────────────────────────────────────────────────────
def create_ar_ledger(out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AR Ledger"
    headers = ['Customer', 'Date', 'Invoice No', 'Description', 'Debit', 'Credit', 'Balance']
    style_sheet(ws, headers)

    # Opening: old customers with aggregate balance 800,000 Dr
    # Current period customers + their payments
    rows = [
        # Old customers (prior period AR)
        ('Old Customers', '2025-12-31', 'OB',     'Opening Balance',        None,   None,  800000),
        ('Old Customers', '2026-01-31', 'GJ-003',  'Write-off bad debt',     None,  15000,  785000),
        # Current period customers — credit sales
        ('Hla Min',   '2026-01-02', 'SJ-001', 'Food & beverage - Shop 1', 250000,  None,  250000),
        ('Hla Min',   '2026-01-05', 'CRJ-003','Payment received',           None, 250000,       0),
        ('Kyaw Zin',  '2026-01-03', 'SJ-002', 'Food & beverage - Shop 1', 180000,  None,  180000),
        ('Kyaw Zin',  '2026-01-10', 'CRJ-005','Payment received',           None, 180000,       0),
        ('Win Htut',  '2026-01-05', 'SJ-003', 'Catering - Shop 2',        450000,  None,  450000),
        ('Win Htut',  '2026-01-14', 'CRJ-007','Payment received',           None, 450000,       0),
        ('Thida Aye', '2026-01-07', 'SJ-004', 'Daily sales - Shop 3',     320000,  None,  320000),
        ('Thida Aye', '2026-01-20', 'CRJ-009','Payment received',           None, 320000,       0),
        ('Mg Mg',     '2026-01-10', 'SJ-005', 'Food & beverage - Shop 1', 195000,  None,  195000),
        ('Su Su',     '2026-01-12', 'SJ-006', 'Bulk order - Shop 2',      600000,  None,  600000),
        ('Aye Aye',   '2026-01-15', 'SJ-007', 'Daily sales - Shop 3',     280000,  None,  280000),
        ('Zaw Lin',   '2026-01-17', 'SJ-008', 'Event catering - General', 750000,  None,  750000),
        ('Zaw Lin',   '2026-01-25', 'CRJ-011','Payment received',           None, 750000,       0),
        ('Nay Chi',   '2026-01-20', 'SJ-009', 'Food & beverage - Shop 1', 215000,  None,  215000),
        ('Tun Tun',   '2026-01-22', 'SJ-010', 'Daily sales - Shop 2',     340000,  None,  340000),
        ('Khin Khin', '2026-01-25', 'SJ-011', 'Daily sales - Shop 3',     290000,  None,  290000),
        ('Aung Aung', '2026-01-28', 'SJ-012', 'Catering - General',       480000,  None,  480000),
        ('Ye Naung',  '2026-01-30', 'SJ-013', 'Food & beverage - Shop 1', 175000,  None,  175000),
    ]
    # AR subsidiary total closing = 785000 + 195000 + 600000 + 280000 + 215000 + 340000 + 290000 + 480000 + 175000
    # = 785000 + 2575000 = 3360000 → must match GL 1100 closing
    add_rows(ws, rows)
    path = out / 'accounts_receivable_ledger.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# 3. ACCOUNTS PAYABLE LEDGER
# ─────────────────────────────────────────────────────────────────────────────
def create_ap_ledger(out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AP Ledger"
    headers = ['Supplier', 'Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    style_sheet(ws, headers)

    rows = [
        # Prior period AP
        ('Prior Suppliers', '2025-12-31', 'OB',     'Opening Balance',               None,   None, 250000),
        # Golden Harvest
        ('Golden Harvest', '2026-01-03', 'PJ-001', 'Rice/flour purchase',            None, 320000, 320000),
        ('Golden Harvest', '2026-01-10', 'PJ-004', 'Oil/seasoning purchase',         None, 145000, 465000),
        ('Golden Harvest', '2026-01-23', 'PJ-009', 'Flour/grains purchase',          None, 310000, 775000),
        ('Golden Harvest', '2026-01-05', 'CPJ-001','Payment',                       320000,   None, 455000),
        # Fresh Market Co
        ('Fresh Market Co', '2026-01-05', 'PJ-002','Vegetables/ingredients',         None, 185000, 185000),
        ('Fresh Market Co', '2026-01-13', 'PJ-005','Meat/protein',                   None, 265000, 450000),
        ('Fresh Market Co', '2026-01-20', 'PJ-008','Produce/dairy',                  None, 223000, 673000),
        ('Fresh Market Co', '2026-01-10', 'CPJ-006','Payment',                      185000,   None, 488000),
        # Pack & Go Ltd
        ('Pack & Go Ltd', '2026-01-08', 'PJ-003', 'Take-away containers',            None,  95000,  95000),
        ('Pack & Go Ltd', '2026-01-15', 'PJ-006', 'Containers',                      None,  72000, 167000),
        ('Pack & Go Ltd', '2026-01-27', 'PJ-010', 'Napkins/disposables',             None,  58000, 225000),
        ('Pack & Go Ltd', '2026-01-15', 'CPJ-008','Payment',                         95000,   None, 130000),
        # Supply Hub
        ('Supply Hub', '2026-01-18', 'PJ-007', 'Coffee/tea',                         None, 198000, 198000),
        ('Supply Hub', '2026-01-29', 'PJ-011', 'Cleaning supplies',                  None,  45000, 243000),
        ('Supply Hub', '2026-01-20', 'CPJ-011','Payment',                           198000,   None,  45000),
    ]
    # AP closing: Prior 250000 + GH 455000 + FM 488000 + PG 130000 + SH 45000 = 1,368,000 → matches GL 2010
    add_rows(ws, rows)
    path = out / 'accounts_payable_ledger.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# 4. CASH LEDGER
# ─────────────────────────────────────────────────────────────────────────────
def create_cash_ledger(out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Ledger"
    headers = ['Bank Account', 'Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    style_sheet(ws, headers)

    rows = [
        ('Main Account', '2025-12-31', 'OB',     'Opening Balance',              None,   None, 1500000),
        ('Main Account', '2026-01-02', 'CRJ-001','Cash sales - Shop 1',         380000,   None, 1880000),
        ('Main Account', '2026-01-03', 'CRJ-002','Cash sales - Shop 2',         290000,   None, 2170000),
        ('Main Account', '2026-01-05', 'CRJ-003','AR receipt - Hla Min',        250000,   None, 2420000),
        ('Main Account', '2026-01-05', 'CPJ-001','Payment - Golden Harvest',      None, 320000, 2100000),
        ('Main Account', '2026-01-06', 'CPJ-002','Rent - Shop 1',                 None, 450000, 1650000),
        ('Main Account', '2026-01-07', 'CRJ-004','Cash sales - Shop 3',         420000,   None, 2070000),
        ('Main Account', '2026-01-07', 'CPJ-003','Rent - Shop 2',                 None, 380000, 1690000),
        ('Main Account', '2026-01-08', 'CPJ-004','Rent - Shop 3',                 None, 350000, 1340000),
        ('Main Account', '2026-01-10', 'CRJ-005','AR receipt - Kyaw Zin',       180000,   None, 1520000),
        ('Main Account', '2026-01-10', 'CPJ-005','Electricity bill',              None, 125000, 1395000),
        ('Main Account', '2026-01-10', 'CPJ-006','Payment - Fresh Market',        None, 185000, 1210000),
        ('Main Account', '2026-01-12', 'CRJ-006','Cash sales - all shops',       550000,   None, 1760000),
        ('Main Account', '2026-01-12', 'CPJ-007','Internet & telephone',           None,  55000, 1705000),
        ('Main Account', '2026-01-14', 'CRJ-007','AR receipt - Win Htut',        450000,   None, 2155000),
        ('Main Account', '2026-01-15', 'CPJ-008','Payment - Pack & Go',            None,  95000, 2060000),
        ('Main Account', '2026-01-15', 'CPJ-009','Transport costs',                None,  85000, 1975000),
        ('Main Account', '2026-01-17', 'CRJ-008','Cash sales - Shop 1',          315000,   None, 2290000),
        ('Main Account', '2026-01-18', 'CPJ-010','Petty cash top-up',              None, 100000, 2190000),
        ('Main Account', '2026-01-20', 'CRJ-009','AR receipt - Thida Aye',       320000,   None, 2510000),
        ('Main Account', '2026-01-20', 'CPJ-011','Payment - Supply Hub',           None, 198000, 2312000),
        ('Main Account', '2026-01-22', 'CRJ-010','Cash sales - Shop 2',           410000,   None, 2722000),
        ('Main Account', '2026-01-22', 'CPJ-012','Insurance premium',              None,  65000, 2657000),
        ('Main Account', '2026-01-25', 'CRJ-011','AR receipt - Zaw Lin',          750000,   None, 3407000),
        ('Main Account', '2026-01-25', 'CPJ-013','Equipment repair',               None,  95000, 3312000),
        ('Main Account', '2026-01-28', 'CRJ-012','Cash sales - all shops',        625000,   None, 3937000),
        ('Main Account', '2026-01-28', 'CPJ-014','Marketing agency',               None, 120000, 3817000),
        ('Main Account', '2026-01-30', 'CRJ-013','Capital introduced',            500000,   None, 4317000),
        ('Main Account', '2026-01-30', 'CPJ-015','Owner drawings',                 None, 200000, 4117000),
        ('Main Account', '2026-01-31', 'PAY-001','Salaries - kitchen',             None, 750000, 3367000),
        ('Main Account', '2026-01-31', 'PAY-002','Salaries - service',             None, 960000, 2407000),
        ('Main Account', '2026-01-31', 'PAY-003','Salaries - management',          None, 600000, 1807000),
        ('Main Account', '2026-01-31', 'PAY-004','Salaries - delivery',            None, 360000, 1447000),
        ('Main Account', '2026-01-31', 'GJ-004', 'Bank charges',                   None,   8500, 1438500),
    ]
    # Closing: 1,438,500 Dr → must match GL 1020 closing
    add_rows(ws, rows)
    path = out / 'cash_ledger.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# 5. FIXED ASSETS LEDGER
# ─────────────────────────────────────────────────────────────────────────────
def create_fixed_assets_ledger(out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fixed Assets"
    headers = ['Asset ID', 'Description', 'Account Code', 'Category', 'Date Acquired',
               'Cost', 'Useful Life (Years)', 'Salvage Value', 'Depreciation Method',
               'Accumulated Depreciation', 'Net Book Value', 'Status']
    style_sheet(ws, headers)

    rows = [
        ('FA-001', 'Shop 1 Building',          15100, 'Buildings',         '2020-01-01', 3000000, 20, 300000, 'Straight-Line',  300000, 2700000, 'Active'),
        ('FA-002', 'Shop 2 Building',          15100, 'Buildings',         '2021-06-01', 2000000, 20, 200000, 'Straight-Line',  200000, 1800000, 'Active'),
        ('FA-003', 'Commercial Kitchen Unit',  15200, 'Plant & Machinery', '2020-03-01', 2000000, 10, 200000, 'Straight-Line',  360000, 1640000, 'Active'),
        ('FA-004', 'Refrigeration Equipment',  15200, 'Plant & Machinery', '2021-01-01', 1000000, 10, 100000, 'Straight-Line',  180000, 820000,  'Active'),
        ('FA-005', 'Dining Furniture Set 1',   15300, 'Furniture',         '2020-06-01',  500000, 10,  50000, 'Straight-Line',  112500, 387500,  'Active'),
        ('FA-006', 'Dining Furniture Set 2',   15300, 'Furniture',         '2022-01-01',  300000, 10,  30000, 'Straight-Line',   54000, 246000,  'Active'),
        ('FA-007', 'POS System & Computers',   15400, 'Office Equipment',  '2021-07-01',  350000,  5,  35000, 'Straight-Line',  157500, 192500,  'Active'),
        ('FA-008', 'CCTV & Security System',   15400, 'Office Equipment',  '2022-03-01',  150000,  5,  15000, 'Straight-Line',   40500, 109500,  'Active'),
    ]
    add_rows(ws, rows)
    path = out / 'fixed_assets_ledger.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r'C:\Users\USER\workspace_accountant\accountant-skill\data\Jan2026')
    out.mkdir(parents=True, exist_ok=True)
    print(f"Creating ledger test data in: {out}")
    create_general_ledger(out)
    create_ar_ledger(out)
    create_ap_ledger(out)
    create_cash_ledger(out)
    create_fixed_assets_ledger(out)
    print("\nDone.")
    print("Run Module 2 with:")
    print("  python scripts/summarize_ledgers.py data/Jan2026 2026-01-01 2026-01-31 data/Jan2026/ledger_summary_Jan2026.xlsx data/Jan2026/chart_of_accounts.xlsx")


if __name__ == '__main__':
    main()
