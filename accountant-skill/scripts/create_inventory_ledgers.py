#!/usr/bin/env python
"""
Create Inventory Sub-Ledgers

This script generates standardized inventory sub-ledger files from existing
tracker data or creates blank templates for new periods.

Usage:
    python scripts/create_inventory_ledgers.py INPUT_DIR OUTPUT_DIR PERIOD_START PERIOD_END

Example:
    python scripts/create_inventory_ledgers.py data/input data/input/ledgers 2026-01-01 2026-01-31

Output:
    - raw_materials_ledger.xlsx (in output_dir)
    - packaging_ledger.xlsx (in output_dir)
    - inventory_items.xlsx (in input_dir/inventory)
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
import math

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.excel_writer import (
    create_workbook, add_sheet, write_title, write_header_row,
    write_data_row, write_total_row, auto_fit_columns, freeze_panes,
    save_workbook, HEADER_FILL, THIN_BORDER, NUMBER_FORMAT_NEG
)
from utils.inventory_mapper import (
    InventoryMapper, InventoryLedger, DEFAULT_INVENTORY_ITEMS,
    calculate_wac, _clean, _clean_numeric
)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Tab colors
TAB_DASHBOARD = '00B050'  # Green
TAB_DETAIL = '4472C4'     # Blue
TAB_SUMMARY = '70AD47'    # Orange


def read_existing_tracker(filepath, item_codes_range):
    """
    Read existing tracker file to extract inventory data.

    Args:
        filepath: Path to the tracker xlsx file
        item_codes_range: Tuple of (min_code, max_code) for items to extract

    Returns:
        Dict of item_code -> dict with 'ground_stock', 'wac', etc.
    """
    if not Path(filepath).exists():
        return {}

    result = {}
    try:
        xl = pd.ExcelFile(filepath)
        min_code, max_code = item_codes_range

        for sheet in xl.sheet_names:
            # Check if sheet name is a valid item code
            try:
                sheet_code = int(float(sheet))
                if not (min_code <= sheet_code <= max_code):
                    continue
            except (ValueError, TypeError):
                continue

            df = pd.read_excel(xl, sheet_name=sheet, header=0)

            # Find the columns
            col_map = {}
            for col in df.columns:
                col_lower = str(col).lower().strip()
                if 'ground stock' in col_lower:
                    col_map['ground_stock'] = col
                elif 'wac' in col_lower or 'weighted average' in col_lower:
                    col_map['wac'] = col
                elif 'closing inventory value' in col_lower:
                    col_map['closing_value'] = col
                elif 'cost of production' in col_lower:
                    col_map['cop'] = col

            # Get the last non-empty row
            ground_stock = 0
            wac = 0
            closing_value = 0

            if 'ground_stock' in col_map:
                col = col_map['ground_stock']
                for val in df[col].dropna().iloc[::-1]:
                    if _clean_numeric(val) != 0:
                        ground_stock = _clean_numeric(val)
                        break

            if 'wac' in col_map:
                col = col_map['wac']
                for val in df[col].dropna().iloc[::-1]:
                    if _clean_numeric(val) != 0:
                        wac = _clean_numeric(val)
                        break

            if 'closing_value' in col_map:
                col = col_map['closing_value']
                for val in df[col].dropna().iloc[::-1]:
                    if _clean_numeric(val) != 0:
                        closing_value = _clean_numeric(val)
                        break

            result[sheet_code] = {
                'ground_stock': ground_stock,
                'wac': wac,
                'closing_value': closing_value
            }
    except Exception as e:
        print(f"Warning: Error reading tracker {filepath}: {e}")

    return result


def create_item_ledger_sheet(wb, item_code, item_name, unit, opening_qty=0, opening_value=0, wac=0):
    """
    Create a single item ledger sheet.

    Args:
        wb: Workbook object
        item_code: Item code (e.g., 12001, 12100)
        item_name: Item name/description
        unit: Unit of measure
        opening_qty: Opening quantity balance
        opening_value: Opening value balance
        wac: Weighted average cost

    Returns:
        Worksheet object
    """
    sheet_name = str(item_code)
    ws = add_sheet(wb, sheet_name, TAB_DETAIL)

    # Title
    row = write_title(ws, f"Inventory Sub-Ledger: {item_name}",
                      f"Item Code: {item_code} | Unit: {unit}")

    # Opening balance section
    ws.cell(row=row, column=1, value="Opening Balance:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, name='Arial')
    ws.cell(row=row, column=2, value=opening_qty)
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=3, value="units")
    ws.cell(row=row, column=5, value="Value:")
    ws.cell(row=row, column=6, value=opening_value)
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=7, value="WAC:")
    ws.cell(row=row, column=8, value=wac)
    ws.cell(row=row, column=8).number_format = '#,##0.0000'
    row += 2

    # Column headers
    headers = [
        'Date', 'Reference', 'Description',
        'Received Qty', 'Issued Qty', 'Balance Qty',
        'Unit Cost', 'Received Value', 'Issued Value', 'Balance Value'
    ]
    row = write_header_row(ws, headers, row)

    # Opening balance row
    values = [
        '', '', 'Opening Balance',
        '', '', opening_qty,
        wac, '', '', opening_value
    ]
    row = write_data_row(ws, values, row, number_cols=[3, 4, 5, 6, 7, 8, 9])

    # Leave empty rows for data entry
    for _ in range(20):
        row = write_data_row(ws, ['', '', '', '', '', '', '', '', '', ''], row)

    # Closing balance row
    row = write_total_row(ws, 'Closing Balance', ['', '', '', '', '', '', '', ''], row - 1)

    # Auto-fit columns
    auto_fit_columns(ws, min_width=10)

    # Freeze panes
    freeze_panes(ws, row=6, col=1)

    return ws


def create_dashboard_sheet(wb, title, items_data):
    """
    Create a dashboard/summary sheet.

    Args:
        wb: Workbook object
        title: Sheet title
        items_data: List of dicts with item details

    Returns:
        Worksheet object
    """
    ws = add_sheet(wb, 'Dashboard', TAB_DASHBOARD)

    # Title
    row = write_title(ws, title, "Inventory Sub-Ledger Summary")

    # Column headers
    headers = [
        'Item Code', 'Item Name', 'Unit',
        'Opening Qty', 'Opening Value',
        'Received Qty', 'Received Value',
        'Issued Qty', 'Issued Value',
        'Closing Qty', 'Closing Value', 'WAC'
    ]
    row = write_header_row(ws, headers, row)

    total_opening_value = 0
    total_received_value = 0
    total_issued_value = 0
    total_closing_value = 0

    for item in items_data:
        values = [
            item.get('code', ''),
            item.get('name', ''),
            item.get('unit', ''),
            item.get('opening_qty', 0),
            item.get('opening_value', 0),
            item.get('received_qty', 0),
            item.get('received_value', 0),
            item.get('issued_qty', 0),
            item.get('issued_value', 0),
            item.get('closing_qty', 0),
            item.get('closing_value', 0),
            item.get('wac', 0)
        ]
        row = write_data_row(ws, values, row, number_cols=[3, 4, 5, 6, 7, 8, 9, 10, 11])

        total_opening_value += item.get('opening_value', 0)
        total_received_value += item.get('received_value', 0)
        total_issued_value += item.get('issued_value', 0)
        total_closing_value += item.get('closing_value', 0)

    # Total row
    row = write_total_row(ws, 'TOTAL', [
        '', '', '',
        '', total_opening_value,
        '', total_received_value,
        '', total_issued_value,
        '', total_closing_value, ''
    ], row, double_line=True)

    # Reconciliation section
    row += 2
    ws.cell(row=row, column=1, value="GL Reconciliation:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, name='Arial')
    row += 1
    ws.cell(row=row, column=1, value="Sub-Ledger Total:")
    ws.cell(row=row, column=3, value=total_closing_value)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value="GL Control Account:")
    ws.cell(row=row, column=3, value="[Enter GL Balance]")
    row += 1
    ws.cell(row=row, column=1, value="Difference:")
    ws.cell(row=row, column=3, value="[Calculated]")

    auto_fit_columns(ws)
    freeze_panes(ws, row=5, col=1)

    return ws


def create_inventory_master_file(output_dir, inventory_mapper):
    """
    Create the inventory_items.xlsx master file.

    Args:
        output_dir: Output directory path
        inventory_mapper: InventoryMapper instance
    """
    wb = create_workbook()
    ws = add_sheet(wb, 'Inventory Items', TAB_SUMMARY)

    row = write_title(ws, "Inventory Items Master", "List of all inventory items")

    headers = ['Item Code', 'Item Name', 'Category', 'Account Code', 'Account Name', 'Unit Measure', 'Status']
    row = write_header_row(ws, headers, row)

    for code, item in sorted(inventory_mapper.items_dict.items()):
        values = [
            code,
            item.get('name', ''),
            item.get('category', ''),
            item.get('account_code', ''),
            '',  # Account name placeholder
            item.get('unit', ''),
            item.get('status', 'Active')
        ]
        row = write_data_row(ws, values, row)

    auto_fit_columns(ws)
    freeze_panes(ws, row=4, col=1)

    output_path = Path(output_dir) / 'inventory_items.xlsx'
    save_workbook(wb, output_path)
    print(f"Created: {output_path}")

    return output_path


def create_raw_materials_ledger(output_dir, inventory_mapper, existing_data=None):
    """
    Create the raw_materials_ledger.xlsx file.

    Args:
        output_dir: Output directory path
        inventory_mapper: InventoryMapper instance
        existing_data: Optional dict from existing tracker
    """
    wb = create_workbook()

    # Get raw materials items
    items = inventory_mapper.get_raw_materials()
    items_data = []

    for item in items:
        code = item['code']
        existing = existing_data.get(code, {}) if existing_data else {}

        opening_qty = existing.get('ground_stock', 0)
        wac = existing.get('wac', 0)
        opening_value = existing.get('closing_value', opening_qty * wac)

        # Create item sheet
        create_item_ledger_sheet(
            wb, code, item['name'], item['unit'],
            opening_qty, opening_value, wac
        )

        items_data.append({
            'code': code,
            'name': item['name'],
            'unit': item['unit'],
            'opening_qty': opening_qty,
            'opening_value': opening_value,
            'received_qty': 0,
            'received_value': 0,
            'issued_qty': 0,
            'issued_value': 0,
            'closing_qty': opening_qty,
            'closing_value': opening_value,
            'wac': wac
        })

    # Create dashboard
    create_dashboard_sheet(wb, "Raw Materials Inventory", items_data)

    # Move dashboard to first position
    wb.move_sheet('Dashboard', offset=-len(items))

    output_path = Path(output_dir) / 'raw_materials_ledger.xlsx'
    save_workbook(wb, output_path)
    print(f"Created: {output_path}")

    return output_path


def create_packaging_ledger(output_dir, inventory_mapper, existing_data=None):
    """
    Create the packaging_ledger.xlsx file.

    Args:
        output_dir: Output directory path
        inventory_mapper: InventoryMapper instance
        existing_data: Optional dict from existing tracker
    """
    wb = create_workbook()

    # Get packaging items
    items = inventory_mapper.get_packaging()
    items_data = []

    for item in items:
        code = item['code']
        existing = existing_data.get(code, {}) if existing_data else {}

        opening_qty = existing.get('ground_stock', 0)
        wac = existing.get('wac', 0)
        opening_value = existing.get('closing_value', opening_qty * wac)

        # Create item sheet
        create_item_ledger_sheet(
            wb, code, item['name'], item['unit'],
            opening_qty, opening_value, wac
        )

        items_data.append({
            'code': code,
            'name': item['name'],
            'unit': item['unit'],
            'opening_qty': opening_qty,
            'opening_value': opening_value,
            'received_qty': 0,
            'received_value': 0,
            'issued_qty': 0,
            'issued_value': 0,
            'closing_qty': opening_qty,
            'closing_value': opening_value,
            'wac': wac
        })

    # Create dashboard
    create_dashboard_sheet(wb, "Packaging Materials Inventory", items_data)

    # Move dashboard to first position
    wb.move_sheet('Dashboard', offset=-len(items))

    output_path = Path(output_dir) / 'packaging_ledger.xlsx'
    save_workbook(wb, output_path)
    print(f"Created: {output_path}")

    return output_path


def main():
    """Main entry point."""
    if len(sys.argv) < 4:
        print(__doc__)
        print("\nUsage: python scripts/create_inventory_ledgers.py DATA_DIR PERIOD_START PERIOD_END [OUTPUT_DIR]")
        print("\nExample:")
        print("  python scripts/create_inventory_ledgers.py data/Jan2026 2026-01-01 2026-01-31")
        sys.exit(1)

    data_dir = Path(sys.argv[1])
    period_start = sys.argv[2]
    period_end = sys.argv[3]
    output_dir = Path(sys.argv[4]) if len(sys.argv) > 4 else data_dir

    # Ensure output directory exists
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Creating inventory sub-ledgers for period {period_start} to {period_end}")
    print(f"Data directory: {data_dir}")
    print(f"Output directory: {output_dir}")
    print()

    # Initialize inventory mapper
    items_file = data_dir / 'inventory_items.xlsx'
    inv_mapper = InventoryMapper(items_file if items_file.exists() else None)

    # Check for existing tracker files
    rm_tracker = data_dir.parent.parent / 'Exisitng Accounting Workflow _ reference files' / 'Books of Prime Entry' / 'Raw Material _Tracker.xlsx'
    pkg_tracker = data_dir.parent.parent / 'Exisitng Accounting Workflow _ reference files' / 'Books of Prime Entry' / 'Packaging_Tracker.xlsx'

    # Also check in data directory
    if not rm_tracker.exists():
        rm_tracker = data_dir / 'Raw Material _Tracker.xlsx'
    if not pkg_tracker.exists():
        pkg_tracker = data_dir / 'Packaging_Tracker.xlsx'

    print("Reading existing tracker data...")
    rm_data = read_existing_tracker(rm_tracker, (12001, 12018)) if rm_tracker.exists() else {}
    pkg_data = read_existing_tracker(pkg_tracker, (12100, 12106)) if pkg_tracker.exists() else {}

    if rm_data:
        print(f"  Raw Materials: Found {len(rm_data)} items from existing tracker")
    if pkg_data:
        print(f"  Packaging: Found {len(pkg_data)} items from existing tracker")
    print()

    # Create inventory master file
    create_inventory_master_file(output_dir, inv_mapper)

    # Create raw materials ledger
    create_raw_materials_ledger(output_dir, inv_mapper, rm_data)

    # Create packaging ledger
    create_packaging_ledger(output_dir, inv_mapper, pkg_data)

    print()
    print("Inventory sub-ledgers created successfully!")
    print()
    print("Output files:")
    print(f"  - {output_dir / 'inventory_items.xlsx'}")
    print(f"  - {output_dir / 'raw_materials_ledger.xlsx'}")
    print(f"  - {output_dir / 'packaging_ledger.xlsx'}")


if __name__ == '__main__':
    main()