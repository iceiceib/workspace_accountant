"""
Module 3: Bank Reconciliation
Shwe Mandalay Cafe / K&K Finance Team

Compares the internal cash book (cash_ledger.xlsx) against the external bank
statement (bank_statement.xlsx) and produces a formal reconciliation workbook.

Usage:
    python reconcile_bank.py <data_dir> <period_start> <period_end> <output_file>

    python reconcile_bank.py \\
        data/Jan2026 \\
        2026-01-01 \\
        2026-01-31 \\
        data/Jan2026/bank_reconciliation_Jan2026.xlsx

Output sheets:
    Dashboard             — summary: adjusted book = adjusted bank (PASS/FAIL)
    Reconciliation        — formal reconciliation statement
    Matched Items         — all transaction pairs successfully matched
    Outstanding Cheques   — book payments not yet cleared by bank
    Deposits in Transit   — book receipts not yet credited by bank
    Bank-Only Items       — bank charges, interest, direct debits/credits
    Adjusting Entries     — required journal entries to update the cash book
    Exceptions            — (only if validation errors found)
"""

import sys
import os
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np

# Add scripts directory to path so utils can be found
sys.path.insert(0, str(Path(__file__).parent))
from utils.excel_reader import read_xlsx, filter_by_period
from utils.excel_writer import (
    create_workbook, add_sheet, write_title, write_header_row,
    write_data_row, write_section_header, write_total_row,
    write_validation_result, auto_fit_columns, freeze_panes,
    save_workbook, NORMAL_FONT, TOTAL_FONT, NEGATIVE_FONT,
    THIN_BORDER, PASS_FILL, FAIL_FILL, WARNING_FILL,
    NUMBER_FORMAT_NEG, DATE_FORMAT, HEADER_FILL, HEADER_FONT
)
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
BANK_ACCOUNT_CODE = '1020'    # Cash at Bank
INTEREST_INCOME_CODE = '4110' # Interest Income
AR_CODE = '1100'              # Accounts Receivable
BANK_CHARGES_CODE = '5920'    # Bank Charges & Fees
INSURANCE_CODE = '5700'       # Insurance Expense
COMMS_CODE = '5220'           # Telephone & Internet (used for subscriptions)
LOAN_CODE = '2100'            # Long-term Loans
MISC_EXPENSE_CODE = '5900'    # Non-Operating Expenses (catch-all)
OTHER_INCOME_CODE = '4199'    # Miscellaneous income placeholder

DATE_PROXIMITY_DAYS = 3       # Max days apart for amount-only match


# ─────────────────────────────────────────────────────────────────────────────
# 1. DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_cash_ledger(data_dir, period_start, period_end):
    """
    Load cash_ledger.xlsx, extract opening balance, and filter period rows.
    Cash ledger perspective: Debit = money IN, Credit = money OUT.
    """
    path = Path(data_dir) / 'cash_ledger.xlsx'
    result = read_xlsx(
        path,
        required_columns=['Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance'],
        optional_columns=['Bank Account']
    )
    if result['error']:
        return None, None, None, result['error']

    df = result['data'].copy()
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
    df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
    df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce')
    df['Reference'] = df['Reference'].astype(str).str.strip()

    # Opening balance: last row before period_start
    pre_period = df[df['Date'] < pd.Timestamp(period_start)]
    if pre_period.empty:
        opening_balance = 0.0
    else:
        opening_balance = float(pre_period.iloc[-1]['Balance'])

    # Period rows only
    period_df = filter_by_period(df, 'Date', period_start, period_end)
    period_df = period_df[period_df['Reference'] != 'OB'].copy()

    # Closing balance: last row in/before period end
    at_end = df[df['Date'] <= pd.Timestamp(period_end)]
    closing_balance = float(at_end.iloc[-1]['Balance']) if not at_end.empty else opening_balance

    # Normalize: _amount > 0 = money in, < 0 = money out
    period_df['_amount'] = period_df['Debit'] - period_df['Credit']
    period_df = period_df.reset_index(drop=True)

    return period_df, opening_balance, closing_balance, None


def load_bank_statement(data_dir, period_start, period_end):
    """
    Load bank_statement.xlsx, extract opening balance, and filter period rows.
    Bank statement perspective: Debit = withdrawal (OUT), Credit = deposit (IN).
    """
    path = Path(data_dir) / 'bank_statement.xlsx'
    result = read_xlsx(
        path,
        required_columns=['Date', 'Reference', 'Description', 'Balance'],
        optional_columns=['Debit', 'Credit']
    )
    if result['error']:
        return None, None, None, result['error']

    df = result['data'].copy()
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Debit'] = pd.to_numeric(df.get('Debit', pd.Series(dtype=float)), errors='coerce').fillna(0)
    df['Credit'] = pd.to_numeric(df.get('Credit', pd.Series(dtype=float)), errors='coerce').fillna(0)
    df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce')
    df['Reference'] = df['Reference'].astype(str).str.strip()

    # Opening balance: last row before period
    pre_period = df[df['Date'] < pd.Timestamp(period_start)]
    if pre_period.empty:
        opening_balance = 0.0
    else:
        opening_balance = float(pre_period.iloc[-1]['Balance'])

    # Period rows
    period_df = filter_by_period(df, 'Date', period_start, period_end)
    period_df = period_df[period_df['Reference'] != 'OB'].copy()

    # Closing balance
    at_end = df[df['Date'] <= pd.Timestamp(period_end)]
    closing_balance = float(at_end.iloc[-1]['Balance']) if not at_end.empty else opening_balance

    # Normalize: _amount > 0 = money in (deposit), < 0 = money out (withdrawal)
    period_df['_amount'] = period_df['Credit'] - period_df['Debit']
    period_df = period_df.reset_index(drop=True)

    return period_df, opening_balance, closing_balance, None


# ─────────────────────────────────────────────────────────────────────────────
# 2. MATCHING
# ─────────────────────────────────────────────────────────────────────────────

def match_transactions(book_df, bank_df):
    """
    Match book (cash ledger) rows against bank statement rows.

    Both sides have been normalized so _amount > 0 = money in, < 0 = money out.

    Matching passes (in priority order):
      Pass 1 — Exact:    same Reference AND same _amount
      Pass 2 — Probable: same _amount, dates within DATE_PROXIMITY_DAYS

    Returns:
        matched_pairs: list of dicts {book_idx, bank_idx, match_type}
        book_only:  DataFrame of unmatched book rows
        bank_only:  DataFrame of unmatched bank rows
    """
    book_matched = set()
    bank_matched = set()
    matched_pairs = []

    book_refs = book_df['Reference'].str.upper()
    bank_refs = bank_df['Reference'].str.upper()

    # Pass 1: Exact reference + exact amount
    for bi in book_df.index:
        b_ref = book_refs[bi]
        b_amt = round(book_df.at[bi, '_amount'], 2)
        for si in bank_df.index:
            if si in bank_matched:
                continue
            s_ref = bank_refs[si]
            s_amt = round(bank_df.at[si, '_amount'], 2)
            if b_ref == s_ref and b_amt == s_amt:
                matched_pairs.append({'book_idx': bi, 'bank_idx': si, 'match_type': 'Exact'})
                book_matched.add(bi)
                bank_matched.add(si)
                break

    # Pass 2: Same amount + date within ±DATE_PROXIMITY_DAYS
    for bi in book_df.index:
        if bi in book_matched:
            continue
        b_amt = round(book_df.at[bi, '_amount'], 2)
        b_date = book_df.at[bi, 'Date']
        for si in bank_df.index:
            if si in bank_matched:
                continue
            s_amt = round(bank_df.at[si, '_amount'], 2)
            s_date = bank_df.at[si, 'Date']
            if b_amt == s_amt:
                try:
                    days_diff = abs((b_date - s_date).days)
                except Exception:
                    days_diff = 9999
                if days_diff <= DATE_PROXIMITY_DAYS:
                    matched_pairs.append({'book_idx': bi, 'bank_idx': si, 'match_type': 'Probable'})
                    book_matched.add(bi)
                    bank_matched.add(si)
                    break

    book_only = book_df[~book_df.index.isin(book_matched)].copy()
    bank_only = bank_df[~bank_df.index.isin(bank_matched)].copy()

    return matched_pairs, book_only, bank_only


def classify_book_only(book_only_df):
    """
    Classify unmatched book items.
    Book Debit (_amount > 0) = Deposit in Transit (receipt recorded but not yet on bank statement).
    Book Credit (_amount < 0) = Outstanding Cheque (payment recorded but not yet cleared).
    """
    deposits_in_transit = book_only_df[book_only_df['_amount'] > 0].copy()
    outstanding_cheques = book_only_df[book_only_df['_amount'] < 0].copy()
    return deposits_in_transit, outstanding_cheques


def classify_bank_only(bank_only_df):
    """
    Classify unmatched bank items.
    Bank Credit (_amount > 0) = credit not in book (interest, direct credits).
    Bank Debit (_amount < 0)  = debit not in book (charges, direct debits).
    """
    bank_credits = bank_only_df[bank_only_df['_amount'] > 0].copy()
    bank_debits = bank_only_df[bank_only_df['_amount'] < 0].copy()
    return bank_credits, bank_debits


# ─────────────────────────────────────────────────────────────────────────────
# 3. ADJUSTING ENTRIES
# ─────────────────────────────────────────────────────────────────────────────

def categorize_bank_item(description, amount):
    """
    Suggest a journal entry for a bank-only item.
    amount > 0 = bank credit (money in) → Dr 1020 / Cr ???
    amount < 0 = bank debit  (money out) → Dr ??? / Cr 1020

    Returns dict: category, dr_account, cr_account, dr_code, cr_code
    """
    desc = str(description).lower()
    abs_amt = abs(amount)

    if amount > 0:
        # Bank credited the account — record as income/receipt in book
        if 'interest' in desc:
            return {'category': 'Bank Interest Earned',
                    'dr_account': 'Cash at Bank', 'dr_code': BANK_ACCOUNT_CODE,
                    'cr_account': 'Interest Income', 'cr_code': INTEREST_INCOME_CODE,
                    'dr_amount': abs_amt, 'cr_amount': abs_amt}
        elif 'payment' in desc or 'collection' in desc or 'transfer in' in desc:
            return {'category': 'Direct Credit — Accounts Receivable',
                    'dr_account': 'Cash at Bank', 'dr_code': BANK_ACCOUNT_CODE,
                    'cr_account': 'Accounts Receivable', 'cr_code': AR_CODE,
                    'dr_amount': abs_amt, 'cr_amount': abs_amt}
        else:
            return {'category': 'Bank Credit — Investigate',
                    'dr_account': 'Cash at Bank', 'dr_code': BANK_ACCOUNT_CODE,
                    'cr_account': 'Other Income', 'cr_code': OTHER_INCOME_CODE,
                    'dr_amount': abs_amt, 'cr_amount': abs_amt}
    else:
        # Bank debited the account — record as expense/payment in book
        if 'charge' in desc or 'fee' in desc or 'service' in desc:
            return {'category': 'Bank Charges & Fees',
                    'dr_account': 'Bank Charges & Fees', 'dr_code': BANK_CHARGES_CODE,
                    'cr_account': 'Cash at Bank', 'cr_code': BANK_ACCOUNT_CODE,
                    'dr_amount': abs_amt, 'cr_amount': abs_amt}
        elif 'insurance' in desc:
            return {'category': 'Insurance — Auto Debit',
                    'dr_account': 'Insurance Expense', 'dr_code': INSURANCE_CODE,
                    'cr_account': 'Cash at Bank', 'cr_code': BANK_ACCOUNT_CODE,
                    'dr_amount': abs_amt, 'cr_amount': abs_amt}
        elif 'subscription' in desc or 'software' in desc or 'saas' in desc:
            return {'category': 'Software / Subscription',
                    'dr_account': 'Telephone & Internet', 'dr_code': COMMS_CODE,
                    'cr_account': 'Cash at Bank', 'cr_code': BANK_ACCOUNT_CODE,
                    'dr_amount': abs_amt, 'cr_amount': abs_amt}
        elif 'loan' in desc or 'repayment' in desc or 'instalment' in desc:
            return {'category': 'Loan Repayment',
                    'dr_account': 'Long-term Loans', 'dr_code': LOAN_CODE,
                    'cr_account': 'Cash at Bank', 'cr_code': BANK_ACCOUNT_CODE,
                    'dr_amount': abs_amt, 'cr_amount': abs_amt}
        elif 'dishonour' in desc or 'nsf' in desc or 'returned' in desc or 'bounce' in desc:
            return {'category': 'Dishonoured Cheque',
                    'dr_account': 'Accounts Receivable', 'dr_code': AR_CODE,
                    'cr_account': 'Cash at Bank', 'cr_code': BANK_ACCOUNT_CODE,
                    'dr_amount': abs_amt, 'cr_amount': abs_amt}
        else:
            return {'category': 'Direct Debit — Investigate',
                    'dr_account': 'Non-Operating Expense', 'dr_code': MISC_EXPENSE_CODE,
                    'cr_account': 'Cash at Bank', 'cr_code': BANK_ACCOUNT_CODE,
                    'dr_amount': abs_amt, 'cr_amount': abs_amt}


def build_adjusting_entries(bank_credits, bank_debits, period_end):
    """Build list of required adjusting journal entries from bank-only items."""
    entries = []
    for _, row in bank_credits.iterrows():
        cat = categorize_bank_item(row['Description'], row['_amount'])
        cat['date'] = row['Date']
        cat['reference'] = row['Reference']
        cat['description'] = row['Description']
        entries.append(cat)
    for _, row in bank_debits.iterrows():
        cat = categorize_bank_item(row['Description'], row['_amount'])
        cat['date'] = row['Date']
        cat['reference'] = row['Reference']
        cat['description'] = row['Description']
        entries.append(cat)
    return entries


# ─────────────────────────────────────────────────────────────────────────────
# 4. RECONCILIATION MATH
# ─────────────────────────────────────────────────────────────────────────────

def build_reconciliation(book_closing, bank_closing,
                         deposits_in_transit, outstanding_cheques,
                         bank_credits, bank_debits):
    """
    Calculate adjusted balances and reconciliation status.

    Bank side:
        Adjusted Bank = Bank Closing
                      + Deposits in Transit (book receipts not yet on bank)
                      - Outstanding Cheques (book payments not yet cleared)

    Book side:
        Adjusted Book = Book Closing
                      + Bank Credits not in Book (interest, direct credits)
                      - Bank Debits not in Book (charges, direct debits)
    """
    total_deposits_in_transit = float(deposits_in_transit['_amount'].sum()) \
        if not deposits_in_transit.empty else 0.0
    total_outstanding_cheques = float(outstanding_cheques['_amount'].sum()) \
        if not outstanding_cheques.empty else 0.0   # already negative

    total_bank_credits = float(bank_credits['_amount'].sum()) \
        if not bank_credits.empty else 0.0
    total_bank_debits = float(bank_debits['_amount'].sum()) \
        if not bank_debits.empty else 0.0       # already negative

    adjusted_bank = bank_closing + total_deposits_in_transit + total_outstanding_cheques
    adjusted_book = book_closing + total_bank_credits + total_bank_debits

    difference = round(adjusted_bank - adjusted_book, 2)
    reconciled = (difference == 0)

    return {
        'book_closing': book_closing,
        'bank_closing': bank_closing,
        'total_deposits_in_transit': total_deposits_in_transit,
        'total_outstanding_cheques': total_outstanding_cheques,
        'adjusted_bank': adjusted_bank,
        'total_bank_credits': total_bank_credits,
        'total_bank_debits': total_bank_debits,
        'adjusted_book': adjusted_book,
        'difference': difference,
        'reconciled': reconciled,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 5. EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_date(val):
    """Format a date value as YYYY-MM-DD string."""
    if pd.isna(val) or val is None:
        return ''
    try:
        return pd.Timestamp(val).strftime('%Y-%m-%d')
    except Exception:
        return str(val)


def _fmt_num(val):
    """Return numeric value or None for blanks."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    return float(val)


def write_dashboard(wb, recon, matched_count, period_start, period_end,
                    deposits_in_transit, outstanding_cheques,
                    bank_credits, bank_debits, exceptions):
    ws = add_sheet(wb, 'Dashboard', tab_color='00B050')
    period_str = f"{period_start}  to  {period_end}"
    row = write_title(ws, 'Bank Reconciliation — Dashboard',
                      'Shwe Mandalay Cafe', period_str)

    def kv(label, value, row, number=False, is_status=False):
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=value)
        cell.border = THIN_BORDER
        if is_status:
            cell.fill = PASS_FILL if value == 'RECONCILED' else FAIL_FILL
            cell.font = Font(bold=True, size=11, name='Arial',
                             color='006100' if value == 'RECONCILED' else '9C0006')
            cell.alignment = Alignment(horizontal='center')
        elif number:
            cell.number_format = NUMBER_FORMAT_NEG
            cell.alignment = Alignment(horizontal='right')
            cell.font = NORMAL_FONT
            if isinstance(value, (int, float)) and value < 0:
                cell.font = NEGATIVE_FONT
        else:
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='left')
        return row + 1

    # Status
    row = write_section_header(ws, 'RECONCILIATION STATUS', row)
    status_text = 'RECONCILED' if recon['reconciled'] else 'NOT RECONCILED'
    row = kv('Status', status_text, row, is_status=True)
    row = kv('Difference', recon['difference'], row, number=True)
    row += 1

    # Balances
    row = write_section_header(ws, 'BALANCES', row)
    row = kv('Cash Book Closing Balance', recon['book_closing'], row, number=True)
    row = kv('Bank Statement Closing Balance', recon['bank_closing'], row, number=True)
    row = kv('Adjusted Book Balance', recon['adjusted_book'], row, number=True)
    row = kv('Adjusted Bank Balance', recon['adjusted_bank'], row, number=True)
    row += 1

    # Counts
    row = write_section_header(ws, 'RECONCILING ITEMS', row)
    row = kv('Transactions Matched', matched_count, row)
    row = kv('Deposits in Transit', len(deposits_in_transit), row)
    row = kv('Outstanding Cheques', len(outstanding_cheques), row)
    row = kv('Bank Credits not in Book', len(bank_credits), row)
    row = kv('Bank Debits not in Book', len(bank_debits), row)
    row = kv('Adjusting Entries Required',
             len(bank_credits) + len(bank_debits), row)
    row += 1

    # Exceptions
    row = write_section_header(ws, 'EXCEPTIONS', row)
    if exceptions:
        for exc in exceptions:
            row = kv('Exception', exc, row)
    else:
        row = kv('Exceptions', 'None', row)

    auto_fit_columns(ws)
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 20
    freeze_panes(ws, row=2, col=1)


def write_reconciliation_statement(wb, recon, deposits_in_transit,
                                   outstanding_cheques, bank_credits, bank_debits,
                                   period_end):
    ws = add_sheet(wb, 'Reconciliation', tab_color='4472C4')
    row = write_title(ws, 'Bank Reconciliation Statement',
                      'Shwe Mandalay Cafe', f'As at {period_end}')

    def section(label, r):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(bold=True, size=11, name='Arial', color='1F4E79')
        c.fill = PatternFill('solid', fgColor='D6E4F0')
        c.border = THIN_BORDER
        return r + 1

    def item(label, amt, r, indent=False, total=False, double=False):
        from openpyxl.styles import Border, Side
        label_cell = ws.cell(row=r, column=1 if not indent else 2, value=label)
        label_cell.font = TOTAL_FONT if total else NORMAL_FONT
        label_cell.border = THIN_BORDER
        if indent:
            ws.cell(row=r, column=1).border = THIN_BORDER

        amount_cell = ws.cell(row=r, column=4, value=_fmt_num(amt))
        amount_cell.number_format = NUMBER_FORMAT_NEG
        amount_cell.alignment = Alignment(horizontal='right')
        if total:
            amount_cell.font = TOTAL_FONT
            from openpyxl.styles import Border, Side
            if double:
                amount_cell.border = Border(bottom=Side(style='double'))
                label_cell.border = Border(bottom=Side(style='double'))
            else:
                amount_cell.border = Border(bottom=Side(style='medium'))
                label_cell.border = Border(bottom=Side(style='medium'))
        else:
            amount_cell.font = NORMAL_FONT
            amount_cell.border = THIN_BORDER
            if isinstance(amt, (int, float)) and amt < 0:
                amount_cell.font = NEGATIVE_FONT

        if indent:
            for col in [2, 3]:
                ws.cell(row=r, column=col).border = THIN_BORDER

        return r + 1

    # ── BANK SIDE ─────────────────────────────────────────────────────────────
    row = section('BANK STATEMENT SIDE', row)
    row = item('Balance per Bank Statement', recon['bank_closing'], row)
    row += 1

    row = section('Add: Deposits in Transit (in book, not yet on bank statement)', row)
    if not deposits_in_transit.empty:
        for _, r_ in deposits_in_transit.iterrows():
            row = item(f"  {_fmt_date(r_['Date'])}  {r_['Reference']}  "
                       f"{r_['Description']}", r_['_amount'], row, indent=True)
    row = item('Total Deposits in Transit', recon['total_deposits_in_transit'],
               row, total=True)
    row += 1

    row = section('Less: Outstanding Cheques (in book, not yet cleared by bank)', row)
    if not outstanding_cheques.empty:
        for _, r_ in outstanding_cheques.iterrows():
            row = item(f"  {_fmt_date(r_['Date'])}  {r_['Reference']}  "
                       f"{r_['Description']}", r_['_amount'], row, indent=True)
    row = item('Total Outstanding Cheques', recon['total_outstanding_cheques'],
               row, total=True)
    row += 1

    row = item('ADJUSTED BANK BALANCE', recon['adjusted_bank'],
               row, total=True, double=True)
    row += 2

    # ── BOOK SIDE ─────────────────────────────────────────────────────────────
    row = section('CASH BOOK SIDE', row)
    row = item('Balance per Cash Book', recon['book_closing'], row)
    row += 1

    row = section('Add: Bank Credits not yet in Cash Book', row)
    if not bank_credits.empty:
        for _, r_ in bank_credits.iterrows():
            row = item(f"  {_fmt_date(r_['Date'])}  {r_['Reference']}  "
                       f"{r_['Description']}", r_['_amount'], row, indent=True)
    row = item('Total Bank Credits not in Book', recon['total_bank_credits'],
               row, total=True)
    row += 1

    row = section('Less: Bank Debits not yet in Cash Book', row)
    if not bank_debits.empty:
        for _, r_ in bank_debits.iterrows():
            row = item(f"  {_fmt_date(r_['Date'])}  {r_['Reference']}  "
                       f"{r_['Description']}", r_['_amount'], row, indent=True)
    row = item('Total Bank Debits not in Book', recon['total_bank_debits'],
               row, total=True)
    row += 1

    row = item('ADJUSTED CASH BOOK BALANCE', recon['adjusted_book'],
               row, total=True, double=True)
    row += 2

    # ── DIFFERENCE ────────────────────────────────────────────────────────────
    diff_row = row
    ws.cell(row=diff_row, column=1, value='Difference (must be ZERO)').font = TOTAL_FONT
    diff_cell = ws.cell(row=diff_row, column=4, value=recon['difference'])
    diff_cell.number_format = NUMBER_FORMAT_NEG
    diff_cell.alignment = Alignment(horizontal='right')
    if recon['reconciled']:
        diff_cell.fill = PASS_FILL
        diff_cell.font = Font(bold=True, size=11, name='Arial', color='006100')
    else:
        diff_cell.fill = FAIL_FILL
        diff_cell.font = Font(bold=True, size=11, name='Arial', color='9C0006')
    row += 2

    status_cell = ws.cell(row=row, column=1,
                          value='RECONCILED' if recon['reconciled'] else 'NOT RECONCILED')
    status_cell.font = Font(bold=True, size=14, name='Arial',
                            color='006100' if recon['reconciled'] else '9C0006')

    for col_letter in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col_letter].width = 50 if col_letter == 'A' else \
            (20 if col_letter in ['B', 'C'] else 18)
    freeze_panes(ws, row=2, col=1)


def write_matched_items(wb, matched_pairs, book_df, bank_df):
    ws = add_sheet(wb, 'Matched Items', tab_color='4472C4')
    row = write_title(ws, 'Matched Transactions',
                      'Transactions successfully matched between cash book and bank statement')
    headers = ['Match Type', 'Book Date', 'Book Reference', 'Book Description',
               'Book Amount', 'Bank Date', 'Bank Reference', 'Bank Description',
               'Bank Amount']
    row = write_header_row(ws, headers, row)

    for pair in matched_pairs:
        bi = pair['book_idx']
        si = pair['bank_idx']
        brow = book_df.loc[bi]
        srow = bank_df.loc[si]
        values = [
            pair['match_type'],
            _fmt_date(brow['Date']),
            brow['Reference'],
            brow['Description'],
            _fmt_num(brow['_amount']),
            _fmt_date(srow['Date']),
            srow['Reference'],
            srow['Description'],
            _fmt_num(srow['_amount']),
        ]
        row = write_data_row(ws, values, row, number_cols=[5, 9])

    auto_fit_columns(ws)
    freeze_panes(ws)


def write_outstanding_cheques(wb, outstanding_cheques):
    ws = add_sheet(wb, 'Outstanding Cheques', tab_color='4472C4')
    row = write_title(ws, 'Outstanding Cheques',
                      'Payments in cash book not yet cleared by bank')
    headers = ['Date', 'Reference', 'Description', 'Amount']
    row = write_header_row(ws, headers, row)

    total = 0.0
    for _, r_ in outstanding_cheques.iterrows():
        amt = r_['_amount']
        total += float(amt)
        row = write_data_row(ws, [
            _fmt_date(r_['Date']), r_['Reference'],
            r_['Description'], _fmt_num(amt)
        ], row, number_cols=[4])

    row = write_total_row(ws, 'Total Outstanding Cheques', [_fmt_num(total)], row,
                         double_line=True)
    auto_fit_columns(ws)
    freeze_panes(ws)


def write_deposits_in_transit(wb, deposits_in_transit):
    ws = add_sheet(wb, 'Deposits in Transit', tab_color='4472C4')
    row = write_title(ws, 'Deposits in Transit',
                      'Receipts in cash book not yet credited by bank')
    headers = ['Date', 'Reference', 'Description', 'Amount']
    row = write_header_row(ws, headers, row)

    total = 0.0
    for _, r_ in deposits_in_transit.iterrows():
        amt = r_['_amount']
        total += float(amt)
        row = write_data_row(ws, [
            _fmt_date(r_['Date']), r_['Reference'],
            r_['Description'], _fmt_num(amt)
        ], row, number_cols=[4])

    row = write_total_row(ws, 'Total Deposits in Transit', [_fmt_num(total)], row,
                         double_line=True)
    auto_fit_columns(ws)
    freeze_panes(ws)


def write_bank_only_items(wb, bank_credits, bank_debits):
    ws = add_sheet(wb, 'Bank-Only Items', tab_color='70AD47')
    row = write_title(ws, 'Bank-Only Items',
                      'Items on bank statement not found in cash book — require adjusting entries')
    headers = ['Date', 'Reference', 'Description', 'Type', 'Amount', 'Action Required']
    row = write_header_row(ws, headers, row)

    if not bank_credits.empty:
        row = write_section_header(ws, 'BANK CREDITS (deposits not in cash book)', row, col_span=6)
        for _, r_ in bank_credits.iterrows():
            cat = categorize_bank_item(r_['Description'], r_['_amount'])
            row = write_data_row(ws, [
                _fmt_date(r_['Date']), r_['Reference'], r_['Description'],
                cat['category'], _fmt_num(r_['_amount']),
                f"Dr {cat['dr_account']} / Cr {cat['cr_account']}"
            ], row, number_cols=[5])

    if not bank_debits.empty:
        row = write_section_header(ws, 'BANK DEBITS (withdrawals not in cash book)', row, col_span=6)
        for _, r_ in bank_debits.iterrows():
            cat = categorize_bank_item(r_['Description'], r_['_amount'])
            row = write_data_row(ws, [
                _fmt_date(r_['Date']), r_['Reference'], r_['Description'],
                cat['category'], _fmt_num(r_['_amount']),
                f"Dr {cat['dr_account']} / Cr {cat['cr_account']}"
            ], row, number_cols=[5])

    auto_fit_columns(ws)
    freeze_panes(ws)


def write_adjusting_entries(wb, adjusting_entries):
    ws = add_sheet(wb, 'Adjusting Entries', tab_color='4472C4')
    row = write_title(ws, 'Required Adjusting Entries',
                      'Journal entries to bring cash book in line with bank statement',
                      'Post these entries in Module 4 (Journal Adjustments)')
    headers = ['Date', 'Reference', 'Description', 'Category',
               'Dr Account', 'Dr Code', 'Dr Amount', 'Cr Account', 'Cr Code', 'Cr Amount']
    row = write_header_row(ws, headers, row)

    total_dr = 0.0
    total_cr = 0.0
    for e in adjusting_entries:
        row = write_data_row(ws, [
            _fmt_date(e['date']), e['reference'], e['description'], e['category'],
            e['dr_account'], e['dr_code'], _fmt_num(e['dr_amount']),
            e['cr_account'], e['cr_code'], _fmt_num(e['cr_amount']),
        ], row, number_cols=[7, 10])
        total_dr += e['dr_amount']
        total_cr += e['cr_amount']

    row = write_total_row(ws, 'Totals', [None, None, None, None,
                                          _fmt_num(total_dr), None, None,
                                          _fmt_num(total_cr)], row,
                         double_line=True)

    # Validation check
    balanced = round(total_dr - total_cr, 2) == 0
    ws.cell(row=row, column=1, value='Double-Entry Check').font = NORMAL_FONT
    ws.cell(row=row, column=6, value='Dr = Cr?').font = NORMAL_FONT
    write_validation_result(ws, row, 7, balanced)
    row += 1

    auto_fit_columns(ws)
    freeze_panes(ws)


def write_exceptions_sheet(wb, exceptions):
    ws = add_sheet(wb, 'Exceptions', tab_color='FF0000')
    row = write_title(ws, 'Exceptions', 'Issues requiring investigation')
    headers = ['#', 'Exception / Warning']
    row = write_header_row(ws, headers, row)
    for i, exc in enumerate(exceptions, 1):
        row = write_data_row(ws, [i, exc], row)
    auto_fit_columns(ws)
    freeze_panes(ws)


# ─────────────────────────────────────────────────────────────────────────────
# 6. MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)

    data_dir = sys.argv[1]
    period_start = sys.argv[2]
    period_end = sys.argv[3]
    output_file = sys.argv[4]

    print(f"\n{'='*60}")
    print(f"  MODULE 3 — BANK RECONCILIATION")
    print(f"  Period : {period_start}  to  {period_end}")
    print(f"  Data   : {data_dir}")
    print(f"  Output : {output_file}")
    print(f"{'='*60}\n")

    exceptions = []

    # ── Load data ─────────────────────────────────────────────────────────────
    print("Loading cash ledger...")
    book_df, book_opening, book_closing, err = load_cash_ledger(data_dir, period_start, period_end)
    if err:
        print(f"  ERROR: {err}")
        sys.exit(1)
    print(f"  Opening balance : {book_opening:,.0f}")
    print(f"  Closing balance : {book_closing:,.0f}")
    print(f"  Period rows     : {len(book_df)}")

    print("\nLoading bank statement...")
    bank_df, bank_opening, bank_closing, err = load_bank_statement(data_dir, period_start, period_end)
    if err:
        print(f"  ERROR: {err}")
        sys.exit(1)
    print(f"  Opening balance : {bank_opening:,.0f}")
    print(f"  Closing balance : {bank_closing:,.0f}")
    print(f"  Period rows     : {len(bank_df)}")

    # Opening balance cross-check
    if round(book_opening, 2) != round(bank_opening, 2):
        msg = (f"Opening balances differ: book={book_opening:,.2f}, "
               f"bank={bank_opening:,.2f}")
        print(f"  WARNING: {msg}")
        exceptions.append(msg)

    # ── Match transactions ────────────────────────────────────────────────────
    print("\nMatching transactions...")
    matched_pairs, book_only, bank_only = match_transactions(book_df, bank_df)
    print(f"  Matched          : {len(matched_pairs)}")
    print(f"  Book-only (unmatched): {len(book_only)}")
    print(f"  Bank-only (unmatched): {len(bank_only)}")

    # ── Classify unmatched ───────────────────────────────────────────────────
    deposits_in_transit, outstanding_cheques = classify_book_only(book_only)
    bank_credits, bank_debits = classify_bank_only(bank_only)

    print(f"\n  Deposits in Transit  : {len(deposits_in_transit)}")
    for _, r_ in deposits_in_transit.iterrows():
        print(f"    {_fmt_date(r_['Date'])}  {r_['Reference']:<12}  "
              f"{r_['_amount']:>12,.0f}  {r_['Description']}")

    print(f"\n  Outstanding Cheques  : {len(outstanding_cheques)}")
    for _, r_ in outstanding_cheques.iterrows():
        print(f"    {_fmt_date(r_['Date'])}  {r_['Reference']:<12}  "
              f"{r_['_amount']:>12,.0f}  {r_['Description']}")

    print(f"\n  Bank Credits not in Book : {len(bank_credits)}")
    for _, r_ in bank_credits.iterrows():
        print(f"    {_fmt_date(r_['Date'])}  {r_['Reference']:<12}  "
              f"{r_['_amount']:>12,.0f}  {r_['Description']}")

    print(f"\n  Bank Debits not in Book  : {len(bank_debits)}")
    for _, r_ in bank_debits.iterrows():
        print(f"    {_fmt_date(r_['Date'])}  {r_['Reference']:<12}  "
              f"{r_['_amount']:>12,.0f}  {r_['Description']}")

    # ── Reconciliation calculation ────────────────────────────────────────────
    recon = build_reconciliation(book_closing, bank_closing,
                                 deposits_in_transit, outstanding_cheques,
                                 bank_credits, bank_debits)

    print(f"\n{'-'*50}")
    print(f"  Cash Book Closing Balance : {recon['book_closing']:>12,.2f}")
    print(f"  + Bank Credits not in Book: {recon['total_bank_credits']:>12,.2f}")
    print(f"  + Bank Debits not in Book : {recon['total_bank_debits']:>12,.2f}")
    print(f"  Adjusted Book Balance     : {recon['adjusted_book']:>12,.2f}")
    print(f"{'-'*50}")
    print(f"  Bank Statement Balance    : {recon['bank_closing']:>12,.2f}")
    print(f"  + Deposits in Transit     : {recon['total_deposits_in_transit']:>12,.2f}")
    print(f"  + Outstanding Cheques     : {recon['total_outstanding_cheques']:>12,.2f}")
    print(f"  Adjusted Bank Balance     : {recon['adjusted_bank']:>12,.2f}")
    print(f"{'-'*50}")
    print(f"  Difference                : {recon['difference']:>12,.2f}")
    print(f"  Status: {'RECONCILED' if recon['reconciled'] else 'NOT RECONCILED'}")
    print(f"{'-'*50}")

    if not recon['reconciled']:
        exceptions.append(f"Reconciliation failed — difference = {recon['difference']:,.2f}")

    # ── Adjusting entries ────────────────────────────────────────────────────
    adjusting_entries = build_adjusting_entries(bank_credits, bank_debits, period_end)
    print(f"\nAdjusting entries required: {len(adjusting_entries)}")
    for e in adjusting_entries:
        print(f"  Dr {e['dr_code']} {e['dr_account']:<25}  "
              f"Cr {e['cr_code']} {e['cr_account']:<25}  "
              f"{e['dr_amount']:>10,.0f}  ({e['description']})")

    # ── Write Excel ──────────────────────────────────────────────────────────
    print(f"\nWriting output to: {output_file}")
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    wb = create_workbook()
    write_dashboard(wb, recon, len(matched_pairs), period_start, period_end,
                    deposits_in_transit, outstanding_cheques,
                    bank_credits, bank_debits, exceptions)
    write_reconciliation_statement(wb, recon, deposits_in_transit, outstanding_cheques,
                                   bank_credits, bank_debits, period_end)
    write_matched_items(wb, matched_pairs, book_df, bank_df)
    write_outstanding_cheques(wb, outstanding_cheques)
    write_deposits_in_transit(wb, deposits_in_transit)
    write_bank_only_items(wb, bank_credits, bank_debits)
    write_adjusting_entries(wb, adjusting_entries)
    if exceptions:
        write_exceptions_sheet(wb, exceptions)

    save_workbook(wb, output_file)

    print(f"\n{'='*60}")
    print(f"  OUTPUT: {output_file}")
    print(f"  Sheets: Dashboard | Reconciliation | Matched Items |")
    print(f"          Outstanding Cheques | Deposits in Transit |")
    print(f"          Bank-Only Items | Adjusting Entries"
          + (" | Exceptions" if exceptions else ""))
    print(f"  Reconciled: {'YES' if recon['reconciled'] else 'NO'}")
    print(f"  Difference: {recon['difference']:,.2f}")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
