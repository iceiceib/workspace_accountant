"""
Creates sample bank_statement.xlsx for testing Module 3 (Bank Reconciliation).
Consistent with cash_ledger.xlsx (Jan 2026) but with intentional differences.

Run:
    python create_bank_statement.py <output_dir>
    # e.g. python create_bank_statement.py data/Jan2026

Reconciling items built into the data:
  Items in CASH BOOK but NOT on bank statement (timing differences):
    CRJ-013  2026-01-30  Capital introduced       500,000  → Deposit in Transit
    CPJ-015  2026-01-30  Owner drawings           200,000  → Outstanding Cheque
    PAY-004  2026-01-31  Salaries - delivery      360,000  → Outstanding Cheque

  Items on BANK STATEMENT but NOT in cash book (need adjusting entries):
    INT-001    2026-01-31  Bank interest earned     15,000  → Dr 1020 / Cr 4110
    DIRECT-01  2026-01-28  Software subscription    18,000  → Dr 5220 / Cr 1020

Resulting balances:
  Cash book closing balance:      1,438,500
  Bank statement closing balance: 1,495,500
  Adjusted book balance:          1,435,500  (book + 15,000 - 18,000)
  Adjusted bank balance:          1,435,500  (bank + 500,000 - 200,000 - 360,000)
  Difference:                             0  → RECONCILED
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))


def style_sheet(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')
        c.border = THIN
        ws.column_dimensions[get_column_letter(i)].width = max(len(str(h)) + 4, 14)


def add_rows(ws, rows, start=2):
    for r, row in enumerate(rows, start):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')


def create_bank_statement(out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    # Column perspective: Debit = withdrawal (money OUT), Credit = deposit (money IN)
    headers = ['Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    style_sheet(ws, headers)

    # Opening balance row — same as cash ledger opening balance
    rows = [
        # (Date, Reference, Description, Debit=withdrawal, Credit=deposit, Balance)
        ('2025-12-31', 'OB',       'Opening Balance',                           None,    None, 1500000),

        # ── January 2026 — transactions that CLEARED the bank ─────────────────
        # (Book-only items CRJ-013, CPJ-015, PAY-004 are intentionally absent)

        ('2026-01-02', 'CRJ-001',  'Cash sales — Shop 1',                       None,  380000, 1880000),
        ('2026-01-03', 'CRJ-002',  'Cash sales — Shop 2',                       None,  290000, 2170000),
        ('2026-01-05', 'CRJ-003',  'AR receipt — Hla Min',                      None,  250000, 2420000),
        ('2026-01-05', 'CPJ-001',  'Payment — Golden Harvest',                320000,    None, 2100000),
        ('2026-01-06', 'CPJ-002',  'Rent — Shop 1',                           450000,    None, 1650000),
        ('2026-01-07', 'CRJ-004',  'Cash sales — Shop 3',                       None,  420000, 2070000),
        ('2026-01-07', 'CPJ-003',  'Rent — Shop 2',                           380000,    None, 1690000),
        ('2026-01-08', 'CPJ-004',  'Rent — Shop 3',                           350000,    None, 1340000),
        ('2026-01-10', 'CRJ-005',  'AR receipt — Kyaw Zin',                     None,  180000, 1520000),
        ('2026-01-10', 'CPJ-005',  'Electricity bill',                        125000,    None, 1395000),
        ('2026-01-10', 'CPJ-006',  'Payment — Fresh Market',                  185000,    None, 1210000),
        ('2026-01-12', 'CRJ-006',  'Cash sales — all shops',                    None,  550000, 1760000),
        ('2026-01-12', 'CPJ-007',  'Internet & telephone',                     55000,    None, 1705000),
        ('2026-01-14', 'CRJ-007',  'AR receipt — Win Htut',                     None,  450000, 2155000),
        ('2026-01-15', 'CPJ-008',  'Payment — Pack & Go',                      95000,    None, 2060000),
        ('2026-01-15', 'CPJ-009',  'Transport costs',                          85000,    None, 1975000),
        ('2026-01-17', 'CRJ-008',  'Cash sales — Shop 1',                       None,  315000, 2290000),
        ('2026-01-18', 'CPJ-010',  'Petty cash top-up',                       100000,    None, 2190000),
        ('2026-01-20', 'CRJ-009',  'AR receipt — Thida Aye',                    None,  320000, 2510000),
        ('2026-01-20', 'CPJ-011',  'Payment — Supply Hub',                    198000,    None, 2312000),
        ('2026-01-22', 'CRJ-010',  'Cash sales — Shop 2',                       None,  410000, 2722000),
        ('2026-01-22', 'CPJ-012',  'Insurance premium',                        65000,    None, 2657000),
        ('2026-01-25', 'CRJ-011',  'AR receipt — Zaw Lin',                      None,  750000, 3407000),
        ('2026-01-25', 'CPJ-013',  'Equipment repair',                         95000,    None, 3312000),
        ('2026-01-28', 'CRJ-012',  'Cash sales — all shops',                    None,  625000, 3937000),
        ('2026-01-28', 'CPJ-014',  'Marketing agency',                        120000,    None, 3817000),
        # ── BANK-ONLY: direct debit not in cash book ─────────────────────────
        ('2026-01-28', 'DIRECT-01','Software subscription (auto-debit)',        18000,    None, 3799000),
        # ── Bank charges (already in book via GJ-004) ─────────────────────────
        ('2026-01-31', 'GJ-004',   'Bank service charges',                      8500,    None, 3790500),
        ('2026-01-31', 'PAY-001',  'Salaries — kitchen staff',                750000,    None, 3040500),
        ('2026-01-31', 'PAY-002',  'Salaries — service staff',                960000,    None, 2080500),
        ('2026-01-31', 'PAY-003',  'Salaries — management',                   600000,    None, 1480500),
        # ── BANK-ONLY: interest credited by bank, not yet in book ─────────────
        ('2026-01-31', 'INT-001',  'Bank interest earned',                       None,   15000, 1495500),
        # Note: CRJ-013 (Capital 500,000)  = NOT HERE → Deposit in Transit
        # Note: CPJ-015 (Drawings 200,000) = NOT HERE → Outstanding Cheque
        # Note: PAY-004 (Salaries 360,000) = NOT HERE → Outstanding Cheque
    ]

    add_rows(ws, rows)

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 50)

    path = out / 'bank_statement.xlsx'
    wb.save(path)
    print(f"  Created: {path}")
    print(f"  Bank statement closing balance: 1,495,500")
    print(f"  Items intentionally absent from bank statement:")
    print(f"    CRJ-013  2026-01-30  Capital introduced    500,000  (Deposit in Transit)")
    print(f"    CPJ-015  2026-01-30  Owner drawings        200,000  (Outstanding Cheque)")
    print(f"    PAY-004  2026-01-31  Salaries - delivery   360,000  (Outstanding Cheque)")
    print(f"  Items on bank statement but NOT in cash book:")
    print(f"    DIRECT-01 2026-01-28  Software subscription  18,000  (Direct Debit)")
    print(f"    INT-001   2026-01-31  Bank interest earned   15,000  (Interest)")
    print(f"  Expected adjusted balance (both sides): 1,435,500")


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else \
        Path(r'C:\Users\USER\workspace_accountant\accountant-skill\data\Jan2026')
    out.mkdir(parents=True, exist_ok=True)
    print(f"Creating bank statement test data in: {out}")
    create_bank_statement(out)
    print("\nDone.")
    print("Now run Module 3:")
    print("  python scripts/reconcile_bank.py data/Jan2026 2026-01-01 2026-01-31"
          " data/Jan2026/bank_reconciliation_Jan2026.xlsx")


if __name__ == '__main__':
    main()
