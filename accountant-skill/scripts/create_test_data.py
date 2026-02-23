"""
Creates sample journal .xlsx files for testing Module 1.
Generates data for January 2026 — Shwe Mandalay Cafe.
Now includes Profit Center (PC) and Cost Center (CC) columns.

PC codes: PC01=Soft Drink, PC02=Drinking Water, PC99=Shared/Corporate
CC codes: CC101=SD Production, CC102=DW Production, CC103=Water Treatment,
          CC104=Preform Production, CC105=Filling & Packaging,
          CC201=Factory Utilities, CC202=Maintenance,
          CC301=Sales & Marketing, CC302=Administration

Run from the scripts directory:
    python create_test_data.py <output_dir>
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_headers(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')
        c.border = THIN
        ws.column_dimensions[get_column_letter(i)].width = max(len(str(h)) + 4, 14)


def add_rows(ws, rows, start_row=2):
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')


# ── Sales Journal ─────────────────────────────────────────────────────────────
def create_sales_journal(output_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Journal"
    headers = ['Date', 'Invoice No', 'Customer', 'Description',
               'Debit Account', 'Credit Account', 'Amount', 'Profit Center']
    style_headers(ws, headers)
    # PC01 = Soft Drink sales  |  PC02 = Drinking Water sales  |  PC99 = Shared/General
    rows = [
        ('2026-01-02', 'SJ-001', 'Hla Min',   'Soft drink sales - Shop 1',      1100, 4010, 250000, 'PC01'),
        ('2026-01-03', 'SJ-002', 'Kyaw Zin',  'Soft drink sales - Shop 1',      1100, 4010, 180000, 'PC01'),
        ('2026-01-05', 'SJ-003', 'Win Htut',  'Drinking water order - Shop 2',  1100, 4020, 450000, 'PC02'),
        ('2026-01-07', 'SJ-004', 'Thida Aye', 'Drinking water - Shop 3',        1100, 4030, 320000, 'PC02'),
        ('2026-01-10', 'SJ-005', 'Mg Mg',     'Soft drink sales - Shop 1',      1100, 4010, 195000, 'PC01'),
        ('2026-01-12', 'SJ-006', 'Su Su',     'Drinking water bulk - Shop 2',   1100, 4020, 600000, 'PC02'),
        ('2026-01-15', 'SJ-007', 'Aye Aye',   'Drinking water - Shop 3',        1100, 4030, 280000, 'PC02'),
        ('2026-01-17', 'SJ-008', 'Zaw Lin',   'Mixed event supply - General',   1100, 4040, 750000, 'PC99'),
        ('2026-01-20', 'SJ-009', 'Nay Chi',   'Soft drink sales - Shop 1',      1100, 4010, 215000, 'PC01'),
        ('2026-01-22', 'SJ-010', 'Tun Tun',   'Drinking water - Shop 2',        1100, 4020, 340000, 'PC02'),
        ('2026-01-25', 'SJ-011', 'Khin Khin', 'Drinking water - Shop 3',        1100, 4030, 290000, 'PC02'),
        ('2026-01-28', 'SJ-012', 'Aung Aung', 'Mixed supply - General',         1100, 4040, 480000, 'PC99'),
        ('2026-01-30', 'SJ-013', 'Ye Naung',  'Soft drink sales - Shop 1',      1100, 4010, 175000, 'PC01'),
    ]
    add_rows(ws, rows)
    path = output_dir / 'sales_journal.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ── Purchases Journal ─────────────────────────────────────────────────────────
def create_purchases_journal(output_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchases Journal"
    headers = ['Date', 'Reference', 'Supplier', 'Description',
               'Debit Account', 'Credit Account', 'Amount', 'Profit Center', 'Cost Center']
    style_headers(ws, headers)
    rows = [
        ('2026-01-03', 'PJ-001', 'Golden Harvest', 'Sugar & flavour — SD production', 5010, 2010, 320000, 'PC01', 'CC101'),
        ('2026-01-05', 'PJ-002', 'Fresh Market Co', 'Mineral salts — DW production',  5010, 2010, 185000, 'PC02', 'CC102'),
        ('2026-01-08', 'PJ-003', 'Pack & Go Ltd',   'PET bottles — SD filling',       5020, 2010,  95000, 'PC01', 'CC105'),
        ('2026-01-10', 'PJ-004', 'Golden Harvest',  'CO2 gas — SD production',        5010, 2010, 145000, 'PC01', 'CC101'),
        ('2026-01-13', 'PJ-005', 'Fresh Market Co', 'Filter media — water treatment', 5010, 2010, 265000, 'PC02', 'CC103'),
        ('2026-01-15', 'PJ-006', 'Pack & Go Ltd',   'Cups & lids — DW filling',       5020, 2010,  72000, 'PC02', 'CC105'),
        ('2026-01-18', 'PJ-007', 'Supply Hub',      'Preform blanks — SD line',       5010, 2010, 198000, 'PC01', 'CC104'),
        ('2026-01-20', 'PJ-008', 'Fresh Market Co', 'Purification chemicals — DW',    5010, 2010, 223000, 'PC02', 'CC103'),
        ('2026-01-23', 'PJ-009', 'Golden Harvest',  'Concentrate — SD production',    5010, 2010, 310000, 'PC01', 'CC101'),
        ('2026-01-27', 'PJ-010', 'Pack & Go Ltd',   'Shrink wrap — shared packaging', 5020, 2010,  58000, 'PC99', 'CC105'),
        ('2026-01-29', 'PJ-011', 'Supply Hub',      'Factory cleaning supplies',       5420, 2010,  45000, 'PC99', 'CC202'),
    ]
    add_rows(ws, rows)
    path = output_dir / 'purchases_journal.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ── Cash Receipts Journal ─────────────────────────────────────────────────────
def create_cash_receipts_journal(output_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Receipts Journal"
    headers = ['Date', 'Receipt No', 'Received From', 'Description',
               'Debit Account', 'Credit Account', 'Amount', 'Bank Account', 'Profit Center']
    style_headers(ws, headers)
    # AR collections inherit the PC of the original sale; cash sales get their own PC
    rows = [
        ('2026-01-02', 'CRJ-001', 'Cash Sales',  'Cash SD sales - Shop 1',       1020, 4010, 380000, 'Main Account', 'PC01'),
        ('2026-01-03', 'CRJ-002', 'Cash Sales',  'Cash DW sales - Shop 2',       1020, 4020, 290000, 'Main Account', 'PC02'),
        ('2026-01-05', 'CRJ-003', 'Hla Min',     'AR receipt - SJ-001 (SD)',     1020, 1100, 250000, 'Main Account', 'PC01'),
        ('2026-01-07', 'CRJ-004', 'Cash Sales',  'Cash DW sales - Shop 3',       1020, 4030, 420000, 'Main Account', 'PC02'),
        ('2026-01-10', 'CRJ-005', 'Kyaw Zin',    'AR receipt - SJ-002 (SD)',     1020, 1100, 180000, 'Main Account', 'PC01'),
        ('2026-01-12', 'CRJ-006', 'Cash Sales',  'Cash mixed sales - all shops', 1020, 4040, 550000, 'Main Account', 'PC99'),
        ('2026-01-14', 'CRJ-007', 'Win Htut',    'AR receipt - SJ-003 (DW)',     1020, 1100, 450000, 'Main Account', 'PC02'),
        ('2026-01-17', 'CRJ-008', 'Cash Sales',  'Cash SD sales - Shop 1',       1020, 4010, 315000, 'Main Account', 'PC01'),
        ('2026-01-20', 'CRJ-009', 'Thida Aye',   'AR receipt - SJ-004 (DW)',     1020, 1100, 320000, 'Main Account', 'PC02'),
        ('2026-01-22', 'CRJ-010', 'Cash Sales',  'Cash DW sales - Shop 2',       1020, 4020, 410000, 'Main Account', 'PC02'),
        ('2026-01-25', 'CRJ-011', 'Zaw Lin',     'AR receipt - SJ-008 (mixed)',  1020, 1100, 750000, 'Main Account', 'PC99'),
        ('2026-01-28', 'CRJ-012', 'Cash Sales',  'Cash mixed sales - all shops', 1020, 4040, 625000, 'Main Account', 'PC99'),
        ('2026-01-30', 'CRJ-013', 'Owner',       'Capital introduced by owner',  1020, 3010, 500000, 'Main Account', 'PC99'),
    ]
    add_rows(ws, rows)
    path = output_dir / 'cash_receipts_journal.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ── Cash Payments Journal ─────────────────────────────────────────────────────
def create_cash_payments_journal(output_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Payments Journal"
    headers = ['Date', 'Payment No', 'Paid To', 'Description',
               'Debit Account', 'Credit Account', 'Amount', 'Bank Account',
               'Profit Center', 'Cost Center']
    style_headers(ws, headers)
    rows = [
        # AP payments — balance sheet entries, no CC required but tag PC for traceability
        ('2026-01-05', 'CPJ-001', 'Golden Harvest',  'Pay PJ-001 sugar/flavour',    2010, 1020, 320000, 'Main Account', 'PC01', ''),
        # Rent — shared factory cost
        ('2026-01-06', 'CPJ-002', 'City Landlord',   'Factory rent - Block A',      5200, 1020, 450000, 'Main Account', 'PC99', 'CC201'),
        ('2026-01-07', 'CPJ-003', 'City Landlord',   'Factory rent - Block B',      5200, 1020, 380000, 'Main Account', 'PC99', 'CC201'),
        ('2026-01-08', 'CPJ-004', 'City Landlord',   'Warehouse rent',              5200, 1020, 350000, 'Main Account', 'PC99', 'CC302'),
        # Utilities
        ('2026-01-10', 'CPJ-005', 'YESC',            'Factory electricity - Jan',   5210, 1020, 125000, 'Main Account', 'PC99', 'CC201'),
        # AP payment
        ('2026-01-10', 'CPJ-006', 'Fresh Market Co', 'Pay PJ-002 mineral salts',    2010, 1020, 185000, 'Main Account', 'PC02', ''),
        # Admin
        ('2026-01-12', 'CPJ-007', 'MPT',             'Office internet & phone',     5220, 1020,  55000, 'Main Account', 'PC99', 'CC302'),
        # AP payment
        ('2026-01-15', 'CPJ-008', 'Pack & Go Ltd',   'Pay PJ-003 PET bottles',      2010, 1020,  95000, 'Main Account', 'PC01', ''),
        # Delivery
        ('2026-01-15', 'CPJ-009', 'Transport Co',    'Product delivery - Jan',      5500, 1020,  85000, 'Main Account', 'PC99', 'CC301'),
        # Petty cash (balance sheet — no CC)
        ('2026-01-18', 'CPJ-010', 'Petty Cash',      'Petty cash top-up',           1030, 1020, 100000, 'Main Account', 'PC99', ''),
        # AP payment
        ('2026-01-20', 'CPJ-011', 'Supply Hub',      'Pay PJ-007 preform blanks',   2010, 1020, 198000, 'Main Account', 'PC01', ''),
        # Insurance
        ('2026-01-22', 'CPJ-012', 'Star Insurance',  'Factory insurance premium',   5700, 1020,  65000, 'Main Account', 'PC99', 'CC302'),
        # Maintenance
        ('2026-01-25', 'CPJ-013', 'Repair Services', 'Production line maintenance', 5400, 1020,  95000, 'Main Account', 'PC99', 'CC202'),
        # Marketing
        ('2026-01-28', 'CPJ-014', 'Marketing Agency','Brand & distribution - Jan',  5600, 1020, 120000, 'Main Account', 'PC99', 'CC301'),
        # Owner drawings (equity — no CC)
        ('2026-01-30', 'CPJ-015', 'Owner',           'Owner drawings - January',    3020, 1020, 200000, 'Main Account', 'PC99', ''),
    ]
    add_rows(ws, rows)
    path = output_dir / 'cash_payments_journal.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ── Payroll Journal ───────────────────────────────────────────────────────────
def create_payroll_journal(output_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Journal"
    headers = ['Date', 'Employee / Department', 'Description',
               'Debit Account', 'Credit Account', 'Debit Amount', 'Credit Amount',
               'Profit Center', 'Cost Center']
    style_headers(ws, headers)
    rows = [
        # Production staff — tagged to their lines
        ('2026-01-31', 'SD Production Team',  'Jan salary - SD production (5 staff)',  5100, 1020, 750000,  750000,  'PC01', 'CC101'),
        ('2026-01-31', 'DW Production Team',  'Jan salary - DW production (4 staff)',  5100, 1020, 960000,  960000,  'PC02', 'CC102'),
        # Shared staff
        ('2026-01-31', 'Management',          'Jan salary - management (2 staff)',      5100, 1020, 600000,  600000,  'PC99', 'CC302'),
        ('2026-01-31', 'Sales & Delivery',    'Jan salary - sales/delivery (3 staff)', 5100, 1020, 360000,  360000,  'PC99', 'CC301'),
        # Accruals
        ('2026-01-31', 'SD Production Team',  'Jan overtime accrual - SD line',        5100, 2030,  85000,   85000,  'PC01', 'CC101'),
        ('2026-01-31', 'All Staff',           'Jan employee benefits accrual',         5110, 2030, 120000,  120000,  'PC99', 'CC302'),
    ]
    add_rows(ws, rows)
    path = output_dir / 'payroll_journal.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ── General Journal ───────────────────────────────────────────────────────────
def create_general_journal(output_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "General Journal"
    headers = ['Date', 'JV No', 'Description',
               'Debit Account', 'Credit Account', 'Debit Amount', 'Credit Amount',
               'Profit Center', 'Cost Center']
    style_headers(ws, headers)
    rows = [
        ('2026-01-31', 'GJ-001', 'Transfer raw materials to COGS - SD Jan usage',  5010, 1200, 450000, 450000, 'PC01', 'CC101'),
        ('2026-01-31', 'GJ-002', 'Prepaid insurance recognized - Jan portion',      5700, 1320,  20000,  20000, 'PC99', 'CC302'),
        ('2026-01-31', 'GJ-003', 'Write off uncollectible AR - old customer',       1110, 1100,  15000,  15000, 'PC99', ''),
        ('2026-01-31', 'GJ-004', 'Bank service charges - Jan 2026',                 5920, 1020,   8500,   8500, 'PC99', 'CC302'),
        ('2026-01-31', 'GJ-005', 'Correction: reclassify office supplies',          5410, 5900,  12000,  12000, 'PC99', 'CC302'),
    ]
    add_rows(ws, rows)
    path = output_dir / 'general_journal.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ── Chart of Accounts ─────────────────────────────────────────────────────────
def create_chart_of_accounts(output_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"
    headers = ['Account Code', 'Account Name', 'Type', 'Sub-Type', 'Normal Balance', 'Status']
    style_headers(ws, headers)
    accounts = [
        (1010, 'Cash on Hand',                               'Asset',     'Current Asset',        'Debit',  'Active'),
        (1020, 'Cash at Bank — Main Account',                'Asset',     'Current Asset',        'Debit',  'Active'),
        (1030, 'Petty Cash',                                 'Asset',     'Current Asset',        'Debit',  'Active'),
        (1100, 'Accounts Receivable',                        'Asset',     'Current Asset',        'Debit',  'Active'),
        (1110, 'Allowance for Doubtful Debts',               'Asset',     'Current Asset',        'Credit', 'Active'),
        (1200, 'Inventory — Raw Materials',                  'Asset',     'Current Asset',        'Debit',  'Active'),
        (1300, 'Prepaid Expenses',                           'Asset',     'Current Asset',        'Debit',  'Active'),
        (1310, 'Prepaid Rent',                               'Asset',     'Current Asset',        'Debit',  'Active'),
        (1320, 'Prepaid Insurance',                          'Asset',     'Current Asset',        'Debit',  'Active'),
        (1610, 'Buildings',                                  'Asset',     'Non-Current Asset',    'Debit',  'Active'),
        (1611, 'Accumulated Depreciation — Buildings',       'Asset',     'Non-Current Asset',    'Credit', 'Active'),
        (1620, 'Plant & Machinery',                          'Asset',     'Non-Current Asset',    'Debit',  'Active'),
        (1621, 'Accumulated Depreciation — Plant & Machinery','Asset',    'Non-Current Asset',    'Credit', 'Active'),
        (1630, 'Furniture & Fixtures',                       'Asset',     'Non-Current Asset',    'Debit',  'Active'),
        (1650, 'Office Equipment',                           'Asset',     'Non-Current Asset',    'Debit',  'Active'),
        (1651, 'Accumulated Depreciation — Office Equipment', 'Asset',    'Non-Current Asset',    'Credit', 'Active'),
        (2010, 'Accounts Payable',                           'Liability', 'Current Liability',    'Credit', 'Active'),
        (2020, 'Accrued Expenses',                           'Liability', 'Current Liability',    'Credit', 'Active'),
        (2030, 'Accrued Wages & Salaries',                   'Liability', 'Current Liability',    'Credit', 'Active'),
        (2040, 'Unearned Revenue',                           'Liability', 'Current Liability',    'Credit', 'Active'),
        (2060, 'Short-term Loans',                           'Liability', 'Current Liability',    'Credit', 'Active'),
        (2100, 'Long-term Loans',                            'Liability', 'Non-Current Liability', 'Credit', 'Active'),
        (3010, "Owner's Capital",                            'Equity',    'Equity',               'Credit', 'Active'),
        (3020, "Owner's Drawings",                           'Equity',    'Equity',               'Debit',  'Active'),
        (3030, 'Retained Earnings',                          'Equity',    'Equity',               'Credit', 'Active'),
        (3040, 'Current Year Earnings',                      'Equity',    'Equity',               'Credit', 'Active'),
        (4010, 'Sales Revenue — Shop 1',                     'Revenue',   'Operating Revenue',    'Credit', 'Active'),
        (4020, 'Sales Revenue — Shop 2',                     'Revenue',   'Operating Revenue',    'Credit', 'Active'),
        (4030, 'Sales Revenue — Shop 3',                     'Revenue',   'Operating Revenue',    'Credit', 'Active'),
        (4040, 'Sales Revenue — General',                    'Revenue',   'Operating Revenue',    'Credit', 'Active'),
        (4100, 'Other Income',                               'Revenue',   'Non-Operating Revenue', 'Credit', 'Active'),
        (4110, 'Interest Income',                            'Revenue',   'Non-Operating Revenue', 'Credit', 'Active'),
        (4200, 'Sales Returns & Allowances',                 'Revenue',   'Revenue Contra',       'Debit',  'Active'),
        (5010, 'Raw Materials Used',                         'Expense',   'COGS',                 'Debit',  'Active'),
        (5020, 'Packaging Costs',                            'Expense',   'COGS',                 'Debit',  'Active'),
        (5030, 'Direct Labour',                              'Expense',   'COGS',                 'Debit',  'Active'),
        (5040, 'Manufacturing Overhead',                     'Expense',   'COGS',                 'Debit',  'Active'),
        (5100, 'Salaries & Wages',                           'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5110, 'Employee Benefits',                          'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5200, 'Rent Expense',                               'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5210, 'Utilities Expense',                          'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5220, 'Telephone & Internet',                       'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5300, 'Depreciation Expense',                       'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5400, 'Repairs & Maintenance',                      'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5410, 'Office Supplies',                            'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5420, 'Cleaning & Sanitation',                      'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5500, 'Transportation & Delivery',                  'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5600, 'Marketing & Advertising',                    'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5700, 'Insurance Expense',                          'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5800, 'Bad Debt Expense',                           'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5900, 'Miscellaneous Expense',                      'Expense',   'Operating Expense',    'Debit',  'Active'),
        (5910, 'Interest Expense',                           'Expense',   'Non-Operating Expense', 'Debit', 'Active'),
        (5920, 'Bank Charges & Fees',                        'Expense',   'Non-Operating Expense', 'Debit', 'Active'),
    ]
    add_rows(ws, accounts)
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)
    path = output_dir / 'chart_of_accounts.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


# ── Profit & Cost Centers ──────────────────────────────────────────────────────
def create_profit_cost_centers(output_dir):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Profit Centers ───────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Profit Centers'
    headers = ['PC Code', 'Name']
    style_headers(ws, headers)
    rows = [
        ('PC01', 'Soft Drink'),
        ('PC02', 'Drinking Water'),
        ('PC99', 'Shared / Corporate'),
    ]
    add_rows(ws, rows)
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25

    # ── Sheet 2: Cost Centers ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Cost Centers')
    headers2 = ['CC Code', 'Name', 'Default PC']
    style_headers(ws2, headers2)
    rows2 = [
        ('CC101', 'Soft Drink Production',       'PC01'),
        ('CC102', 'Drinking Water Production',   'PC02'),
        ('CC103', 'Water Treatment',             'PC02'),
        ('CC104', 'Preform Production',          ''),
        ('CC105', 'Filling & Packaging',         ''),
        ('CC201', 'Factory Utilities',           'PC99'),
        ('CC202', 'Maintenance',                 'PC99'),
        ('CC301', 'Sales & Marketing',           'PC99'),
        ('CC302', 'Administration',              'PC99'),
    ]
    add_rows(ws2, rows2)
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 14

    path = output_dir / 'profit_cost_centers.xlsx'
    wb.save(path)
    print(f"  Created: {path}")


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r'C:\Users\USER\workspace_accountant\accountant-skill\data\Jan2026')
    out.mkdir(parents=True, exist_ok=True)
    print(f"Creating test data in: {out}")
    create_chart_of_accounts(out)
    create_profit_cost_centers(out)
    create_sales_journal(out)
    create_purchases_journal(out)
    create_cash_receipts_journal(out)
    create_cash_payments_journal(out)
    create_payroll_journal(out)
    create_general_journal(out)
    print("\nDone. All test data files created.")
    print("Run Module 1 with:")
    print("  python scripts/summarize_journals.py data/Jan2026 2026-01-01 2026-01-31 data/Jan2026/books_of_prime_entry_Jan2026.xlsx data/Jan2026/chart_of_accounts.xlsx data/Jan2026/profit_cost_centers.xlsx")


if __name__ == '__main__':
    main()
