"""
Add Financial Notes Sheet to Financial Statements (IFRS-Compliant Format)
Creates Financial Notes sheet with proper formatting and IFRS structure.
Includes Opening Balance, Movements, and Closing Balance from Trial Balance.

Usage:
    python scripts/add_financial_notes_v3.py <financial_statements_file> <trial_balance_file>

Example:
    python scripts/add_financial_notes_v3.py data/output/Feb2025/financial_statements_Feb2025.xlsx data/output/Feb2025/trial_balance_Feb2025.xlsx
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styling constants - use simple styles to avoid conflicts
NOTE_TITLE_FONT = Font(bold=True, size=11, name='Arial', color='1F4E79')
SUBHEADER_FONT = Font(bold=True, size=11, name='Arial')
NORMAL_FONT = Font(size=11, name='Arial')
TOTAL_FONT = Font(bold=True, size=11, name='Arial')
TITLE_FONT = Font(bold=True, size=14, name='Arial')
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')

# Simple borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
MEDIUM_BOTTOM = Border(bottom=Side(style='medium'))
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))

NUMBER_FORMAT = '#,##0;(#,##0);-'
INDENT = '  '

# Column indices
COL_B = 2
COL_C = 3
COL_D = 4
COL_E = 5
COL_F = 6


def read_account_data(ws):
    """Read accounts from Income Statement or Balance Sheet."""
    accounts = {}
    for row in range(1, ws.max_row + 1):
        name_val = ws.cell(row=row, column=1).value
        amount_val = ws.cell(row=row, column=2).value

        if not name_val:
            continue

        name_str = str(name_val).strip()

        # Skip headers, totals, and calculated lines
        skip_keywords = ['income statement', 'balance sheet', 'for the period', 'as at',
                        'total ', 'gross ', 'operating profit', 'net ', 'assets', 'liabilities', 'equity']
        if any(name_str.lower().startswith(kw) or name_str.lower() == kw for kw in skip_keywords):
            continue

        try:
            amount = float(amount_val) if amount_val is not None else 0
        except (ValueError, TypeError):
            continue

        accounts[name_str] = amount

    return accounts


def get_amount(accounts_dict, *names):
    """Get account amount by searching for matching names."""
    for name in names:
        if name in accounts_dict:
            return accounts_dict[name]
        for acc_name, amt in accounts_dict.items():
            if acc_name.lower().strip() == name.lower().strip():
                return amt
    return 0


def read_trial_balance(tb_path):
    """
    Read Trial Balance file to get opening/movements/closing data.
    Returns dict: {code: {'opening': x, 'debit': x, 'credit': x, 'closing': x, 'name': str}}
    """
    wb = load_workbook(tb_path)

    # Find the trial balance sheet
    ws = None
    for sheet_name in ['Trial Balance', 'Adjusted TB', 'TB']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break

    if ws is None:
        print("  Warning: Trial Balance sheet not found")
        return {}

    # Find header row (looking for 'Opening Balance' or 'Account Code')
    header_row = None
    col_map = {}

    for row in range(1, min(ws.max_row + 1, 10)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                val_str = str(val).lower().strip()
                if 'account code' in val_str or 'code' in val_str:
                    header_row = row
                    # Map columns
                    for c in range(1, ws.max_column + 1):
                        h = ws.cell(row=row, column=c).value
                        if h:
                            h_str = str(h).lower().strip()
                            if 'code' in h_str:
                                col_map['code'] = c
                            elif 'name' in h_str:
                                col_map['name'] = c
                            elif 'opening' in h_str:
                                col_map['opening'] = c
                            elif 'debit' in h_str and col_map.get('debit') is None:
                                col_map['debit'] = c
                            elif 'credit' in h_str and col_map.get('credit') is None:
                                col_map['credit'] = c
                            elif 'ending' in h_str or 'closing' in h_str:
                                col_map['closing'] = c
                    break
        if header_row:
            break

    if header_row is None:
        print("  Warning: Could not find header row in Trial Balance")
        return {}

    print(f"  Trial Balance columns found: {col_map}")

    # Read data
    accounts = {}
    for row in range(header_row + 1, ws.max_row + 1):
        code_val = ws.cell(row=row, column=col_map.get('code', 1)).value

        if not code_val:
            continue

        try:
            code = int(float(str(code_val).strip()))
        except (ValueError, TypeError):
            continue

        name = ws.cell(row=row, column=col_map.get('name', 2)).value
        opening = ws.cell(row=row, column=col_map.get('opening', 3)).value
        debit = ws.cell(row=row, column=col_map.get('debit', 4)).value
        credit = ws.cell(row=row, column=col_map.get('credit', 5)).value
        closing = ws.cell(row=row, column=col_map.get('closing', 6)).value

        def to_float(v):
            try:
                return float(v) if v is not None else 0.0
            except:
                return 0.0

        accounts[code] = {
            'code': code,
            'name': str(name).strip() if name else f'Account {code}',
            'opening': to_float(opening),
            'debit': to_float(debit),
            'credit': to_float(credit),
            'closing': to_float(closing)
        }

    wb.close()
    return accounts


def clean_balance_sheet(ws):
    """
    Remove ALL accumulated depreciation rows from Balance Sheet (IFRS-compliant).
    Assets should be shown at Net Book Value only.
    Accumulated depreciation details belong in Financial Notes (Note 16).

    Note: TOTAL ASSETS is already the Net Book Value (already factored in accumulated depreciation),
    so we just remove the rows without recalculating.
    """
    rows_to_delete = []
    acc_dep_total = 0

    for row in range(1, ws.max_row + 1):
        name_val = ws.cell(row=row, column=1).value
        amt_val = ws.cell(row=row, column=2).value
        if not name_val:
            continue
        name_str = str(name_val).strip().lower()

        # Remove ALL accumulated depreciation rows
        if 'accumulated depreciation' in name_str:
            rows_to_delete.append(row)
            try:
                acc_dep_total += float(amt_val) if amt_val else 0
            except (ValueError, TypeError):
                pass

    # Delete rows from bottom to top
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)

    print(f"    Removed {len(rows_to_delete)} accumulated depreciation rows from Balance Sheet")
    print(f"    Accumulated depreciation total: {acc_dep_total:,.2f} (now shown in Financial Notes)")
    print(f"    TOTAL ASSETS remains unchanged (Net Book Value already calculated)")


def format_balance_sheet(ws):
    """Apply proper formatting to Balance Sheet."""
    # Format title and headers
    for row in range(1, ws.max_row + 1):
        name_val = ws.cell(row=row, column=1).value
        amt_val = ws.cell(row=row, column=2).value
        name_str = str(name_val).strip().lower() if name_val else ''

        # Title row
        if name_str == 'balance sheet':
            ws.cell(row=row, column=1).font = TITLE_FONT
        # Period row
        elif 'as at' in name_str:
            ws.cell(row=row, column=1).font = Font(italic=True, size=11, name='Arial')
        # Section headers (ASSETS, LIABILITIES, EQUITY)
        elif name_str in ['assets', 'liabilities', 'equity']:
            ws.cell(row=row, column=1).font = SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='D6E4F0')
        # Total rows
        elif name_str.startswith('total'):
            ws.cell(row=row, column=1).font = TOTAL_FONT
            ws.cell(row=row, column=1).border = DOUBLE_BOTTOM
            if amt_val:
                cell = ws.cell(row=row, column=2)
                cell.font = TOTAL_FONT
                cell.border = DOUBLE_BOTTOM
                cell.number_format = NUMBER_FORMAT
        # Regular items
        elif name_val and amt_val:
            ws.cell(row=row, column=1).font = NORMAL_FONT
            cell = ws.cell(row=row, column=2)
            cell.font = NORMAL_FONT
            cell.number_format = NUMBER_FORMAT
            cell.alignment = Alignment(horizontal='right')

    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18

    # Tab color
    ws.sheet_properties.tabColor = '4472C4'


def format_income_statement(ws):
    """Apply proper formatting to Income Statement."""
    for row in range(1, ws.max_row + 1):
        name_val = ws.cell(row=row, column=1).value
        amt_val = ws.cell(row=row, column=2).value
        name_str = str(name_val).strip().lower() if name_val else ''

        # Title row
        if name_str == 'income statement':
            ws.cell(row=row, column=1).font = TITLE_FONT
        # Period row
        elif 'for the period' in name_str:
            ws.cell(row=row, column=1).font = Font(italic=True, size=11, name='Arial')
        # Total/summary rows
        elif any(name_str.startswith(kw) for kw in ['gross profit', 'operating profit', 'net profit', 'net loss']):
            ws.cell(row=row, column=1).font = TOTAL_FONT
            if amt_val:
                cell = ws.cell(row=row, column=2)
                cell.font = TOTAL_FONT
                cell.number_format = NUMBER_FORMAT
                cell.border = MEDIUM_BOTTOM
        # Regular items
        elif name_val and amt_val:
            ws.cell(row=row, column=1).font = NORMAL_FONT
            cell = ws.cell(row=row, column=2)
            cell.font = NORMAL_FONT
            cell.number_format = NUMBER_FORMAT
            cell.alignment = Alignment(horizontal='right')

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18

    # Tab color
    ws.sheet_properties.tabColor = '4472C4'


def write_financial_notes(wb, is_accounts, bs_accounts, period_str, tb_accounts=None):
    """Create Financial Notes sheet with IFRS-compliant structure."""
    if tb_accounts is None:
        tb_accounts = {}
    if 'Financial Notes' in wb.sheetnames:
        del wb['Financial Notes']

    ws = wb.create_sheet('Financial Notes')
    row = 1
    all_accounts = {**is_accounts, **bs_accounts}

    # Account code mapping for trial balance lookup
    account_code_map = {
        'Sales Revenue': 40000,
        'Cost of Goods Sold': 50000,
        'Operating Expenses': 60000,
        'Marketing & Advertising': 60000,
        'Office Salaries': 61000,
        'Meal Allowance': 62000,
        'Utilities': 63000,
        'Transportation & Distribution': 64000,
        'Factory Buildings & Office Supplies': 65000,
        'Depreciation Expenses - SG&A': 66000,
        'Inventory Write-off': 67000,
        'Other Expenses': 68000,
        'Key Management Compensation': 69000,
        'Depreciation Expenses - COGS': 53300,
        'Interest Income': 70000,
        'Other Income': 70000,
        'Cash in hand': 10000,
        'Cash in Hand': 10000,
        'Cash at Bank': 10100,
        'Accounts Receivable': 11000,
        'Inventory - Raw Material': 12000,
        'Inventory - Packaging': 12100,
        'Inventory - Finished Goods': 12200,
        'Work-in-progress': 12400,
        'Advanced Payments': 13000,
        'Deferred Preliminary Expenses': 14000,
        'Land': 15000,
        'Buildings & Structures': 15100,
        'Machinery & Equipment': 15200,
        'Office & Facility Equipment': 15300,
        'Electrical & Utility Systems': 15400,
        'Construction in Progress': 15500,
        'Motor Vehicles': 15600,
        'Accumulated Depreciation - Buildings & Structures': 15110,
        'Accumulated Depreciation - Machinery & Equipment': 15210,
        'Accumulated Depreciation - Office & Facility Equipment': 15310,
        'Accumulated Depreciation - Electrical & Utility Systems': 15410,
        'Accumulated Depreciation - Motor Vehicles': 15510,
        'Paid-up Capital': 31000,
        'Retained Earnings': 32000,
        'Accounts Payable': 20000,
        'Short-term Loans': 21000,
        'Utility Bills': 22000,
        'Wages Payable': 22200,
        'Bank Loan': 25000,
    }

    def get_tb_data(account_name):
        """Get trial balance data for an account."""
        code = account_code_map.get(account_name)
        if code and code in tb_accounts:
            return tb_accounts[code]
        return None

    has_tb = len(tb_accounts) > 0

    # Column positions - shift right if we have trial balance data
    COL_OPENING = COL_D if has_tb else None
    COL_MOVEMENT = COL_E if has_tb else None
    COL_CLOSING = COL_F if has_tb else COL_D

    def write_note_header_full(note_num, title, show_cols=None):
        """Write note header with column headers."""
        nonlocal row
        ws.cell(row=row, column=COL_B, value=f'Note {note_num}').font = NOTE_TITLE_FONT
        ws.cell(row=row, column=COL_C, value=title).font = NOTE_TITLE_FONT
        row += 1

        if has_tb:
            # Column headers
            ws.cell(row=row, column=COL_C, value='').font = SUBHEADER_FONT
            ws.cell(row=row, column=COL_OPENING, value='Opening').font = SUBHEADER_FONT
            ws.cell(row=row, column=COL_MOVEMENT, value='Movement').font = SUBHEADER_FONT
            ws.cell(row=row, column=COL_CLOSING, value='Closing').font = SUBHEADER_FONT
            for col in [COL_C, COL_OPENING, COL_MOVEMENT, COL_CLOSING]:
                ws.cell(row=row, column=col).border = MEDIUM_BOTTOM
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        else:
            ws.cell(row=row, column=COL_C, value=period_str).font = SUBHEADER_FONT
            ws.cell(row=row, column=COL_C).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=COL_C).border = MEDIUM_BOTTOM
            ws.cell(row=row, column=COL_D).border = MEDIUM_BOTTOM
        row += 1

    def write_note_header(note_num, title):
        nonlocal row
        ws.cell(row=row, column=COL_B, value=f'Note {note_num}').font = NOTE_TITLE_FONT
        ws.cell(row=row, column=COL_C, value=title).font = NOTE_TITLE_FONT
        row += 1
        # Period header
        ws.cell(row=row, column=COL_C, value=period_str).font = SUBHEADER_FONT
        ws.cell(row=row, column=COL_C).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=COL_C).border = MEDIUM_BOTTOM
        ws.cell(row=row, column=COL_D).border = MEDIUM_BOTTOM
        row += 1

    def write_line_item_full(name, opening=0, movement=0, closing=0, indent=False):
        """Write line item with opening/movement/closing columns."""
        nonlocal row
        display_name = INDENT + name if indent else name
        ws.cell(row=row, column=COL_C, value=display_name).font = NORMAL_FONT
        ws.cell(row=row, column=COL_C).border = THIN_BORDER

        if has_tb:
            cell_o = ws.cell(row=row, column=COL_OPENING, value=opening if opening != 0 else None)
            cell_o.font = NORMAL_FONT
            cell_o.border = THIN_BORDER
            cell_o.number_format = NUMBER_FORMAT
            cell_o.alignment = Alignment(horizontal='right')

            cell_m = ws.cell(row=row, column=COL_MOVEMENT, value=movement if movement != 0 else None)
            cell_m.font = NORMAL_FONT
            cell_m.border = THIN_BORDER
            cell_m.number_format = NUMBER_FORMAT
            cell_m.alignment = Alignment(horizontal='right')

            cell_c = ws.cell(row=row, column=COL_CLOSING, value=closing if closing != 0 else None)
            cell_c.font = NORMAL_FONT
            cell_c.border = THIN_BORDER
            cell_c.number_format = NUMBER_FORMAT
            cell_c.alignment = Alignment(horizontal='right')
        else:
            cell = ws.cell(row=row, column=COL_D, value=closing if closing != 0 else None)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.number_format = NUMBER_FORMAT
            cell.alignment = Alignment(horizontal='right')

        row += 1
        return row - 1

    def write_line_item(name, amount, indent=False):
        nonlocal row
        display_name = INDENT + name if indent else name
        ws.cell(row=row, column=COL_C, value=display_name).font = NORMAL_FONT
        ws.cell(row=row, column=COL_C).border = THIN_BORDER
        cell = ws.cell(row=row, column=COL_D, value=amount)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.number_format = NUMBER_FORMAT
        cell.alignment = Alignment(horizontal='right')
        row += 1
        return row - 1  # Return the row number just written

    def write_total_row(start_row, end_row):
        nonlocal row
        ws.cell(row=row, column=COL_C, value='Total').font = TOTAL_FONT
        ws.cell(row=row, column=COL_C).border = MEDIUM_BOTTOM
        cell = ws.cell(row=row, column=COL_D, value=f'=SUM(D{start_row}:D{end_row})')
        cell.font = TOTAL_FONT
        cell.border = MEDIUM_BOTTOM
        cell.number_format = NUMBER_FORMAT
        cell.alignment = Alignment(horizontal='right')
        row += 2  # Leave blank row after total

    def write_item_with_tb(account_name):
        """Write line item with trial balance data."""
        tb_data = get_tb_data(account_name)
        if tb_data:
            opening = tb_data.get('opening', 0)
            debit = tb_data.get('debit', 0)
            credit = tb_data.get('credit', 0)
            closing = tb_data.get('closing', 0)
            # Movement = Closing - Opening (actual change in balance)
            movement = closing - opening
            return write_line_item_full(account_name, opening, movement, closing, indent=True)
        else:
            # Fall back to balance sheet amount
            amt = get_amount(all_accounts, account_name)
            return write_line_item_full(account_name, 0, 0, amt, indent=True)

    # ========== NOTE 1: REVENUE ==========
    if has_tb:
        write_note_header_full(1, 'Revenue')
        write_item_with_tb('Sales Revenue')
        row += 2
    else:
        write_note_header(1, 'Revenue')
        r1 = write_line_item('Sales Revenue', get_amount(all_accounts, 'Sales Revenue'), indent=True)
        write_total_row(r1, r1)
        row += 1

    # ========== NOTE 2: COST OF GOODS SOLD ==========
    if has_tb:
        write_note_header_full(2, 'Cost of Goods Sold')
        write_item_with_tb('Cost of Goods Sold')
        row += 2
    else:
        write_note_header(2, 'Cost of Goods Sold')
        r1 = write_line_item('Cost of Goods Sold', get_amount(all_accounts, 'Cost of Goods Sold'), indent=True)
        write_total_row(r1, r1)
        row += 1

    # ========== NOTE 3: SG&A EXPENSES ==========
    if has_tb:
        write_note_header_full(3, 'SG&A Expenses')
        sga_items = ['Marketing & Advertising', 'Office Salaries', 'Meal Allowance',
                     'Utilities', 'Transportation & Distribution', 'Factory Buildings & Office Supplies',
                     'Depreciation Expenses - SG&A', 'Inventory Write-off', 'Other Expenses', 'Key Management Compensation']
        has_sga = False
        for item in sga_items:
            tb_data = get_tb_data(item)
            if tb_data and (tb_data.get('debit', 0) != 0 or tb_data.get('credit', 0) != 0):
                write_item_with_tb(item)
                has_sga = True
        if not has_sga:
            write_item_with_tb('Operating Expenses')
        row += 2
    else:
        write_note_header(3, 'SG&A Expenses')
        sga_items = [
            'Marketing & Advertising',
            'Office Salaries',
            'Meal Allowance',
            'Utilities',
            'Transportation & Distribution',
            'Factory Buildings & Office Supplies',
            'Depreciation Expenses - SG&A',
            'Inventory Write-off',
            'Other Expenses',
            'Key Management Compensation',
        ]

        item_rows = []
        for item in sga_items:
            amt = get_amount(all_accounts, item)
            if amt != 0:
                r = write_line_item(item, amt, indent=True)
                item_rows.append(r)

        if not item_rows:
            r = write_line_item('Operating Expenses', get_amount(all_accounts, 'Operating Expenses'), indent=True)
            item_rows.append(r)

        write_total_row(min(item_rows), max(item_rows))
        row += 1

    # ========== NOTE 4: DEPRECIATION & AMORTIZATION ==========
    if has_tb:
        write_note_header_full(4, 'Depreciation & Amortization')
        write_item_with_tb('Depreciation Expenses - COGS')
        write_item_with_tb('Depreciation Expenses - SG&A')
        row += 2
    else:
        write_note_header(4, 'Depreciation & Amortization')
        r1 = write_line_item('Depreciation Expenses - COGS', get_amount(all_accounts, 'Depreciation Expenses - COGS'), indent=True)
        r2 = write_line_item('Depreciation Expenses - SG&A', get_amount(all_accounts, 'Depreciation Expenses - SG&A'), indent=True)
        write_total_row(r1, r2)
        row += 1

    # ========== NOTE 5: OTHER INCOME ==========
    if has_tb:
        write_note_header_full(5, 'Other Income')
        write_item_with_tb('Interest Income')
        row += 2
    else:
        write_note_header(5, 'Other Income')
        r1 = write_line_item('Interest Income', get_amount(all_accounts, 'Interest Income', 'Other Income (Interest)'), indent=True)
        write_total_row(r1, r1)
        row += 1

    # ========== NOTE 11: CASH ==========
    if has_tb:
        write_note_header_full(11, 'Cash and Cash Equivalents')
        write_item_with_tb('Cash in hand')
        write_item_with_tb('Cash at Bank')
        row += 2
    else:
        write_note_header(11, 'Cash and Cash Equivalents')
        r1 = write_line_item('Cash in hand', get_amount(all_accounts, 'Cash in hand', 'Cash in Hand'), indent=True)
        r2 = write_line_item('Cash at Bank', get_amount(all_accounts, 'Cash at Bank', 'Cash at bank'), indent=True)
        write_total_row(r1, r2)
        row += 1

    # ========== NOTE 12: ACCOUNTS RECEIVABLE ==========
    if has_tb:
        write_note_header_full(12, 'Accounts Receivable')
        write_item_with_tb('Accounts Receivable')
        row += 2
    else:
        write_note_header(12, 'Accounts Receivable')
        r1 = write_line_item('Accounts Receivable', get_amount(all_accounts, 'Accounts Receivable'), indent=True)
        write_total_row(r1, r1)
    row += 1

    # ========== NOTE 13: INVENTORY ==========
    if has_tb:
        write_note_header_full(13, 'Inventory')
        inv_items = ['Inventory - Raw Material', 'Inventory - Packaging',
                     'Work-in-progress', 'Inventory - Finished Goods']
        for item in inv_items:
            tb_data = get_tb_data(item)
            if tb_data:
                write_item_with_tb(item)
        row += 2
    else:
        write_note_header(13, 'Inventory')
        inv_items = [
            ('Inventory - Raw Material', 'Inventory - Raw Material'),
            ('Inventory - Packaging', 'Inventory - Packaging'),
            ('Inventory - Work-in-progress', 'Work-in-progress'),
            ('Inventory - Finished Goods', 'Inventory - Finished Goods', 'Inventory - Finished Good'),
        ]

        item_rows = []
        for display, *search in inv_items:
            amt = get_amount(all_accounts, *search)
            if amt != 0:
                r = write_line_item(display, amt, indent=True)
                item_rows.append(r)

        if item_rows:
            write_total_row(min(item_rows), max(item_rows))
        else:
            row += 2
    row += 1

    # ========== NOTE 14: ADVANCE PAYMENTS ==========
    if has_tb:
        write_note_header_full(14, 'Advance Payments')
        write_item_with_tb('Advanced Payments')
        row += 2
    else:
        write_note_header(14, 'Advance Payments')
        r1 = write_line_item('Advanced Payments', get_amount(all_accounts, 'Advanced Payments', 'Advance Payments'), indent=True)
        write_total_row(r1, r1)
    row += 1

    # ========== NOTE 15: DEFERRED PRELIMINARY EXPENSES ==========
    if has_tb:
        write_note_header_full(15, 'Deferred Preliminary Expenses')
        write_item_with_tb('Deferred Preliminary Expenses')
        row += 2
    else:
        write_note_header(15, 'Deferred Preliminary Expenses')
        r1 = write_line_item('Deferred Preliminary Expenses', get_amount(all_accounts, 'Deferred Preliminary Expenses'), indent=True)
        write_total_row(r1, r1)
    row += 1

    # ========== NOTE 16: PROPERTY, PLANT AND EQUIPMENT ==========
    if has_tb:
        write_note_header_full(16, 'Property, Plant and Equipment')
        ppe_items = ['Land', 'Buildings & Structures', 'Machinery & Equipment',
                     'Office & Facility Equipment', 'Electrical & Utility Systems',
                     'Construction in Progress', 'Motor Vehicles']
        for item in ppe_items:
            tb_data = get_tb_data(item)
            if tb_data:
                write_item_with_tb(item)
        # Accumulated depreciation as single negative line
        acc_dep_codes = [15110, 15210, 15310, 15410, 15510]
        acc_dep_total = sum(tb_accounts.get(code, {}).get('closing', 0) for code in acc_dep_codes)
        if acc_dep_total != 0:
            write_line_item_full('Less: Accumulated Depreciation', 0, 0, -acc_dep_total, indent=True)
        row += 2
    else:
        write_note_header(16, 'Property, Plant and Equipment')

        # Header row for NBV
        ws.cell(row=row, column=COL_C, value='').font = SUBHEADER_FONT
        ws.cell(row=row, column=COL_D, value='Net Book Value').font = SUBHEADER_FONT
        ws.cell(row=row, column=COL_C).border = THIN_BORDER
        ws.cell(row=row, column=COL_D).border = THIN_BORDER
        ws.cell(row=row, column=COL_D).alignment = Alignment(horizontal='right')
        row += 1

        ppe_items = [
            ('Land', 'Land'),
            ('Buildings & Structures', 'Buildings & Structures'),
            ('Machinery & Equipment', 'Machinery & Equipment'),
            ('Office & Facility Equipment', 'Office & Facility Equipment'),
            ('Electrical & Utility Systems', 'Electrical & Utility Systems'),
            ('Construction in Progress', 'Construction in Progress'),
            ('Motor Vehicles', 'Motor Vehicles'),
        ]

        item_rows = []
        for display, search in ppe_items:
            amt = get_amount(all_accounts, search, display)
            if amt != 0:
                r = write_line_item(display, amt, indent=True)
                item_rows.append(r)

        # Accumulated depreciation as single negative line
        acc_dep_items = ['Accumulated Depreciation - Buildings & Structures',
                         'Accumulated Depreciation - Machinery & Equipment',
                         'Accumulated Depreciation - Office & Facility Equipment',
                         'Accumulated Depreciation - Electrical & Utility Systems',
                         'Accumulated Depreciation - Motor Vehicles',
                         'Accumulated Depreciation - Vehicles']
        acc_dep_total = sum(get_amount(all_accounts, item) for item in acc_dep_items)

        if acc_dep_total != 0:
            r = write_line_item('Less: Accumulated Depreciation', -acc_dep_total, indent=True)
            item_rows.append(r)

        if item_rows:
            write_total_row(min(item_rows), max(item_rows))
    row += 1

    # ========== NOTE 17: PAID-UP CAPITAL ==========
    if has_tb:
        write_note_header_full(17, 'Paid-up Capital')
        write_item_with_tb('Paid-up Capital')
        row += 2
    else:
        write_note_header(17, 'Paid-up Capital')
        r1 = write_line_item('Paid-up Capital', get_amount(all_accounts, 'Paid-up Capital', 'Paid up Capital'), indent=True)
        write_total_row(r1, r1)
    row += 1

    # ========== NOTE 18: RETAINED EARNINGS ==========
    if has_tb:
        write_note_header_full(18, 'Retained Earnings')
        write_item_with_tb('Retained Earnings')
        row += 2
    else:
        write_note_header(18, 'Retained Earnings')
        r1 = write_line_item('Retained Earnings', get_amount(all_accounts, 'Retained Earnings'), indent=True)
        write_total_row(r1, r1)
    row += 1

    # ========== NOTE 19: LIABILITIES ==========
    if has_tb:
        write_note_header_full(19, 'Liabilities')
        liab_items = ['Accounts Payable', 'Short-term Loans', 'Utility Bills',
                      'Wages Payable', 'Bank Loan']
        for item in liab_items:
            tb_data = get_tb_data(item)
            if tb_data:
                write_item_with_tb(item)
        row += 2
    else:
        write_note_header(19, 'Liabilities')
        liab_items = [
            ('Accounts Payable', 'Accounts Payable'),
            ('Short-term Loans', 'Short-term Loans'),
            ('Utility Bills', 'Utility Bills'),
            ('Wages Payable', 'Wages Payable'),
            ('Bank Loan', 'Bank Loan'),
        ]

        item_rows = []
        for display, search in liab_items:
            amt = get_amount(all_accounts, search)
            if amt != 0:
                r = write_line_item(display, amt, indent=True)
                item_rows.append(r)

        if item_rows:
            write_total_row(min(item_rows), max(item_rows))
        else:
            r = write_line_item('Total Liabilities', 0, indent=True)
            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 18
    if has_tb:
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18

    # Tab color
    ws.sheet_properties.tabColor = '70AD47'


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    filepath = Path(sys.argv[1])
    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    # Check for trial balance file
    tb_accounts = {}
    if len(sys.argv) >= 3:
        tb_path = Path(sys.argv[2])
        if tb_path.exists():
            print(f"Reading Trial Balance: {tb_path.name}")
            tb_accounts = read_trial_balance(tb_path)
            print(f"  Trial Balance accounts: {len(tb_accounts)}")

    print(f"Processing: {filepath.name}")

    wb = load_workbook(filepath)

    is_accounts = {}
    bs_accounts = {}

    if 'Income Statement' in wb.sheetnames:
        is_accounts = read_account_data(wb['Income Statement'])
        print(f"  Income Statement: {len(is_accounts)} accounts")
        # Format Income Statement
        format_income_statement(wb['Income Statement'])

    if 'Balance Sheet' in wb.sheetnames:
        bs_accounts = read_account_data(wb['Balance Sheet'])
        print(f"  Balance Sheet: {len(bs_accounts)} accounts")
        # Clean Balance Sheet - remove accumulated depreciation items
        clean_balance_sheet(wb['Balance Sheet'])
        # Format Balance Sheet
        format_balance_sheet(wb['Balance Sheet'])

    if not is_accounts and not bs_accounts:
        print("ERROR: No account data found")
        sys.exit(1)

    period_str = filepath.stem.replace('financial_statements_', '')
    write_financial_notes(wb, is_accounts, bs_accounts, period_str, tb_accounts)

    wb.save(filepath)
    print(f"  Saved with Financial Notes sheet")


if __name__ == '__main__':
    main()