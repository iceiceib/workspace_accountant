#!/usr/bin/env python
"""
Batch process all months from Feb 2025 to Sep 2025.
Extracts data from General Ledger and creates output files for each month.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import subprocess
import sys

# Paths
SOURCE_DIR = Path('Exisitng Accounting Workflow _ reference files')
OUTPUT_INPUT = Path('data/input')
OUTPUT_OUTPUT = Path('data/output')

# Months to process
MONTHS = [
    ('2025-02-01', '2025-02-28', 'Feb2025'),
    ('2025-03-01', '2025-03-31', 'Mar2025'),
    ('2025-04-01', '2025-04-30', 'Apr2025'),
    ('2025-05-01', '2025-05-31', 'May2025'),
    ('2025-06-01', '2025-06-30', 'Jun2025'),
    ('2025-07-01', '2025-07-31', 'Jul2025'),
    ('2025-08-01', '2025-08-31', 'Aug2025'),
    ('2025-09-01', '2025-09-30', 'Sep2025'),
]

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def write_simple_excel(data, filepath, sheet_name='Sheet1'):
    """Write a simple Excel file from list of dicts or DataFrame."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if isinstance(data, list):
        if len(data) > 0:
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame()
    else:
        df = data

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, (int, float)) and not pd.isna(value):
                cell.number_format = '#,##0.00'

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save(filepath)


def process_month(gl_df, start_date, end_date, period_name):
    """Process a single month and create all output files."""
    print(f"\n{'='*60}")
    print(f"Processing: {period_name}")
    print(f"Period: {start_date} to {end_date}")
    print('='*60)

    # Create output directory
    output_dir = OUTPUT_OUTPUT / period_name
    output_dir.mkdir(parents=True, exist_ok=True)

    # Filter GL for this period
    period_gl = gl_df[(gl_df['Date'] >= start_date) & (gl_df['Date'] <= end_date)].copy()
    print(f"  GL transactions: {len(period_gl)}")

    if len(period_gl) == 0:
        print(f"  SKIPPING - No transactions found")
        return False

    # ==========================================================================
    # Create Cash Receipts Journal
    # ==========================================================================
    cash_receipts = []

    # Sales Revenue entries
    sales_df = period_gl[period_gl['COA Account Number'] == 40000]
    for idx, row in sales_df.iterrows():
        desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
        amount = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0
        cash_receipts.append({
            'Date': row['Date'],
            'Receipt No': f'CR-{row["Date"].strftime("%m%d")}-{len(cash_receipts)+1:03d}',
            'Received From': 'Cash Sales',
            'Description': desc if desc else 'Sale of Drinking Water',
            'Amount': amount,
            'Bank Account': 'Main',
            'Debit Account': 10100,
            'Credit Account': 40000,
        })

    # Other cash receipts (Interest, Capital)
    other_receipts = period_gl[(period_gl['COA Account Number'] == 70000) |
                                (period_gl['COA Account Number'] == 31000)]
    for idx, row in other_receipts.iterrows():
        desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
        credit_acct = int(row['COA Account Number'])
        amount = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0
        cash_receipts.append({
            'Date': row['Date'],
            'Receipt No': f'CR-{row["Date"].strftime("%m%d")}-{len(cash_receipts)+1:03d}',
            'Received From': desc[:50] if desc else 'Other Receipt',
            'Description': desc,
            'Amount': amount,
            'Bank Account': 'Main',
            'Debit Account': 10100,
            'Credit Account': credit_acct,
        })

    print(f"  Cash Receipts: {len(cash_receipts)} entries")

    # ==========================================================================
    # Create Cash Payments Journal
    # ==========================================================================
    cash_payments = []

    # Purchases
    for acct in [50010, 50110]:
        purch_df = period_gl[period_gl['COA Account Number'] == acct]
        for idx, row in purch_df.iterrows():
            desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else f'Purchase Account {acct}'
            amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
            cash_payments.append({
                'Date': row['Date'],
                'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
                'Paid To': desc[:50],
                'Description': desc,
                'Amount': amount,
                'Bank Account': 'Main',
                'Debit Account': acct,
                'Credit Account': 10100,
            })

    # Operating Expenses
    expense_accounts = [53000, 53100, 53200, 65000, 14000]
    for acct in expense_accounts:
        exp_df = period_gl[period_gl['COA Account Number'] == acct]
        for idx, row in exp_df.iterrows():
            desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
            amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
            if amount > 0:
                cash_payments.append({
                    'Date': row['Date'],
                    'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
                    'Paid To': desc[:50] if desc else f'Payment for account {acct}',
                    'Description': desc,
                    'Amount': amount,
                    'Bank Account': 'Main',
                    'Debit Account': acct,
                    'Credit Account': 10100,
                })

    # Construction in Progress & Machinery
    for acct in [15500, 15200]:
        cip_df = period_gl[period_gl['COA Account Number'] == acct]
        for idx, row in cip_df.iterrows():
            desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
            amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
            if amount > 0:
                cash_payments.append({
                    'Date': row['Date'],
                    'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
                    'Paid To': desc[:50],
                    'Description': desc,
                    'Amount': amount,
                    'Bank Account': 'Main',
                    'Debit Account': acct,
                    'Credit Account': 10100,
                })

    # Advanced payments
    adv_df = period_gl[period_gl['COA Account Number'] == 13000]
    for idx, row in adv_df.iterrows():
        amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
        if amount > 0:
            desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else 'Advance Payment'
            cash_payments.append({
                'Date': row['Date'],
                'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
                'Paid To': desc[:50],
                'Description': desc,
                'Amount': amount,
                'Bank Account': 'Main',
                'Debit Account': 13000,
                'Credit Account': 10100,
            })

    print(f"  Cash Payments: {len(cash_payments)} entries")

    # ==========================================================================
    # Create General Journal
    # ==========================================================================
    general_journal = []

    # Inventory adjustments
    inv_accounts = [50000, 50020, 50100, 50120, 50200, 50220, 12000, 12100, 12200]
    inv_df = period_gl[period_gl['COA Account Number'].isin(inv_accounts)]
    for idx, row in inv_df.iterrows():
        desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
        debit = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
        credit = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0
        if debit > 0 or credit > 0:
            general_journal.append({
                'Date': row['Date'],
                'JV No': f'JV-{row["Date"].strftime("%m")}-{len(general_journal)+1:03d}',
                'Description': desc,
                'Debit Account': int(row['COA Account Number']),
                'Credit Account': int(row['COA Account Number']),
                'Debit Amount': debit,
                'Credit Amount': credit,
            })

    # Depreciation
    dep_accounts = [15110, 15210, 15410, 66000, 53300]
    dep_df = period_gl[period_gl['COA Account Number'].isin(dep_accounts)]
    for idx, row in dep_df.iterrows():
        desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else 'Depreciation'
        debit = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
        credit = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0
        if debit > 0 or credit > 0:
            general_journal.append({
                'Date': row['Date'],
                'JV No': f'JV-{row["Date"].strftime("%m")}-{len(general_journal)+1:03d}',
                'Description': desc,
                'Debit Account': int(row['COA Account Number']),
                'Credit Account': int(row['COA Account Number']),
                'Debit Amount': debit,
                'Credit Amount': credit,
            })

    print(f"  General Journal: {len(general_journal)} entries")

    # ==========================================================================
    # Write Input Files (temporarily update journals)
    # ==========================================================================
    journals_dir = OUTPUT_INPUT / 'journals'

    write_simple_excel(cash_receipts, journals_dir / 'cash_receipts_journal.xlsx', 'Cash Receipts')
    write_simple_excel(cash_payments, journals_dir / 'cash_payments_journal.xlsx', 'Cash Payments')
    write_simple_excel(general_journal, journals_dir / 'general_journal.xlsx', 'General Journal')

    # Update General Ledger
    gl_output = period_gl[['Date', 'COA Account Number', 'Account Name', 'Descritpion', 'Debit (MMK)', 'Credit (MMK)', 'Account Balance (MMK)']].copy()
    gl_output.columns = ['Date', 'Account Code', 'Account Name', 'Description', 'Debit', 'Credit', 'Balance']
    gl_output = gl_output.sort_values(['Account Code', 'Date'])
    write_simple_excel(gl_output, OUTPUT_INPUT / 'ledgers' / 'general_ledger.xlsx', 'General Ledger')

    # ==========================================================================
    # Run Modules
    # ==========================================================================
    print(f"\n  Running Module 1 (Summarize Journals)...")
    cmd = f'python scripts/summarize_journals.py data/input/journals {start_date} {end_date} data/output/{period_name}/books_of_prime_entry_{period_name}.xlsx data/input/master'
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"    ERROR: {result.stderr}")
    else:
        print(f"    Done")

    print(f"  Running Module 2 (Summarize Ledgers)...")
    cmd = f'python scripts/summarize_ledgers.py data/input/ledgers {start_date} {end_date} data/output/{period_name}/ledger_summary_{period_name}.xlsx data/input/master'
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"    ERROR: {result.stderr}")
    else:
        print(f"    Done")

    print(f"  Running Module 5 (Trial Balance)...")
    cmd = f'python scripts/generate_trial_balance.py data/input/ledgers data/output/{period_name} {start_date} {end_date} data/output/{period_name}/trial_balance_{period_name}.xlsx'
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"    ERROR: {result.stderr}")
    else:
        print(f"    Done")

    print(f"  Running Module 6 (Financial Statements)...")
    cmd = f'python scripts/generate_financials.py data/input/ledgers data/output/{period_name} {start_date} {end_date} data/output/{period_name}/financial_statements_{period_name}.xlsx'
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"    ERROR: {result.stderr}")
    else:
        print(f"    Done")

    return True


def main():
    print("="*60)
    print("BATCH PROCESSING: Feb 2025 - Sep 2025")
    print("="*60)

    # Read the full General Ledger once
    print("\nLoading General Ledger...")
    gl_df = pd.read_excel(SOURCE_DIR / 'Ledger Accounts' / 'General_Ledger_edited.xlsx', header=3)
    gl_df = gl_df.dropna(how='all')
    gl_df['Date'] = pd.to_datetime(gl_df['Date'], errors='coerce')
    gl_df = gl_df[gl_df['Date'].notna()]
    print(f"  Total GL rows: {len(gl_df)}")

    # Process each month
    results = []
    for start_date, end_date, period_name in MONTHS:
        success = process_month(gl_df, start_date, end_date, period_name)
        results.append((period_name, success))

    # Summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    for period_name, success in results:
        status = "DONE" if success else "SKIPPED"
        print(f"  {period_name}: {status}")

    print("\nOutput files saved to: data/output/<PERIOD>/")


if __name__ == '__main__':
    main()