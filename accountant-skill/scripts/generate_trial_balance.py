"""
Module 5: Generate Trial Balance
Shwe Mandalay Cafe / K&K Finance Team

Produces a Trial Balance workbook containing:
  - Unadjusted Trial Balance  (GL closing balances before adjusting entries)
  - Adjustments               (ADJ- entries from Module 4, individual + per-account)
  - Adjusted Trial Balance    (balances after applying all adjustments)
  - TB Worksheet              (6-column combined view: Unadj | Adj entries | Adjusted)
  - Dashboard                 (Dr=Cr validation for both TB columns)

Usage:
    python generate_trial_balance.py <data_dir> <period_start> <period_end> <output_file>

    python generate_trial_balance.py \\
        data/Jan2026 \\
        2026-01-01 \\
        2026-01-31 \\
        data/Jan2026/trial_balance_Jan2026.xlsx

Output sheets:
    Dashboard     -- totals, Dr=Cr validation for unadjusted and adjusted TB
    Unadjusted TB -- GL closing balances before adjustments
    Adjustments   -- ADJ- entries (individual journal list + per-account summary)
    Adjusted TB   -- balances after applying all adjusting entries
    TB Worksheet  -- 6-column combined worksheet view
    Exceptions    -- (only if errors found)
"""

import sys
import os
from pathlib import Path

import pandas as pd
import numpy as np

sys.path.insert(0, str(Path(__file__).parent))
from utils.excel_reader import read_all_sheets
from utils.excel_writer import (
    create_workbook, add_sheet, write_title, write_header_row,
    write_data_row, write_section_header, write_total_row,
    write_validation_result, auto_fit_columns, freeze_panes,
    save_workbook, NORMAL_FONT, TOTAL_FONT, THIN_BORDER,
    PASS_FILL, FAIL_FILL
)
from utils.coa_mapper import COAMapper
from openpyxl.styles import Font, Alignment


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _n(val):
    """Return numeric value or None; blanks out zero / NaN for TB display."""
    if val is None:
        return None
    try:
        v = float(val)
    except (ValueError, TypeError):
        return None
    if np.isnan(v) or abs(v) < 0.005:
        return None
    return v


def _norm_code(val):
    """Normalize float-string account codes: '1020.0' -> '1020'."""
    try:
        return str(int(float(str(val).strip())))
    except (ValueError, TypeError):
        return str(val).strip()


def _find_col(df, candidates):
    """Return the first column name from candidates that exists in df.columns."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _normalize_cols(df):
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def _fmt_date(val):
    try:
        return pd.Timestamp(val).strftime('%Y-%m-%d')
    except Exception:
        return str(val) if val else ''


# ---------------------------------------------------------------------------
# 1. Load GL Balances (unadjusted closing balances)
# ---------------------------------------------------------------------------

def load_gl_balances(data_dir, period_start, period_end, coa):
    """
    Read general_ledger.xlsx and compute the closing balance per account as at
    period_end. These are the UNADJUSTED balances for the trial balance.

    Supports:
      - Single-sheet format: all accounts in one sheet with Account Code column
      - Multi-sheet format:  one sheet per account (sheet name contains account code)

    Returns:
        dict[code_int -> {name, type, sub_type, normal_balance,
                          opening, period_dr, period_cr, closing}]
        and error string or None
    """
    path = Path(data_dir) / 'general_ledger.xlsx'
    if not path.exists():
        return {}, f"general_ledger.xlsx not found in {data_dir}"

    result = read_all_sheets(path)
    if result['error']:
        return {}, result['error']

    sheets = result['data']
    if not sheets:
        return {}, "general_ledger.xlsx is empty"

    balances = {}

    if len(sheets) == 1:
        # ── Single-sheet format ──────────────────────────────────────────────
        df = _normalize_cols(list(sheets.values())[0])

        code_col    = _find_col(df, ['account code', 'code', 'acct code', 'account_code', 'no.'])
        date_col    = _find_col(df, ['date', 'trans date', 'entry date'])
        debit_col   = _find_col(df, ['debit', 'dr'])
        credit_col  = _find_col(df, ['credit', 'cr'])
        balance_col = _find_col(df, ['balance', 'running balance', 'bal'])

        if code_col is None:
            return {}, "General Ledger: 'Account Code' column not found."
        if date_col is None:
            return {}, "General Ledger: 'Date' column not found."

        df[code_col] = pd.to_numeric(df[code_col], errors='coerce')
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

        for code_val in df[code_col].dropna().unique():
            code     = int(code_val)
            info     = coa.get_account(code)
            normal_bal = info['normal_balance'].lower() if info else 'debit'
            acct_name  = info['name']     if info else f'Account {code}'
            acct_type  = info['type']     if info else 'Unknown'
            sub_type   = info['sub_type'] if info else ''

            acct_df  = df[df[code_col] == code_val].copy()

            # Opening balance: last Balance value on rows before period_start
            pre_rows = acct_df[acct_df[date_col] < pd.Timestamp(period_start)]
            opening  = 0.0
            if not pre_rows.empty and balance_col:
                last_val = pd.to_numeric(pre_rows.iloc[-1].get(balance_col), errors='coerce')
                if not pd.isna(last_val):
                    opening = float(last_val)

            # Period rows
            period_df = acct_df[
                acct_df[date_col].notna() &
                (acct_df[date_col] >= pd.Timestamp(period_start)) &
                (acct_df[date_col] <= pd.Timestamp(period_end))
            ]
            period_dr = pd.to_numeric(
                period_df[debit_col],  errors='coerce').fillna(0).sum() if debit_col  else 0.0
            period_cr = pd.to_numeric(
                period_df[credit_col], errors='coerce').fillna(0).sum() if credit_col else 0.0

            # Try Balance column for closing; fall back to computed value
            closing = None
            if balance_col and not period_df.empty:
                last_val = pd.to_numeric(period_df.iloc[-1].get(balance_col), errors='coerce')
                if not pd.isna(last_val):
                    closing = float(last_val)
            if closing is None:
                closing = (opening + period_dr - period_cr
                           if normal_bal == 'debit'
                           else opening - period_dr + period_cr)

            # Skip completely-zero accounts
            if abs(opening) < 0.01 and abs(period_dr) < 0.01 and abs(period_cr) < 0.01:
                continue

            balances[code] = {
                'name':           acct_name,
                'type':           acct_type,
                'sub_type':       sub_type,
                'normal_balance': normal_bal,
                'opening':        round(opening, 2),
                'period_dr':      round(period_dr, 2),
                'period_cr':      round(period_cr, 2),
                'closing':        round(closing, 2),
            }

    else:
        # ── Multi-sheet format ───────────────────────────────────────────────
        for sheet_name, raw_df in sheets.items():
            df = _normalize_cols(raw_df)
            try:
                code = int(''.join(c for c in sheet_name if c.isdigit()))
            except ValueError:
                continue

            info = coa.get_account(code)
            if info is None:
                continue

            date_col    = _find_col(df, ['date', 'trans date'])
            debit_col   = _find_col(df, ['debit', 'dr'])
            credit_col  = _find_col(df, ['credit', 'cr'])
            balance_col = _find_col(df, ['balance', 'bal'])
            if date_col is None:
                continue

            normal_bal = info['normal_balance'].lower()
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

            pre_rows = df[df[date_col] < pd.Timestamp(period_start)]
            opening  = 0.0
            if not pre_rows.empty and balance_col:
                last_val = pd.to_numeric(pre_rows.iloc[-1].get(balance_col), errors='coerce')
                if not pd.isna(last_val):
                    opening = float(last_val)

            period_df = df[
                df[date_col].notna() &
                (df[date_col] >= pd.Timestamp(period_start)) &
                (df[date_col] <= pd.Timestamp(period_end))
            ]
            period_dr = pd.to_numeric(
                period_df[debit_col],  errors='coerce').fillna(0).sum() if debit_col  else 0.0
            period_cr = pd.to_numeric(
                period_df[credit_col], errors='coerce').fillna(0).sum() if credit_col else 0.0

            closing = None
            if balance_col and not period_df.empty:
                last_val = pd.to_numeric(period_df.iloc[-1].get(balance_col), errors='coerce')
                if not pd.isna(last_val):
                    closing = float(last_val)
            if closing is None:
                closing = (opening + period_dr - period_cr
                           if normal_bal == 'debit'
                           else opening - period_dr + period_cr)

            if abs(opening) < 0.01 and abs(period_dr) < 0.01 and abs(period_cr) < 0.01:
                continue

            balances[code] = {
                'name':           info['name'],
                'type':           info['type'],
                'sub_type':       info['sub_type'],
                'normal_balance': normal_bal,
                'opening':        round(opening, 2),
                'period_dr':      round(period_dr, 2),
                'period_cr':      round(period_cr, 2),
                'closing':        round(closing, 2),
            }

    return balances, None


# ---------------------------------------------------------------------------
# 2. Load Adjusting Entries from Module 4 output
# ---------------------------------------------------------------------------

def load_adj_entries(data_dir):
    """
    Read the 'All Entries' sheet from adjusting_entries_*.xlsx.
    Uses dynamic header-row detection because write_title() places a title block
    above the actual column headers.

    Returns: (list of entry dicts, error string or None)
    """
    data_dir   = Path(data_dir)
    candidates = list(data_dir.glob('adjusting_entries*.xlsx'))
    if not candidates:
        return [], "No adjusting_entries*.xlsx found -- adjustments skipped."

    adj_file = candidates[0]

    try:
        # Detect header row: scan for a row containing 'Dr Code' or 'Entry No.'
        df_raw = pd.read_excel(adj_file, sheet_name='All Entries', header=None)
        header_row_idx = None
        for i, row_vals in df_raw.iterrows():
            row_strs = [str(v).strip() for v in row_vals.values]
            if 'Dr Code' in row_strs or 'Entry No.' in row_strs:
                header_row_idx = i
                break
        if header_row_idx is None:
            return [], (f"Could not find header row in 'All Entries' sheet "
                        f"of {adj_file.name}")

        df = pd.read_excel(adj_file, sheet_name='All Entries', header=header_row_idx)
    except Exception as e:
        return [], f"Could not read 'All Entries' from {adj_file.name}: {e}"

    df.columns = [str(c).strip() for c in df.columns]

    needed  = ['Dr Code', 'Debit Amount', 'Cr Code', 'Credit Amount']
    missing = [c for c in needed if c not in df.columns]
    if missing:
        return [], (f"'All Entries' sheet missing columns: {missing}. "
                    f"Found: {list(df.columns)}")

    entries = []
    for _, row in df.iterrows():
        try:
            dr_amt = float(row['Debit Amount'])
            cr_amt = float(row['Credit Amount'])
        except (ValueError, TypeError):
            continue
        if pd.isna(dr_amt) or pd.isna(cr_amt) or dr_amt <= 0:
            continue

        # Skip grand-total / validation rows (Dr Code is not a numeric account code)
        dr_code_str = _norm_code(row['Dr Code'])
        cr_code_str = _norm_code(row['Cr Code'])
        if not dr_code_str.isdigit() or not cr_code_str.isdigit():
            continue

        entries.append({
            'ref':         str(row.get('Entry No.', '')).strip(),
            'date':        row.get('Date', ''),
            'type':        str(row.get('Type', '')).strip(),
            'description': str(row.get('Category / Description', '')).strip(),
            'dr_code':     dr_code_str,
            'dr_name':     str(row.get('Debit Account', '')).strip(),
            'dr_amount':   round(dr_amt, 2),
            'cr_code':     cr_code_str,
            'cr_name':     str(row.get('Credit Account', '')).strip(),
            'cr_amount':   round(cr_amt, 2),
        })

    return entries, None


# ---------------------------------------------------------------------------
# 3. Build Trial Balance Rows
# ---------------------------------------------------------------------------

def _tb_display(balance, normal_bal):
    """
    Return (dr_col_value, cr_col_value) for TB display.
    Positive balance shows on the normal side; negative (abnormal) on opposite.
    """
    b = round(balance, 2)
    if normal_bal == 'debit':
        return (b, None) if b >= 0 else (None, abs(b))
    else:   # credit-normal
        return (None, b) if b >= 0 else (abs(b), None)


def build_tb_rows(gl_balances, adj_entries, coa):
    """
    Combine GL balances + adjusting entries to produce trial balance rows.

    For each account code that appears in either the GL or the adjusting entries:
      unadj_bal  = GL closing balance (before adjustments)
      adj_dr_tot = sum of all ADJ debits to this account
      adj_cr_tot = sum of all ADJ credits to this account
      adj_bal    = unadj_bal +/- net adjustments (sign depends on normal balance)

    Returns: list of row dicts sorted by account code.
    """
    # Aggregate adjustments per account code
    adj_dr_map = {}   # code_int -> total Dr
    adj_cr_map = {}   # code_int -> total Cr
    for e in adj_entries:
        for code_str, side in [(e['dr_code'], 'dr'), (e['cr_code'], 'cr')]:
            try:
                code_int = int(float(code_str))
            except (ValueError, TypeError):
                continue
            if side == 'dr':
                adj_dr_map[code_int] = adj_dr_map.get(code_int, 0.0) + e['dr_amount']
            else:
                adj_cr_map[code_int] = adj_cr_map.get(code_int, 0.0) + e['cr_amount']

    all_codes = set(gl_balances.keys()) | set(adj_dr_map.keys()) | set(adj_cr_map.keys())

    rows = []
    for code in sorted(all_codes):
        info    = coa.get_account(code)
        gl_data = gl_balances.get(code)

        if info:
            name       = info['name']
            acct_type  = info['type']
            sub_type   = info['sub_type']
            normal_bal = info['normal_balance'].lower()
        elif gl_data:
            name       = gl_data['name']
            acct_type  = gl_data['type']
            sub_type   = gl_data['sub_type']
            normal_bal = gl_data['normal_balance']
        else:
            name       = f'Account {code}'
            acct_type  = 'Unknown'
            sub_type   = ''
            normal_bal = 'debit'

        unadj_bal = gl_data['closing'] if gl_data else 0.0
        a_dr      = adj_dr_map.get(code, 0.0)
        a_cr      = adj_cr_map.get(code, 0.0)

        # Adjusted balance
        if normal_bal == 'debit':
            adj_bal = unadj_bal + a_dr - a_cr
        else:
            adj_bal = unadj_bal + a_cr - a_dr

        # Skip truly zero accounts (no opening, no movements, no adjustments)
        if (abs(unadj_bal) < 0.005 and abs(adj_bal) < 0.005
                and abs(a_dr) < 0.005 and abs(a_cr) < 0.005):
            continue

        unadj_dr, unadj_cr = _tb_display(unadj_bal, normal_bal)
        adj_dr,   adj_cr   = _tb_display(adj_bal,   normal_bal)

        rows.append({
            'code':           code,
            'name':           name,
            'type':           acct_type,
            'sub_type':       sub_type,
            'normal_balance': normal_bal,
            # Unadjusted
            'unadj_bal':      unadj_bal,
            'unadj_dr':       unadj_dr,
            'unadj_cr':       unadj_cr,
            # Adjusting entries (raw totals for this account)
            'adj_entries_dr': a_dr if a_dr > 0.005 else None,
            'adj_entries_cr': a_cr if a_cr > 0.005 else None,
            # Adjusted
            'adj_bal':        adj_bal,
            'adj_dr':         adj_dr,
            'adj_cr':         adj_cr,
        })

    return rows


# ---------------------------------------------------------------------------
# 4. Write Dashboard
# ---------------------------------------------------------------------------

def write_dashboard(wb, tb_rows, adj_entries, period_start, period_end, exceptions):
    ws         = add_sheet(wb, 'Dashboard', tab_color='00B050')
    period_str = f"{period_start}  to  {period_end}"
    row        = write_title(ws, 'Trial Balance -- Dashboard',
                             'Shwe Mandalay Cafe', period_str)

    # Totals
    unadj_dr_tot = sum(r['unadj_dr'] for r in tb_rows if r['unadj_dr'])
    unadj_cr_tot = sum(r['unadj_cr'] for r in tb_rows if r['unadj_cr'])
    adj_dr_tot   = sum(r['adj_dr']   for r in tb_rows if r['adj_dr'])
    adj_cr_tot   = sum(r['adj_cr']   for r in tb_rows if r['adj_cr'])
    adj_entr_dr  = sum(e['dr_amount'] for e in adj_entries)
    adj_entr_cr  = sum(e['cr_amount'] for e in adj_entries)

    unadj_ok   = abs(unadj_dr_tot - unadj_cr_tot) < 0.01
    adj_ok     = abs(adj_dr_tot   - adj_cr_tot)   < 0.01
    entries_ok = abs(adj_entr_dr  - adj_entr_cr)  < 0.01

    # Summary
    row = write_section_header(ws, 'TRIAL BALANCE SUMMARY', row, col_span=5)
    row = write_header_row(ws, ['Description', 'Total Debit', 'Total Credit',
                                 'Difference', 'Result'], row)
    for label, dr, cr in [
        ('Unadjusted Trial Balance', unadj_dr_tot, unadj_cr_tot),
        ('Adjusted Trial Balance',   adj_dr_tot,   adj_cr_tot),
    ]:
        diff    = dr - cr
        balanced = abs(diff) < 0.01
        row = write_data_row(ws, [label, _n(dr), _n(cr), _n(diff), ''],
                             row, number_cols=[2, 3, 4])
        write_validation_result(ws, row - 1, 5, balanced)
        ws.cell(row=row - 1, column=5).border = THIN_BORDER

    row += 1

    # Contents
    accts_affected = len(set(e['dr_code'] for e in adj_entries) |
                         set(e['cr_code'] for e in adj_entries))
    row = write_section_header(ws, 'CONTENTS', row, col_span=5)
    row = write_header_row(ws, ['Item', 'Count'], row)
    row = write_data_row(ws, ['GL accounts in Trial Balance',    len(tb_rows)],      row)
    row = write_data_row(ws, ['Adjusting entries applied',       len(adj_entries)],  row)
    row = write_data_row(ws, ['Accounts affected by adjustments', accts_affected],   row)
    row += 1

    # Validation
    row = write_section_header(ws, 'VALIDATION CHECKS', row, col_span=5)
    row = write_header_row(ws, ['Check', 'Total Debit', 'Total Credit',
                                 'Difference', 'Result'], row)
    for label, dr, cr, ok in [
        ('Unadjusted TB: Dr = Cr',   unadj_dr_tot, unadj_cr_tot, unadj_ok),
        ('Adjusted TB:   Dr = Cr',   adj_dr_tot,   adj_cr_tot,   adj_ok),
        ('Adj. entries:  Dr = Cr',   adj_entr_dr,  adj_entr_cr,  entries_ok),
    ]:
        diff = dr - cr
        row  = write_data_row(ws, [label, _n(dr), _n(cr), _n(diff), ''],
                              row, number_cols=[2, 3, 4])
        write_validation_result(ws, row - 1, 5, ok)
        ws.cell(row=row - 1, column=5).border = THIN_BORDER

    row += 1

    if exceptions:
        row = write_section_header(ws, 'WARNINGS', row, col_span=5)
        for exc in exceptions:
            row = write_data_row(ws, ['Warning', exc], row)

    auto_fit_columns(ws)
    ws.column_dimensions['A'].width = 40
    freeze_panes(ws, row=2, col=1)


# ---------------------------------------------------------------------------
# 5. Write Unadjusted Trial Balance
# ---------------------------------------------------------------------------

def write_unadjusted_tb(wb, tb_rows, period_start, period_end):
    ws  = add_sheet(wb, 'Unadjusted TB', tab_color='4472C4')
    row = write_title(ws, 'Unadjusted Trial Balance',
                      'GL closing balances before adjusting entries',
                      f"{period_start}  to  {period_end}")

    headers = ['Account Code', 'Account Name', 'Type', 'Normal Balance', 'Debit', 'Credit']
    row     = write_header_row(ws, headers, row)

    total_dr = total_cr = 0.0
    for r in tb_rows:
        row = write_data_row(ws, [
            r['code'], r['name'], r['type'], r['normal_balance'].capitalize(),
            _n(r['unadj_dr']), _n(r['unadj_cr']),
        ], row, number_cols=[5, 6])
        total_dr += r['unadj_dr'] or 0.0
        total_cr += r['unadj_cr'] or 0.0

    row = write_total_row(ws, 'TOTAL',
                          [None, None, None, _n(total_dr), _n(total_cr)],
                          row, double_line=True)
    row += 1

    balanced = abs(total_dr - total_cr) < 0.01
    ws.cell(row=row, column=1, value='Dr = Cr Check').font = NORMAL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=5,
            value=f"Dr {total_dr:,.2f} = Cr {total_cr:,.2f}").font = NORMAL_FONT
    ws.cell(row=row, column=5).border = THIN_BORDER
    write_validation_result(ws, row, 6, balanced)
    ws.cell(row=row, column=6).border = THIN_BORDER

    auto_fit_columns(ws)
    ws.column_dimensions['B'].width = 38
    freeze_panes(ws)


# ---------------------------------------------------------------------------
# 6. Write Adjustments Sheet
# ---------------------------------------------------------------------------

def write_adjustments_sheet(wb, adj_entries, tb_rows, period_end):
    ws  = add_sheet(wb, 'Adjustments', tab_color='4472C4')
    row = write_title(ws, 'Adjustments -- ADJ- Journal Entries',
                      'Period-end adjusting entries applied to the Trial Balance',
                      f'Period ending {period_end}')

    if not adj_entries:
        ws.cell(row=row, column=1,
                value='No adjusting entries found.').font = NORMAL_FONT
        return

    # ── Section 1: Individual entries ────────────────────────────────────────
    row = write_section_header(ws, 'INDIVIDUAL ADJUSTING ENTRIES', row, col_span=10)
    headers = ['Entry No.', 'Date', 'Type', 'Description',
               'Debit Account', 'Dr Code', 'Debit Amount',
               'Credit Account', 'Cr Code', 'Credit Amount']
    row = write_header_row(ws, headers, row)

    total_dr = total_cr = 0.0
    for e in adj_entries:
        row = write_data_row(ws, [
            e['ref'], _fmt_date(e['date']), e['type'], e['description'],
            e['dr_name'], e['dr_code'], _n(e['dr_amount']),
            e['cr_name'], e['cr_code'], _n(e['cr_amount']),
        ], row, number_cols=[7, 10])
        total_dr += e['dr_amount']
        total_cr += e['cr_amount']

    row = write_total_row(ws, 'TOTAL',
                          [None, None, None, None, None, _n(total_dr),
                           None, None, _n(total_cr)],
                          row, double_line=True)
    row += 2

    # ── Section 2: Per-account summary ───────────────────────────────────────
    row = write_section_header(ws, 'PER-ACCOUNT ADJUSTMENT SUMMARY', row, col_span=6)
    headers2 = ['Account Code', 'Account Name', 'Type',
                'Total Dr Adj', 'Total Cr Adj', 'Net Effect on Balance']
    row = write_header_row(ws, headers2, row)

    for r in tb_rows:
        a_dr = r['adj_entries_dr'] or 0.0
        a_cr = r['adj_entries_cr'] or 0.0
        if abs(a_dr) < 0.005 and abs(a_cr) < 0.005:
            continue
        # Net effect on the account's balance (from normal-balance perspective)
        nb  = r['normal_balance']
        net = (a_dr - a_cr) if nb == 'debit' else (a_cr - a_dr)

        row = write_data_row(ws, [
            r['code'], r['name'], r['type'],
            _n(a_dr), _n(a_cr), _n(net),
        ], row, number_cols=[4, 5, 6])

    auto_fit_columns(ws)
    ws.column_dimensions['D'].width = 38
    freeze_panes(ws)


# ---------------------------------------------------------------------------
# 7. Write Adjusted Trial Balance
# ---------------------------------------------------------------------------

def write_adjusted_tb(wb, tb_rows, period_start, period_end):
    ws  = add_sheet(wb, 'Adjusted TB', tab_color='4472C4')
    row = write_title(ws, 'Adjusted Trial Balance',
                      'GL balances after applying all adjusting entries',
                      f"{period_start}  to  {period_end}")

    headers = ['Account Code', 'Account Name', 'Type', 'Normal Balance', 'Debit', 'Credit']
    row     = write_header_row(ws, headers, row)

    total_dr = total_cr = 0.0
    for r in tb_rows:
        row = write_data_row(ws, [
            r['code'], r['name'], r['type'], r['normal_balance'].capitalize(),
            _n(r['adj_dr']), _n(r['adj_cr']),
        ], row, number_cols=[5, 6])
        total_dr += r['adj_dr'] or 0.0
        total_cr += r['adj_cr'] or 0.0

    row = write_total_row(ws, 'TOTAL',
                          [None, None, None, _n(total_dr), _n(total_cr)],
                          row, double_line=True)
    row += 1

    balanced = abs(total_dr - total_cr) < 0.01
    ws.cell(row=row, column=1, value='Dr = Cr Check').font = NORMAL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=5,
            value=f"Dr {total_dr:,.2f} = Cr {total_cr:,.2f}").font = NORMAL_FONT
    ws.cell(row=row, column=5).border = THIN_BORDER
    write_validation_result(ws, row, 6, balanced)
    ws.cell(row=row, column=6).border = THIN_BORDER

    auto_fit_columns(ws)
    ws.column_dimensions['B'].width = 38
    freeze_panes(ws)


# ---------------------------------------------------------------------------
# 8. Write TB Worksheet (6-column combined view)
# ---------------------------------------------------------------------------

def write_tb_worksheet(wb, tb_rows, period_start, period_end):
    ws  = add_sheet(wb, 'TB Worksheet', tab_color='70AD47')
    row = write_title(ws, 'Trial Balance Worksheet',
                      'Unadjusted  |  Adjustments  |  Adjusted',
                      f"{period_start}  to  {period_end}")

    headers = ['Code', 'Account Name', 'Type',
               'Unadj Debit', 'Unadj Credit',
               'Adj Entries Dr', 'Adj Entries Cr',
               'Adjusted Debit', 'Adjusted Credit']
    row = write_header_row(ws, headers, row)

    t = [0.0] * 6   # [unadj_dr, unadj_cr, adj_e_dr, adj_e_cr, final_dr, final_cr]

    for r in tb_rows:
        u_dr = r['unadj_dr']       or 0.0
        u_cr = r['unadj_cr']       or 0.0
        a_dr = r['adj_entries_dr'] or 0.0
        a_cr = r['adj_entries_cr'] or 0.0
        f_dr = r['adj_dr']         or 0.0
        f_cr = r['adj_cr']         or 0.0

        row = write_data_row(ws, [
            r['code'], r['name'], r['type'],
            _n(u_dr), _n(u_cr),
            _n(a_dr), _n(a_cr),
            _n(f_dr), _n(f_cr),
        ], row, number_cols=[4, 5, 6, 7, 8, 9])

        t[0] += u_dr; t[1] += u_cr
        t[2] += a_dr; t[3] += a_cr
        t[4] += f_dr; t[5] += f_cr

    row = write_total_row(ws, 'TOTAL',
                          [None, None,
                           _n(t[0]), _n(t[1]),
                           _n(t[2]), _n(t[3]),
                           _n(t[4]), _n(t[5])],
                          row, double_line=True)
    row += 1

    # Inline validation row
    unadj_ok = abs(t[0] - t[1]) < 0.01
    adj_ok   = abs(t[4] - t[5]) < 0.01
    ws.cell(row=row, column=1, value='Unadj. Dr=Cr').font = NORMAL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    write_validation_result(ws, row, 4, unadj_ok)
    ws.cell(row=row, column=4).border = THIN_BORDER
    ws.cell(row=row, column=5, value='Adjusted Dr=Cr').font = NORMAL_FONT
    ws.cell(row=row, column=5).border = THIN_BORDER
    write_validation_result(ws, row, 8, adj_ok)
    ws.cell(row=row, column=8).border = THIN_BORDER

    auto_fit_columns(ws)
    ws.column_dimensions['B'].width = 38
    freeze_panes(ws)


# ---------------------------------------------------------------------------
# 9. Write Exceptions Sheet
# ---------------------------------------------------------------------------

def write_exceptions_sheet(wb, exceptions):
    ws  = add_sheet(wb, 'Exceptions', tab_color='FF0000')
    row = write_title(ws, 'Exceptions & Warnings')
    row = write_header_row(ws, ['#', 'Exception / Warning'], row)
    for i, exc in enumerate(exceptions, 1):
        row = write_data_row(ws, [i, exc], row)
    auto_fit_columns(ws)
    freeze_panes(ws)


# ---------------------------------------------------------------------------
# 10. Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)

    data_dir     = sys.argv[1]
    period_start = sys.argv[2]
    period_end   = sys.argv[3]
    output_file  = sys.argv[4]

    print(f"\n{'='*60}")
    print(f"  MODULE 5 -- GENERATE TRIAL BALANCE")
    print(f"  Period : {period_start}  to  {period_end}")
    print(f"  Data   : {data_dir}")
    print(f"  Output : {output_file}")
    print(f"{'='*60}\n")

    coa_path = Path(data_dir) / 'chart_of_accounts.xlsx'
    coa      = COAMapper(str(coa_path)) if coa_path.exists() else COAMapper()
    exceptions = []

    # ── 1. GL balances (unadjusted) ──────────────────────────────────────────
    print("Loading GL balances...")
    gl_balances, err = load_gl_balances(data_dir, period_start, period_end, coa)
    if err:
        print(f"  ERROR: {err}")
        exceptions.append(err)
        gl_balances = {}
    else:
        print(f"  GL accounts loaded: {len(gl_balances)}")

    # ── 2. Adjusting entries ─────────────────────────────────────────────────
    print("\nLoading adjusting entries...")
    adj_entries, warn = load_adj_entries(data_dir)
    if warn:
        print(f"  WARNING: {warn}")
        exceptions.append(warn)
    else:
        total_adj = sum(e['dr_amount'] for e in adj_entries)
        print(f"  Adjusting entries  : {len(adj_entries)}")
        print(f"  Total adj amount   : {total_adj:,.2f}")
        for e in adj_entries:
            print(f"    {e['ref']:<10}  Dr {e['dr_code']}  "
                  f"Cr {e['cr_code']}  {e['dr_amount']:>10,.2f}")

    # ── 3. Build TB rows ─────────────────────────────────────────────────────
    print("\nBuilding trial balance...")
    tb_rows = build_tb_rows(gl_balances, adj_entries, coa)
    print(f"  TB accounts: {len(tb_rows)}")

    unadj_dr = sum(r['unadj_dr'] or 0.0 for r in tb_rows)
    unadj_cr = sum(r['unadj_cr'] or 0.0 for r in tb_rows)
    adj_dr   = sum(r['adj_dr']   or 0.0 for r in tb_rows)
    adj_cr   = sum(r['adj_cr']   or 0.0 for r in tb_rows)

    print(f"\n  Unadjusted TB:")
    print(f"    Total Debit  : {unadj_dr:>15,.2f}")
    print(f"    Total Credit : {unadj_cr:>15,.2f}")
    print(f"    Difference   : {unadj_dr - unadj_cr:>15,.2f}")
    print(f"    Balanced     : {'YES' if abs(unadj_dr - unadj_cr) < 0.01 else 'NO'}")

    print(f"\n  Adjusted TB:")
    print(f"    Total Debit  : {adj_dr:>15,.2f}")
    print(f"    Total Credit : {adj_cr:>15,.2f}")
    print(f"    Difference   : {adj_dr - adj_cr:>15,.2f}")
    print(f"    Balanced     : {'YES' if abs(adj_dr - adj_cr) < 0.01 else 'NO'}")

    if abs(unadj_dr - unadj_cr) > 0.01:
        exceptions.append(
            f"Unadjusted TB NOT balanced: "
            f"Dr {unadj_dr:,.2f} vs Cr {unadj_cr:,.2f} "
            f"(diff {unadj_dr - unadj_cr:,.2f})")

    if abs(adj_dr - adj_cr) > 0.01:
        exceptions.append(
            f"Adjusted TB NOT balanced: "
            f"Dr {adj_dr:,.2f} vs Cr {adj_cr:,.2f} "
            f"(diff {adj_dr - adj_cr:,.2f})")

    # ── 4. Write Excel ───────────────────────────────────────────────────────
    print(f"\nWriting output to: {output_file}")
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    wb = create_workbook()
    write_dashboard(wb, tb_rows, adj_entries, period_start, period_end, exceptions)
    write_unadjusted_tb(wb, tb_rows, period_start, period_end)
    write_adjustments_sheet(wb, adj_entries, tb_rows, period_end)
    write_adjusted_tb(wb, tb_rows, period_start, period_end)
    write_tb_worksheet(wb, tb_rows, period_start, period_end)
    if exceptions:
        write_exceptions_sheet(wb, exceptions)

    save_workbook(wb, output_file)

    sheets_written = ('Dashboard | Unadjusted TB | Adjustments | '
                      'Adjusted TB | TB Worksheet'
                      + (' | Exceptions' if exceptions else ''))

    print(f"\n{'='*60}")
    print(f"  OUTPUT  : {output_file}")
    print(f"  Sheets  : {sheets_written}")
    print(f"  Unadjusted TB  Dr = Cr : {'YES' if abs(unadj_dr - unadj_cr) < 0.01 else 'NO'}")
    print(f"  Adjusted   TB  Dr = Cr : {'YES' if abs(adj_dr   - adj_cr)   < 0.01 else 'NO'}")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
