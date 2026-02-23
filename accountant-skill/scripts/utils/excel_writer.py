"""
Excel Writer Utility — Professional .xlsx output with consistent formatting.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# Standard colors
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
TITLE_FONT = Font(bold=True, size=14, name='Arial')
SUBTITLE_FONT = Font(bold=True, size=12, name='Arial')
PERIOD_FONT = Font(italic=True, size=11, name='Arial')
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
SECTION_FONT = Font(bold=True, size=11, name='Arial', color='1F4E79')
NORMAL_FONT = Font(size=11, name='Arial')
TOTAL_FONT = Font(bold=True, size=11, name='Arial')
NEGATIVE_FONT = Font(size=11, name='Arial', color='FF0000')
PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
WARNING_FILL = PatternFill('solid', fgColor='FFEB9C')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
BOTTOM_BORDER = Border(bottom=Side(style='medium'))
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))

NUMBER_FORMAT = '#,##0'
NUMBER_FORMAT_NEG = '#,##0;(#,##0);"-"'
PERCENT_FORMAT = '0.0%'
DATE_FORMAT = 'YYYY-MM-DD'


def create_workbook():
    """Create a new workbook with default settings."""
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def add_sheet(wb, name, tab_color=None):
    """Add a sheet with optional tab color. Returns the worksheet."""
    ws = wb.create_sheet(title=name[:31])  # Excel max sheet name = 31 chars
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    return ws


def write_title(ws, title, subtitle=None, period=None, start_row=1):
    """Write a report title block at the top of a sheet."""
    row = start_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center')
    row += 1

    if subtitle:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=subtitle)
        cell.font = SUBTITLE_FONT
        cell.alignment = Alignment(horizontal='center')
        row += 1

    if period:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=period)
        cell.font = PERIOD_FONT
        cell.alignment = Alignment(horizontal='center')
        row += 1

    row += 1  # blank row after title
    return row


def write_header_row(ws, headers, row, start_col=1):
    """Write a formatted header row."""
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    return row + 1


def write_data_row(ws, values, row, start_col=1, number_cols=None, font=None, border=None):
    """Write a data row with formatting."""
    if number_cols is None:
        number_cols = []
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.font = font or NORMAL_FONT
        cell.border = border or THIN_BORDER
        col_idx = start_col + i
        if col_idx in number_cols or (isinstance(val, (int, float)) and i > 0):
            cell.number_format = NUMBER_FORMAT_NEG
            cell.alignment = Alignment(horizontal='right')
            if isinstance(val, (int, float)) and val < 0:
                cell.font = NEGATIVE_FONT
        else:
            cell.alignment = Alignment(horizontal='left')
    return row + 1


def write_section_header(ws, text, row, col_span=8, start_col=1):
    """Write a section header row (e.g., 'REVENUE', 'OPERATING EXPENSES')."""
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + col_span - 1)
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER
    return row + 1


def write_total_row(ws, label, values, row, start_col=1, double_line=False):
    """Write a total/subtotal row with bold formatting and border."""
    border = DOUBLE_BOTTOM if double_line else BOTTOM_BORDER
    cell = ws.cell(row=row, column=start_col, value=label)
    cell.font = TOTAL_FONT
    cell.border = border
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + 1 + i, value=val)
        cell.font = TOTAL_FONT
        cell.number_format = NUMBER_FORMAT_NEG
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        if isinstance(val, (int, float)) and val < 0:
            cell.font = Font(bold=True, size=11, name='Arial', color='FF0000')
    return row + 1


def auto_fit_columns(ws, min_width=12, max_width=50):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


def freeze_panes(ws, row=2, col=2):
    """Freeze panes at the given row and column."""
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


def write_validation_result(ws, row, col, passed):
    """Write a PASS/FAIL cell."""
    cell = ws.cell(row=row, column=col, value='PASS' if passed else 'FAIL')
    cell.font = Font(bold=True, size=11, name='Arial', color='006100' if passed else '9C0006')
    cell.fill = PASS_FILL if passed else FAIL_FILL
    cell.alignment = Alignment(horizontal='center')
    return row


def save_workbook(wb, filepath):
    """Save workbook to file."""
    wb.save(filepath)
    return filepath
