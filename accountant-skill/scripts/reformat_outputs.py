"""
Reformat Excel Output Files
Applies consistent formatting to existing output files without modifying data.

Usage:
    python scripts/reformat_outputs.py <output_dir>

Example:
    python scripts/reformat_outputs.py data/output/Feb2025
"""
import sys
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Standard colors
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
TITLE_FONT = Font(bold=True, size=14, name='Arial')
SUBTITLE_FONT = Font(bold=True, size=12, name='Arial')
PERIOD_FONT = Font(italic=True, size=11, name='Arial')
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
SECTION_FONT = Font(bold=True, size=11, name='Arial', color='1F4E79')
NORMAL_FONT = Font(size=11, name='Arial')
TOTAL_FONT = Font(bold=True, size=11, name='Arial')
NEGATIVE_FONT = Font(size=11, name='Arial', color='FF0000')
PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
WARNING_FILL = PatternFill('solid', fgColor='FFEB9C')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
BOTTOM_BORDER = Border(bottom=Side(style='medium'))
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))

NUMBER_FORMAT = '#,##0'
NUMBER_FORMAT_NEG = '#,##0;(#,##0);"-"'
PERCENT_FORMAT = '0.0%'

# Tab colors by sheet name
TAB_COLORS = {
    'dashboard': '00B050',      # Green
    'summary': '00B050',        # Green
    'income statement': '4472C4',  # Blue
    'balance sheet': '4472C4',     # Blue
    'cash flow': '4472C4',         # Blue
    'gl balances': '4472C4',       # Blue
    'adjusted tb': '4472C4',       # Blue
    'unadjusted tb': '4472C4',     # Blue
    'adjustments': '4472C4',       # Blue
    'tb worksheet': '70AD47',      # Orange-green
    'ar by customer': '4472C4',    # Blue
    'ap by supplier': '4472C4',    # Blue
    'cash by bank': '4472C4',      # Blue
    'fixed assets': '4472C4',      # Blue
    'control acct recon': 'FF0000', # Red
    'inventory summary': '70AD47',  # Orange-green
    'double-entry checks': '4472C4', # Blue
    'cross-module flow': '70AD47',   # Orange-green
    'financial validation': '4472C4', # Blue
    'exceptions': 'FF0000',          # Red
}


def is_header_row(row_values):
    """Detect if a row is a header row based on content."""
    if not row_values:
        return False
    # Check for common header keywords
    header_keywords = ['code', 'name', 'date', 'amount', 'debit', 'credit', 'balance',
                       'account', 'type', 'status', 'description', 'total', 'result',
                       'check', 'category', 'period', 'item', 'reference']
    text_values = [str(v).lower().strip() for v in row_values if v]
    if not text_values:
        return False
    # If most values contain header keywords, it's a header
    keyword_count = sum(1 for v in text_values if any(kw in v for kw in header_keywords))
    return keyword_count >= len(text_values) * 0.5 and len(text_values) >= 2


def is_number_column(ws, col_idx, start_row, end_row):
    """Determine if a column contains mostly numbers."""
    number_count = 0
    total_count = 0
    for row in range(start_row, min(end_row + 1, start_row + 20)):  # Check up to 20 rows
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            total_count += 1
            if isinstance(cell.value, (int, float)):
                number_count += 1
    return total_count > 0 and number_count / total_count > 0.5


def reformat_sheet(ws, sheet_name):
    """Apply formatting to a single sheet."""
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row < 2 or max_col < 1:
        return

    # Set tab color
    sheet_lower = sheet_name.lower()
    for key, color in TAB_COLORS.items():
        if key in sheet_lower:
            ws.sheet_properties.tabColor = color
            break

    # Find header row (usually row 4-6 after title block)
    header_row = None
    for r in range(1, min(10, max_row + 1)):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if is_header_row(row_values):
            header_row = r
            break

    if header_row is None:
        header_row = 4  # Default assumption

    # Format header row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

    # Identify number columns
    number_cols = []
    for col in range(1, max_col + 1):
        if is_number_column(ws, col, header_row + 1, max_row):
            number_cols.append(col)

    # Format data rows
    for row in range(header_row + 1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue

            # Apply font and border
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

            # Format numbers
            if col in number_cols and isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT_NEG
                cell.alignment = Alignment(horizontal='right')
                if cell.value < 0:
                    cell.font = NEGATIVE_FONT
            else:
                cell.alignment = Alignment(horizontal='left')

    # Format total rows (rows with "TOTAL" in first column)
    for row in range(header_row + 1, max_row + 1):
        first_cell = ws.cell(row=row, column=1)
        if first_cell.value and 'TOTAL' in str(first_cell.value).upper():
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = TOTAL_FONT
                if col in number_cols:
                    cell.border = DOUBLE_BOTTOM
                else:
                    cell.border = BOTTOM_BORDER

    # Auto-fit columns
    for col in range(1, max_col + 1):
        max_length = 12
        col_letter = get_column_letter(col)
        for row in range(1, min(max_row + 1, 50)):  # Check first 50 rows
            cell = ws.cell(row=row, column=col)
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = min(length, 50)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Freeze panes
    ws.freeze_panes = f'B{header_row + 1}'


def reformat_file(filepath):
    """Reformat a single Excel file."""
    print(f"  Reformatting: {filepath.name}")

    try:
        wb = load_workbook(filepath)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            reformat_sheet(ws, sheet_name)

        wb.save(filepath)
        print(f"    Done: {len(wb.sheetnames)} sheets processed")
        return True

    except Exception as e:
        print(f"    ERROR: {e}")
        return False


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("Please specify an output directory.")
        sys.exit(1)

    output_dir = Path(sys.argv[1])

    if not output_dir.exists():
        print(f"ERROR: Directory not found: {output_dir}")
        sys.exit(1)

    # Find all .xlsx files
    xlsx_files = list(output_dir.glob('*.xlsx'))

    if not xlsx_files:
        print(f"No .xlsx files found in {output_dir}")
        sys.exit(0)

    print(f"\nReformatting {len(xlsx_files)} files in: {output_dir}")
    print("-" * 50)

    success = 0
    failed = 0

    for filepath in sorted(xlsx_files):
        if reformat_file(filepath):
            success += 1
        else:
            failed += 1

    print("-" * 50)
    print(f"Complete: {success} succeeded, {failed} failed")


if __name__ == '__main__':
    main()