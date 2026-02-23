"""
Module 6: Generate Financial Statements
Shwe Mandalay Cafe / K&K Finance Team

Reads the Adjusted Trial Balance (from Module 5 output) and produces:
  - Dashboard           (key metrics, BS check, CF check)
  - Income Statement    (P&L with sections and margins)
  - Balance Sheet       (Assets = Liabilities + Equity check)
  - Cash Flow           (Indirect method, reconciles to cash balance)
  - Exceptions          (only if errors found)

Usage:
    python generate_financials.py <data_dir> <period_start> <period_end> <output_file>

    python generate_financials.py \\
        data/Jan2026 \\
        2026-01-01 \\
        2026-01-31 \\
        data/Jan2026/financial_statements_Jan2026.xlsx
"""

import sys
import os
from pathlib import Path
from glob import glob

import pandas as pd
import numpy as np

sys.path.insert(0, str(Path(__file__).parent))
from utils.excel_writer import (
    create_workbook, add_sheet, write_title, write_header_row,
    write_data_row, write_section_header, write_total_row,
    write_validation_result, auto_fit_columns, freeze_panes,
    save_workbook, NORMAL_FONT, TOTAL_FONT, THIN_BORDER,
    PASS_FILL, FAIL_FILL, NEGATIVE_FONT, PERCENT_FORMAT,
    NUMBER_FORMAT_NEG
)
from utils.coa_mapper import COAMapper, CONTRA_ACCOUNTS
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm_code(val):
    """Normalize float-string account codes: '1020.0' -> '1020'."""
    try:
        return str(int(float(str(val).strip())))
    except (ValueError, TypeError):
        return str(val).strip()


def _find_col(df, candidates):
    """Return first column name from candidates that exists in df.columns (case-insensitive)."""
    lower_cols = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c.lower() in lower_cols:
            return lower_cols[c.lower()]
    return None


def _normalize_cols(df):
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def _n(val):
    """Return float or None; treat zero/NaN as None for display."""
    if val is None:
        return None
    try:
        v = float(val)
    except (ValueError, TypeError):
        return None
    if np.isnan(v) or abs(v) < 0.005:
        return None
    return v


def _fmt_date(d):
    try:
        return pd.Timestamp(d).strftime('%Y-%m-%d')
    except Exception:
        return str(d) if d else ''


def _is_numeric_code(val):
    """True if val represents a valid numeric account code."""
    try:
        int(float(str(val).strip()))
        return True
    except (ValueError, TypeError):
        return False


# ---------------------------------------------------------------------------
# 1. Load Adjusted TB from Module 5 output
# ---------------------------------------------------------------------------

def load_adjusted_tb(data_dir):
    """
    Read the 'Adjusted TB' sheet from trial_balance_*.xlsx.

    The sheet has a title block before the column headers; scan for the row
    that contains 'Account Code' to find the true header row.

    Returns: (list of account dicts, error string or None)
    Each dict: {code, name, type, normal_balance, debit, credit, balance}
    """
    data_dir = Path(data_dir)
    candidates = list(data_dir.glob('trial_balance*.xlsx'))
    if not candidates:
        return [], "No trial_balance*.xlsx found in data directory."

    tb_file = candidates[0]

    try:
        df_raw = pd.read_excel(tb_file, sheet_name='Adjusted TB', header=None)
    except Exception as e:
        return [], f"Could not read 'Adjusted TB' sheet from {tb_file.name}: {e}"

    # Find header row: scan for row containing 'Account Code'
    header_row_idx = None
    for i, row_vals in df_raw.iterrows():
        row_strs = [str(v).strip().lower() for v in row_vals.values]
        if 'account code' in row_strs:
            header_row_idx = i
            break

    if header_row_idx is None:
        return [], f"Could not find 'Account Code' header in 'Adjusted TB' sheet of {tb_file.name}"

    try:
        df = pd.read_excel(tb_file, sheet_name='Adjusted TB', header=header_row_idx)
    except Exception as e:
        return [], f"Could not re-read 'Adjusted TB' with header at row {header_row_idx}: {e}"

    df = _normalize_cols(df)

    code_col   = _find_col(df, ['account code', 'code', 'acct code'])
    name_col   = _find_col(df, ['account name', 'name', 'description'])
    type_col   = _find_col(df, ['type', 'account type'])
    norm_col   = _find_col(df, ['normal balance', 'normal_balance'])
    debit_col  = _find_col(df, ['debit', 'dr'])
    credit_col = _find_col(df, ['credit', 'cr'])

    if code_col is None:
        return [], f"'Account Code' column not found in normalized TB sheet. Cols: {list(df.columns)}"

    accounts = []
    for _, row in df.iterrows():
        code_raw = row.get(code_col, None)
        if code_raw is None or pd.isna(code_raw):
            continue
        code_str = _norm_code(code_raw)
        if not _is_numeric_code(code_str):
            continue  # Skip TOTAL row and non-numeric rows

        code_int = int(code_str)
        name = str(row[name_col]).strip() if name_col and not pd.isna(row.get(name_col, '')) else f'Account {code_int}'
        acct_type = str(row[type_col]).strip() if type_col and not pd.isna(row.get(type_col, '')) else ''
        normal_bal = str(row[norm_col]).strip().lower() if norm_col and not pd.isna(row.get(norm_col, '')) else 'debit'

        debit_val  = 0.0
        credit_val = 0.0
        if debit_col is not None:
            try:
                v = float(row[debit_col])
                if not np.isnan(v):
                    debit_val = v
            except (ValueError, TypeError):
                pass
        if credit_col is not None:
            try:
                v = float(row[credit_col])
                if not np.isnan(v):
                    credit_val = v
            except (ValueError, TypeError):
                pass

        # Balance: debit-normal accounts carry debit balance; credit-normal carry credit balance
        if 'debit' in normal_bal:
            balance = debit_val - credit_val
        else:
            balance = credit_val - debit_val

        accounts.append({
            'code':           code_int,
            'name':           name,
            'type':           acct_type,
            'normal_balance': normal_bal,
            'debit':          debit_val,
            'credit':         credit_val,
            'balance':        round(balance, 2),
        })

    return accounts, None


# ---------------------------------------------------------------------------
# 2. Load GL Data (for Cash Flow opening balances)
# ---------------------------------------------------------------------------

def load_gl_data(data_dir, period_start, coa):
    """
    Read general_ledger.xlsx and return opening balances per account code.
    Returns: dict {code_int: opening_balance}
    """
    path = Path(data_dir) / 'general_ledger.xlsx'
    if not path.exists():
        return {}, "general_ledger.xlsx not found (Cash Flow opening balances will be 0)"

    try:
        # Try reading all sheets
        xf = pd.ExcelFile(path)
        sheets = {name: xf.parse(name) for name in xf.sheet_names}
    except Exception as e:
        return {}, f"Could not read general_ledger.xlsx: {e}"

    openings = {}

    if len(sheets) == 1:
        df = _normalize_cols(list(sheets.values())[0])
        code_col    = _find_col(df, ['account code', 'code', 'acct code', 'no.'])
        date_col    = _find_col(df, ['date', 'trans date', 'entry date'])
        balance_col = _find_col(df, ['balance', 'running balance', 'bal'])

        if code_col and date_col:
            df[code_col] = pd.to_numeric(df[code_col], errors='coerce')
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

            for code_val in df[code_col].dropna().unique():
                code_int = int(code_val)
                acct_df  = df[df[code_col] == code_val].copy()
                pre_rows = acct_df[acct_df[date_col] < pd.Timestamp(period_start)]
                opening  = 0.0
                if not pre_rows.empty and balance_col:
                    last_val = pd.to_numeric(pre_rows.iloc[-1].get(balance_col), errors='coerce')
                    if not pd.isna(last_val):
                        opening = float(last_val)
                openings[code_int] = opening
    else:
        for sheet_name, raw_df in sheets.items():
            df = _normalize_cols(raw_df)
            try:
                code_int = int(''.join(c for c in sheet_name if c.isdigit()))
            except ValueError:
                continue
            date_col    = _find_col(df, ['date', 'trans date'])
            balance_col = _find_col(df, ['balance', 'bal'])
            if date_col is None:
                continue
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
            pre_rows = df[df[date_col] < pd.Timestamp(period_start)]
            opening  = 0.0
            if not pre_rows.empty and balance_col:
                last_val = pd.to_numeric(pre_rows.iloc[-1].get(balance_col), errors='coerce')
                if not pd.isna(last_val):
                    opening = float(last_val)
            openings[code_int] = opening

    return openings, None


# ---------------------------------------------------------------------------
# 3. Load Adjusting Entries (for non-cash items in Cash Flow)
# ---------------------------------------------------------------------------

def load_adj_entries(data_dir):
    """
    Read 'All Entries' sheet from adjusting_entries_*.xlsx.
    Returns: (list of entry dicts, warning string or None)
    """
    data_dir   = Path(data_dir)
    candidates = list(data_dir.glob('adjusting_entries*.xlsx'))
    if not candidates:
        return [], "No adjusting_entries*.xlsx found -- non-cash adjustments skipped."

    adj_file = candidates[0]

    try:
        df_raw = pd.read_excel(adj_file, sheet_name='All Entries', header=None)
        header_row_idx = None
        for i, row_vals in df_raw.iterrows():
            row_strs = [str(v).strip() for v in row_vals.values]
            if 'Dr Code' in row_strs or 'Entry No.' in row_strs:
                header_row_idx = i
                break
        if header_row_idx is None:
            return [], f"Could not find header in 'All Entries' sheet of {adj_file.name}"

        df = pd.read_excel(adj_file, sheet_name='All Entries', header=header_row_idx)
    except Exception as e:
        return [], f"Could not read 'All Entries' from {adj_file.name}: {e}"

    df.columns = [str(c).strip() for c in df.columns]

    needed  = ['Dr Code', 'Debit Amount', 'Cr Code', 'Credit Amount']
    missing = [c for c in needed if c not in df.columns]
    if missing:
        return [], f"'All Entries' missing columns: {missing}"

    entries = []
    for _, row in df.iterrows():
        try:
            dr_amt = float(row['Debit Amount'])
            cr_amt = float(row['Credit Amount'])
        except (ValueError, TypeError):
            continue
        if pd.isna(dr_amt) or pd.isna(cr_amt) or dr_amt <= 0:
            continue

        dr_code_str = _norm_code(row['Dr Code'])
        cr_code_str = _norm_code(row['Cr Code'])
        if not dr_code_str.isdigit() or not cr_code_str.isdigit():
            continue

        entries.append({
            'ref':         str(row.get('Entry No.', '')).strip(),
            'description': str(row.get('Category / Description', '')).strip(),
            'dr_code':     dr_code_str,
            'dr_amount':   round(dr_amt, 2),
            'cr_code':     cr_code_str,
            'cr_amount':   round(cr_amt, 2),
        })

    return entries, None


# ---------------------------------------------------------------------------
# 4. Build Income Statement Data
# ---------------------------------------------------------------------------

CONTRA_REVENUE = {4200, 4210}  # Sales Returns, Discounts (debit-normal)
NON_CASH_EXPENSE_CODES = {5300, 5800, 5930}  # Depreciation, Bad Debt, Loss on Disposal
GAIN_ON_DISPOSAL_CODES = {4120}              # Gain on Disposal (non-cash income)


def _get_balance(acct):
    """Return the signed balance for IS/BS use: positive = normal-side balance."""
    return acct['balance']


def build_is_data(accounts, coa):
    """
    Group accounts into Income Statement sections.

    Returns a dict of lists and computed totals/margins.
    Amounts are always positive for 'normal' items; negatives indicate deductions.
    """
    revenue      = []
    contra_rev   = []
    cogs         = []
    opex         = []
    other_income = []
    other_exp    = []
    tax          = []

    for acct in accounts:
        code = acct['code']
        info = coa.get_account(code)
        if info is None:
            continue
        t = info['type']
        if t not in ('Revenue', 'Expense'):
            continue

        bal = acct['balance']  # positive = credit for revenue, positive = debit for expense

        if t == 'Revenue':
            if code in CONTRA_REVENUE or (4200 <= code <= 4299):
                # Contra revenue: balance is debit = shown as negative deduction
                contra_rev.append((code, acct['name'], -abs(bal)))
            elif 4100 <= code <= 4199:
                other_income.append((code, acct['name'], bal))
            else:
                revenue.append((code, acct['name'], bal))
        else:  # Expense
            if 5000 <= code <= 5099:
                cogs.append((code, acct['name'], bal))
            elif 5100 <= code <= 5899:
                opex.append((code, acct['name'], bal))
            elif code == 5950:
                tax.append((code, acct['name'], bal))
            else:  # 5900-5949 + 5951-5999
                other_exp.append((code, acct['name'], bal))

    # Sort each section by code
    for lst in [revenue, contra_rev, cogs, opex, other_income, other_exp, tax]:
        lst.sort(key=lambda x: x[0])

    gross_revenue      = sum(a for _, _, a in revenue)
    total_contra       = sum(a for _, _, a in contra_rev)   # already negative
    net_revenue        = gross_revenue + total_contra
    total_cogs         = sum(a for _, _, a in cogs)
    gross_profit       = net_revenue - total_cogs
    total_opex         = sum(a for _, _, a in opex)
    operating_profit   = gross_profit - total_opex
    total_other_income = sum(a for _, _, a in other_income)
    total_other_exp    = sum(a for _, _, a in other_exp)
    net_other          = total_other_income - total_other_exp
    total_tax          = sum(a for _, _, a in tax)
    profit_before_tax  = operating_profit + net_other
    net_profit         = profit_before_tax - total_tax

    gross_margin     = (gross_profit    / net_revenue) if abs(net_revenue) > 0.01 else 0.0
    operating_margin = (operating_profit / net_revenue) if abs(net_revenue) > 0.01 else 0.0
    net_margin       = (net_profit      / net_revenue) if abs(net_revenue) > 0.01 else 0.0

    return {
        'revenue':          revenue,
        'contra_rev':       contra_rev,
        'cogs':             cogs,
        'opex':             opex,
        'other_income':     other_income,
        'other_exp':        other_exp,
        'tax':              tax,
        'gross_revenue':    round(gross_revenue,    2),
        'total_contra':     round(total_contra,     2),
        'net_revenue':      round(net_revenue,      2),
        'total_cogs':       round(total_cogs,       2),
        'gross_profit':     round(gross_profit,     2),
        'total_opex':       round(total_opex,       2),
        'operating_profit': round(operating_profit, 2),
        'total_other_income': round(total_other_income, 2),
        'total_other_exp':  round(total_other_exp,  2),
        'net_other':        round(net_other,         2),
        'profit_before_tax': round(profit_before_tax, 2),
        'total_tax':        round(total_tax,         2),
        'net_profit':       round(net_profit,        2),
        'gross_margin':     round(gross_margin,      4),
        'operating_margin': round(operating_margin,  4),
        'net_margin':       round(net_margin,         4),
    }


# ---------------------------------------------------------------------------
# 5. Build Balance Sheet Data
# ---------------------------------------------------------------------------

def build_bs_data(accounts, coa, net_profit):
    """
    Group accounts into Balance Sheet sections.

    Sign rules:
    - Assets: debit-normal positive; contra assets (credit-normal) shown negative
    - Liabilities/Equity: credit-normal positive; contra equity (Drawings) shown negative
    - Net Profit added as extra equity line
    """
    noncurrent_assets = []
    current_assets    = []
    equity            = []
    noncurrent_liab   = []
    current_liab      = []

    for acct in accounts:
        code = acct['code']
        info = coa.get_account(code)
        if info is None:
            continue
        t = info['type']
        if t not in ('Asset', 'Liability', 'Equity'):
            continue

        bal   = acct['balance']
        is_contra = code in CONTRA_ACCOUNTS

        if t == 'Asset':
            # Contra assets have credit normal balance; bal is positive = credit balance = negative on BS
            display_amt = -abs(bal) if is_contra else bal
            indent = is_contra
            entry  = (code, acct['name'], display_amt, indent)
            if 1600 <= code <= 1999:
                noncurrent_assets.append(entry)
            else:
                current_assets.append(entry)

        elif t == 'Liability':
            display_amt = bal   # credit-normal, positive = credit balance = positive on BS
            entry = (code, acct['name'], display_amt, False)
            if 2100 <= code <= 2999:
                noncurrent_liab.append(entry)
            else:
                current_liab.append(entry)

        elif t == 'Equity':
            # Contra equity (Drawings 3020) has debit normal → bal positive = debit = negative on BS
            display_amt = -abs(bal) if is_contra else bal
            indent = is_contra
            equity.append((code, acct['name'], display_amt, indent))

    # Sort by code
    for lst in [noncurrent_assets, current_assets, equity, noncurrent_liab, current_liab]:
        lst.sort(key=lambda x: x[0])

    # Add net profit to equity
    equity.append((3040, 'Current Period Net Profit/(Loss)', round(net_profit, 2), False))

    # Compute totals
    total_noncurrent_assets = sum(a for _, _, a, _ in noncurrent_assets)
    total_current_assets    = sum(a for _, _, a, _ in current_assets)
    total_assets            = total_noncurrent_assets + total_current_assets

    total_equity            = sum(a for _, _, a, _ in equity)
    total_noncurrent_liab   = sum(a for _, _, a, _ in noncurrent_liab)
    total_current_liab      = sum(a for _, _, a, _ in current_liab)
    total_liabilities       = total_noncurrent_liab + total_current_liab
    total_equity_and_liab   = total_equity + total_liabilities

    bs_check = round(total_assets - total_equity_and_liab, 2)

    return {
        'noncurrent_assets':       noncurrent_assets,
        'current_assets':          current_assets,
        'equity':                  equity,
        'noncurrent_liab':         noncurrent_liab,
        'current_liab':            current_liab,
        'total_noncurrent_assets': round(total_noncurrent_assets, 2),
        'total_current_assets':    round(total_current_assets,    2),
        'total_assets':            round(total_assets,            2),
        'total_equity':            round(total_equity,            2),
        'total_noncurrent_liab':   round(total_noncurrent_liab,  2),
        'total_current_liab':      round(total_current_liab,     2),
        'total_liabilities':       round(total_liabilities,      2),
        'total_equity_and_liab':   round(total_equity_and_liab,  2),
        'bs_check':                bs_check,
    }


# ---------------------------------------------------------------------------
# 6. Build Cash Flow Data (Indirect Method)
# ---------------------------------------------------------------------------

# Working capital accounts: (code, description, is_asset)
# is_asset=True  → increase is negative CF (cash tied up)
# is_asset=False → increase is positive CF (credit-normal: liabilities and contra-assets)
WC_ACCOUNTS = [
    (1100, 'Accounts Receivable',          True),
    (1110, 'Allowance for Doubtful Debts', False),  # contra-asset (credit-normal)
    (1200, 'Inventory - Raw Materials',    True),
    (1210, 'Inventory - Packaging',        True),
    (1220, 'Inventory - Finished Goods',   True),
    (1300, 'Prepaid Expenses',             True),
    (1310, 'Prepaid Insurance',            True),
    (1320, 'Prepaid Rent',                 True),
    (1400, 'Advances to Employees',        True),
    (2010, 'Accounts Payable',             False),
    (2020, 'Accrued Expenses',             False),
    (2030, 'Accrued Wages',                False),
    (2040, 'Unearned Revenue',             False),
    (2050, 'Tax Payable',                  False),
]

CASH_CODES = [1010, 1020, 1021, 1022, 1030]

# Fixed asset codes (non-contra) for investing activities
FIXED_ASSET_CODES = [1600, 1610, 1620, 1630, 1640, 1650, 1660]

# Financing codes
FINANCING_CODES = [
    (3010, 'Owner Capital Introduced',   False),  # increase = positive CF
    (3020, 'Owner Drawings',             True),   # contra: increase in drawings = negative CF
    (2060, 'Short-term Loan Proceeds',   False),
    (2100, 'Long-term Loan Proceeds',    False),
    (2110, 'Mortgage Payable',           False),
]


def build_cf_data(accounts, gl_openings, adj_entries, net_profit):
    """
    Build Cash Flow Statement using the indirect method.

    Returns a dict with operating, investing, financing sections.
    """
    # Create quick lookup for account balances
    acct_map = {a['code']: a for a in accounts}

    def closing_bal(code):
        """Get closing balance (positive = normal side)."""
        a = acct_map.get(code)
        return a['balance'] if a else 0.0

    def opening_bal(code):
        return gl_openings.get(code, 0.0)

    # ── Operating: Non-cash items ────────────────────────────────────────────
    non_cash_items = []

    # Depreciation: scan adjusting entries for Dr 5300
    depr_total = sum(e['dr_amount'] for e in adj_entries if e['dr_code'] == '5300')
    if abs(depr_total) > 0.005:
        non_cash_items.append(('Add: Depreciation', round(depr_total, 2)))

    # Bad Debt Expense: scan for Dr 5800
    bad_debt_total = sum(e['dr_amount'] for e in adj_entries if e['dr_code'] == '5800')
    if abs(bad_debt_total) > 0.005:
        non_cash_items.append(('Add: Bad Debt Expense', round(bad_debt_total, 2)))

    # Loss on Disposal of Assets: any Dr 5930
    loss_disposal = sum(e['dr_amount'] for e in adj_entries if e['dr_code'] == '5930')
    if abs(loss_disposal) > 0.005:
        non_cash_items.append(('Add: Loss on Disposal', round(loss_disposal, 2)))

    # Gain on Disposal (4120): non-cash income to deduct
    gain_disposal = closing_bal(4120)
    if abs(gain_disposal) > 0.005:
        non_cash_items.append(('Less: Gain on Disposal', -round(gain_disposal, 2)))

    # ── Operating: Working capital changes ──────────────────────────────────
    working_capital = []
    for (code, desc, is_asset) in WC_ACCOUNTS:
        close = closing_bal(code)
        open_ = opening_bal(code)
        change = close - open_
        if abs(change) < 0.005:
            continue
        # Asset increase = cash outflow (negative); Liability increase = cash inflow (positive)
        cf_effect = -change if is_asset else change
        working_capital.append((f'Change in {desc}', round(cf_effect, 2)))

    total_non_cash = sum(a for _, a in non_cash_items)
    total_wc       = sum(a for _, a in working_capital)
    net_operating  = round(net_profit + total_non_cash + total_wc, 2)

    # ── Investing: Fixed asset purchases / disposals ─────────────────────────
    investing_items = []
    for code in FIXED_ASSET_CODES:
        close = closing_bal(code)
        open_ = opening_bal(code)
        change = close - open_
        if abs(change) < 0.005:
            continue
        a = acct_map.get(code)
        name = a['name'] if a else f'Fixed Asset {code}'
        # Increase in asset = purchase = cash outflow (negative)
        cf_effect = -change
        label = f'Purchase of {name}' if cf_effect < 0 else f'Proceeds from {name}'
        investing_items.append((label, round(cf_effect, 2)))

    net_investing = round(sum(a for _, a in investing_items), 2)

    # ── Financing: Capital, Drawings, Loans ──────────────────────────────────
    financing_items = []
    for (code, desc, is_contra) in FINANCING_CODES:
        close = closing_bal(code)
        open_ = opening_bal(code)
        change = close - open_
        if abs(change) < 0.005:
            continue
        # For normal credit accounts: increase = positive CF
        # For contra (Drawings): increase in balance = negative CF
        if is_contra:
            cf_effect = -abs(change)
        else:
            cf_effect = change
        financing_items.append((desc, round(cf_effect, 2)))

    net_financing = round(sum(a for _, a in financing_items), 2)

    # ── Reconciliation ───────────────────────────────────────────────────────
    net_change_in_cash = round(net_operating + net_investing + net_financing, 2)

    # Opening cash: sum of GL opening balances for cash accounts
    opening_cash = round(sum(opening_bal(c) for c in CASH_CODES), 2)

    # Closing cash: sum of closing balances for cash accounts from Adjusted TB
    closing_cash = round(sum(closing_bal(c) for c in CASH_CODES), 2)

    # CF check: opening + net_change should equal closing
    cf_check = round(opening_cash + net_change_in_cash - closing_cash, 2)

    return {
        'net_profit':         round(net_profit,        2),
        'non_cash_items':     non_cash_items,
        'working_capital':    working_capital,
        'total_non_cash':     round(total_non_cash,    2),
        'total_wc':           round(total_wc,          2),
        'net_operating':      net_operating,
        'investing_items':    investing_items,
        'net_investing':      net_investing,
        'financing_items':    financing_items,
        'net_financing':      net_financing,
        'net_change_in_cash': net_change_in_cash,
        'opening_cash':       opening_cash,
        'closing_cash':       closing_cash,
        'cf_check':           cf_check,
    }


# ---------------------------------------------------------------------------
# 7. Sheet Writers
# ---------------------------------------------------------------------------

def write_dashboard(wb, is_data, bs_data, cf_data, period_start, period_end, exceptions):
    ws  = add_sheet(wb, 'Dashboard', tab_color='00B050')
    row = write_title(ws, 'Financial Statements -- Dashboard',
                      'Shwe Mandalay Cafe', f"{period_start}  to  {period_end}")

    # ── Key Metrics ──────────────────────────────────────────────────────────
    row = write_section_header(ws, 'KEY FINANCIAL METRICS', row, col_span=4)
    row = write_header_row(ws, ['Metric', 'Amount', 'Margin / Check', ''], row)

    metrics = [
        ('Gross Revenue',          is_data['gross_revenue'],     None),
        ('Net Revenue',            is_data['net_revenue'],       None),
        ('Gross Profit',           is_data['gross_profit'],      is_data['gross_margin']),
        ('Operating Profit',       is_data['operating_profit'],  is_data['operating_margin']),
        ('Net Profit / (Loss)',    is_data['net_profit'],        is_data['net_margin']),
        ('Total Assets',           bs_data['total_assets'],      None),
        ('Total Liabilities',      bs_data['total_liabilities'], None),
        ('Total Equity',           bs_data['total_equity'],      None),
    ]

    for label, amount, margin in metrics:
        row_vals = [label, _n(amount), '', '']
        row = write_data_row(ws, row_vals, row, number_cols=[2])
        if margin is not None:
            ws.cell(row=row - 1, column=3, value=margin).number_format = PERCENT_FORMAT
            ws.cell(row=row - 1, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=row - 1, column=3).font = NORMAL_FONT

    row += 1

    # ── Validation Checks ────────────────────────────────────────────────────
    row = write_section_header(ws, 'VALIDATION CHECKS', row, col_span=4)
    row = write_header_row(ws, ['Check', 'Expected', 'Actual Difference', 'Result'], row)

    bs_ok = abs(bs_data['bs_check']) < 0.01
    cf_ok = abs(cf_data['cf_check']) < 0.01

    checks = [
        ('Balance Sheet: Assets = Equity + Liabilities', 0, bs_data['bs_check'], bs_ok),
        ('Cash Flow: Opening + Net Change = Closing Cash', 0, cf_data['cf_check'], cf_ok),
    ]

    for label, expected, diff, ok in checks:
        row = write_data_row(ws, [label, _n(expected), _n(diff) if abs(diff) > 0.005 else 0, ''],
                             row, number_cols=[2, 3])
        write_validation_result(ws, row - 1, 4, ok)
        ws.cell(row=row - 1, column=4).border = THIN_BORDER

    row += 1

    # ── Cash Summary ─────────────────────────────────────────────────────────
    row = write_section_header(ws, 'CASH FLOW SUMMARY', row, col_span=4)
    row = write_header_row(ws, ['Item', 'Amount', '', ''], row)
    cf_summary = [
        ('Opening Cash Balance',           cf_data['opening_cash']),
        ('Net Cash from Operations',       cf_data['net_operating']),
        ('Net Cash from Investing',        cf_data['net_investing']),
        ('Net Cash from Financing',        cf_data['net_financing']),
        ('Net Change in Cash',             cf_data['net_change_in_cash']),
        ('Closing Cash Balance (per BS)',  cf_data['closing_cash']),
    ]
    for label, amount in cf_summary:
        row = write_data_row(ws, [label, _n(amount), '', ''], row, number_cols=[2])

    row += 1

    if exceptions:
        row = write_section_header(ws, 'EXCEPTIONS & WARNINGS', row, col_span=4)
        for exc in exceptions:
            row = write_data_row(ws, [exc, '', '', ''], row)

    auto_fit_columns(ws)
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    freeze_panes(ws, row=2, col=1)


def write_income_statement(wb, is_data, period_start, period_end):
    ws  = add_sheet(wb, 'Income Statement', tab_color='4472C4')
    row = write_title(ws, 'Income Statement (Profit & Loss)',
                      'Shwe Mandalay Cafe',
                      f"For the period ended {period_end}")

    headers = ['Account', 'Description', 'Amount', 'Subtotal']
    row = write_header_row(ws, headers, row)

    def _line(code, name, amount, indent=False):
        nonlocal row
        disp_name = f'  {name}' if indent else name
        row = write_data_row(ws, [str(code), disp_name, _n(amount), None],
                             row, number_cols=[3, 4])

    def _subtotal(label, amount):
        nonlocal row
        row = write_total_row(ws, label, [None, None, _n(amount)], row)

    def _grand_total(label, amount):
        nonlocal row
        row = write_total_row(ws, label, [None, None, _n(amount)], row, double_line=True)

    # ── Revenue ──────────────────────────────────────────────────────────────
    row = write_section_header(ws, 'REVENUE', row, col_span=4)
    for code, name, amt in is_data['revenue']:
        _line(code, name, amt)
    for code, name, amt in is_data['contra_rev']:
        _line(code, f'Less: {name}', amt, indent=True)
    _subtotal('NET REVENUE', is_data['net_revenue'])
    row += 1

    # ── COGS ─────────────────────────────────────────────────────────────────
    row = write_section_header(ws, 'COST OF GOODS SOLD', row, col_span=4)
    for code, name, amt in is_data['cogs']:
        _line(code, name, amt)
    _subtotal('TOTAL COGS', -is_data['total_cogs'])
    row += 1

    # ── Gross Profit ─────────────────────────────────────────────────────────
    _grand_total('GROSS PROFIT', is_data['gross_profit'])
    row = write_data_row(ws, ['', 'Gross Profit Margin', None, is_data['gross_margin']],
                         row, number_cols=[3, 4])
    ws.cell(row=row - 1, column=4).number_format = PERCENT_FORMAT
    row += 1

    # ── Operating Expenses ───────────────────────────────────────────────────
    row = write_section_header(ws, 'OPERATING EXPENSES', row, col_span=4)
    for code, name, amt in is_data['opex']:
        _line(code, name, amt)
    _subtotal('TOTAL OPERATING EXPENSES', -is_data['total_opex'])
    row += 1

    # ── Operating Profit ─────────────────────────────────────────────────────
    _grand_total('OPERATING PROFIT', is_data['operating_profit'])
    row = write_data_row(ws, ['', 'Operating Profit Margin', None, is_data['operating_margin']],
                         row, number_cols=[3, 4])
    ws.cell(row=row - 1, column=4).number_format = PERCENT_FORMAT
    row += 1

    # ── Other Income / (Expenses) ────────────────────────────────────────────
    if is_data['other_income'] or is_data['other_exp']:
        row = write_section_header(ws, 'OTHER INCOME / (EXPENSES)', row, col_span=4)
        for code, name, amt in is_data['other_income']:
            _line(code, name, amt)
        for code, name, amt in is_data['other_exp']:
            _line(code, name, -amt)  # expenses shown as negative
        _subtotal('NET OTHER INCOME/(EXPENSES)', is_data['net_other'])
        row += 1

    # ── Profit Before Tax ────────────────────────────────────────────────────
    _grand_total('PROFIT BEFORE TAX', is_data['profit_before_tax'])
    row += 1

    # ── Tax ──────────────────────────────────────────────────────────────────
    if is_data['tax']:
        row = write_section_header(ws, 'TAX EXPENSE', row, col_span=4)
        for code, name, amt in is_data['tax']:
            _line(code, name, -amt)
        _subtotal('TOTAL TAX', -is_data['total_tax'])
        row += 1

    # ── Net Profit ───────────────────────────────────────────────────────────
    _grand_total('NET PROFIT / (LOSS)', is_data['net_profit'])
    row = write_data_row(ws, ['', 'Net Profit Margin', None, is_data['net_margin']],
                         row, number_cols=[3, 4])
    ws.cell(row=row - 1, column=4).number_format = PERCENT_FORMAT

    auto_fit_columns(ws)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    freeze_panes(ws)


def write_balance_sheet(wb, bs_data, period_end):
    ws  = add_sheet(wb, 'Balance Sheet', tab_color='4472C4')
    row = write_title(ws, 'Balance Sheet (Statement of Financial Position)',
                      'Shwe Mandalay Cafe',
                      f"As at {period_end}")

    headers = ['Account', 'Description', 'Amount', 'Total']
    row = write_header_row(ws, headers, row)

    def _line(code, name, amount, indent=False):
        nonlocal row
        disp_name = f'  {name}' if indent else name
        row = write_data_row(ws, [str(code), disp_name, _n(amount), None],
                             row, number_cols=[3, 4])

    def _subtotal(label, amount):
        nonlocal row
        row = write_total_row(ws, label, [None, None, _n(amount)], row)

    def _grand_total(label, amount):
        nonlocal row
        row = write_total_row(ws, label, [None, None, _n(amount)], row, double_line=True)

    # ── Non-Current Assets ───────────────────────────────────────────────────
    row = write_section_header(ws, 'NON-CURRENT ASSETS', row, col_span=4)
    for code, name, amt, indent in bs_data['noncurrent_assets']:
        _line(code, name, amt, indent)
    _subtotal('TOTAL NON-CURRENT ASSETS', bs_data['total_noncurrent_assets'])
    row += 1

    # ── Current Assets ───────────────────────────────────────────────────────
    row = write_section_header(ws, 'CURRENT ASSETS', row, col_span=4)
    for code, name, amt, indent in bs_data['current_assets']:
        _line(code, name, amt, indent)
    _subtotal('TOTAL CURRENT ASSETS', bs_data['total_current_assets'])
    row += 1

    _grand_total('TOTAL ASSETS', bs_data['total_assets'])
    row += 2

    # ── Equity ───────────────────────────────────────────────────────────────
    row = write_section_header(ws, 'EQUITY', row, col_span=4)
    for code, name, amt, indent in bs_data['equity']:
        _line(code, name, amt, indent)
    _subtotal('TOTAL EQUITY', bs_data['total_equity'])
    row += 1

    # ── Non-Current Liabilities ──────────────────────────────────────────────
    row = write_section_header(ws, 'NON-CURRENT LIABILITIES', row, col_span=4)
    if bs_data['noncurrent_liab']:
        for code, name, amt, indent in bs_data['noncurrent_liab']:
            _line(code, name, amt, indent)
    else:
        row = write_data_row(ws, ['', 'None', None, None], row)
    _subtotal('TOTAL NON-CURRENT LIABILITIES', bs_data['total_noncurrent_liab'])
    row += 1

    # ── Current Liabilities ──────────────────────────────────────────────────
    row = write_section_header(ws, 'CURRENT LIABILITIES', row, col_span=4)
    if bs_data['current_liab']:
        for code, name, amt, indent in bs_data['current_liab']:
            _line(code, name, amt, indent)
    else:
        row = write_data_row(ws, ['', 'None', None, None], row)
    _subtotal('TOTAL CURRENT LIABILITIES', bs_data['total_current_liab'])
    row += 1

    _subtotal('TOTAL LIABILITIES', bs_data['total_liabilities'])
    _grand_total('TOTAL EQUITY & LIABILITIES', bs_data['total_equity_and_liab'])
    row += 1

    # BS check row
    bs_ok  = abs(bs_data['bs_check']) < 0.01
    check_val = bs_data['bs_check']
    row = write_data_row(ws, ['', 'CHECK: Assets - (Equity + Liabilities)',
                               _n(check_val) if abs(check_val) > 0.005 else 0, ''],
                         row, number_cols=[3])
    write_validation_result(ws, row - 1, 4, bs_ok)
    ws.cell(row=row - 1, column=4).border = THIN_BORDER

    auto_fit_columns(ws)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    freeze_panes(ws)


def write_cash_flow(wb, cf_data, period_start, period_end):
    ws  = add_sheet(wb, 'Cash Flow', tab_color='4472C4')
    row = write_title(ws, 'Cash Flow Statement (Indirect Method)',
                      'Shwe Mandalay Cafe',
                      f"For the period ended {period_end}")

    headers = ['Item', 'Description', 'Amount', 'Net Total']
    row = write_header_row(ws, headers, row)

    def _line(label, amount):
        nonlocal row
        row = write_data_row(ws, ['', label, _n(amount), None], row, number_cols=[3, 4])

    def _subtotal(label, amount):
        nonlocal row
        row = write_total_row(ws, label, [None, None, _n(amount)], row)

    def _grand_total(label, amount):
        nonlocal row
        row = write_total_row(ws, label, [None, None, _n(amount)], row, double_line=True)

    # ── Operating Activities ──────────────────────────────────────────────────
    row = write_section_header(ws, 'OPERATING ACTIVITIES', row, col_span=4)
    _line('Net Profit/(Loss)', cf_data['net_profit'])
    row += 1

    if cf_data['non_cash_items']:
        row = write_data_row(ws, ['', 'Adjustments for non-cash items:', None, None], row)
        ws.cell(row=row - 1, column=2).font = Font(bold=True, italic=True, size=11, name='Arial')
        for desc, amt in cf_data['non_cash_items']:
            _line(f'  {desc}', amt)
        row += 1

    if cf_data['working_capital']:
        row = write_data_row(ws, ['', 'Changes in working capital:', None, None], row)
        ws.cell(row=row - 1, column=2).font = Font(bold=True, italic=True, size=11, name='Arial')
        for desc, amt in cf_data['working_capital']:
            _line(f'  {desc}', amt)
        row += 1

    _subtotal('NET CASH FROM OPERATING ACTIVITIES', cf_data['net_operating'])
    row += 1

    # ── Investing Activities ──────────────────────────────────────────────────
    row = write_section_header(ws, 'INVESTING ACTIVITIES', row, col_span=4)
    if cf_data['investing_items']:
        for desc, amt in cf_data['investing_items']:
            _line(desc, amt)
    else:
        _line('No investing activities this period', 0)
    _subtotal('NET CASH FROM INVESTING ACTIVITIES', cf_data['net_investing'])
    row += 1

    # ── Financing Activities ──────────────────────────────────────────────────
    row = write_section_header(ws, 'FINANCING ACTIVITIES', row, col_span=4)
    if cf_data['financing_items']:
        for desc, amt in cf_data['financing_items']:
            _line(desc, amt)
    else:
        _line('No financing activities this period', 0)
    _subtotal('NET CASH FROM FINANCING ACTIVITIES', cf_data['net_financing'])
    row += 1

    # ── Reconciliation ────────────────────────────────────────────────────────
    _grand_total('NET INCREASE/(DECREASE) IN CASH', cf_data['net_change_in_cash'])
    row += 1

    row = write_data_row(ws, ['', 'Cash & Cash Equivalents at Start of Period',
                               _n(cf_data['opening_cash']), None],
                         row, number_cols=[3])

    _grand_total('Cash & Cash Equivalents at End of Period', cf_data['closing_cash'])
    row += 1

    # CF check
    cf_ok     = abs(cf_data['cf_check']) < 0.01
    check_val = cf_data['cf_check']
    row = write_data_row(ws, ['', 'CHECK: Opening + Net Change - Closing =',
                               _n(check_val) if abs(check_val) > 0.005 else 0, ''],
                         row, number_cols=[3])
    write_validation_result(ws, row - 1, 4, cf_ok)
    ws.cell(row=row - 1, column=4).border = THIN_BORDER

    auto_fit_columns(ws)
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    freeze_panes(ws)


def write_exceptions_sheet(wb, exceptions):
    ws  = add_sheet(wb, 'Exceptions', tab_color='FF0000')
    row = write_title(ws, 'Exceptions & Warnings',
                      'Financial Statements -- Module 6')
    row = write_header_row(ws, ['#', 'Exception / Warning'], row)
    for i, exc in enumerate(exceptions, 1):
        row = write_data_row(ws, [i, exc], row)
    auto_fit_columns(ws)
    ws.column_dimensions['B'].width = 60
    freeze_panes(ws)


# ---------------------------------------------------------------------------
# 8. Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)

    data_dir     = sys.argv[1]
    period_start = sys.argv[2]
    period_end   = sys.argv[3]
    output_file  = sys.argv[4]

    print(f"\n{'='*60}")
    print(f"  MODULE 6 -- GENERATE FINANCIAL STATEMENTS")
    print(f"  Period : {period_start}  to  {period_end}")
    print(f"  Data   : {data_dir}")
    print(f"  Output : {output_file}")
    print(f"{'='*60}\n")

    coa_path = Path(data_dir) / 'chart_of_accounts.xlsx'
    coa      = COAMapper(str(coa_path)) if coa_path.exists() else COAMapper()
    exceptions = []

    # ── 1. Load Adjusted TB ──────────────────────────────────────────────────
    print("Loading Adjusted Trial Balance...")
    accounts, err = load_adjusted_tb(data_dir)
    if err:
        print(f"  ERROR: {err}")
        exceptions.append(f"FATAL: {err}")
        print("Cannot continue without trial balance. Exiting.")
        sys.exit(1)
    print(f"  Accounts loaded: {len(accounts)}")

    # ── 2. Load GL Data (opening balances for CF) ────────────────────────────
    print("\nLoading GL opening balances for Cash Flow...")
    gl_openings, warn = load_gl_data(data_dir, period_start, coa)
    if warn:
        print(f"  WARNING: {warn}")
        exceptions.append(warn)
    else:
        print(f"  GL accounts with openings: {len(gl_openings)}")

    # ── 3. Load Adjusting Entries (for non-cash items) ───────────────────────
    print("\nLoading adjusting entries...")
    adj_entries, warn = load_adj_entries(data_dir)
    if warn:
        print(f"  WARNING: {warn}")
        exceptions.append(warn)
    else:
        print(f"  Adjusting entries: {len(adj_entries)}")

    # ── 4. Build Financial Statement Data ────────────────────────────────────
    print("\nBuilding Income Statement...")
    is_data = build_is_data(accounts, coa)
    print(f"  Net Revenue        : {is_data['net_revenue']:>15,.2f}")
    print(f"  Gross Profit       : {is_data['gross_profit']:>15,.2f}  ({is_data['gross_margin']:.1%})")
    print(f"  Operating Profit   : {is_data['operating_profit']:>15,.2f}  ({is_data['operating_margin']:.1%})")
    print(f"  Net Profit         : {is_data['net_profit']:>15,.2f}  ({is_data['net_margin']:.1%})")

    print("\nBuilding Balance Sheet...")
    bs_data = build_bs_data(accounts, coa, is_data['net_profit'])
    print(f"  Total Assets       : {bs_data['total_assets']:>15,.2f}")
    print(f"  Total Equity       : {bs_data['total_equity']:>15,.2f}")
    print(f"  Total Liabilities  : {bs_data['total_liabilities']:>15,.2f}")
    print(f"  BS Check           : {bs_data['bs_check']:>15,.2f}  ({'PASS' if abs(bs_data['bs_check']) < 0.01 else 'FAIL'})")

    if abs(bs_data['bs_check']) > 0.01:
        msg = (f"Balance Sheet does NOT balance: "
               f"Assets={bs_data['total_assets']:,.2f}, "
               f"Equity+Liab={bs_data['total_equity_and_liab']:,.2f}, "
               f"Diff={bs_data['bs_check']:,.2f}")
        exceptions.append(msg)
        print(f"  EXCEPTION: {msg}")

    print("\nBuilding Cash Flow Statement...")
    cf_data = build_cf_data(accounts, gl_openings, adj_entries, is_data['net_profit'])
    print(f"  Net Operating CF   : {cf_data['net_operating']:>15,.2f}")
    print(f"  Net Investing CF   : {cf_data['net_investing']:>15,.2f}")
    print(f"  Net Financing CF   : {cf_data['net_financing']:>15,.2f}")
    print(f"  Net Change in Cash : {cf_data['net_change_in_cash']:>15,.2f}")
    print(f"  Opening Cash       : {cf_data['opening_cash']:>15,.2f}")
    print(f"  Closing Cash (BS)  : {cf_data['closing_cash']:>15,.2f}")
    print(f"  CF Check           : {cf_data['cf_check']:>15,.2f}  ({'PASS' if abs(cf_data['cf_check']) < 0.01 else 'FAIL'})")

    if abs(cf_data['cf_check']) > 0.01:
        msg = (f"Cash Flow does NOT reconcile: "
               f"Opening({cf_data['opening_cash']:,.2f}) + "
               f"NetChange({cf_data['net_change_in_cash']:,.2f}) != "
               f"Closing({cf_data['closing_cash']:,.2f}), "
               f"Diff={cf_data['cf_check']:,.2f}")
        exceptions.append(msg)
        print(f"  EXCEPTION: {msg}")

    # ── 5. Write Excel ───────────────────────────────────────────────────────
    print(f"\nWriting output to: {output_file}")
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    wb = create_workbook()
    write_dashboard(wb, is_data, bs_data, cf_data, period_start, period_end, exceptions)
    write_income_statement(wb, is_data, period_start, period_end)
    write_balance_sheet(wb, bs_data, period_end)
    write_cash_flow(wb, cf_data, period_start, period_end)
    if exceptions:
        write_exceptions_sheet(wb, exceptions)

    save_workbook(wb, output_file)

    sheets = 'Dashboard | Income Statement | Balance Sheet | Cash Flow'
    if exceptions:
        sheets += ' | Exceptions'

    bs_check_label = 'Assets = Equity + Liab: YES' if abs(bs_data['bs_check']) < 0.01 else f'FAIL (diff={bs_data["bs_check"]:,.2f})'
    cf_check_label = 'CF reconciles: YES' if abs(cf_data['cf_check']) < 0.01 else f'FAIL (diff={cf_data["cf_check"]:,.2f})'

    print(f"\n{'='*60}")
    print(f"  OUTPUT  : {output_file}")
    print(f"  Sheets  : {sheets}")
    print(f"  {bs_check_label}")
    print(f"  {cf_check_label}")
    print(f"  Net Profit (IS)    : {is_data['net_profit']:>15,.2f}")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
