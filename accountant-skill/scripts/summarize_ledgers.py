"""
Module 2: Summarize Ledgers
Reads all 6 ledger files for a period and produces a consolidated summary
with control account reconciliation and movement analysis.

Usage:
    python summarize_ledgers.py <input_dir> <period_start> <period_end> <output_file> [coa_file]

Example:
    python summarize_ledgers.py data/Jan2026 2026-01-01 2026-01-31 data/Jan2026/ledger_summary_Jan2026.xlsx data/Jan2026/chart_of_accounts.xlsx
"""
import sys
import os
import pandas as pd
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))
from utils.excel_reader import read_xlsx, read_all_sheets
from utils.excel_writer import (create_workbook, add_sheet, write_title, write_header_row,
                                 write_data_row, write_section_header, write_total_row,
                                 auto_fit_columns, freeze_panes, save_workbook,
                                 write_validation_result, WARNING_FILL, NORMAL_FONT, TOTAL_FONT,
                                 THIN_BORDER)
from openpyxl.styles import PatternFill, Font
from utils.coa_mapper import COAMapper


# ── Control account codes ─────────────────────────────────────────────────────
AR_GL_ACCOUNT   = 1100
AP_GL_ACCOUNT   = 2010
CASH_GL_ACCOUNTS = [1020, 1021, 1022]
MOVEMENT_FLAG_PCT = 50.0   # Flag accounts with >50% balance movement

# ── Ledger file name patterns ─────────────────────────────────────────────────
LEDGER_FILES = {
    'general_ledger': ['general_ledger', 'general ledger', 'gl'],
    'ar_ledger':      ['accounts_receivable_ledger', 'ar_ledger', 'receivable'],
    'ap_ledger':      ['accounts_payable_ledger', 'ap_ledger', 'payable'],
    'cash_ledger':    ['cash_ledger', 'cash ledger', 'bank_ledger'],
    'fixed_assets':   ['fixed_assets_ledger', 'fixed_asset_ledger', 'fixed assets'],
    'equity_ledger':  ['equity_ledger', 'equity ledger'],
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def find_ledger_file(input_dir, patterns):
    input_dir = Path(input_dir)
    for f in input_dir.glob('*.xlsx'):
        fname = f.stem.lower().replace('-', '_').replace(' ', '_')
        for p in patterns:
            if p.replace(' ', '_') in fname:
                return f
    return None


def to_num(series):
    return pd.to_numeric(series, errors='coerce').fillna(0.0)


def normalize_cols(df):
    """Lowercase + strip all column names."""
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def find_col(df, candidates):
    """Return the first column name from candidates that exists in df."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def get_opening_and_period(df, date_col, debit_col, credit_col, balance_col,
                            period_start, period_end, normal_balance):
    """
    Split a ledger DataFrame into opening balance and period activity.

    Opening balance = the Balance value of the last row before period_start
                      OR a row whose description says 'opening'.
    Returns: (opening_balance, period_debits, period_credits, closing_balance, row_count)
    """
    df = df.copy()
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

    # Identify opening-balance rows (no date, date before period, or description = "opening")
    desc_col = find_col(df, ['description', 'narration', 'details', 'memo'])
    is_opening = (
        df[date_col].isna() |
        (df[date_col] < pd.Timestamp(period_start))
    )
    if desc_col:
        is_opening = is_opening | df[desc_col].astype(str).str.lower().str.contains('opening', na=False)

    pre_rows = df[is_opening]
    period_rows = df[
        df[date_col].notna() &
        (df[date_col] >= pd.Timestamp(period_start)) &
        (df[date_col] <= pd.Timestamp(period_end))
    ]

    # Opening balance: prefer Balance column; fall back to 0
    opening = 0.0
    if balance_col and balance_col in df.columns and len(pre_rows) > 0:
        bal_series = to_num(pre_rows[balance_col]).replace(0, float('nan')).dropna()
        if len(bal_series) > 0:
            opening = float(bal_series.iloc[-1])

    debits  = to_num(period_rows[debit_col]).sum()  if debit_col  else 0.0
    credits = to_num(period_rows[credit_col]).sum() if credit_col else 0.0

    nb = normal_balance.lower()
    closing = (opening + debits - credits) if nb == 'debit' else (opening + credits - debits)

    return opening, debits, credits, closing, len(period_rows)


# ── General Ledger ────────────────────────────────────────────────────────────

def process_general_ledger(filepath, period_start, period_end, coa):
    """
    Read general_ledger.xlsx (single sheet with Account Code column,
    or multi-sheet with one sheet per account).

    Returns: dict[int → account_summary], error_string
    """
    result = read_all_sheets(filepath)
    if result['error']:
        return None, result['error']

    sheets = result['data']
    accounts = {}

    # --- Single-sheet format: one big table with Account Code column ---
    if len(sheets) == 1:
        df = normalize_cols(list(sheets.values())[0])

        code_col    = find_col(df, ['account code', 'code', 'acct code', 'account_code', 'no.'])
        date_col    = find_col(df, ['date', 'trans date', 'entry date'])
        debit_col   = find_col(df, ['debit', 'dr'])
        credit_col  = find_col(df, ['credit', 'cr'])
        balance_col = find_col(df, ['balance', 'running balance', 'bal'])

        if code_col is None:
            return None, "General Ledger: 'Account Code' column not found."
        if date_col is None:
            return None, "General Ledger: 'Date' column not found."

        df[code_col] = pd.to_numeric(df[code_col], errors='coerce')

        for code in df[code_col].dropna().unique():
            code = int(code)
            info = coa.get_account(code)
            if not info:
                continue
            acct_df = df[df[code_col] == code]
            opening, debits, credits, closing, count = get_opening_and_period(
                acct_df, date_col, debit_col, credit_col, balance_col,
                period_start, period_end, info['normal_balance']
            )
            accounts[code] = {
                'name': info['name'], 'type': info['type'],
                'sub_type': info['sub_type'], 'normal_balance': info['normal_balance'],
                'opening': opening, 'debits': debits,
                'credits': credits, 'closing': closing, 'rows': count,
            }

    # --- Multi-sheet format: each sheet = one account ---
    else:
        for sheet_name, raw_df in sheets.items():
            df = normalize_cols(raw_df)
            try:
                code = int(''.join(c for c in sheet_name if c.isdigit()))
            except ValueError:
                continue
            info = coa.get_account(code)
            if not info:
                continue
            date_col    = find_col(df, ['date', 'trans date'])
            debit_col   = find_col(df, ['debit', 'dr'])
            credit_col  = find_col(df, ['credit', 'cr'])
            balance_col = find_col(df, ['balance', 'bal'])
            if date_col is None:
                continue
            opening, debits, credits, closing, count = get_opening_and_period(
                df, date_col, debit_col, credit_col, balance_col,
                period_start, period_end, info['normal_balance']
            )
            accounts[code] = {
                'name': info['name'], 'type': info['type'],
                'sub_type': info['sub_type'], 'normal_balance': info['normal_balance'],
                'opening': opening, 'debits': debits,
                'credits': credits, 'closing': closing, 'rows': count,
            }

    return accounts, None


# ── Subsidiary Ledgers ────────────────────────────────────────────────────────

def process_subsidiary_ledger(filepath, period_start, period_end, entity_col_candidates, normal_balance):
    """
    Read an AR or AP subsidiary ledger.
    Returns: dict[entity_name → summary], error
    """
    result = read_all_sheets(filepath)
    if result['error']:
        return None, result['error']

    sheets = result['data']
    entities = {}

    for sheet_name, raw_df in sheets.items():
        df = normalize_cols(raw_df)
        entity_col  = find_col(df, [c.lower() for c in entity_col_candidates])
        date_col    = find_col(df, ['date', 'trans date'])
        debit_col   = find_col(df, ['debit', 'dr'])
        credit_col  = find_col(df, ['credit', 'cr'])
        balance_col = find_col(df, ['balance', 'bal'])

        if date_col is None:
            continue

        if entity_col:
            # Single sheet with entity column — group by entity
            for entity, group in df.groupby(entity_col):
                entity = str(entity).strip()
                if entity.lower() in ('nan', ''):
                    continue
                opening, debits, credits, closing, count = get_opening_and_period(
                    group, date_col, debit_col, credit_col, balance_col,
                    period_start, period_end, normal_balance
                )
                if entity in entities:
                    entities[entity]['debits']  += debits
                    entities[entity]['credits'] += credits
                    entities[entity]['closing']  = (
                        entities[entity]['opening'] + entities[entity]['debits'] - entities[entity]['credits']
                        if normal_balance == 'debit'
                        else entities[entity]['opening'] + entities[entity]['credits'] - entities[entity]['debits']
                    )
                else:
                    entities[entity] = {'opening': opening, 'debits': debits,
                                         'credits': credits, 'closing': closing, 'rows': count}
        else:
            # Sheet name is the entity
            entity = sheet_name.strip()
            opening, debits, credits, closing, count = get_opening_and_period(
                df, date_col, debit_col, credit_col, balance_col,
                period_start, period_end, normal_balance
            )
            entities[entity] = {'opening': opening, 'debits': debits,
                                  'credits': credits, 'closing': closing, 'rows': count}

    return entities, None


def process_cash_ledger(filepath, period_start, period_end):
    """
    Read cash_ledger.xlsx. Groups by Bank Account if that column exists.
    Returns: dict[bank_name → summary], error
    """
    result = read_all_sheets(filepath)
    if result['error']:
        return None, result['error']

    sheets = result['data']
    banks = {}

    for sheet_name, raw_df in sheets.items():
        df = normalize_cols(raw_df)
        bank_col    = find_col(df, ['bank account', 'bank_account', 'bank', 'account'])
        date_col    = find_col(df, ['date', 'trans date'])
        debit_col   = find_col(df, ['debit', 'dr'])
        credit_col  = find_col(df, ['credit', 'cr'])
        balance_col = find_col(df, ['balance', 'bal'])

        if date_col is None:
            continue

        if bank_col:
            for bank, group in df.groupby(bank_col):
                bank = str(bank).strip()
                if bank.lower() in ('nan', ''):
                    continue
                opening, debits, credits, closing, count = get_opening_and_period(
                    group, date_col, debit_col, credit_col, balance_col,
                    period_start, period_end, 'debit'
                )
                banks[bank] = {'opening': opening, 'debits': debits,
                                'credits': credits, 'closing': closing, 'rows': count}
        else:
            bank = sheet_name.strip()
            opening, debits, credits, closing, count = get_opening_and_period(
                df, date_col, debit_col, credit_col, balance_col,
                period_start, period_end, 'debit'
            )
            banks[bank] = {'opening': opening, 'debits': debits,
                            'credits': credits, 'closing': closing, 'rows': count}

    return banks, None


def process_fixed_assets(filepath):
    """Read the fixed asset register. Returns list of row dicts."""
    result = read_xlsx(
        filepath,
        required_columns=['Asset ID', 'Description', 'Cost'],
        optional_columns=['Account Code', 'Date Acquired', 'Useful Life (Years)',
                          'Salvage Value', 'Depreciation Method',
                          'Accumulated Depreciation', 'Net Book Value', 'Status', 'Category']
    )
    if result['error']:
        return None, result['error']
    return result['data'].to_dict('records'), None


# ── Output sheets ─────────────────────────────────────────────────────────────

def write_dashboard(wb, gl_accounts, ar_entities, ap_entities, cash_banks,
                    assets, exceptions, period_start, period_end):
    ws = add_sheet(wb, 'Dashboard', tab_color='00B050')
    row = write_title(ws, 'SHWE MANDALAY CAFE', 'Ledger Summary — Dashboard',
                      f"{period_start} to {period_end}")

    # Ledger status table
    row = write_section_header(ws, 'LEDGER FILES', row, col_span=3)
    row = write_header_row(ws, ['Ledger', 'Accounts / Entities', 'Status'], row)
    statuses = [
        ('General Ledger',    len(gl_accounts),  'OK' if gl_accounts else 'ERROR'),
        ('AR Ledger',         len(ar_entities),  'OK' if ar_entities else 'NOT FOUND'),
        ('AP Ledger',         len(ap_entities),  'OK' if ap_entities else 'NOT FOUND'),
        ('Cash Ledger',       len(cash_banks),   'OK' if cash_banks  else 'NOT FOUND'),
        ('Fixed Assets',      len(assets),       'OK' if assets      else 'NOT FOUND'),
    ]
    for label, count, status in statuses:
        row = write_data_row(ws, [label, count, status], row)

    # Control account checks
    row += 2
    row = write_section_header(ws, 'CONTROL ACCOUNT RECONCILIATION', row, col_span=5)
    row = write_header_row(ws, ['Account', 'GL Balance', 'Subsidiary Total', 'Difference', 'Result'], row)

    ar_total   = sum(e['closing'] for e in ar_entities.values())
    ap_total   = sum(e['closing'] for e in ap_entities.values())
    cash_total = sum(b['closing'] for b in cash_banks.values())

    gl_ar   = gl_accounts.get(AR_GL_ACCOUNT,  {}).get('closing', None)
    gl_ap   = gl_accounts.get(AP_GL_ACCOUNT,  {}).get('closing', None)
    gl_cash = sum(gl_accounts.get(c, {}).get('closing', 0) for c in CASH_GL_ACCOUNTS)

    checks = [
        ('Accounts Receivable (1100)', gl_ar,   ar_total,   ar_entities),
        ('Accounts Payable (2010)',    gl_ap,   ap_total,   ap_entities),
        ('Cash at Bank (1020)',        gl_cash, cash_total, cash_banks),
    ]
    for label, gl_bal, sub_total, entities in checks:
        if gl_bal is None or not entities:
            row = write_data_row(ws, [label, gl_bal or 'N/A', sub_total, 'N/A', 'SKIP'], row)
            continue
        diff   = gl_bal - sub_total
        passed = abs(diff) < 0.01
        status = 'MATCH' if passed else 'MISMATCH'
        row = write_data_row(ws, [label, gl_bal, sub_total, diff, status], row)

    auto_fit_columns(ws)
    freeze_panes(ws)
    return ws


def write_gl_balances(wb, gl_accounts, period_start, period_end):
    ws = add_sheet(wb, 'GL Balances', tab_color='4472C4')
    row = write_title(ws, 'General Ledger — Account Balances',
                      period=f"{period_start} to {period_end}")
    headers = ['Code', 'Account Name', 'Type', 'Sub-Type', 'Normal Bal',
               'Opening', 'Debits', 'Credits', 'Closing', 'Movement %', 'Flag']
    row = write_header_row(ws, headers, row)

    REVIEW_FILL = WARNING_FILL

    for code in sorted(gl_accounts.keys()):
        a = gl_accounts[code]
        opening = a['opening']
        closing = a['closing']

        # Movement analysis
        flag = ''
        movement_str = '-'
        if opening != 0:
            pct = abs(closing - opening) / abs(opening) * 100
            movement_str = f"{pct:.1f}%"
            if pct > MOVEMENT_FLAG_PCT:
                flag = 'REVIEW'

        vals = [code, a['name'], a['type'], a['sub_type'], a['normal_balance'],
                opening, a['debits'], a['credits'], closing, movement_str, flag]
        row = write_data_row(ws, vals, row)

        # Highlight review rows
        if flag == 'REVIEW':
            for col_idx in range(1, len(vals) + 1):
                cell = ws.cell(row=row - 1, column=col_idx)
                cell.fill = REVIEW_FILL

    # Totals
    total_dr = sum(a['closing'] for a in gl_accounts.values() if a['normal_balance'].lower() == 'debit')
    total_cr = sum(a['closing'] for a in gl_accounts.values() if a['normal_balance'].lower() == 'credit')
    row += 1
    write_total_row(ws, 'Total Debit Balances',
                     [None, None, None, None, None, None, None, total_dr, None, None], row)
    row += 1
    write_total_row(ws, 'Total Credit Balances',
                     [None, None, None, None, None, None, None, total_cr, None, None], row)

    auto_fit_columns(ws)
    freeze_panes(ws)


def write_subsidiary_sheet(wb, sheet_name, entities, title, entity_label,
                            tab_color, period_start, period_end):
    ws = add_sheet(wb, sheet_name, tab_color=tab_color)
    row = write_title(ws, title, period=f"{period_start} to {period_end}")
    row = write_header_row(ws, [entity_label, 'Opening Balance', 'Debits', 'Credits', 'Closing Balance'], row)

    total_closing = 0.0
    for entity in sorted(entities.keys()):
        e = entities[entity]
        row = write_data_row(ws, [entity, e['opening'], e['debits'], e['credits'], e['closing']], row)
        total_closing += e['closing']

    row += 1
    write_total_row(ws, 'TOTAL', [None, None, None, total_closing], row, double_line=True)
    auto_fit_columns(ws)
    freeze_panes(ws)
    return total_closing


def write_control_account_sheet(wb, gl_accounts, ar_total, ap_total, cash_total,
                                  ar_entities, ap_entities, cash_banks):
    ws = add_sheet(wb, 'Control Acct Recon', tab_color='FF0000')
    row = write_title(ws, 'Control Account Reconciliation')
    row = write_header_row(ws, ['Account', 'GL Balance', 'Subsidiary Total', 'Difference', 'Result'], row)

    gl_ar   = gl_accounts.get(AR_GL_ACCOUNT, {}).get('closing', 0)
    gl_ap   = gl_accounts.get(AP_GL_ACCOUNT, {}).get('closing', 0)
    gl_cash = sum(gl_accounts.get(c, {}).get('closing', 0) for c in CASH_GL_ACCOUNTS)

    checks = [
        ('AR — Accts Receivable (1100)', gl_ar,   ar_total,   bool(ar_entities)),
        ('AP — Accts Payable (2010)',     gl_ap,   ap_total,   bool(ap_entities)),
        ('Cash at Bank (1020–1022)',      gl_cash, cash_total, bool(cash_banks)),
    ]
    all_ok = True
    for label, gl_bal, sub_total, has_data in checks:
        if not has_data:
            row = write_data_row(ws, [label, gl_bal, 'N/A', 'N/A', 'SKIP'], row)
            continue
        diff   = gl_bal - sub_total
        passed = abs(diff) < 0.01
        if not passed:
            all_ok = False
        row_data = [label, gl_bal, sub_total, diff, 'MATCH' if passed else 'MISMATCH']
        row = write_data_row(ws, row_data, row)
        # Color the result cell
        result_cell = ws.cell(row=row - 1, column=5)
        result_cell.fill = PatternFill('solid', fgColor='C6EFCE' if passed else 'FFC7CE')
        result_cell.font = Font(bold=True, color='006100' if passed else '9C0006')

    auto_fit_columns(ws)
    freeze_panes(ws)
    return all_ok


def write_fixed_assets_sheet(wb, assets):
    ws = add_sheet(wb, 'Fixed Assets', tab_color='4472C4')
    row = write_title(ws, 'Fixed Assets Register Summary')

    if not assets:
        ws.cell(row=row, column=1, value='No fixed asset data provided.')
        return

    headers = ['Asset ID', 'Description', 'Account Code', 'Date Acquired',
               'Cost', 'Accum. Depreciation', 'Net Book Value', 'Method', 'Status']
    row = write_header_row(ws, headers, row)

    total_cost = total_accum = total_nbv = 0.0
    for a in assets:
        cost  = float(a.get('Cost', 0) or 0)
        accum = float(a.get('Accumulated Depreciation', 0) or 0)
        nbv   = float(a.get('Net Book Value', cost - accum) or 0)
        total_cost  += cost
        total_accum += accum
        total_nbv   += nbv
        row = write_data_row(ws, [
            a.get('Asset ID', ''), a.get('Description', ''),
            a.get('Account Code', ''), str(a.get('Date Acquired', '')),
            cost, accum, nbv,
            a.get('Depreciation Method', ''), a.get('Status', 'Active')
        ], row)

    row += 1
    write_total_row(ws, 'TOTAL', [None, None, None, total_cost, total_accum, total_nbv, None, None],
                     row, double_line=True)
    auto_fit_columns(ws)
    freeze_panes(ws)


def write_exceptions_sheet(wb, exceptions):
    if not exceptions:
        return
    ws = add_sheet(wb, 'Exceptions', tab_color='FF0000')
    row = write_title(ws, 'Exceptions & Warnings')
    row = write_header_row(ws, ['Ledger', 'Issue'], row)
    for exc in exceptions:
        row = write_data_row(ws, [exc['ledger'], exc['issue']], row)
    auto_fit_columns(ws)


# ── Main ──────────────────────────────────────────────────────────────────────

def main(input_dir, period_start, period_end, output_file, coa_file=None):
    input_dir = Path(input_dir)
    coa = COAMapper(coa_file) if coa_file else COAMapper()
    exceptions = []

    # ─ 1. General Ledger ────────────────────────────────────────────────────
    gl_file = find_ledger_file(input_dir, LEDGER_FILES['general_ledger'])
    if not gl_file:
        print("ERROR: general_ledger.xlsx not found in input directory.")
        sys.exit(1)

    gl_accounts, gl_err = process_general_ledger(gl_file, period_start, period_end, coa)
    if gl_err:
        print(f"ERROR: {gl_err}")
        sys.exit(1)
    print(f"  General Ledger : {len(gl_accounts)} accounts")

    # ─ 2. AR Ledger ─────────────────────────────────────────────────────────
    ar_entities = {}
    ar_file = find_ledger_file(input_dir, LEDGER_FILES['ar_ledger'])
    if ar_file:
        ar_entities, err = process_subsidiary_ledger(
            ar_file, period_start, period_end,
            ['Customer', 'Client', 'Debtor', 'Received From'], 'debit')
        if err:
            exceptions.append({'ledger': 'AR Ledger', 'issue': err})
    else:
        exceptions.append({'ledger': 'AR Ledger', 'issue': 'File not found'})
    print(f"  AR Ledger      : {len(ar_entities)} customers")

    # ─ 3. AP Ledger ─────────────────────────────────────────────────────────
    ap_entities = {}
    ap_file = find_ledger_file(input_dir, LEDGER_FILES['ap_ledger'])
    if ap_file:
        ap_entities, err = process_subsidiary_ledger(
            ap_file, period_start, period_end,
            ['Supplier', 'Vendor', 'Creditor', 'Paid To'], 'credit')
        if err:
            exceptions.append({'ledger': 'AP Ledger', 'issue': err})
    else:
        exceptions.append({'ledger': 'AP Ledger', 'issue': 'File not found'})
    print(f"  AP Ledger      : {len(ap_entities)} suppliers")

    # ─ 4. Cash Ledger ───────────────────────────────────────────────────────
    cash_banks = {}
    cash_file = find_ledger_file(input_dir, LEDGER_FILES['cash_ledger'])
    if cash_file:
        cash_banks, err = process_cash_ledger(cash_file, period_start, period_end)
        if err:
            exceptions.append({'ledger': 'Cash Ledger', 'issue': err})
    else:
        exceptions.append({'ledger': 'Cash Ledger', 'issue': 'File not found'})
    print(f"  Cash Ledger    : {len(cash_banks)} bank account(s)")

    # ─ 5. Fixed Assets ──────────────────────────────────────────────────────
    assets = []
    fa_file = find_ledger_file(input_dir, LEDGER_FILES['fixed_assets'])
    if fa_file:
        assets, err = process_fixed_assets(fa_file)
        if err:
            exceptions.append({'ledger': 'Fixed Assets', 'issue': err})
    else:
        exceptions.append({'ledger': 'Fixed Assets', 'issue': 'File not found (optional)'})
    print(f"  Fixed Assets   : {len(assets)} assets")

    # ─ Build workbook ───────────────────────────────────────────────────────
    wb = create_workbook()

    write_dashboard(wb, gl_accounts, ar_entities, ap_entities, cash_banks,
                     assets, exceptions, period_start, period_end)
    write_gl_balances(wb, gl_accounts, period_start, period_end)

    ar_total   = 0.0
    ap_total   = 0.0
    cash_total = 0.0

    if ar_entities:
        ar_total = write_subsidiary_sheet(
            wb, 'AR by Customer', ar_entities,
            'Accounts Receivable Ledger — by Customer', 'Customer',
            '4472C4', period_start, period_end)

    if ap_entities:
        ap_total = write_subsidiary_sheet(
            wb, 'AP by Supplier', ap_entities,
            'Accounts Payable Ledger — by Supplier', 'Supplier',
            '4472C4', period_start, period_end)

    if cash_banks:
        cash_total = write_subsidiary_sheet(
            wb, 'Cash by Bank', cash_banks,
            'Cash Ledger — by Bank Account', 'Bank Account',
            '4472C4', period_start, period_end)

    write_fixed_assets_sheet(wb, assets)
    all_ok = write_control_account_sheet(
        wb, gl_accounts, ar_total, ap_total, cash_total,
        ar_entities, ap_entities, cash_banks)
    write_exceptions_sheet(wb, exceptions)

    # Move Dashboard to front
    wb.move_sheet('Dashboard', offset=-(len(wb.sheetnames) - 1))

    save_workbook(wb, output_file)

    # ─ Print summary ────────────────────────────────────────────────────────
    gl_ar   = gl_accounts.get(AR_GL_ACCOUNT,  {}).get('closing', 0)
    gl_ap   = gl_accounts.get(AP_GL_ACCOUNT,  {}).get('closing', 0)
    gl_cash = sum(gl_accounts.get(c, {}).get('closing', 0) for c in CASH_GL_ACCOUNTS)

    print(f"\nSaved to: {output_file}")
    print(f"Control Checks:")
    print(f"  AR  — GL: {gl_ar:,.0f}  |  Subsidiary: {ar_total:,.0f}  |  {'MATCH' if abs(gl_ar - ar_total) < 0.01 else 'MISMATCH'}")
    print(f"  AP  — GL: {gl_ap:,.0f}  |  Subsidiary: {ap_total:,.0f}  |  {'MATCH' if abs(gl_ap - ap_total) < 0.01 else 'MISMATCH'}")
    print(f"  Cash— GL: {gl_cash:,.0f}  |  Subsidiary: {cash_total:,.0f}  |  {'MATCH' if abs(gl_cash - cash_total) < 0.01 else 'MISMATCH'}")
    if exceptions:
        print(f"Warnings: {len(exceptions)} (see Exceptions sheet)")


if __name__ == '__main__':
    if len(sys.argv) < 5:
        print("Usage: python summarize_ledgers.py <input_dir> <period_start> <period_end> <output_file> [coa_file]")
        sys.exit(1)
    coa = sys.argv[5] if len(sys.argv) > 5 else None
    main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], coa)
