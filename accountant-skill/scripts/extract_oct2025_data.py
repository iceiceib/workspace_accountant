#!/usr/bin/env python
"""
Extract October 2025 data from existing files and create input files.
Uses detailed transaction entries from GL for proper journal entries.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import sys

# Paths
SOURCE_DIR = Path('Exisitng Accounting Workflow _ reference files')
OUTPUT_INPUT = Path('data/input')
OUTPUT_OUTPUT = Path('data/output/Oct2025')

# Create directories
(OUTPUT_INPUT / 'journals').mkdir(parents=True, exist_ok=True)
(OUTPUT_INPUT / 'ledgers').mkdir(parents=True, exist_ok=True)
(OUTPUT_INPUT / 'master').mkdir(parents=True, exist_ok=True)
OUTPUT_OUTPUT.mkdir(parents=True, exist_ok=True)

# Period
PERIOD_START = '2025-10-01'
PERIOD_END = '2025-10-31'

print("="*60)
print("EXTRACTING OCTOBER 2025 DATA")
print("="*60)

# =============================================================================
# 1. READ SOURCE FILES
# =============================================================================
print("\n1. Reading source files...")

# Read General Ledger
gl_df = pd.read_excel(SOURCE_DIR / 'Ledger Accounts' / 'General_Ledger_edited.xlsx', header=3)
gl_df = gl_df.dropna(how='all')
gl_df['Date'] = pd.to_datetime(gl_df['Date'], errors='coerce')
gl_df = gl_df[gl_df['Date'].notna()]
print(f"   General Ledger: {len(gl_df)} total rows")

# Filter for October 2025
oct_gl = gl_df[(gl_df['Date'] >= PERIOD_START) & (gl_df['Date'] <= PERIOD_END)].copy()
print(f"   October 2025 GL: {len(oct_gl)} rows")

# =============================================================================
# 2. CREATE CASH RECEIPTS JOURNAL FROM SALES REVENUE ENTRIES
# =============================================================================
print("\n2. Creating Cash Receipts Journal...")

# Sales Revenue entries represent individual cash sales
sales_df = oct_gl[oct_gl['COA Account Number'] == 40000].copy()
cash_receipts = []

for idx, row in sales_df.iterrows():
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
    amount = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0

    # Extract product info from description
    product = 'Drinking Water'
    if '140 ml' in desc:
        product = '140 ml Drinking Water'
    elif '175 ml' in desc:
        product = '175 ml Drinking Water'

    cash_receipts.append({
        'Date': row['Date'],
        'Receipt No': f'CR-{row["Date"].strftime("%m%d")}-{len(cash_receipts)+1:03d}',
        'Received From': 'Cash Sales',
        'Description': desc if desc else f'Sale of {product}',
        'Amount': amount,
        'Bank Account': 'Main',
        'Debit Account': 10100,  # Cash at Bank
        'Credit Account': 40000,  # Sales Revenue
    })

# Add other cash receipts (Interest, Capital)
other_receipts = oct_gl[(oct_gl['COA Account Number'] == 70000) |  # Interest Income
                         (oct_gl['COA Account Number'] == 31000)]  # Capital

for idx, row in other_receipts.iterrows():
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
    credit_acct = int(row['COA Account Number'])
    amount = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0

    cash_receipts.append({
        'Date': row['Date'],
        'Receipt No': f'CR-{row["Date"].strftime("%m%d")}-{len(cash_receipts)+1:03d}',
        'Received From': desc[:50] if desc else 'Other Receipt',
        'Description': desc,
        'Amount': amount,
        'Bank Account': 'Main',
        'Debit Account': 10100,
        'Credit Account': credit_acct,
    })

print(f"   Cash Receipts: {len(cash_receipts)} entries, Total: {sum(cr['Amount'] for cr in cash_receipts):,.0f} MMK")

# =============================================================================
# 3. CREATE CASH PAYMENTS JOURNAL
# =============================================================================
print("\n3. Creating Cash Payments Journal...")

cash_payments = []

# Purchases - Raw Materials
purchases_rm = oct_gl[oct_gl['COA Account Number'] == 50010]
for idx, row in purchases_rm.iterrows():
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else 'Purchase Raw Materials'
    amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
    cash_payments.append({
        'Date': row['Date'],
        'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
        'Paid To': desc[:50],
        'Description': desc,
        'Amount': amount,
        'Bank Account': 'Main',
        'Debit Account': 50010,
        'Credit Account': 10100,
    })

# Purchases - Packaging
purchases_pkg = oct_gl[oct_gl['COA Account Number'] == 50110]
for idx, row in purchases_pkg.iterrows():
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else 'Purchase Packaging'
    amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
    cash_payments.append({
        'Date': row['Date'],
        'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
        'Paid To': desc[:50],
        'Description': desc,
        'Amount': amount,
        'Bank Account': 'Main',
        'Debit Account': 50110,
        'Credit Account': 10100,
    })

# Operating Expenses (paid in cash)
expense_accounts = [53000, 53100, 53200, 65000, 14000]  # Labor, Maintenance, Utilities, Supplies, Deferred
for acct in expense_accounts:
    exp_df = oct_gl[oct_gl['COA Account Number'] == acct]
    for idx, row in exp_df.iterrows():
        desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
        amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
        if amount > 0:
            cash_payments.append({
                'Date': row['Date'],
                'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
                'Paid To': desc[:50] if desc else f'Payment for account {acct}',
                'Description': desc,
                'Amount': amount,
                'Bank Account': 'Main',
                'Debit Account': acct,
                'Credit Account': 10100,
            })

# Construction in Progress (capital expenditure)
cip_df = oct_gl[oct_gl['COA Account Number'] == 15500]
for idx, row in cip_df.iterrows():
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else 'Construction'
    amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
    if amount > 0:
        cash_payments.append({
            'Date': row['Date'],
            'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
            'Paid To': desc[:50],
            'Description': desc,
            'Amount': amount,
            'Bank Account': 'Main',
            'Debit Account': 15500,
            'Credit Account': 10100,
        })

# Machinery purchases
mach_df = oct_gl[oct_gl['COA Account Number'] == 15200]
for idx, row in mach_df.iterrows():
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else 'Machinery'
    amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
    if amount > 0:
        cash_payments.append({
            'Date': row['Date'],
            'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
            'Paid To': desc[:50],
            'Description': desc,
            'Amount': amount,
            'Bank Account': 'Main',
            'Debit Account': 15200,
            'Credit Account': 10100,
        })

# Advanced payments (13000)
adv_df = oct_gl[oct_gl['COA Account Number'] == 13000]
for idx, row in adv_df.iterrows():
    amount = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
    if amount > 0:
        desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else 'Advance Payment'
        cash_payments.append({
            'Date': row['Date'],
            'Payment No': f'CP-{row["Date"].strftime("%m%d")}-{len(cash_payments)+1:03d}',
            'Paid To': desc[:50],
            'Description': desc,
            'Amount': amount,
            'Bank Account': 'Main',
            'Debit Account': 13000,
            'Credit Account': 10100,
        })

print(f"   Cash Payments: {len(cash_payments)} entries, Total: {sum(cp['Amount'] for cp in cash_payments):,.0f} MMK")

# =============================================================================
# 4. CREATE GENERAL JOURNAL (Adjusting Entries)
# =============================================================================
print("\n4. Creating General Journal...")

general_journal = []

# Inventory adjustments (Opening/Closing)
inv_accounts = [50000, 50020, 50100, 50120, 50200, 50220, 12000, 12100, 12200]
inv_df = oct_gl[oct_gl['COA Account Number'].isin(inv_accounts)]

for idx, row in inv_df.iterrows():
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else ''
    debit = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
    credit = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0

    if debit > 0 or credit > 0:
        general_journal.append({
            'Date': row['Date'],
            'JV No': f'JV-10-{len(general_journal)+1:03d}',
            'Description': desc,
            'Debit Account': int(row['COA Account Number']),
            'Credit Account': int(row['COA Account Number']),
            'Debit Amount': debit,
            'Credit Amount': credit,
        })

# Depreciation entries
dep_accounts = [15110, 15210, 15410, 66000, 53300]
dep_df = oct_gl[oct_gl['COA Account Number'].isin(dep_accounts)]

for idx, row in dep_df.iterrows():
    desc = str(row['Descritpion']) if pd.notna(row['Descritpion']) else 'Depreciation'
    debit = float(row['Debit (MMK)']) if pd.notna(row['Debit (MMK)']) else 0.0
    credit = float(row['Credit (MMK)']) if pd.notna(row['Credit (MMK)']) else 0.0

    if debit > 0 or credit > 0:
        general_journal.append({
            'Date': row['Date'],
            'JV No': f'JV-10-{len(general_journal)+1:03d}',
            'Description': desc,
            'Debit Account': int(row['COA Account Number']),
            'Credit Account': int(row['COA Account Number']),
            'Debit Amount': debit,
            'Credit Amount': credit,
        })

print(f"   General Journal: {len(general_journal)} entries")

# =============================================================================
# 5. CREATE GENERAL LEDGER
# =============================================================================
print("\n5. Creating General Ledger...")

# Use the full GL for October
gl_output = oct_gl[['Date', 'COA Account Number', 'Account Name', 'Descritpion', 'Debit (MMK)', 'Credit (MMK)', 'Account Balance (MMK)']].copy()
gl_output.columns = ['Date', 'Account Code', 'Account Name', 'Description', 'Debit', 'Credit', 'Balance']
gl_output = gl_output.sort_values(['Account Code', 'Date'])

print(f"   General Ledger: {len(gl_output)} rows")

# =============================================================================
# 6. WRITE OUTPUT FILES
# =============================================================================
print("\n6. Writing output files...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def write_simple_excel(data, filepath, sheet_name='Sheet1'):
    """Write a simple Excel file from list of dicts or DataFrame."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if isinstance(data, list):
        if len(data) > 0:
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame()
    else:
        df = data

    # Write headers
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, (int, float)) and not pd.isna(value):
                cell.number_format = '#,##0.00'

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filepath)

# Write journals
write_simple_excel(cash_receipts, OUTPUT_INPUT / 'journals' / 'cash_receipts_journal.xlsx', 'Cash Receipts')
print(f"   Written: {OUTPUT_INPUT / 'journals' / 'cash_receipts_journal.xlsx'}")

write_simple_excel(cash_payments, OUTPUT_INPUT / 'journals' / 'cash_payments_journal.xlsx', 'Cash Payments')
print(f"   Written: {OUTPUT_INPUT / 'journals' / 'cash_payments_journal.xlsx'}")

write_simple_excel(general_journal, OUTPUT_INPUT / 'journals' / 'general_journal.xlsx', 'General Journal')
print(f"   Written: {OUTPUT_INPUT / 'journals' / 'general_journal.xlsx'}")

# Write ledger
write_simple_excel(gl_output, OUTPUT_INPUT / 'ledgers' / 'general_ledger.xlsx', 'General Ledger')
print(f"   Written: {OUTPUT_INPUT / 'ledgers' / 'general_ledger.xlsx'}")

# =============================================================================
# 7. SUMMARY
# =============================================================================
print("\n" + "="*60)
print("SUMMARY")
print("="*60)
print(f"Period: {PERIOD_START} to {PERIOD_END}")
print(f"\nJournals created:")
print(f"  - Cash Receipts: {len(cash_receipts)} entries, Total: {sum(cr['Amount'] for cr in cash_receipts):,.0f} MMK")
print(f"  - Cash Payments: {len(cash_payments)} entries, Total: {sum(cp['Amount'] for cp in cash_payments):,.0f} MMK")
print(f"  - General Journal: {len(general_journal)} entries")
print(f"\nLedger: {len(gl_output)} transactions")
print(f"\nFiles saved to: {OUTPUT_INPUT}")