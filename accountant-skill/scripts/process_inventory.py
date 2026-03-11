#!/usr/bin/env python
"""
Process Inventory Transactions

This script processes inventory movements from purchases and production usage,
calculates Weighted Average Cost (WAC), and generates inventory reports.

Usage:
    python scripts/process_inventory.py DATA_DIR PERIOD_START PERIOD_END

Example:
    python scripts/process_inventory.py data/Jan2026 2026-01-01 2026-01-31

Input files:
    - purchases_journal.xlsx (or raw_materials_ledger.xlsx with purchases)
    - production_usage.xlsx (manual input or calculated)

Output files:
    - raw_materials_ledger.xlsx (updated)
    - packaging_ledger.xlsx (updated)
    - inventory_summary_[PERIOD].xlsx
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
import math

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.excel_reader import read_xlsx, filter_by_period
from utils.excel_writer import (
    create_workbook, add_sheet, write_title, write_header_row,
    write_data_row, write_total_row, write_section_header,
    auto_fit_columns, freeze_panes, save_workbook,
    NUMBER_FORMAT_NEG, THIN_BORDER
)
from utils.inventory_mapper import (
    InventoryMapper, InventoryLedger, calculate_wac,
    _clean, _clean_numeric
)
from openpyxl.styles import Font, PatternFill, Alignment


# Tab colors
TAB_DASHBOARD = '00B050'  # Green
TAB_DETAIL = '4472C4'     # Blue


def read_inventory_ledger(filepath, item_code):
    """
    Read an item ledger from the inventory ledger file.

    Returns:
        Dict with opening_qty, opening_value, transactions list
    """
    result = {
        'opening_qty': 0,
        'opening_value': 0,
        'wac': 0,
        'transactions': []
    }

    if not Path(filepath).exists():
        return result

    try:
        xl = pd.ExcelFile(filepath)
        sheet_name = str(item_code)

        if sheet_name not in xl.sheet_names:
            return result

        df = pd.read_excel(xl, sheet_name=sheet_name, header=None)

        # Find the header row (contains 'Date')
        header_row = None
        for i, row in df.iterrows():
            if 'Date' in str(row[0]) or any('date' in str(v).lower() for v in row if pd.notna(v)):
                header_row = i
                break

        if header_row is None:
            return result

        # Get column mapping
        headers = df.iloc[header_row].tolist()
        col_map = {}
        for i, h in enumerate(headers):
            if pd.isna(h):
                continue
            h_lower = str(h).lower().strip()
            if 'date' in h_lower:
                col_map['date'] = i
            elif 'reference' in h_lower:
                col_map['reference'] = i
            elif 'description' in h_lower:
                col_map['description'] = i
            elif 'received qty' in h_lower or 'received_qty' in h_lower:
                col_map['received_qty'] = i
            elif 'issued qty' in h_lower or 'issued_qty' in h_lower:
                col_map['issued_qty'] = i
            elif 'balance qty' in h_lower:
                col_map['balance_qty'] = i
            elif 'unit cost' in h_lower:
                col_map['unit_cost'] = i
            elif 'received value' in h_lower:
                col_map['received_value'] = i
            elif 'issued value' in h_lower:
                col_map['issued_value'] = i
            elif 'balance value' in h_lower:
                col_map['balance_value'] = i

        # Read opening balance (row before header)
        if header_row > 0:
            for i in range(header_row - 1, -1, -1):
                row = df.iloc[i]
                if pd.notna(row[2]) and 'opening' in str(row[2]).lower():
                    if 'balance_qty' in col_map:
                        result['opening_qty'] = _clean_numeric(row[col_map['balance_qty']])
                    if 'balance_value' in col_map:
                        result['opening_value'] = _clean_numeric(row[col_map['balance_value']])
                    break

        # Read transactions
        for i in range(header_row + 2, len(df)):  # Skip opening row
            row = df.iloc[i]
            if all(pd.isna(v) for v in row if isinstance(v, (int, float)) or pd.notna(v)):
                continue

            txn = {}
            for key, col_idx in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if key in ['received_qty', 'issued_qty', 'balance_qty', 'unit_cost',
                           'received_value', 'issued_value', 'balance_value']:
                    txn[key] = _clean_numeric(val)
                else:
                    txn[key] = _clean(val)

            if txn.get('date') or txn.get('received_qty') or txn.get('issued_qty'):
                result['transactions'].append(txn)

    except Exception as e:
        print(f"Warning: Error reading ledger {filepath} sheet {item_code}: {e}")

    return result


def read_purchases_for_inventory(data_dir, period_start, period_end, inventory_mapper):
    """
    Read purchases journal to extract inventory purchases.

    Returns:
        Dict of item_code -> list of purchase transactions
    """
    purchases_file = Path(data_dir) / 'purchases_journal.xlsx'
    if not purchases_file.exists():
        print(f"Warning: Purchases journal not found: {purchases_file}")
        return {}

    result = read_xlsx(
        purchases_file,
        required_columns=['Date'],
        optional_columns=['Reference', 'Supplier', 'Description', 'Debit Account',
                         'Credit Account', 'Amount', 'Quantity', 'Unit Cost', 'Item Code']
    )

    if result.get('error'):
        print(f"Error reading purchases journal: {result['error']}")
        return {}

    df = result['data']

    # Filter by period
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        start = pd.to_datetime(period_start)
        end = pd.to_datetime(period_end)
        df = df[(df['Date'] >= start) & (df['Date'] <= end)]

    # Group by item code
    purchases_by_item = {}

    for _, row in df.iterrows():
        # Try to identify item code from debit account or item code column
        item_code = None

        if 'Item Code' in df.columns and pd.notna(row.get('Item Code')):
            try:
                item_code = int(float(row['Item Code']))
            except (ValueError, TypeError):
                pass

        if item_code is None and 'Debit Account' in df.columns and pd.notna(row.get('Debit Account')):
            debit_acct = row['Debit Account']
            try:
                debit_acct = int(float(debit_acct))
                # Check if this is an inventory account
                if 12000 <= debit_acct <= 12999:
                    # This is an inventory purchase, but we need item code
                    # For now, we'll need manual mapping or assume the description has it
                    pass
            except (ValueError, TypeError):
                pass

        if item_code is None:
            continue

        if item_code not in purchases_by_item:
            purchases_by_item[item_code] = []

        purchases_by_item[item_code].append({
            'date': row.get('Date'),
            'reference': _clean(row.get('Reference', '')),
            'description': _clean(row.get('Description', '')),
            'supplier': _clean(row.get('Supplier', '')),
            'quantity': _clean_numeric(row.get('Quantity', 1)),
            'unit_cost': _clean_numeric(row.get('Unit Cost', row.get('Amount', 0))),
            'value': _clean_numeric(row.get('Amount', 0))
        })

    return purchases_by_item


def read_production_usage(data_dir, period_start, period_end):
    """
    Read production usage file (if exists).

    This file should contain:
    - Date
    - Item Code
    - Quantity Used
    - Production Batch Reference

    Returns:
        Dict of item_code -> list of usage transactions
    """
    usage_file = Path(data_dir) / 'production_usage.xlsx'
    if not usage_file.exists():
        print(f"Note: Production usage file not found: {usage_file}")
        print("      Inventory issued to production will need manual entry.")
        return {}

    result = read_xlsx(
        usage_file,
        required_columns=['Date', 'Item Code'],
        optional_columns=['Reference', 'Description', 'Quantity Used', 'Batch']
    )

    if result.get('error'):
        print(f"Error reading production usage: {result['error']}")
        return {}

    df = result['data']

    # Filter by period
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        start = pd.to_datetime(period_start)
        end = pd.to_datetime(period_end)
        df = df[(df['Date'] >= start) & (df['Date'] <= end)]

    usage_by_item = {}

    for _, row in df.iterrows():
        try:
            item_code = int(float(row['Item Code']))
        except (ValueError, TypeError):
            continue

        if item_code not in usage_by_item:
            usage_by_item[item_code] = []

        qty_col = 'Quantity Used' if 'Quantity Used' in df.columns else 'Quantity'
        usage_by_item[item_code].append({
            'date': row.get('Date'),
            'reference': _clean(row.get('Reference', row.get('Batch', ''))),
            'description': _clean(row.get('Description', 'Issued to production')),
            'quantity': _clean_numeric(row.get(qty_col, 0))
        })

    return usage_by_item


def process_inventory_item(ledger, purchases, usage, item_code, item_name, unit):
    """
    Process all transactions for a single inventory item.

    Returns:
        InventoryLedger object with all transactions applied
    """
    # Apply purchases
    for p in purchases:
        try:
            ledger.receive(
                date=p['date'],
                reference=p['reference'],
                qty=p['quantity'],
                unit_cost=p['unit_cost'],
                description=p.get('description', p.get('supplier', 'Purchase'))
            )
        except Exception as e:
            print(f"Error processing purchase for {item_code}: {e}")

    # Apply usage
    for u in usage:
        try:
            ledger.issue(
                date=u['date'],
                reference=u['reference'],
                qty=u['quantity'],
                description=u['description']
            )
        except ValueError as e:
            print(f"Warning: {e} for item {item_code} ({item_name})")
        except Exception as e:
            print(f"Error processing usage for {item_code}: {e}")

    return ledger


def update_inventory_ledger_file(filepath, items_data):
    """
    Update an inventory ledger file with processed data.

    Args:
        filepath: Path to the ledger xlsx file
        items_data: List of dicts with item summary and transactions
    """
    from openpyxl import load_workbook

    if not Path(filepath).exists():
        print(f"Warning: Ledger file not found: {filepath}")
        return

    try:
        wb = load_workbook(filepath)

        for item in items_data:
            sheet_name = str(item.get('item_code', item.get('code', '')))
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # Find the data start row (after header)
            data_row = None
            for row in range(1, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=1).value
                if cell_val and 'date' in str(cell_val).lower():
                    data_row = row + 2  # Skip header and opening row
                    break

            if data_row is None:
                continue

            # Write transactions
            for txn in item.get('transactions', []):
                ws.cell(row=data_row, column=1, value=txn.get('date', ''))
                ws.cell(row=data_row, column=2, value=txn.get('reference', ''))
                ws.cell(row=data_row, column=3, value=txn.get('description', ''))
                ws.cell(row=data_row, column=4, value=txn.get('received_qty', ''))
                ws.cell(row=data_row, column=5, value=txn.get('issued_qty', ''))
                ws.cell(row=data_row, column=6, value=txn.get('balance_qty', ''))
                ws.cell(row=data_row, column=7, value=txn.get('unit_cost', ''))
                ws.cell(row=data_row, column=8, value=txn.get('received_value', ''))
                ws.cell(row=data_row, column=9, value=txn.get('issued_value', ''))
                ws.cell(row=data_row, column=10, value=txn.get('balance_value', ''))
                data_row += 1

        # Update Dashboard
        if 'Dashboard' in wb.sheetnames:
            ws = wb['Dashboard']

            # Find the data start row
            data_row = None
            for row in range(1, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=1).value
                if cell_val and 'item code' in str(cell_val).lower():
                    data_row = row + 1
                    break

            if data_row:
                for item in items_data:
                    ws.cell(row=data_row, column=4, value=item.get('opening_qty', 0))
                    ws.cell(row=data_row, column=5, value=item.get('opening_value', 0))
                    ws.cell(row=data_row, column=6, value=item.get('received_qty', 0))
                    ws.cell(row=data_row, column=7, value=item.get('received_value', 0))
                    ws.cell(row=data_row, column=8, value=item.get('issued_qty', 0))
                    ws.cell(row=data_row, column=9, value=item.get('issued_value', 0))
                    ws.cell(row=data_row, column=10, value=item.get('closing_qty', 0))
                    ws.cell(row=data_row, column=11, value=item.get('closing_value', 0))
                    ws.cell(row=data_row, column=12, value=item.get('wac', 0))
                    data_row += 1

        wb.save(filepath)
        print(f"Updated: {filepath}")

    except Exception as e:
        print(f"Error updating ledger {filepath}: {e}")


def create_inventory_summary(output_path, period_start, period_end, rm_items, pkg_items):
    """
    Create an inventory summary report.

    Args:
        output_path: Path to save the summary file
        period_start: Period start date
        period_end: Period end date
        rm_items: List of raw materials summary dicts
        pkg_items: List of packaging summary dicts
    """
    wb = create_workbook()
    ws = add_sheet(wb, 'Summary', TAB_DASHBOARD)

    period_str = f"For the period {period_start} to {period_end}"
    row = write_title(ws, "Inventory Summary Report", period=period_str)

    # Raw Materials Section
    row = write_section_header(ws, "RAW MATERIALS", row)

    headers = ['Item Code', 'Item Name', 'Unit', 'Opening Qty', 'Opening Value',
               'Purchases Qty', 'Purchases Value', 'Issued Qty', 'Issued Value',
               'Closing Qty', 'Closing Value', 'WAC']
    row = write_header_row(ws, headers, row)

    total_rm_opening = 0
    total_rm_purchases = 0
    total_rm_issued = 0
    total_rm_closing = 0

    for item in rm_items:
        values = [
            item.get('item_code', ''), item.get('item_name', ''), item.get('unit', ''),
            item.get('opening_qty', 0), item.get('opening_value', 0),
            item.get('received_qty', 0), item.get('received_value', 0),
            item.get('issued_qty', 0), item.get('issued_value', 0),
            item.get('closing_qty', 0), item.get('closing_value', 0),
            item.get('wac', 0)
        ]
        row = write_data_row(ws, values, row, number_cols=[3, 4, 5, 6, 7, 8, 9, 10, 11])

        total_rm_opening += item.get('opening_value', 0)
        total_rm_purchases += item.get('received_value', 0)
        total_rm_issued += item.get('issued_value', 0)
        total_rm_closing += item.get('closing_value', 0)

    row = write_total_row(ws, 'Total Raw Materials',
                          ['', '', '', total_rm_opening, '', total_rm_purchases,
                           '', total_rm_issued, '', total_rm_closing, ''], row)

    # Packaging Section
    row += 1
    row = write_section_header(ws, "PACKAGING MATERIALS", row)
    row = write_header_row(ws, headers, row)

    total_pkg_opening = 0
    total_pkg_purchases = 0
    total_pkg_issued = 0
    total_pkg_closing = 0

    for item in pkg_items:
        values = [
            item.get('item_code', ''), item.get('item_name', ''), item.get('unit', ''),
            item.get('opening_qty', 0), item.get('opening_value', 0),
            item.get('received_qty', 0), item.get('received_value', 0),
            item.get('issued_qty', 0), item.get('issued_value', 0),
            item.get('closing_qty', 0), item.get('closing_value', 0),
            item.get('wac', 0)
        ]
        row = write_data_row(ws, values, row, number_cols=[3, 4, 5, 6, 7, 8, 9, 10, 11])

        total_pkg_opening += item.get('opening_value', 0)
        total_pkg_purchases += item.get('received_value', 0)
        total_pkg_issued += item.get('issued_value', 0)
        total_pkg_closing += item.get('closing_value', 0)

    row = write_total_row(ws, 'Total Packaging',
                          ['', '', '', total_pkg_opening, '', total_pkg_purchases,
                           '', total_pkg_issued, '', total_pkg_closing, ''], row)

    # Grand Total
    row += 1
    grand_opening = total_rm_opening + total_pkg_opening
    grand_purchases = total_rm_purchases + total_pkg_purchases
    grand_issued = total_rm_issued + total_pkg_issued
    grand_closing = total_rm_closing + total_pkg_closing

    row = write_total_row(ws, 'GRAND TOTAL',
                          ['', '', '', grand_opening, '', grand_purchases,
                           '', grand_issued, '', grand_closing, ''], row, double_line=True)

    # Cost of Production Summary
    row += 2
    row = write_section_header(ws, "COST OF PRODUCTION", row)

    ws.cell(row=row, column=1, value="Raw Materials Used:")
    ws.cell(row=row, column=3, value=total_rm_issued)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value="Packaging Used:")
    ws.cell(row=row, column=3, value=total_pkg_issued)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value="Total Cost of Materials:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=grand_issued)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)

    auto_fit_columns(ws)
    freeze_panes(ws, row=5, col=1)

    save_workbook(wb, output_path)
    print(f"Created: {output_path}")


def main():
    """Main entry point."""
    if len(sys.argv) < 4:
        print(__doc__)
        print("\nUsage: python scripts/process_inventory.py DATA_DIR PERIOD_START PERIOD_END")
        print("\nExample:")
        print("  python scripts/process_inventory.py data/Jan2026 2026-01-01 2026-01-31")
        sys.exit(1)

    data_dir = Path(sys.argv[1])
    period_start = sys.argv[2]
    period_end = sys.argv[3]

    print(f"Processing inventory for period {period_start} to {period_end}")
    print(f"Data directory: {data_dir}")
    print()

    # Initialize inventory mapper
    items_file = data_dir / 'inventory_items.xlsx'
    inv_mapper = InventoryMapper(items_file if items_file.exists() else None)

    # Read purchases
    print("Reading purchases...")
    purchases = read_purchases_for_inventory(data_dir, period_start, period_end, inv_mapper)
    if purchases:
        print(f"  Found purchases for {len(purchases)} items")
    else:
        print("  No purchases found (purchases will need manual entry)")

    # Read production usage
    print("Reading production usage...")
    usage = read_production_usage(data_dir, period_start, period_end)
    if usage:
        print(f"  Found usage records for {len(usage)} items")
    else:
        print("  No production usage found (usage will need manual entry)")

    # Process raw materials
    print("\nProcessing raw materials...")
    rm_items = []
    rm_ledger_file = data_dir / 'raw_materials_ledger.xlsx'

    for item in inv_mapper.get_raw_materials():
        code = item['code']

        # Read existing ledger data
        ledger_data = read_inventory_ledger(rm_ledger_file, code)

        # Create ledger object
        ledger = InventoryLedger(
            item_code=code,
            item_name=item['name'],
            unit=item['unit'],
            opening_qty=ledger_data['opening_qty'],
            opening_value=ledger_data['opening_value']
        )

        # Process transactions
        item_purchases = purchases.get(code, [])
        item_usage = usage.get(code, [])
        ledger = process_inventory_item(ledger, item_purchases, item_usage, code, item['name'], item['unit'])

        # Get summary
        summary = ledger.get_period_summary()
        summary['transactions'] = [t for t in ledger.transactions]
        rm_items.append(summary)

    print(f"  Processed {len(rm_items)} raw material items")

    # Process packaging
    print("\nProcessing packaging materials...")
    pkg_items = []
    pkg_ledger_file = data_dir / 'packaging_ledger.xlsx'

    for item in inv_mapper.get_packaging():
        code = item['code']

        # Read existing ledger data
        ledger_data = read_inventory_ledger(pkg_ledger_file, code)

        # Create ledger object
        ledger = InventoryLedger(
            item_code=code,
            item_name=item['name'],
            unit=item['unit'],
            opening_qty=ledger_data['opening_qty'],
            opening_value=ledger_data['opening_value']
        )

        # Process transactions
        item_purchases = purchases.get(code, [])
        item_usage = usage.get(code, [])
        ledger = process_inventory_item(ledger, item_purchases, item_usage, code, item['name'], item['unit'])

        # Get summary
        summary = ledger.get_period_summary()
        summary['transactions'] = [t for t in ledger.transactions]
        pkg_items.append(summary)

    print(f"  Processed {len(pkg_items)} packaging items")

    # Update ledger files
    print("\nUpdating ledger files...")
    if rm_ledger_file.exists():
        update_inventory_ledger_file(rm_ledger_file, rm_items)
    if pkg_ledger_file.exists():
        update_inventory_ledger_file(pkg_ledger_file, pkg_items)

    # Create inventory summary
    period_name = period_end[:7].replace('-', '')  # e.g., "202601"
    summary_file = data_dir / f'inventory_summary_{period_name}.xlsx'
    create_inventory_summary(summary_file, period_start, period_end, rm_items, pkg_items)

    print()
    print("Inventory processing complete!")


if __name__ == '__main__':
    main()