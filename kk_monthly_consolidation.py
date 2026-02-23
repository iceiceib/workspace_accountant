"""
KK Business Monthly Sales Consolidation Script
Combines Soft Drink and Purified Water monthly sales into one summary report.

Usage:
    python kk_monthly_consolidation.py --month Dec2025
    python kk_monthly_consolidation.py --month Jan2026
    python kk_monthly_consolidation.py --softdrink path/to/file1.xlsx --water path/to/file2.xlsx

Input files expected:
    KK_SoftDrink_{month}.xlsx   (e.g., KK_SoftDrink_Dec2025.xlsx)
    KK_Water_{month}.xlsx       (e.g., KK_Water_Dec2025.xlsx)

Output:
    KK_Monthly_Summary_{month}.xlsx
"""

import pandas as pd
import argparse
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime


def read_sales_file(filepath, business_type):
    """Read a monthly sales Excel file and add business type column."""
    if not os.path.exists(filepath):
        print(f"WARNING: File not found: {filepath}")
        return None

    df = pd.read_excel(filepath)

    # Standardize column names (strip whitespace)
    df.columns = df.columns.str.strip()

    # Add business type identifier
    df['Business'] = business_type

    # Parse date column
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], format='mixed', dayfirst=True)

    # Ensure numeric columns
    for col in ['Volume Sold', 'Sale Price', 'Total Sale', 'Cash', 'Bank', 'Discount/FOC']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    print(f"  Loaded {len(df)} rows from {os.path.basename(filepath)} ({business_type})")
    return df


def create_summary_report(combined_df, output_path, month_label):
    """Create a professional Excel summary report with multiple sheets."""
    wb = Workbook()

    # --- Style definitions ---
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2E7D32')  # Dark green
    subheader_fill = PatternFill('solid', fgColor='66BB6A')  # Light green
    subheader_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    title_font = Font(name='Arial', bold=True, size=14, color='1B5E20')
    subtitle_font = Font(name='Arial', bold=True, size=11, color='424242')
    data_font = Font(name='Arial', size=10)
    total_font = Font(name='Arial', bold=True, size=11)
    total_fill = PatternFill('solid', fgColor='E8F5E9')
    mmk_format = '#,##0'
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD')
    )

    def style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    def style_data_cell(ws, row, col, is_currency=False):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.border = thin_border
        if is_currency:
            cell.number_format = mmk_format
            cell.alignment = Alignment(horizontal='right')
        return cell

    def auto_width(ws, min_width=10, max_width=25):
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col_cells), default=0)
            width = min(max(max_len + 3, min_width), max_width)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width

    # =============================================
    # SHEET 1: OVERVIEW (Grand Summary)
    # =============================================
    ws1 = wb.active
    ws1.title = 'Overview'

    # Title
    ws1.merge_cells('A1:F1')
    ws1['A1'] = f'KK Business - Monthly Sales Summary'
    ws1['A1'].font = title_font
    ws1.merge_cells('A2:F2')
    ws1['A2'] = f'Period: {month_label}'
    ws1['A2'].font = subtitle_font
    ws1.merge_cells('A3:F3')
    ws1['A3'] = f'Generated: {datetime.now().strftime("%d-%b-%Y %H:%M")}'
    ws1['A3'].font = Font(name='Arial', size=9, color='757575')

    # Grand totals by business
    row = 5
    headers = ['Business', 'Total Volume', 'Total Sales (MMK)', 'Cash (MMK)', 'Bank (MMK)', 'Discount/FOC (MMK)']
    for c, h in enumerate(headers, 1):
        ws1.cell(row=row, column=c, value=h)
    style_header_row(ws1, row, len(headers))

    biz_summary = combined_df.groupby('Business').agg({
        'Volume Sold': 'sum',
        'Total Sale': 'sum',
        'Cash': 'sum',
        'Bank': 'sum',
        'Discount/FOC': 'sum'
    }).reset_index()

    for i, biz_row in biz_summary.iterrows():
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=biz_row['Business']).font = data_font
        ws1.cell(row=r, column=1).border = thin_border
        style_data_cell(ws1, r, 2, True).value = biz_row['Volume Sold']
        style_data_cell(ws1, r, 3, True).value = biz_row['Total Sale']
        style_data_cell(ws1, r, 4, True).value = biz_row['Cash']
        style_data_cell(ws1, r, 5, True).value = biz_row['Bank']
        style_data_cell(ws1, r, 6, True).value = biz_row['Discount/FOC']

    # Grand total row with formulas
    total_row = row + 1 + len(biz_summary)
    ws1.cell(row=total_row, column=1, value='GRAND TOTAL').font = total_font
    ws1.cell(row=total_row, column=1).fill = total_fill
    ws1.cell(row=total_row, column=1).border = thin_border
    for c in range(2, 7):
        col_letter = get_column_letter(c)
        cell = ws1.cell(row=total_row, column=c)
        cell.value = f'=SUM({col_letter}{row+1}:{col_letter}{total_row-1})'
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = mmk_format
        cell.border = thin_border

    auto_width(ws1)

    # =============================================
    # SHEET 2: SALES BY SKU/PRODUCT
    # =============================================
    ws2 = wb.create_sheet('By Product')

    ws2.merge_cells('A1:G1')
    ws2['A1'] = f'Sales by Product - {month_label}'
    ws2['A1'].font = title_font

    row = 3
    headers = ['Business', 'SKU Code', 'Total Volume', 'Avg Sale Price', 'Total Sales (MMK)', 'Cash (MMK)', 'Bank (MMK)']
    for c, h in enumerate(headers, 1):
        ws2.cell(row=row, column=c, value=h)
    style_header_row(ws2, row, len(headers))

    sku_summary = combined_df.groupby(['Business', 'SKU Code']).agg({
        'Volume Sold': 'sum',
        'Sale Price': 'mean',
        'Total Sale': 'sum',
        'Cash': 'sum',
        'Bank': 'sum'
    }).reset_index().sort_values(['Business', 'Total Sale'], ascending=[True, False])

    r = row + 1
    data_start = r
    for _, sku_row in sku_summary.iterrows():
        ws2.cell(row=r, column=1, value=sku_row['Business']).font = data_font
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=2, value=sku_row['SKU Code']).font = data_font
        ws2.cell(row=r, column=2).border = thin_border
        style_data_cell(ws2, r, 3, True).value = sku_row['Volume Sold']
        style_data_cell(ws2, r, 4, True).value = round(sku_row['Sale Price'], 2)
        style_data_cell(ws2, r, 5, True).value = sku_row['Total Sale']
        style_data_cell(ws2, r, 6, True).value = sku_row['Cash']
        style_data_cell(ws2, r, 7, True).value = sku_row['Bank']
        r += 1

    # Total row with formulas
    ws2.cell(row=r, column=1, value='TOTAL').font = total_font
    ws2.cell(row=r, column=1).fill = total_fill
    ws2.cell(row=r, column=1).border = thin_border
    ws2.cell(row=r, column=2).fill = total_fill
    ws2.cell(row=r, column=2).border = thin_border
    for c in [3, 5, 6, 7]:
        col_letter = get_column_letter(c)
        cell = ws2.cell(row=r, column=c)
        cell.value = f'=SUM({col_letter}{data_start}:{col_letter}{r-1})'
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = mmk_format
        cell.border = thin_border
    cell = ws2.cell(row=r, column=4)
    cell.fill = total_fill
    cell.border = thin_border

    auto_width(ws2)

    # =============================================
    # SHEET 3: SALES BY CHANNEL
    # =============================================
    ws3 = wb.create_sheet('By Channel')

    ws3.merge_cells('A1:F1')
    ws3['A1'] = f'Sales by Channel - {month_label}'
    ws3['A1'].font = title_font

    row = 3
    headers = ['Business', 'Sales Channel', 'Total Volume', 'Total Sales (MMK)', 'Cash (MMK)', 'Bank (MMK)']
    for c, h in enumerate(headers, 1):
        ws3.cell(row=row, column=c, value=h)
    style_header_row(ws3, row, len(headers))

    channel_summary = combined_df.groupby(['Business', 'Sales Channel']).agg({
        'Volume Sold': 'sum',
        'Total Sale': 'sum',
        'Cash': 'sum',
        'Bank': 'sum'
    }).reset_index().sort_values(['Business', 'Total Sale'], ascending=[True, False])

    r = row + 1
    data_start = r
    for _, ch_row in channel_summary.iterrows():
        ws3.cell(row=r, column=1, value=ch_row['Business']).font = data_font
        ws3.cell(row=r, column=1).border = thin_border
        ws3.cell(row=r, column=2, value=ch_row['Sales Channel']).font = data_font
        ws3.cell(row=r, column=2).border = thin_border
        style_data_cell(ws3, r, 3, True).value = ch_row['Volume Sold']
        style_data_cell(ws3, r, 4, True).value = ch_row['Total Sale']
        style_data_cell(ws3, r, 5, True).value = ch_row['Cash']
        style_data_cell(ws3, r, 6, True).value = ch_row['Bank']
        r += 1

    ws3.cell(row=r, column=1, value='TOTAL').font = total_font
    ws3.cell(row=r, column=1).fill = total_fill
    ws3.cell(row=r, column=1).border = thin_border
    ws3.cell(row=r, column=2).fill = total_fill
    ws3.cell(row=r, column=2).border = thin_border
    for c in [3, 4, 5, 6]:
        col_letter = get_column_letter(c)
        cell = ws3.cell(row=r, column=c)
        cell.value = f'=SUM({col_letter}{data_start}:{col_letter}{r-1})'
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = mmk_format
        cell.border = thin_border

    auto_width(ws3)

    # =============================================
    # SHEET 4: DAILY SALES BREAKDOWN
    # =============================================
    ws4 = wb.create_sheet('By Date')

    ws4.merge_cells('A1:F1')
    ws4['A1'] = f'Daily Sales Breakdown - {month_label}'
    ws4['A1'].font = title_font

    row = 3
    headers = ['Date', 'Business', 'Total Volume', 'Total Sales (MMK)', 'Cash (MMK)', 'Bank (MMK)']
    for c, h in enumerate(headers, 1):
        ws4.cell(row=row, column=c, value=h)
    style_header_row(ws4, row, len(headers))

    daily_summary = combined_df.groupby(['Date', 'Business']).agg({
        'Volume Sold': 'sum',
        'Total Sale': 'sum',
        'Cash': 'sum',
        'Bank': 'sum'
    }).reset_index().sort_values(['Date', 'Business'])

    r = row + 1
    data_start = r
    for _, day_row in daily_summary.iterrows():
        date_val = day_row['Date']
        if isinstance(date_val, pd.Timestamp):
            date_str = date_val.strftime('%d-%b-%Y')
        else:
            date_str = str(date_val)
        ws4.cell(row=r, column=1, value=date_str).font = data_font
        ws4.cell(row=r, column=1).border = thin_border
        ws4.cell(row=r, column=2, value=day_row['Business']).font = data_font
        ws4.cell(row=r, column=2).border = thin_border
        style_data_cell(ws4, r, 3, True).value = day_row['Volume Sold']
        style_data_cell(ws4, r, 4, True).value = day_row['Total Sale']
        style_data_cell(ws4, r, 5, True).value = day_row['Cash']
        style_data_cell(ws4, r, 6, True).value = day_row['Bank']
        r += 1

    ws4.cell(row=r, column=1, value='TOTAL').font = total_font
    ws4.cell(row=r, column=1).fill = total_fill
    ws4.cell(row=r, column=1).border = thin_border
    ws4.cell(row=r, column=2).fill = total_fill
    ws4.cell(row=r, column=2).border = thin_border
    for c in [3, 4, 5, 6]:
        col_letter = get_column_letter(c)
        cell = ws4.cell(row=r, column=c)
        cell.value = f'=SUM({col_letter}{data_start}:{col_letter}{r-1})'
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = mmk_format
        cell.border = thin_border

    auto_width(ws4)

    # =============================================
    # SHEET 5: RAW DATA (combined)
    # =============================================
    ws5 = wb.create_sheet('Raw Data')
    ws5['A1'] = f'Combined Raw Data - {month_label}'
    ws5['A1'].font = title_font

    raw_cols = ['Business', 'No', 'Month', 'Date', 'Account Name', 'SKU Code',
                'Sales Channel', 'Volume Sold', 'Sale Price', 'Total Sale',
                'Cash', 'Bank', 'Discount/FOC', 'Balance Check',
                'Transaction Type', 'Payment Type']
    available_cols = [c for c in raw_cols if c in combined_df.columns]

    for c, h in enumerate(available_cols, 1):
        ws5.cell(row=3, column=c, value=h)
    style_header_row(ws5, 3, len(available_cols))

    for i, (_, data_row) in enumerate(combined_df.iterrows()):
        for c, col_name in enumerate(available_cols, 1):
            val = data_row[col_name]
            if isinstance(val, pd.Timestamp):
                val = val.strftime('%d-%b-%Y')
            cell = ws5.cell(row=4 + i, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col_name in ['Volume Sold', 'Sale Price', 'Total Sale', 'Cash', 'Bank', 'Discount/FOC', 'Balance Check']:
                cell.number_format = mmk_format

    auto_width(ws5)

    wb.save(output_path)
    print(f"\n  Report saved: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description='KK Business Monthly Sales Consolidation')
    parser.add_argument('--month', type=str, help='Month label (e.g., Dec2025, Jan2026)')
    parser.add_argument('--softdrink', type=str, help='Path to soft drink sales file')
    parser.add_argument('--water', type=str, help='Path to water sales file')
    parser.add_argument('--input-dir', type=str, default='.', help='Directory containing input files (default: current dir)')
    parser.add_argument('--output-dir', type=str, default='.', help='Directory for output file (default: current dir)')
    args = parser.parse_args()

    print("=" * 60)
    print("  KK Business - Monthly Sales Consolidation")
    print("=" * 60)

    if args.softdrink and args.water:
        softdrink_path = args.softdrink
        water_path = args.water
        month_label = args.month or 'Unknown'
    elif args.month:
        softdrink_path = os.path.join(args.input_dir, f'KK_SoftDrink_{args.month}.xlsx')
        water_path = os.path.join(args.input_dir, f'KK_Water_{args.month}.xlsx')
        month_label = args.month
    else:
        print("\nERROR: Please provide either --month or both --softdrink and --water")
        print("\nUsage examples:")
        print("  python kk_monthly_consolidation.py --month Dec2025")
        print("  python kk_monthly_consolidation.py --softdrink data/KK_SoftDrink_Dec2025.xlsx --water data/KK_Water_Dec2025.xlsx")
        sys.exit(1)

    print(f"\n  Month: {month_label}")
    print(f"  Soft Drink file: {softdrink_path}")
    print(f"  Water file: {water_path}")
    print("-" * 60)

    print("\nReading input files...")
    dfs = []

    sd_df = read_sales_file(softdrink_path, 'Soft Drink')
    if sd_df is not None:
        dfs.append(sd_df)

    water_df = read_sales_file(water_path, 'Purified Water')
    if water_df is not None:
        dfs.append(water_df)

    if not dfs:
        print("\nERROR: No valid input files found. Exiting.")
        sys.exit(1)

    combined = pd.concat(dfs, ignore_index=True)
    print(f"\n  Combined: {len(combined)} total rows")
    print(f"  Total Sales: {combined['Total Sale'].sum():,.0f} MMK")

    output_filename = f'KK_Monthly_Summary_{month_label}.xlsx'
    output_path = os.path.join(args.output_dir, output_filename)

    print("\nGenerating summary report...")
    create_summary_report(combined, output_path, month_label)

    print("\n" + "=" * 60)
    print("  CONSOLIDATION COMPLETE!")
    print(f"  Output: {output_path}")
    print("=" * 60)


if __name__ == '__main__':
    main()
